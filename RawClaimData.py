
import pandas as pd
import numpy as np
import os
import io
import msoffcrypto
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
import datetime as dt
from datetime import timedelta
import warnings
import itertools
warnings.filterwarnings('ignore')
sns.set(rc={'figure.figsize':(5,5)})
plt.rcParams["axes.formatter.limits"] = (-99, 99)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle, Protection

from ColumnNameManagement import *

class RawClaimData():

  def __init__(self, benefit_index_path):
    col_setup = [
        'policy_id', # This is the policy_id for future database development, f'{policy_number}__{policy_start_date:%Y%m}'
        'policy_number',
        'insurer',
        'client_name',
        'policy_start_date',
        'policy_end_date',
        'policy_data_date',
        'claim_id',
        'incur_date',
        'discharge_date',
        'submission_date',
        'pay_date',
        'claim_status',
        'claim_remark_1',
        'claim_remark_2',
        'claim_remark_3',
        'claim_remark_4',
        'claim_remark_5',
        'cert_true_copy',
        'claimant',
        'gender',
        'age',
        'dep_type',
        'class',
        'member_status',
        'benefit_type',
        'benefit',
        'diagnosis_code',
        'diagnosis',
        'diagnosis_organ',
        'speciality',
        'procedure_code',
        'procedure',
        'room_type',
        'hospital_code',
        'hospital_name',
        'hospital_type',
        'provider_code',
        'provider',
        'physician_code',
        'physician',
        'chronic',
        'opd',
        'currency',
        'incurred_amount',
        'paid_amount',
        'panel',
        'payment_method',
        'suboffice',
        'deptment',
        'region',
        'common_diagnosis_flag',
    ]

    self.col_setup = col_setup
    self.df = pd.DataFrame(columns=self.col_setup)
    self.upload_time = dt.datetime.now()
    self.upload_log = pd.DataFrame(columns=['policy_id', 'upload_time', 'insurer', 'shortfall_supplement'])
    self.benefit_index = pd.read_excel('benefit_indexing.xlsx')
    self.speciality_index = pd.read_excel('speciality_indexing.xlsx')
    self.minor_surg_index = pd.read_excel('minor_surgery_indexing.xlsx')
    self.ibnr = pd.read_csv('ibnr_ref.csv', dtype={'data_month': int, 'ibnr': float})
    self.provider_grouping_path = 'provider_mapping.csv'

  def date_conversion(self, date_series):
        try:
            date_series = pd.to_datetime(date_series, format="%d/%m/%Y")
        except:
            try:
                date_series = pd.to_datetime(date_series, format="%Y-%m-%d")
            except:
                try:
                    date_series = pd.to_datetime(date_series, format="%m/%d/%Y")
                except:
                    try:
                        date_series = pd.to_datetime(date_series, format="%Y%m%d")
                    except:
                        try:
                            date_series = pd.to_datetime(date_series, format="%Y-%m-%d %H:%M:%S")
                        except:
                            date_series = pd.to_datetime(date_series, format="ISO8601")
        return date_series


  def __axa_raw_claim(self, raw_claim_path, password=None, policy_start_date=None, policy_data_date=None, client_name=None, region='HK', col_mapper=None):

    axa_df = pd.DataFrame(columns=self.col_setup)


    rename_col_clin = {
        # 'policy_id', # This is the policy_id for future database development, f'{policy_number}__{policy_start_date:%Y%m}'
        'POLICY HOLDER NO': 'policy_number',
        # 'insurer', # AXA
        # 'client_name', # self input
        # 'policy_start_date', # from file name
        # 'policy_end_date', # from file name
        'CLAIM NUMBER': 'claim_id',
        'DATE OF ADMISSION': 'incur_date',
        # 'discharge_date', # clinic set to incur_date
        'DATE OF RECEIPT': 'submission_date',
        'DATE OF PAYMENT': 'pay_date',
        'STATUS': 'claim_status',
        'REMARK CODE1': 'claim_remark_1',
        'REMARK CODE2': 'claim_remark_2',
        'REMARK CODE3': 'claim_remark_3',
        # 'cert_true_copy', # deduce from remark code
        'MEMBER': 'claimant',
        'CERTIFICATE NO': 'claimant',
        'SEX': 'gender',
        # 'age', deduce from DOB
        'DEPENDENT NO': 'dep_type',
        'HEALTH CLASS LEVEL  /OPT. HOSPITAL CLASS': 'class',
        # 'member_status', # not available
        # 'benefit_type', # sheetname
        'BENEFIT CODE': 'benefit', # need process
        # 'diagnosis', # not available
        # 'chronic', # not available
        'CURRENCY': 'currency',
        'AMOUNT BILLED': 'incurred_amount',
        'AMOUNT PAID': 'paid_amount',
        # 'panel', deduce from benefit code
        'AFFILIATED CO': 'suboffice',
        # 'region',
        'AGE': 'age',
        'DIAGNOSIS': 'diagnosis',

        'DATE OF BIRTH': 'birth_date'
    }

    dtype_axa_clin = {
        # 'policy_id', # This is the policy_id for future database development, f'{policy_number}__{policy_start_date:%Y%m}'
        'POLICY HOLDER NO': str,
        # 'insurer', # AXA
        # 'client_name', # self input
        # 'policy_start_date', # from file name
        # 'policy_end_date', # from file name
        'CLAIM NUMBER': str,
        'DATE OF ADMISSION': str,
        # 'discharge_date', # clinic set to 1
        'DATE OF RECEIPT': str,
        'DATE OF PAYMENT': str,
        'STATUS': str,
        'REMARK CODE1': str,
        'REMARK CODE2': str,
        'REMARK CODE3': str,
        # 'cert_true_copy', # deduce from remark code
        'MEMBER': str,
        'CERTIFICATE NO': str,
        'SEX': str,
        # 'age', deduce from DOB
        'DEPENDENT NO': str,
        'HEALTH CLASS LEVEL  /OPT. HOSPITAL CLASS': str,
        # 'member_status', # not available
        # 'benefit_type', # sheetname
        'BENEFIT CODE': str, # need process
        # 'diagnosis', # not available
        # 'chronic', # not available
        'CURRENCY': str,
        'AMOUNT BILLED': float,
        'AMOUNT PAID': float,
        # 'panel', deduce from benefit code
        'AFFILIATED CO': str,
        # 'region',
        'AGE': int,
        'DIAGNOSIS': str,

        'DATE OF BIRTH': str
    }

    date_col_clin = ['incur_date', 'submission_date', 'pay_date', 'birth_date']

    rename_col_hosp = {
        # 'policy_id', # This is the policy_id for future database development, f'{policy_number}__{policy_start_date:%Y%m}'
        'POLICY HOLDER NO': 'policy_number',
        # 'insurer', # AXA
        # 'client_name', # self input
        # 'policy_start_date', # from file name
        # 'policy_end_date', # from file name
        'CLAIM NUMBER': 'claim_id',
        'DATE OF ADMISSION': 'incur_date',
        'DATE OF DISCHARGE': 'discharge_date',
        'DATE OF RECEIPT': 'submission_date',
        'DATE OF PAYMENT': 'pay_date',
        'STATUS': 'claim_status',
        'REMARK': 'claim_remark_1',
        # 'cert_true_copy', # deduce from remark code
        'MEMBER': 'claimant',
        'CERTIFICATE NO': 'claimant',
        'SEX': 'gender',
        # 'age', deduce from DOB
        'DEPENDENT NO': 'dep_type',
        'HEALTH CLASS LEVEL  /OPT. HOSPITAL CLASS': 'class',
        # 'member_status', # not available
        # 'benefit_type', # sheetname
        'BENEFIT CODE': 'benefit', # need process
        # 'diagnosis', # not available
        # 'chronic', # not available
        'CURRENCY': 'currency',
        'AMOUNT BILLED': 'incurred_amount',
        'AMOUNT PAID': 'paid_amount',
        # 'panel', deduce from benefit code
        'AFFILIATED CO': 'suboffice',
        # 'region',
        'AGE': 'age',
        'DIAGNOSIS': 'diagnosis',

        'DATE OF BIRTH': 'birth_date'
    }

    dtype_axa_hosp = {
        # 'policy_id', # This is the policy_id for future database development, f'{policy_number}__{policy_start_date:%Y%m}'
        'POLICY HOLDER NO': str,
        # 'insurer', # AXA
        # 'client_name', # self input
        # 'policy_start_date', # from file name
        # 'policy_end_date', # from file name
        'CLAIM NUMBER': str,
        'DATE OF ADMISSION': str,
        'DATE OF DISCHARGE': str,
        'DATE OF RECEIPT': str,
        'DATE OF PAYMENT': str,
        'STATUS': str,
        'REMARK': str,
        # 'cert_true_copy', # deduce from remark code
        'MEMBER': str,
        'CERTICFICATE NO': str,
        'SEX': str,
        # 'age', deduce from DOB
        'DEPENDENT NO': str,
        'HEALTH CLASS LEVEL  /OPT. HOSPITAL CLASS': str,
        # 'member_status', # not available
        # 'benefit_type', # sheetname
        'BENEFIT CODE': str, # need process
        # 'diagnosis', # not available
        # 'chronic', # not available
        'CURRENCY': str,
        'AMOUNT BILLED': float,
        'AMOUNT PAID': float,
        # 'panel', deduce from benefit code
        'AFFILIATED CO': str,
        # 'region',
        'AGE': int,
        'DIAGNOSIS': str,

        'DATE OF BIRTH': str
    }
    date_col_hosp = ['incur_date', 'discharge_date', 'submission_date', 'pay_date', 'birth_date']

    unlocked_file = io.BytesIO()

    _policy_number = raw_claim_path.split('/')[-1].split('_')[1]
    _policy_start = raw_claim_path.split('/')[-1].split('_')[3]
    _policy_end = raw_claim_path.split('/')[-1].split('_')[4].split(' - ')[0]
    _client_name = client_name
    _insurer = 'AXA'

    print(_policy_start)

    with open(raw_claim_path, "rb") as file:
      excel_file = msoffcrypto.OfficeFile(file)
      excel_file.load_key(password = password)
      excel_file.decrypt(unlocked_file)
      from openpyxl import load_workbook
      wb = load_workbook(filename = unlocked_file)

      b_type_l = wb.sheetnames


      if 'CLIN' in b_type_l:
        df_c = pd.read_excel(unlocked_file, sheet_name='CLIN', dtype=dtype_axa_clin)

        # length = length + len(df_c)
        # print(len(df_c))
        df_c.rename(columns=rename_col_clin, inplace=True)


        if len(df_c) > 0:
          for col in date_col_clin:
            if col in df_c.columns.tolist():
              df_c[col] = df_c[col].loc[df_c[col] != '0']
              df_c[col] = pd.to_datetime(df_c[col], format='%Y%m%d')
            else:
              df_c[col] = np.nan



          if policy_start_date == None:
            start_d_ = pd.to_datetime(_policy_start, format='%Y%m%d')
          else:
            start_d_ = pd.to_datetime(policy_start_date, format='%Y%m%d')

          policy_no_ = df_c['policy_number'].iloc[0]
          df_c['policy_id'] = f'{policy_no_}_{start_d_:%Y%m}'

          df_c['insurer'] = 'AXA'
          df_c['client_name'] = client_name
          df_c['policy_start_date'] = start_d_
          df_c['policy_end_date'] = start_d_ + pd.DateOffset(years=1)
          # df_c['policy_end_date'] = pd.to_datetime(_policy_end, format='%Y%m%d')
          df_c['discharge_date'] = df_c['incur_date']
          df_c['cert_true_copy'] = np.nan
          print(df_c['birth_date'].values[0])
          if 'age' in df_c.columns.tolist() and df_c['age'].values[0] != np.nan:
            df_c['age'] = df_c['age'].astype(int)
          elif pd.isna(df_c['birth_date'].values[0]) == True:
            df_c['age'] = np.nan
          elif 'birth_date' in df_c.columns.tolist() and df_c['birth_date'].values[0] != np.nan:
            df_c['age'] = ((df_c['policy_start_date'] - df_c['birth_date']).dt.days/365.25).astype(int)
          else:
            df_c['age'] = np.nan
          df_c['member_status'] = np.nan
          df_c['benefit_type'] = 'Clinic'
          df_c['diagnosis'] = np.nan
          df_c['chronic'] = np.nan
          df_c['panel'] = 'Non-Panel'
          df_c['panel'].loc[df_c.benefit.str.contains('1')] = 'Panel'

        for col in self.col_setup:
          if col not in df_c.columns.tolist():
            df_c[col] = np.nan


      if 'DENT' in b_type_l:
        df_d = pd.read_excel(unlocked_file, sheet_name='DENT', dtype=dtype_axa_clin)
        df_d.rename(columns=rename_col_clin, inplace=True)
        if len(df_d) > 0:
          for col in date_col_clin:
            if col in df_d.columns.tolist():
              df_d[col] = pd.to_datetime(df_d[col], format='%Y%m%d')
            else:
              df_d[col] = np.nan

          if policy_start_date == None:
            start_d_ = pd.to_datetime(_policy_start, format='%Y%m%d')
          else:
            start_d_ = pd.to_datetime(policy_start_date, format='%Y%m%d')

          policy_no_ = df_d['policy_number'].iloc[0]
          df_d['policy_id'] = f'{policy_no_}_{start_d_:%Y%m}'

          df_d['insurer'] = 'AXA'
          df_d['client_name'] = client_name
          df_d['policy_start_date'] = start_d_
          df_d['policy_end_date'] = start_d_ + pd.DateOffset(years=1)
          #df_d['policy_end_date'] = pd.to_datetime(_policy_end, format='%Y%m%d')
          df_d['discharge_date'] = df_d['incur_date']
          df_d['cert_true_copy'] = np.nan
          if 'age' in df_d.columns.to_list() and df_d['age'].values[0] != np.nan:
            df_d['age'] = df_d['age'].astype(int)
          elif pd.isna(df_d['birth_date'].values[0]) == True:
            df_d['age'] = np.nan
          elif 'birth_date' in df_d.columns.tolist() and df_d['birth_date'].values[0] != np.nan:
            df_d['age'] = ((df_d['policy_start_date'] - df_d['birth_date']).dt.days/365.25).astype(int)
          else:
            df_d['age'] = np.nan
          df_d['member_status'] = np.nan
          df_d['benefit_type'] = 'Dental'
          df_d['diagnosis'] = np.nan
          df_d['chronic'] = np.nan
          df_d['panel'] = 'Non-Panel'
          df_d['panel'].loc[df_d.benefit.str.contains('1')] = 'Panel'

        for col in self.col_setup:
          if col not in df_d.columns.tolist():
            df_d[col] = np.nan


      if 'HOSP' in b_type_l:

        df_h = pd.read_excel(unlocked_file, sheet_name='HOSP', dtype=dtype_axa_hosp)
        df_h.rename(columns=rename_col_hosp, inplace=True)
        if len(df_h) > 0:
          for col in date_col_hosp:
            if col in df_h.columns.tolist():
              df_h[col] = pd.to_datetime(df_h[col], format='%Y%m%d')
            else:
              df_h[col] = np.nan

          if policy_start_date == None:
            start_d_ = pd.to_datetime(_policy_start, format='%Y%m%d')
          else:
            start_d_ = pd.to_datetime(policy_start_date, format='%Y%m%d')

          policy_no_ = df_h['policy_number'].iloc[0]
          df_h['policy_id'] = f'{policy_no_}_{start_d_:%Y%m}'

          df_h['insurer'] = 'AXA'
          df_h['client_name'] = client_name
          df_h['policy_start_date'] = start_d_
          df_h['policy_end_date'] = start_d_ + pd.DateOffset(years=1)
          # df_h['policy_end_date'] = pd.to_datetime(_policy_end, format='%Y%m%d')
          # df_h['discharge_date'] = df_h['incur_date']
          df_h['cert_true_copy'] = np.nan

          if 'age' in df_h.columns.to_list() and df_h['age'].values[0] != np.nan:
            df_h['age'] = df_h['age'].astype(int)
          elif pd.isna(df_h['birth_date'].values[0]) == True:
            df_h['age'] = np.nan
          elif 'birth_date' in df_h.columns.tolist() and df_h['birth_date'].values[0] != np.nan:
            df_h['age'] = ((df_h['policy_start_date'] - df_h['birth_date']).dt.days/365.25).astype(int)
          else:
            df_h['age'] = np.nan

          df_h['member_status'] = np.nan
          df_h['benefit_type'] = 'Hospital'
          df_h['diagnosis'] = np.nan
          df_h['chronic'] = np.nan
          df_h['panel'] = 'Non-Panel'
          df_h['panel'].loc[df_h.benefit.str.contains('1')] = 'Panel'

        for col in self.col_setup:
          if col not in df_h.columns.tolist():
            df_h[col] = np.nan

      axa_index = self.benefit_index[['gum_benefit', 'axa_benefit_code']]
      # print(df_d.columns)
      t_df = pd.concat([df_c[self.col_setup], df_h[self.col_setup], df_d[self.col_setup]], axis=0, ignore_index=True)
      t_df = pd.merge(left=t_df, right=axa_index, left_on='benefit', right_on='axa_benefit_code', how='left')
      t_df.benefit = t_df.gum_benefit
      t_df.suboffice.fillna('00', inplace=True)
      t_df.diagnosis.fillna('No diagnosis provided', inplace=True)
      t_df['region'] = region
      t_df['policy_data_date'] = pd.to_datetime(policy_data_date, format='%Y%m%d')
      t_df['age'] = t_df['age'].astype(int)
      t_df = t_df[self.col_setup]

    if pd.isna(t_df.gender).any() == False:
      t_df.gender.loc[t_df.gender.str.contains('f', case=False)] = 'F'
      t_df.gender.loc[t_df.gender.str.contains('m', case=False)] = 'M'

    axa_dep_type = {'1': 'EE', '2': 'SP', '3': 'CH', '4': 'CH', '5': 'CH', '6': 'CH', '7': 'CH', '8': 'CH'}
    t_df.dep_type.replace(axa_dep_type, inplace=True)
    # self.df = pd.concat([self.df, t_df], axis=0)
    return t_df

  def __bupa_raw_claim(self, raw_claim_path, password=None, policy_start_date=None, policy_data_date=None, client_name=None, region='HK', col_mapper=None):
    if col_mapper == None:
      bupa_rename_col = {
          # 'policy_id', # This is the policy_id for future database development, f'{policy_number}__{policy_start_date:%Y%m}'
          # 'policy_number', # Contract NUmber A1
          # 'insurer', # Bupa
          # 'client_name', # Customer Name A2
          # 'policy_start_date', # Period Cell A3
          # 'policy_end_date',
          'Voucher number': 'claim_id',
          'Voucher Number': 'claim_id',
          'Incur Date': 'incur_date',
          'Incurred From Date': 'incur_date',
          # 'discharge_date', convert from days cover
          'Receipt date': 'submission_date',
          'Pay date': 'pay_date',
          'Pay Date': 'pay_date',
          'Status Code': 'claim_status',
          'Reject Code Description': 'claim_remark_1',
          # 'cert_true_copy',
          'Claimant': 'claimant',
          'Sex': 'gender',
          'Age': 'age',
          'Member type': 'dep_type',
          'Member Type': 'dep_type',
          'Class': 'class',
          'Class ID': 'class',
          'Member Status': 'member_status',
          'Benefit Type': 'benefit_type',
          'Benefit': 'benefit',
          'Benefit Code Description': 'benefit',
          'Diagnosis Code': 'diagnosis_code',
          'Diagnosis description': 'diagnosis',
          'Diagnosis': 'diagnosis',
          'Procedure Code': 'procedure_code',
          'Procedure Description': 'procedure',
          'Hospital Name': 'hospital_name',
          'Provider Code': 'provider_code',
          'Provider Name': 'provider',
          'Physician Code': 'physician_code',
          'Physician Name': 'physician',
          'Chronic flag': 'chronic',
          'OPD indicator': 'opd',
          'Diagnosis Chronic Flag': 'chronic',
          # 'currency',
          'Presented': 'incurred_amount',
          'Adjusted': 'paid_amount',
          'Network Category': 'panel',
          'Network': 'panel',
          'E Claims Indicator': 'payment_method',
          
          # 'suboffice', # deduced from contract number columns last 2 digits
          # 'region',

          'Days Cover': 'days_cover',
          'Contract Number': 'contract_number',
      }
      dtype_bupa = {
          # 'policy_id', # This is the policy_id for future database development, f'{policy_number}__{policy_start_date:%Y%m}'
          # 'policy_number', # Contract NUmber A1
          # 'insurer', # Bupa
          # 'client_name', # Customer Name A2
          # 'policy_start_date', # Period Cell A3
          # 'policy_end_date',
          'Voucher number': str,
          'Voucher Number': str,
          'Incur Date': str,
          'Incurred From Date': str,
          # 'discharge_date', convert from days cover
          'Receipt date': str,
          'Pay date': str,
          'Pay Date': str,
          'Status Code': str,
          # 'claim_status',
          # 'claim_remark_1',
          'Reject Code Description': str,
          # 'cert_true_copy',
          'Claimant': str,
          'Sex': str,
          'Age': int,
          'Member type': str,
          'Member Type': str,
          'Class': str,
          'Class ID': str,
          'Member Status': str,
          'Benefit Type': str,
          'Benefit': str,
          'Benefit Code Description': str,
          'Diagnosis Code': str,
          'Diagnosis description': str,
          'Diagnosis': str,
          'Procedure Code': str,
          'Procedure Description': str,
          'Hospital Name': str,
          'Provider Code': str,
          'Provider Name': str,
          'Physician Code': str,
          'Physician Name': str,
          'Chronic flag': str,
          'OPD indicator': str,
          'Diagnosis Chronic Flag': str,
          # 'currency',
          'Presented': float,
          'Adjusted': float,
          'Network Category': str,
          'Network': str,
          'E Claims Indicator': str,
          # 'suboffice', # deduced from contract number columns last 2 digits
          # 'region',

          'Days Cover': int,
          'Contract Number': str,
      }

    date_cols = ['incur_date', 'submission_date', 'pay_date']
    df_ = pd.read_excel(raw_claim_path, sheet_name='Details')
    client_ = df_.columns[0].split('   ')[-1]
    policy_no_ = str(df_.iloc[0, 0].split('   ')[-1])
  

    if policy_start_date != None:
      start_d_ = pd.to_datetime(policy_start_date, format='%Y%m%d')
    else:
      start_d_ = pd.to_datetime(df_.iloc[1, 0].split(': ')[1].split(' ')[-5], format='%Y-%m-%d')
    end_d_ = pd.to_datetime(df_.iloc[1, 0].split(': ')[1].split(' ')[-1], format='%Y-%m-%d')
    df_ = pd.read_excel(raw_claim_path, sheet_name='Details', skiprows=7, dtype=dtype_bupa)
    df_.rename(columns=bupa_rename_col, inplace=True)

    for col in date_cols:
      if col in df_.columns.tolist():
        df_[col] = df_[col].loc[df_[col] != '0']
        df_[col] = pd.to_datetime(df_[col], format='%Y%m%d')
      else:
        df_[col] = np.nan

    if client_name != None:
      df_['client_name'] = client_name
    else:
      df_['client_name'] = client_

    df_['policy_number'] = policy_no_
    df_['policy_id'] = f'{policy_no_}_{start_d_:%Y%m}'
    df_['insurer'] = 'Bupa'
    df_['policy_start_date'] = start_d_
    df_['policy_end_date'] = end_d_
    df_['policy_data_date'] = pd.to_datetime(policy_data_date, format='%Y%m%d')
    df_['discharge_date'] = df_['incur_date']

    if 'days_cover' in df_.columns:
      df_['discharge_date'].loc[df_.benefit_type == 'H'] = df_['discharge_date'] + pd.to_timedelta(df_.days_cover, unit='D')
    df_['suboffice'] = df_['contract_number'].str[-2:]

    df_['region'] = region

    bupa_index = self.benefit_index[['gum_benefit', 'bupa_benefit_desc']]
    df_ = pd.merge(left=df_, right=bupa_index, left_on='benefit', right_on='bupa_benefit_desc', how='left')
    df_.benefit = df_.gum_benefit

    bupa_b_type = {'C': 'Clinic', 'H': 'Hospital', 'D': 'Dental', 'G': 'Optical', 'T': 'Hospital', 'W': 'Hospital', 'M': 'Maternity'}
    df_['benefit_type'].replace(bupa_b_type, inplace=True)
    bupa_dep_type = {'E': 'EE', 'S': 'SP', 'C': 'CH'}
    df_['dep_type'].replace(bupa_dep_type, inplace=True)
    bupa_panel_type = {'H': 'Panel', 'N': 'Non-Panel', 'O': 'Panel'}
    df_['panel'].replace(bupa_panel_type, inplace=True)
    df_.gender.loc[df_.gender.str.contains('f', case=False)] = 'F'
    df_.gender.loc[df_.gender.str.contains('m', case=False)] = 'M'
    df_.diagnosis.fillna('No diagnosis provided', inplace=True)

    for col in self.col_setup:
      if col not in df_.columns:
        df_[col] = np.nan

    df_ = df_[self.col_setup]



    return df_

  def __aia_raw_claim(self, raw_claim_path, password=None, policy_start_date=None, policy_data_date=None, client_name=None, region='HK', col_mapper=None):
    if col_mapper == None:
      dtype_aia = {
          # 'policy_id',
          'Policy No': str,
          'Policy': str,
          # 'insurer',
          'client_name': str,
          'Plan Start Date': str,
          'Plan End Date': str,
          'Claim ID': str,
          'ClaimNo.': str,
          'Claim No.': str,
          'Admission Date': str,
          'Incurred Date': str,
          'Date From': str,
          'Discharge date': str,
          'File Date': str,
          'Received Date': str,
          'Date Notified': str,
          'Claim Payment Date': str,
          'Process Date': str,
          # 'claim_status',
          'Remark1': str,
          'Remark2': str,
          # 'cert_true_copy',
          'Member ID': str,
          'Claimant No.': str,
          'Claimant Name': str,
          'Gender': str,
          'Sex': str, # M/F
          'Age': str,
          'Relationships': str, # MEMBER/ SPOUSE/ CHILD
          'Dep Type': str, # M/ S/ C
          'Plan': str,
          'BenPlan': str,
          # 'member_status',
          'Claim Type': str,
          'Product Name': str,
          'Claim Type 2': str,
          'Ben Desc': str,
          'Diagnosis': str,
          # 'chronic',
          'Billed Currency': str,
          'Currency Symbol': str,
          'Submitted Claim Amount': float,
          'Presented Amount': float,
          'Amount Billed': float,
          'Paid Claim': float,
          'Reimbursed Amount': float,
          'Amount Paid': float,
          'Provider Type': str,
          'Provider Name': str,
          'Network/Non-network': str,
          'SubOffice': str,
          # 'Entity': str,
          # 'region',
      }
      aia_rename_col = {
          # 'policy_id',
          'Policy No': 'policy_number',
          'Policy': 'policy_number',
          # 'insurer',
          'client_name': 'Client Name',
          'Plan Start Date': 'policy_start_date',
          'Plan End Date': 'policy_end_date',
          'Claim ID': 'claim_id',
          'ClaimNo.': 'claim_id',
          'Claim No.': 'claim_id',
          'Admission Date': 'incur_date',
          'Incurred Date': 'incur_date',
          'Date From': 'incur_date',
          # 'Discharge date': 'discharge_date',
          # 'File Date': 'submission_date',
          'Received Date': 'submission_date',
          'Date Notified': 'submission_date',
          'Claim Payment Date': 'pay_date',
          'Process Date': 'pay_date',
          'Date Paid': 'pay_date',
          # 'claim_status',
          'Remark1': 'claim_remark_1',
          'Remark2': 'claim_remark_2',
          # 'cert_true_copy',
          'Member ID': 'claimant',
          'Claimant No.': 'claimant',
          'Claimant Name': 'claimant',
          'Gender': 'gender',
          'Sex': 'gender', # M/F
          'Age': 'age',
          'Relationships': 'dep_type', # MEMBER/ SPOUSE/ CHILD
          'Dep Type': 'dep_type', # M/ S/ C
          'Plan': 'class',
          'BenPlan': 'class',
          # 'member_status',
          'Claim Type': 'benefit_type',
          'Product Name': 'benefit_type',
          'Claim Type 2': 'benefit',
          'Ben Desc': 'benefit',
          'Diagnosis': 'diagnosis',
          'Diagnosis Description': 'diagnosis',
          'Procedure': 'procedure',
          'ProviderName': 'provider',
          # 'chronic',
          'Billed Currency': 'currency',
          'Currency Symbol': 'currency', 
          'Submitted Claim Amount': 'incurred_amount',
          'Presented Amount': 'incurred_amount',
          'Amount Billed': 'incurred_amount',
          'Paid Claim': 'paid_amount',
          'Reimbursed Amount': 'paid_amount',
          'Amount Paid': 'paid_amount',
          'Provider Type': 'panel',
          'Provider Name': 'provider',
          'Network/Non-network': 'panel',
          'SubOffice': 'suboffice',
          # 'Entity': 'suboffice',
          # 'region',
      }


    date_cols = ['policy_start_date', 'policy_end_date', 'incur_date', 'discharge_date', 'submission_date', 'pay_date']

    if raw_claim_path.split('.')[-1] == 'xls':
      print('Please save this excel into xlsx format. *NOT Strict Open XML Spreadsheet')
      return


    if password == None:
      t_df = pd.read_excel(raw_claim_path, dtype=dtype_aia)
    else:
      unlocked_file = io.BytesIO()
      with open(raw_claim_path, "rb") as file:
        excel_file = msoffcrypto.OfficeFile(file)
        excel_file.load_key(password = password)
        excel_file.decrypt(unlocked_file)
        from openpyxl import load_workbook
        wb = load_workbook(filename = unlocked_file)
      t_df = pd.read_excel(unlocked_file, dtype=dtype_aia)

    t_df.rename(columns=aia_rename_col, inplace=True)
    t_df['insurer'] = 'AIA'
    t_df['policy_data_date'] = pd.to_datetime(policy_data_date, format='%Y%m%d')

    for col in date_cols:
      if col in t_df.columns.tolist():
        print(t_df[col].values[0])
        try:
          t_df[col] = pd.to_datetime(t_df[col], format='ISO8601')
        except:
          try:
            t_df[col] = pd.to_datetime(t_df[col], format='%Y%m%d')
          except:
            try:
              t_df[col] = pd.to_datetime(t_df[col], format='%y-%b')
            except:
              try:
                t_df[col] = pd.to_datetime(t_df[col], format='%Y-%m-%d')
              except:
                t_df[col] = pd.to_datetime(t_df[col])
      else:
        t_df[col] = np.nan

    if policy_start_date != None:
      t_df['policy_start_date'] = pd.to_datetime(policy_start_date, format='%Y%m%d')

    if client_name != None:
      t_df['client_name'] = client_name

    __start_date = t_df['policy_start_date'].iloc[0]
    __policy_number = t_df.policy_number.values[0]
    t_df['policy_id'] = f'{__policy_number}_{__start_date:%Y%m}'


    t_df['claim_status'] = np.nan
    t_df['cert_true_copy'] = np.nan
    t_df['member_status'] = np.nan
    t_df['chronic'] = np.nan
    t_df['region'] = region

    for col in self.col_setup:
      if col not in t_df.columns.tolist():
        t_df[col] = np.nan
      elif t_df[col].dtype == 'object':
        t_df[col] = t_df[col].str.strip()


    aia_index = self.benefit_index[['gum_benefit', 'aia_benefit']]
    t_df.benefit = t_df.benefit.str.strip()
    t_df = pd.merge(left=t_df, right=aia_index, left_on='benefit', right_on='aia_benefit', how='left')
    t_df.benefit = t_df.gum_benefit


    t_df = t_df[self.col_setup]

    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('aso', case=False)] = 'ASO'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('HMO-HS', case=False)] = 'Hospital'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('hosp', case=False)] = 'Hospital'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('clin', case=False)] = 'Clinic'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('dent', case=False)] = 'Dental'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('supp', case=False)] = 'Hospital'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('mat', case=False)] = 'Maternity'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('opt', case=False)] = 'Optical'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('well', case=False)] = 'Wellness'

    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('inpatient', case=False)] = 'Hospital'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('outpatient', case=False)] = 'Clinic'

    t_df.diagnosis.fillna('No diagnosis provided', inplace=True)
    # if len(t_df.diagnosis.unique()) > 2:
    #   t_df['benefit_type'].loc[(t_df['diagnosis'].str.contains('dental', case=False)) & (t_df['benefit'].str.contains('clinical', case=False) == False)] = 'Dental'
    #   t_df['benefit'].loc[t_df['diagnosis'].str.contains('dental', case=False) & (t_df['benefit'].str.contains('clinical', case=False) == False)] = 'Dental'
    #bupa_b_type = {'C': 'Clinic', 'H': 'Hospital', 'D': 'Dental', 'G': 'Optical', 'T': 'Hospital', 'W': 'Hospital', 'M': 'Maternity'}
    aia_dep_type = {'M': 'EE', 'S':'SP', 'C': 'CH', 'E': 'EE'}
    t_df['dep_type'].replace(aia_dep_type, inplace=True)
    t_df['dep_type'].loc[t_df['dep_type'].str.contains('mem', case=False)] = 'EE'
    t_df['dep_type'].loc[t_df['dep_type'].str.contains('emp', case=False)] = 'EE'
    t_df['dep_type'].loc[t_df['dep_type'].str.contains('spo', case=False)] = 'SP'
    t_df['dep_type'].loc[t_df['dep_type'].str.contains('chi', case=False)] = 'CH'
    t_df.gender.loc[t_df.gender.str.contains('f', case=False)] = 'F'
    t_df.gender.loc[t_df.gender.str.contains('m', case=False)] = 'M'
    t_df.panel.replace({'Non-network': 'Non-Panel', 'Network': 'Panel'}, inplace=True)
    t_df.panel.fillna('Non-Panel', inplace=True)
    t_df.suboffice.fillna('00', inplace=True)

    return t_df

  def blue_cross_raw_claim(self, raw_claim_path, password=None, policy_start_date=None, policy_data_date=None, client_name=None, region='HK', col_mapper=None):
    if col_mapper == None:
      dtype_bluecross_raw = {
          'INPUT CLAIM TYPE': str,
          'POLICY NO.': str,
          'MEMBER': str,
          'DEPENDANT CODE': str,
          'LEVEL CODE': str,
          'BENEFIT NAME': str,
          'CLAIM DATE': str,
          'CLAIM AMOUNT': float,
          'PAID AMOUNT': float,
          'REJECT REASON': str,
          'PAID BY CHQ/AUTOPAY': str,
      }
      bluecross_rename_col = {
        # 'policy_id', # This is the policy_id for future database development, f'{policy_number}__{policy_start_date:%Y%m}'
        'POLICY NO.': 'policy_number',
        # 'insurer',
        # 'client_name',
        # 'policy_start_date',
        # 'policy_end_date',
        # 'claim_id',
        'CLAIM DATE': 'incur_date',
        # 'discharge_date',
        # 'submission_date',
        # 'pay_date',
        # 'claim_status',
        'REJECT REASON': 'claim_remark_1',
        # 'cert_true_copy',
        'MEMBER': 'claimant',
        # 'gender',
        # 'age',
        'DEPENDANT CODE': 'dep_type',
        #'class',
        # 'member_status',
        # 'benefit_type',
        'BENEFIT NAME': 'benefit',
        # 'diagnosis',
        # 'procedure',
        # 'hospital_name',
        # 'chronic',
        # 'currency',PAID AMOUNT
        'CLAIM AMOUNT': 'incurred_amount',
        'PAID AMOUNT': 'paid_amount',
        'INPUT CLAIM TYPE': 'panel',
        'PAID BY CHQ/AUTOPAY': 'payment_method',
        # 'suboffice',
        # 'region',

        'LEVEL CODE': 'level_code'

      }
      date_cols = ['incur_date']
    
    
    t_df = pd.read_excel(raw_claim_path, dtype=dtype_bluecross_raw)
    t_df.rename(columns=bluecross_rename_col, inplace=True)

    for date_col in date_cols:
      if date_col in t_df.columns.tolist():
        t_df[date_col] = pd.to_datetime(t_df[date_col], format='%Y%m%d')
      else:
        t_df[date_col] = np.nan

    if policy_start_date != None:
      t_df['policy_start_date'] = pd.to_datetime(policy_start_date, format='%Y%m%d')
    else:
      t_df['policy_start_date'] = t_df['incur_date'].sort_values().iloc[0]

    t_df['policy_end_date'] = t_df['policy_start_date'] + pd.DateOffset(years=1)
    t_df['insurer'] = 'Blue Cross'
    t_df['client_name'] = client_name
    t_df['policy_start_date'] = pd.to_datetime(t_df['policy_start_date'])
    t_df['policy_data_date'] = pd.to_datetime(policy_data_date, format='%Y%m%d')
    _start_date = t_df['policy_start_date'].iloc[0]
    t_df['policy_id'] = f'{t_df.policy_number.values[0]}_{_start_date:%Y%m}'
    t_df['claim_status'] = np.nan
    t_df['claim_status'].loc[t_df['claim_remark_1'].isna() == False] = 'R'
    t_df['panel'].replace({'E': 'Panel', 'I': 'Non-Panel', 'O': 'Non-Panel'}, inplace=True)
    t_df['region'] = region
    t_df['class'] = t_df['level_code'].str.split(' ').str[-1].str[0]
    t_df['benefit_type'] = t_df['level_code'].str.split(' ').str[0]
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('HS', case=True)] = 'Hospital'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('OP', case=True)] = 'Clinic'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('DENT', case=True)] = 'Dental'
    t_df['suboffice'] = "00"
    t_df['dep_type'].fillna('EE', inplace=True)
    t_df['dep_type'].loc[t_df['dep_type'].str.contains('H|W', case=False)] = 'SP'
    t_df['dep_type'].loc[t_df['dep_type'].str.contains('C', case=False)] = 'CH'
    

    bluecross_benefit = self.benefit_index[['gum_benefit', 'blue_cross_benefit']]
    t_df = pd.merge(left=t_df, right=bluecross_benefit, left_on='benefit', right_on='blue_cross_benefit', how='left')
    t_df.benefit = t_df.gum_benefit
    if 'diagnosis' not in t_df.columns.tolist():
      t_df['diagnosis'] = 'No diagnosis provided'
    else:
      t_df['diagnosis'].fillna('No diagnosis provided', inplace=True)



    for col in self.col_setup:
      if col not in t_df.columns.tolist():
        t_df[col] = np.nan
    t_df = t_df[self.col_setup]

    return t_df

  def axa_raw_single(self, raw_claim_path, password=None, policy_start_date=None, policy_data_date=None, client_name=None, region='HK', col_mapper=None):
    if col_mapper == None:
      dtype_axa_raw = {
          'POLICY_HOLDER_NO': str,
          'CERT_NO': str,
          'PATIENT': str,
          'DEP_NO': str,
          'ROLE': str,
          'AGE': int,
          'GENDER': str,
          'CLAIM_NO': str,
          'ADMISSION/CONSULTATION_DATE': str,
          'DISCHARGE_DATE': str,
          'PAYMENT_DATE': str,
          'ROOM_TYPE': str,
          'HOSPITAL': str,
          'HOSPITAL_TYPE': str,
          'DOCTOR_CODE': str,
          'DOCTOR': str,
          'DIAGNOSIS_CODE': str,
          'DIAGNOSIS': str,
          'DIAGNOSIS_ORGAN': str,
          'PROCEDURE_CODE': str,
          'PROCEDURE': str,
          'STATUS': str,
          'NETWORK': str,
          'BENEFIT_CODE': str,
          'CURRENCY': str,
          'CLAIMS_AMOUNT(HKD)': float,
          'CLAIM_PAID_AMOUNT(HKD)': float,
          'CLASS': str,
          'AFFILIATED_CODE': str,
          'DEPARTMENT_CODE': str,
          'REMARK_1': str,
          'CLAIM_TYPE': str,
          'PAY_METHOD': str,
      }
      axa_rename_col = {
        # 'policy_id', # This is the policy_id for future database development, f'{policy_number}__{policy_start_date:%Y%m}'
        'POLICY_HOLDER_NO': 'policy_number',
        # 'insurer',
        # 'client_name',
        # 'policy_start_date',
        # 'policy_end_date',
        'CLAIM_NO': 'claim_id',
        'ADMISSION/CONSULTATION_DATE': 'incur_date',
        'DISCHARGE_DATE': 'discharge_date',
        # 'submission_date',
        'PAYMENT_DATE': 'pay_date',
        'STATUS': 'claim_status',
        'REMARK_1': 'claim_remark_1',
        'REMARK_2': 'claim_remark_2',
        'REMARK_3': 'claim_remark_3',
        'REMARK_4': 'claim_remark_4',
        'REMARK_5': 'claim_remark_5',
        # 'cert_true_copy',
        'PATIENT': 'claimant',
        'GENDER': 'gender',
        'AGE': 'age',
        'ROLE': 'dep_type',
        'CLASS': 'class',
        # 'member_status',
        'CLAIM_TYPE': 'benefit_type',
        'BENEFIT_CODE': 'benefit',
        'DIAGNOSIS_CODE': 'diagnosis_code',
        'DIAGNOSIS': 'diagnosis',
        'DIAGNOSIS_ORGAN': 'diagnosis_organ',
        'PROCEDURE_CODE': 'procedure_code',
        'PROCEDURE': 'procedure',
        'ROOM_TYPE': 'room_type',
        'HOSPITAL': 'hospital_name',
        'HOSPITAL_TYPE': 'hospital_type',
        'DOCTOR_CODE': 'physician_code',
        'DOCTOR': 'physician',
        # 'chronic',
        'CURRENCY': 'currency',
        'CLAIMS_AMOUNT(HKD)': 'incurred_amount',
        'CLAIM_PAID_AMOUNT(HKD)': 'paid_amount',
        'NETWORK': 'panel',
        'AFFILIATED_CODE': 'suboffice',
        'DEPARTMENT_CODE': 'department',
        'PAY_METHOD': 'payment_method',
        # 'region',
      }
    date_cols = ['incur_date', 'discharge_date', 'pay_date']
    t_df = pd.read_excel(raw_claim_path, dtype=dtype_axa_raw)
    t_df.rename(columns=axa_rename_col, inplace=True)
    for col in date_cols:
      if col in t_df.columns.tolist():
        t_df[col] = pd.to_datetime(t_df[col])
      else:
        t_df[col] = np.nan
    if policy_start_date != None:
      t_df['policy_start_date'] = pd.to_datetime(policy_start_date, format='%Y%m%d')
    else:
      t_df['policy_start_date'] = t_df['incur_date'].sort_values().iloc[0]
    t_df['policy_end_date'] = t_df['policy_start_date'] + pd.DateOffset(years=1)
    t_df['policy_data_date'] = pd.to_datetime(policy_data_date, format='%Y%m%d')
    t_df['insurer'] = 'AXA'
    t_df['client_name'] = client_name
    _start_date = t_df['policy_start_date'].iloc[0]
    t_df['policy_id'] = f'{t_df.policy_number.values[0]}_{_start_date:%Y%m}'
    t_df['dep_type'].replace({'EMPLOYEE': 'EE', 'SPOUSE': 'SP', 'CHILD': 'CH'}, inplace=True)
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('inp', case=False)] = 'Hospital'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('out', case=False)] = 'Clinic'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('den', case=False)] = 'Dental'
    t_df['panel'].replace({'Network': 'Panel', 'Non-Network': 'Non-Panel'}, inplace=True)

    axa_benefit = self.benefit_index[['gum_benefit', 'axa_benefit_code']]
    t_df = pd.merge(left=t_df, right=axa_benefit, left_on='benefit', right_on='axa_benefit_code', how='left')
    t_df.benefit = t_df.gum_benefit
    t_df['region'] = region
    

    for col in self.col_setup:
      if col not in t_df.columns.tolist():
        t_df[col] = np.nan


    t_df.diagnosis.fillna('No diagnosis provided', inplace=True)
    t_df.suboffice.fillna('00', inplace=True)
    t_df = t_df[self.col_setup]

    return t_df


  def lfh_raw_claim(self, raw_claim_path, password=None, policy_start_date=None, policy_data_date=None, client_name=None, region='HK', col_mapper=None):
    if col_mapper == None:
      dtype_lfh_raw = {
          'policy_number': str,
          # 'Policy Holder': str,
          # 'Dependant': str,
          # 'Relationship
          'Claim no': str,
          'Cert no': str,
          'Benefit code': str,
          'Benefit Name': str,
          'Plan': str,
          'Claim date': str,
          'Claim amt': float,
          'Paid date': str,
          'Paid amt': float,
      }
      lfh_rename_col = {
        # 'policy_id', # This is the policy_id for future database development, f'{policy_number}__{policy_start_date:%Y%m}'
        #'Claim no': 'policy_number',
        #'insurer',
        #'client_name',
        #'policy_start_date',
        #'policy_end_date',
        'Claim no': 'claim_id',
        'Claim date': 'incur_date',
        # 'discharge_date',
        # 'submission_date',
        'Paid date': 'pay_date',
        #'claim_status',
        #'claim_remark_1',
        #'cert_true_copy',
        'Cert no': 'claimant',
        #'gender',
        #'age',
        #'dep_type',
        'Plan': 'class',
        #'member_status',
        'Benefit Name': 'benefit_type',
        #'benefit',
        #'diagnosis',
        #'procedure',
        #'hospital_name',
        #'chronic',
        #'currency',
        'Claim amt': 'incurred_amount',
        'Paid amt': 'paid_amount',
        # 'panel',
        # 'suboffice',
        # 'region',
      }
    date_cols = ['incur_date', 'pay_date']
    t_df = pd.read_excel(raw_claim_path, dtype=dtype_lfh_raw)
    t_df.rename(columns=lfh_rename_col, inplace=True)
    for col in date_cols:
      if col in t_df.columns.tolist():
        t_df[col] = pd.to_datetime(t_df[col])
      else:
        t_df[col] = np.nan
    if policy_start_date != None:
      t_df['policy_start_date'] = pd.to_datetime(policy_start_date, format='%Y%m%d')
    else:
      t_df['policy_start_date'] = t_df['incur_date'].sort_values().iloc[0]
    t_df['policy_end_date'] = t_df['policy_start_date'] + pd.DateOffset(years=1)
    t_df['policy_data_date'] = pd.to_datetime(policy_data_date, format='%Y%m%d')
    t_df['insurer'] = 'LFH'
    t_df['client_name'] = client_name
    _start_date = t_df['policy_start_date'].iloc[0]
    t_df['policy_id'] = f'{t_df.policy_number.values[0]}_{_start_date:%Y%m}'
    ##############################################################################################################
    # Since LFH does not have the dep_type, we will assume that all the claims are EE
    # LFH does not have the benefit, we will assume that all the benefits are the same as the benefit_type
    ##############################################################################################################
    t_df['dep_type'] = 'EE'
    t_df['benefit'] = t_df['benefit_type']
    t_df['region'] = region
    t_df['suboffice'] = '00'
    if 'diagnosis' not in t_df.columns.tolist():
      t_df['diagnosis'] = 'No diagnosis provided'
    else:
      t_df['diagnosis'].fillna('No diagnosis provided', inplace=True)
    
    for col in self.col_setup:
      if col not in t_df.columns.tolist():
        t_df[col] = np.nan
    t_df = t_df[self.col_setup]

    return t_df

  def hsbc_raw_claim(self, raw_claim_path, password=None, policy_start_date=None, policy_data_date=None, client_name=None, region='HK', col_mapper=None):
    if col_mapper == None:
      dtype_hsbc_raw = {
          "POLICY NO.": str,
          'Masked ID': str,
          'DEPENDANT CODE': str,
          'Next Anniversary Date': str,
          'Gender': str,
          'Medical Tier': str,
          'Termination Date': str,
          'Effective Date': str,
          'Claim Type': str,
          'Benefit Type': str,
          'BENEFIT DESCRIPTION': str,
          'CLAIM STATUS': str,
          'CLAIM INCURRED DATE': str,
          'CLAIM INCURRED AMOUNT': float,
          'CLAIM PROCESSED DATE': str,
          'CLAIM PAID AMOUNT': float,
          'Claim Pay': str,
          'Payment method': str,
          'Currency': str,
          'REMARK DESCRIPTION': str,
      }
      hsbc_rename_col = {
        "POLICY NO.": "policy_number",
        "Masked ID": "claimant",
        "DEPENDANT CODE": "dep_type",
        "Next Anniversary Date": "policy_start_date",
        "Gender": "gender",
        "Medical Tier": "class",
        # "Termination Date"	
        # "Effective Date" 
        "Claim Type": "benefit_type",
        "Benefit Type": "benefit_type",
        "BENEFIT DESCRIPTION": "benefit",
        "CLAIM STATUS": "claim_status",	
        "CLAIM INCURRED DATE": "incur_date",	
        "CLAIM INCURRED AMOUNT": "incurred_amount",	
        "CLAIM PROCESSED DATE": "pay_date",	
        "CLAIM PAID AMOUNT": "paid_amount",
        # "Claim Pay": 	
        # "Payment method"	
        "Currency": "currency",
        "REMARK DESCRIPTION": "claim_remark_1",}
    # 'policy_id', # This is the policy_id for future
    date_cols = ['incur_date', 'pay_date']
    t_df = pd.read_excel(raw_claim_path, dtype=dtype_hsbc_raw)
    t_df.rename(columns=hsbc_rename_col, inplace=True)

    for col in date_cols:
      if col in t_df.columns.tolist():
        t_df[col] = pd.to_datetime(t_df[col])
      else:
        t_df[col] = np.nan
    if policy_start_date != None:
      t_df['policy_start_date'] = pd.to_datetime(policy_start_date, format='%Y%m%d')
    else:
      t_df['policy_start_date'] = t_df['incur_date'].sort_values().iloc[0]
    t_df['policy_end_date'] = t_df['policy_start_date'] + pd.DateOffset(years=1)
    t_df['policy_data_date'] = pd.to_datetime(policy_data_date, format='%Y%m%d')
    t_df['insurer'] = 'HSBC'
    t_df['client_name'] = client_name
    _start_date = t_df['policy_start_date'].iloc[0]
    t_df['policy_id'] = f'{t_df.policy_number.values[0]}_{_start_date:%Y%m}'
    
    t_df['dep_type'].replace({"1": "EE", "2": "SP"}, inplace=True)
    t_df['dep_type'].loc[t_df['dep_type'].isin(["EE", "SP"]) == False] = "CH"
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('HOSP', case=False)] = 'Hospital'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('CLIN', case=False)] = 'Clinic'
    t_df['benefit_type'].loc[t_df['benefit_type'].str.contains('SMM', case=False)] = 'Hospital'
    hsbc_index = self.benefit_index[['gum_benefit', 'hsbc_benefit']]
    t_df = pd.merge(left=t_df, right=hsbc_index, left_on='benefit', right_on='hsbc_benefit', how='left')
    t_df.benefit = t_df.gum_benefit
    t_df['region'] = region
    t_df['suboffice'] = '00'
    if 'diagnosis' not in t_df.columns.tolist():
      t_df['diagnosis'] = 'No diagnosis provided'
    else:
      t_df['diagnosis'].fillna('No diagnosis provided', inplace=True)
    
    for col in self.col_setup:
      if col not in t_df.columns.tolist():
        t_df[col] = np.nan
    t_df = t_df[self.col_setup]
    return t_df

  def bolttech_raw_claim(self, raw_claim_path, password=None, policy_start_date=None, policy_data_date=None, client_name=None, region='HK', col_mapper=None):
    if col_mapper == None:
      dtype_bolttech_raw = {
        'Policy_No': str,
        'SubsidiaryName': str,
        'Relationship': str,
        'PLANDESC': str,
        'ClaimOcc': str,
        'Claim_Status': str,
        'Network': str,
        'Product_Code': str,
        'ServiceCode': str,
        'ServiceDesc': str,
        'IncurredAmount': float,
        'ActualPaid': float,
        'RejectCode1': str,
      }
      bolttech_rename_col = {
        'Policy_No': 'policy_number',
        'SubsidiaryName': 'client_name',
        'Relationship': 'dep_type',
        'PLANDESC': 'class',
        'ClaimOcc': 'claim_id',
        'Claim_Status': 'claim_status',
        'Network': 'panel',
        'Product_Code': 'benefit_type',
        'ServiceDesc': 'benefit',
        'IncurredAmount': 'incurred_amount',
        'ActualPaid': 'paid_amount',
        'RejectCode1': 'claim_remark_1',
      }

    t_df = pd.read_excel(raw_claim_path, sheet_name="Raw_Claim", dtype=dtype_bolttech_raw)
    t_df.rename(columns=bolttech_rename_col, inplace=True)
    
    t_df['policy_start_date'] = policy_start_date
    t_df['policy_start_date'] = pd.to_datetime(t_df['policy_start_date'], format='%Y%m%d')
    t_df['policy_end_date'] = t_df['policy_start_date'] + pd.DateOffset(years=1)
    t_df['policy_data_date'] = pd.to_datetime(policy_data_date, format='%Y%m%d')
    _start_date = t_df['policy_start_date'].iloc[0]
    t_df['policy_id'] = f"{t_df['policy_number'].values[0]}_{_start_date:%Y%m}"
    t_df['insurer'] = 'Bolttech'
    t_df['benefit_type'].replace({'HOSP': 'Hospital', 'OUTP': 'Clinic', "MATN": 'Maternity'}, inplace=True)
    t_df['panel'].replace({'Network': 'Panel', 'Non-Network': 'Non-Panel'}, inplace=True)
    t_df['incur_date'] = t_df['policy_start_date']
    t_df['pay_date'] = t_df['policy_start_date']
    t_df['claimant'] = t_df['claim_id']
    t_df['region'] = region
    t_df['suboffice'] = '00'
    bolttech_index = self.benefit_index[['gum_benefit', 'bolttech_benefit']]
    t_df = pd.merge(left=t_df, right=bolttech_index, left_on='benefit', right_on='bolttech_benefit', how='left')
    t_df.benefit = t_df.gum_benefit
    if 'diagnosis' not in t_df.columns.tolist():
      t_df['diagnosis'] = 'No diagnosis provided'
    else:
      t_df['diagnosis'].fillna('No diagnosis provided', inplace=True)
    for col in self.col_setup:
      if col not in t_df.columns.tolist():
        t_df[col] = np.nan
    t_df = t_df[self.col_setup]

    
    return t_df


  def sunlife_raw_claim(self, raw_claim_path, password=None, policy_start_date=None, policy_data_date=None, client_name=None, region='HK', col_mapper=None):
    if col_mapper == None:
      dtype_sunlife_raw = {
          'claim_number': str,
          'claim_status': str,
          'group_policy_number': str,
          'company_code': str,
          'company_name_en': str,
          'employee_code': str,
          'dependent_code': int,
          'employee_name': str,
          'dependent_name': str,
          'membership_class': str,
          'incurred_date': str,
          'received_date': str,
          'payment_date': str,
          'claim_type': str,
          'benefit_type': str,
          'benefit_name': str,
          'benefit_category': str,
          'incurred_amount': float,
          'cob_amount': float,
          'uncovered_amount': float,
          'paid_amount_exclude_smm': float,
          'paid_amount_smm': float,
          'total_paid_amount': float,
          'payment_gateway': str,
      }
      sunlife_rename_col = {
        'claim_number': 'claim_id',
        'claim_status': 'claim_status',
        'group_policy_number': 'policy_number',
        'company_name_en': 'client_name',
        'employee_code': 'claimant',
        'dependent_code': 'dep_type',
        # 'employee_name': 'claimant',
        # 'dependent_name': 'claimant',
        'membership_class': 'class',
        'incurred_date': 'incur_date',
        'received_date': 'submission_date',
        'payment_date': 'pay_date',
        'claim_type': 'panel',
        # 'benefit_type': 'benefit_type',
        'benefit_name': 'benefit',
        'benefit_category': 'benefit_type',
        'incurred_amount': 'incurred_amount',
        'total_paid_amount': 'paid_amount',
      }
    
    date_cols = ['incur_date', 'submission_date', 'pay_date']
    t_df = pd.read_excel(raw_claim_path, dtype=dtype_sunlife_raw)
    t_df.drop(columns=['benefit_type'], inplace=True)
    t_df.rename(columns=sunlife_rename_col, inplace=True)
    
    t_df['insurer'] = 'Sunlife'
    t_df['policy_start_date'] = pd.to_datetime(policy_start_date)
    t_df['policy_end_date'] = t_df['policy_start_date'] + pd.DateOffset(years=1)
    t_df['policy_data_date'] = pd.to_datetime(policy_data_date, format='%Y%m%d')
    _start_date = t_df['policy_start_date'].iloc[0]
    t_df['policy_id'] = f"{t_df['policy_number'].values[0]}_{_start_date:%Y%m}"
    t_df['benefit_type'].replace({'HS': 'Hospital', 'OPCC': 'Clinic', 'UNCOVERED': 'not_covered'}, inplace=True)
    t_df['panel'].replace({'panel': 'Panel', 'reimbursement': 'Non-Panel'}, inplace=True)
    t_df['dep_type'].replace({0: 'EE', 1: 'SP', 2: 'CH', 3: 'CH', 4: 'CH', 5: 'CH', 6: 'CH', 7: 'CH'}, inplace=True)
    t_df['region'] = region
    t_df['suboffice'] = '00'
    for col in date_cols:
      if col in t_df.columns.tolist():
        t_df[col] = self.date_conversion(t_df[col])
      else:
        t_df[col] = np.nan
  
    sunlife_index = self.benefit_index[['gum_benefit', 'sunlife_benefit']]
    t_df = pd.merge(left=t_df, right=sunlife_index, left_on='benefit', right_on='sunlife_benefit', how='left')
    t_df.benefit = t_df.gum_benefit
    if 'diagnosis' not in t_df.columns.tolist():
      t_df['diagnosis'] = 'No diagnosis provided'
    else:
      t_df['diagnosis'].fillna('No diagnosis provided', inplace=True)
    for col in self.col_setup:
      if col not in t_df.columns.tolist():
        t_df[col] = np.nan
    t_df = t_df[self.col_setup]
    return t_df


  def __consol_raw_claim(self, raw_claim_path):
    dtype_con_raw = {
      'policy_id': str,
      'policy_number': str,
      'insurer': str,
      'client_name': str,
      'policy_start_date': str,
      'policy_end_date': str,
      'claim_id': str,
      'incur_date': str,
      'discharge_date': str,
      'submission_date': str,
      'pay_date': str,
      'claim_status': str,
      'claim_remark_1': str,
      'cert_true_copy': str,
      'claimant': str,
      'gender': str,
      'age': str,
      'dep_type': str,
      'class': str,
      'member_status': str,
      'benefit_type': str,
      'benefit': str,
      'diagnosis': str,
      'chronic': str,
      'currency': str,
      'incurred_amount': float,
      'paid_amount': float,
      'panel': str,
      'suboffice': str,
      'region': str,
    }
    date_cols = ['policy_start_date', 'policy_end_date', 'incur_date', 'discharge_date', 'submission_date', 'pay_date']

    if raw_claim_path.split('.')[-1] == 'csv':
      t_df = pd.read_csv(raw_claim_path, dtype= dtype_con_raw)
    else:
      t_df = pd.read_excel(raw_claim_path, dtype= dtype_con_raw)

    for col in date_cols:
      if len(t_df[col].unique()) > 1:
        try:
          t_df[col].loc[~(t_df[col].isna())] = pd.to_datetime(t_df[col].loc[~(t_df[col].isna())], format='%m/%d/%y')
        except:
          t_df[col].loc[~(t_df[col].isna())] = pd.to_datetime(t_df[col].loc[~(t_df[col].isna())], format='%Y-%m-%d')

    return t_df


  def add_raw_data(self, raw_claim_path, insurer=None, password=None, policy_start_date=None, policy_data_date=None, client_name=None, region='HK', col_mapper=None):

    if insurer == 'AXA':
      temp_df = self.__axa_raw_claim(raw_claim_path, password, policy_start_date, policy_data_date, client_name, region, col_mapper)
    elif insurer == 'AIA':
      temp_df = self.__aia_raw_claim(raw_claim_path, password, policy_start_date, policy_data_date, client_name, region, col_mapper)
    elif insurer == 'Bupa':
      temp_df = self.__bupa_raw_claim(raw_claim_path, password, policy_start_date, policy_data_date, client_name, region, col_mapper)
    elif insurer == 'Blue Cross':
      temp_df = self.blue_cross_raw_claim(raw_claim_path, password, policy_start_date, policy_data_date, client_name, region, col_mapper)
    elif insurer == 'AXA single':
      temp_df = self.axa_raw_single(raw_claim_path, password, policy_start_date, policy_data_date, client_name, region, col_mapper)
    elif insurer == 'LFH':
      temp_df = self.lfh_raw_claim(raw_claim_path, password, policy_start_date, policy_data_date, client_name, region, col_mapper)
    elif insurer == 'HSBC':
      temp_df = self.hsbc_raw_claim(raw_claim_path, password, policy_start_date, policy_data_date, client_name, region, col_mapper)
    elif insurer == 'Bolttech':
      temp_df = self.bolttech_raw_claim(raw_claim_path, password, policy_start_date, policy_data_date, client_name, region, col_mapper)
    elif insurer == 'Sunlife':
      temp_df = self.sunlife_raw_claim(raw_claim_path, password, policy_start_date, policy_data_date, client_name, region, col_mapper)
    else:
      # print('Please make sure that the colums of the DataFrame is aligned with the standard format')
      temp_df = self.__consol_raw_claim(raw_claim_path)
    self.df = pd.concat([self.df, temp_df], axis=0, ignore_index=True)

    self.policy_id_list = self.df['policy_id'].unique().tolist()
    self.policy_list = self.df['policy_number'].unique().tolist()
    self.clinet_name_list = self.df['client_name'].unique().tolist()
    self.benefit_type_list = self.df['benefit_type'].unique().tolist()
    self.benefit_list = self.df['benefit'].unique().tolist()

  def bupa_shortfall_supplement(self, shortfall_processed_df):
    # shortfall_panel = shortfall_processed_df.loc[shortfall_processed_df['panel'] == 'Panel']
    shortfall_panel = shortfall_processed_df.loc[shortfall_processed_df['panel'] == 'Overall']
    for n00 in np.arange(len(shortfall_panel)):
      __policy_id = shortfall_panel['policy_id'].iloc[n00]
      __class = shortfall_panel['class'].iloc[n00]
      __benefit = shortfall_panel['benefit'].iloc[n00]
    
      __no_of_claims = shortfall_panel['no_of_claims'].iloc[n00] - \
        self.df['incurred_amount'].loc[(self.df['policy_id'] == __policy_id) & (self.df['class'] == __class) & (self.df['benefit'] == __benefit) & (self.df['panel'] == 'Non-Panel')].dropna().count() - \
        self.df['incurred_amount'].loc[(self.df['policy_id'] == __policy_id) & (self.df['class'] == __class) & (self.df['benefit'] == __benefit) & (self.df['panel'] == 'Panel')].dropna().count()
      
      if __no_of_claims > 0:
        __incurred_amount = shortfall_panel['incurred_amount'].iloc[n00] - \
          self.df['incurred_amount'].loc[(self.df['policy_id'] == __policy_id) & (self.df['class'] == __class) & (self.df['benefit'] == __benefit) & (self.df['panel'] == 'Non-Panel')].dropna().sum() - \
          self.df['incurred_amount'].loc[(self.df['policy_id'] == __policy_id) & (self.df['class'] == __class) & (self.df['benefit'] == __benefit) & (self.df['panel'] == 'Panel')].dropna().sum()
        
        __paid_amount = shortfall_panel['paid_amount'].iloc[n00] - \
          self.df['paid_amount'].loc[(self.df['policy_id'] == __policy_id) & (self.df['class'] == __class) & (self.df['benefit'] == __benefit) & (self.df['panel'] == 'Non-Panel')].dropna().sum() -\
          self.df['paid_amount'].loc[(self.df['policy_id'] == __policy_id) & (self.df['class'] == __class) & (self.df['benefit'] == __benefit) & (self.df['panel'] == 'Panel')].dropna().sum()

        t_incur_per_claim = __incurred_amount / __no_of_claims
        t_paid_per_claim = __paid_amount / __no_of_claims

        self.df['incurred_amount'].loc[(self.df['policy_id'] == __policy_id) & (self.df['class'] == __class) & (self.df['benefit'] == __benefit) & (self.df['panel'] == 'Panel') & (pd.isna(self.df['incurred_amount']) == True)] = t_incur_per_claim
        self.df['paid_amount'].loc[(self.df['policy_id'] == __policy_id) & (self.df['class'] == __class) & (self.df['benefit'] == __benefit) & (self.df['panel'] == 'Panel') & (pd.isna(self.df['paid_amount']) == True)] = t_paid_per_claim
    return



  def preprocessing(self, policy_id=None, rejected_claim=True, aso=True, smm=True, diagnosis=False, provider_grouping=False, common_diagnosis=True, group_optical=False, research_mode=False, age_band=False):

    self.df['data_month'] = (((self.df['policy_data_date'] - self.df['policy_start_date']).dt.days / 30.5).astype(int) + 1).astype(int)
    print(self.df.data_month.unique())
    self.df = pd.merge(left=self.df, right=self.ibnr, how='left', left_on='data_month', right_on='data_month')
    # self.df['factoring'] = self.df['year'] == self.policy_start_date.year
    # Factoring is determined in mcr_pages function
    self.factoring_cols = ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']

    self.df['class'] = self.df['class'].apply(lambda v: v.strip() if isinstance(v, str) else v)
    # self.df.drop(columns=['data_month_ibnr'], inplace=True)

    if aso == True:
      self.df = self.df.loc[self.df.benefit_type != 'ASO']

    if rejected_claim == True:
      reject_claim_words = ['submit', 'resumit', 'submission', 'receipt', 'signature', 'photo', 'provide', 'form', 'Amount claimed for was already processed by another insurer']
      self.df.claim_remark_1.fillna('no_remark', inplace=True)
      # Bupa claim remark = Reject Code so it must be rejected
      self.df.claim_status.loc[(self.df.insurer == 'Bupa') & (self.df.claim_remark_1 != 'no_remark')] = 'R'
      self.df.claim_status.loc[self.df.claim_remark_1.str.contains('|'.join(reject_claim_words), case=False) & ((self.df.paid_amount == 0) | (self.df.paid_amount.isna()))] = 'R'
      self.df.claim_status.loc[(self.df.insurer == 'AXA') & (self.df.claim_status == 'R') & (self.df.paid_amount != 0)] = 'PR'
      self.df.claim_status.loc[(self.df.insurer == 'HSBC') & (self.df.claim_status != "Approved")] = 'R'
      self.df.claim_status.loc[(self.df.insurer == "Sunlife") & (self.df.benefit_type == "not_covered")] = 'R'
      self.df = self.df.loc[(self.df.claim_status != 'R')]
      # self.df = self.df.loc[(self.df.claim_remark_1 == 'no_remark')]
      

    if smm == True:
      self.df.benefit.fillna('no_benefit', inplace=True)
      self.df.incurred_amount.loc[self.df.benefit == 'Top-up/SMM'] = 0
      self.df.incurred_amount.loc[self.df.benefit.str.contains('secondary claim', case=False)] = self.df.paid_amount.loc[self.df.benefit.str.contains('secondary claim', case=False)]
      self.df.incurred_amount.loc[self.df.benefit.str.contains('daily cash benefit', case=False)] = self.df.paid_amount.loc[self.df.benefit.str.contains('daily cash benefit', case=False)]

    if diagnosis == True:
      self.df.diagnosis = self.df.diagnosis.str.lower()
      self.df.diagnosis.loc[self.df.diagnosis.str.contains('acute upper respiratory infec|common cold', case=False)] = 'acute upper respiratory infection & common cold'
      self.df.diagnosis.loc[self.df.diagnosis.str.contains(', unspec', case=False)] = self.df.diagnosis.loc[self.df.diagnosis.str.contains(', unspec', case=False)].replace(to_replace={', unspec': '', ', unspecified': ''})
      self.df.diagnosis.loc[(self.df.diagnosis.str.contains('chlamydia', case=False)) & (self.df.insurer == 'AIA') ] = 'viral warts'
      self.df.diagnosis.loc[self.df.diagnosis.str.contains('dermatitis|eczema', case=False)] = 'dermatitis & eczema'
      self.df.diagnosis.loc[self.df.diagnosis.str.contains('gastritis|gastroenteritis|intestinal infec', case=False)] = 'gastritis and gastroenteritis'
      self.df.diagnosis.loc[self.df.diagnosis.str.contains('benign neoplasm of colon|polyp of colon|anal and rectal polyp', case=False)] = 'benign neoplasm of colon & polyp'
      self.df.diagnosis.loc[self.df.diagnosis.str.contains('benign neoplasm of stomach|polyp of stomach and duodenum', case=False)] = 'benign neoplasm of stomach & polyp'
      self.df.diagnosis.loc[self.df.diagnosis.str.contains('influenza', case=False)] = 'influenza'
      self.df.diagnosis.loc[self.df.diagnosis.str.contains('lumbar sprain|sprain and strain of lumbar spine|backache|low back pain|sprain of unspec site of back', case=False)] = 'lumbago'
      self.df.diagnosis.loc[self.df.diagnosis.str.contains('pain in joint|pain in ankle and joints of foot|sprains and strains of ankle and foot', case=False)] = 'pain in joint and ankle'
      self.df.diagnosis.loc[self.df.diagnosis.str.contains('neck sprain|cervicalgia', case=False)] = 'cervicalgia (neck pain)'
      self.df.diagnosis.loc[self.df.diagnosis.str.contains('cataract', case=False)] = 'cataract'
      # self.df.diagnosis.loc[self.df.diagnosis.str.contains('bronchitis|bronchiolitis', case=False)] = 'bronchitis and bronchiolitis' # bronchitis and bronchiolitis are not the same speicality
      self.df.diagnosis.loc[((self.df.diagnosis.str.contains('hypertension', case=False)) & (self.df.diagnosis.str.contains('ocular', case=False) == False))|((self.df.diagnosis.str.contains('hypertensive', case=False)) & (self.df.diagnosis.str.contains('retinopathy', case=False) == False))] = 'hypertension & hypertensive disease'
      # print('changed')
    
    if provider_grouping == True:
      provider_grouping_df = pd.read_csv(self.provider_grouping_path, dtype={'provider_name': str, 'mapped_name': str})
      self.df['provider'].fillna('No Provider Provided', inplace=True)
      self.df = pd.merge(
          left=self.df,
          right=provider_grouping_df,
          left_on='provider',
          right_on='provider_name',
          how='left'
      )
      self.df['mapped_name'].loc[self.df['mapped_name'].isna()] = self.df['provider'].loc[self.df['mapped_name'].isna()]
      self.df['provider'] = self.df['mapped_name']
      self.df.drop(columns=['provider_name', 'mapped_name'], inplace=True)
    
    if common_diagnosis == True:
      self.df['common_diagnosis_flag'].fillna("others", inplace=True)
      self.df['procedure'].fillna("no procedures provided", inplace=True)
      self.df['common_diagnosis_flag'].loc[(self.df.diagnosis.str.contains('viral warts', case=False))|self.df.procedure.str.contains('warts', case=False)] = 'viral_warts'
      self.df['common_diagnosis_flag'].loc[(self.df.procedure.str.contains('endoscopy|esophagoscopy|gastroscopy|colonoscopy|esophagogastroduodenoscopy|anoscopy|proctoscopy|sigmoidoscopy|proctosigmoidoscopy', case=False) == True) & 
                                           (self.df.procedure.str.contains('nasal', case=False) == False)] = 'endoscopy'
      endo_id_list = self.df['claim_id'].loc[self.df['common_diagnosis_flag'] == 'endoscopy'].unique().tolist()
      viral_warts_id_list = self.df['claim_id'].loc[self.df['common_diagnosis_flag'] == 'viral_warts'].unique().tolist()
      self.df['common_diagnosis_flag'].loc[(self.df['common_diagnosis_flag'] == 'others') & (self.df['claim_id'].isin(endo_id_list) == True)] = 'endoscopy'
      self.df['common_diagnosis_flag'].loc[(self.df['common_diagnosis_flag'] == 'others') & (self.df['claim_id'].isin(viral_warts_id_list) == True)] = 'viral_warts'
      

    if group_optical == False:
      self.df['benefit_type'].loc[self.df['benefit_type'] == 'Optical'] = 'Clinic'

    self.df.suboffice.fillna('00', inplace=True)
    self.df = pd.merge(
        left=self.df,
        right=self.speciality_index,
        left_on='diagnosis',
        right_on='diagnosis',
        how='left'
    )
    self.df['speciality_x'] = self.df['speciality_y']
    self.df.drop(columns=['speciality_y'], inplace=True)
    self.df.rename(columns={'speciality_x': 'speciality'}, inplace=True)
    self.df['speciality'].fillna('no_index', inplace=True)

    self.df = pd.merge(
        left=self.df,
        right=self.minor_surg_index,
        left_on='procedure',
        right_on='minor_surgeries',
        how='left'
    )

    if research_mode == True:
      self.df.claimant = self.df.policy_id + "_" + self.df.claimant
      self.df.claim_id = self.df.policy_id + "_" + self.df.claim_id
      # self.df['class'] = self.df.policy_id + "_" + self.df['class']

    if age_band == True:
      self.df['age'].fillna("No Age Provided", inplace=True)
      self.df['age'].loc[self.df.age.astype(str).str.contains('-')] = self.df['age'].loc[self.df.age.astype(str).str.contains('-')].str.split('-').str[0].astype(int)
      self.df['age'].loc[self.df.age.astype(str).str.contains('\+')] = self.df['age'].loc[self.df.age.astype(str).str.contains('\+')].str.split('+').str[0].astype(int)
      self.df['age_band'] = self.df['age'].copy(deep=True)
      bins = range(-1, 101, 10)
      labels = [f"{i+1}-{i+10}" for i in bins[:-1]]
      self.df['age_band'].loc[self.df['age'].astype(str).str.isnumeric() == True] = pd.cut(self.df['age'].loc[self.df['age'].astype(str).str.isnumeric() == True].astype(int), bins=bins, labels=labels)
      self.df['age_band'] = self.df['age_band'].astype(str)

    self.df['policy_end_date'] = self.df['policy_start_date'] + pd.DateOffset(years=1) - pd.DateOffset(days=1)
    self.df['day_procedure_flag'] = ""
    # pat_target = r"Surgeon Fee - Minor|day centre|Surgeon Fee - Non Classified \(AXA\)"
    pat_target = r"Surgeon Fee|day centre"
    pat_room   = r"Daily Room & Board"

    # 2. Claim-level booleans (broadcast back to each row)
    id_has_target = self.df.groupby('claim_id')['benefit'].transform(
        lambda s: s.str.contains(pat_target, case=False, na=False).any()
    )

    id_has_room = self.df.groupby('claim_id')['benefit'].transform(
        lambda s: s.str.contains(pat_room, case=False, na=False).any()
    )

    # 3. Build a 4-category flag
    conditions = [
        id_has_target & ~id_has_room,   # 1. target AND no room
        id_has_target & id_has_room,    # 2. target AND room
        ~id_has_target & id_has_room,   # 3. no target AND room
        ~id_has_target & ~id_has_room   # 4. no target AND no room
    ]

    choices = [
        "day_procedure",           # (1)
        "hospital_have_surgery",   # (2)
        "hospital_no_surgery",     # (3)
        "no_hospital_no_surgery"   # (4)
    ]

    self.df['day_procedure_flag'] = np.select(conditions, choices, default="no_hospital_no_surgery")
    self.df['day_procedure_flag'].loc[self.df.benefit_type != "Hospital"] = "clinic_and_other"

    return None

  def mcr_policy_info(self, by=None):
    policy_info_col = by + ['insurer', 'client_name', 'policy_start_date', 'policy_end_date', 'policy_data_date', 'data_month', 'ibnr', 'factoring']
    policy_info_df = self.mcr_df[policy_info_col].drop_duplicates().set_index(by).sort_index()
    self.mcr_factor_df = policy_info_df[['data_month', 'ibnr', 'factoring']].copy(deep=True)

    self.mcr_df['has_diagnosis'] = self.mcr_df['diagnosis'].str.contains('no diagnosis provided', case=False) == False
    diag_prov = self.mcr_df[by + ['has_diagnosis']].groupby(by=by, dropna=False).agg({'has_diagnosis': 'any'})
    policy_info_df = pd.merge(policy_info_df, diag_prov, how='left', left_index=True, right_index=True)
    self.mcr_df.drop(columns=['has_diagnosis'], inplace=True)

    self.mcr_df['has_hospital'] = self.mcr_df['hospital_name'].str.contains('No Hospital Name', case=False) == False
    hosp_prov = self.mcr_df[by + ['has_hospital']].groupby(by=by, dropna=False).agg({'has_hospital': 'any'})
    policy_info_df = pd.merge(policy_info_df, hosp_prov, how='left', left_index=True, right_index=True)
    self.mcr_df.drop(columns=['has_hospital'], inplace=True)

    self.mcr_df['has_provider'] = self.mcr_df['provider'].str.contains('No Provider Provided', case=False) == False
    provider_prov = self.mcr_df[by + ['has_provider']].groupby(by=by, dropna=False).agg({'has_provider': 'any'})
    policy_info_df = pd.merge(policy_info_df, provider_prov, how='left', left_index=True, right_index=True)
    self.mcr_df.drop(columns=['has_provider'], inplace=True)

    self.mcr_df['has_procedure'] = self.mcr_df['procedure'].str.contains('no procedures provided', case=False) == False
    procedure_prov = self.mcr_df[by + ['has_procedure']].groupby(by=by, dropna=False).agg({'has_procedure': 'any'})
    policy_info_df = pd.merge(policy_info_df, procedure_prov, how='left', left_index=True, right_index=True)
    self.mcr_df.drop(columns=['has_procedure'], inplace=True)

    self.mcr_factor_df.data_month = self.mcr_factor_df.data_month.astype(int)
    self.policy_info = policy_info_df
    return policy_info_df

  def mcr_p20_policy(self, by=None, annualize=False, ibnr=False, research_mode=False):
    __p20_policy_df_col = by + ['incurred_amount', 'paid_amount', 'claimant']
    __p20_policy_group_col = by.copy()
    
    self.mcr_df.policy_start_date = pd.to_datetime(self.mcr_df.policy_start_date)
    p20_policy_df = self.mcr_df[__p20_policy_df_col].groupby(by=__p20_policy_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique'}).rename(columns={'claimant': 'no_of_claimants'})
    p20_policy_df['usage_ratio'] = p20_policy_df['paid_amount'] / p20_policy_df['incurred_amount']
    p20_policy_df = p20_policy_df.unstack().stack(dropna=False)

    if annualize == True or ibnr == True:
      p20_policy_df = pd.merge(p20_policy_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p20_policy_df_factoring = p20_policy_df.copy(deep=True)
      if annualize == True:
        p20_policy_df_factoring['incurred_amount'] = p20_policy_df_factoring['incurred_amount'] * (12 / p20_policy_df_factoring['data_month'])
        p20_policy_df_factoring['paid_amount']     = p20_policy_df_factoring['paid_amount']     * (12 / p20_policy_df_factoring['data_month'])
        p20_policy_df_factoring['no_of_claimants'] = p20_policy_df_factoring['no_of_claimants']
      if ibnr == True:
        p20_policy_df_factoring['incurred_amount'] = p20_policy_df_factoring['incurred_amount'] * (1 + p20_policy_df_factoring['ibnr'])
        p20_policy_df_factoring['paid_amount']     = p20_policy_df_factoring['paid_amount']     * (1 + p20_policy_df_factoring['ibnr'])
        p20_policy_df_factoring['no_of_claimants'] = p20_policy_df_factoring['no_of_claimants']

      p20_policy_df_diff = p20_policy_df_factoring[['incurred_amount', 'paid_amount', 'no_of_claimants']] - p20_policy_df[['incurred_amount', 'paid_amount', 'no_of_claimants']]
      temp_by = by.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p20_policy_df_diff = p20_policy_df_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p20_policy_df.reset_index().set_index(temp_by, inplace=True)
      p20_policy_df.loc[p20_policy_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claimants']] = p20_policy_df.loc[p20_policy_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claimants']].add(p20_policy_df_diff)
      temp_by = by.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p20_policy_df = p20_policy_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p20_policy_df['usage_ratio'] = p20_policy_df['paid_amount'] / p20_policy_df['incurred_amount']

    __p20_policy_group_col.remove('policy_id') if 'policy_id' in __p20_policy_group_col else __p20_policy_group_col
    p20_policy_df = p20_policy_df.reset_index().set_index(__p20_policy_group_col).dropna()
    p20_policy_df = p20_policy_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claimants']]

    if research_mode == True:
      p20_policy_df = p20_policy_df.reset_index().groupby(by=__p20_policy_group_col, dropna=False).sum()
      p20_policy_df['usage_ratio'] = p20_policy_df['paid_amount'] / p20_policy_df['incurred_amount']
      p20_policy_df = p20_policy_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claimants']]

    self.p20_policy = p20_policy_df
    return p20_policy_df

  def mcr_p20_benefit(self, by=None, benefit_type_order=['Hospital', 'Clinic', 'Dental', 'Optical', 'Maternity', 'Total'], annualize=False, ibnr=False, research_mode=False):

    by = by or []
    cols_df   = by + ['benefit_type', 'incurred_amount', 'paid_amount', 'claim_id', 'claimant']
    cols_cnt  = by + ['benefit_type', 'incur_date']
    grp_cols  = by.copy() + ['benefit_type']
    
    # 3) sum incurred/paid and count unique claimants
    p20 = (
        self.mcr_df[cols_df]
        .groupby(grp_cols, dropna=False)
        .agg(
            incurred_amount=('incurred_amount', 'sum'),
            paid_amount    =('paid_amount',     'sum'),
            no_of_claimants=('claimant',      'nunique'),
            no_of_claim_id=('claim_id', 'nunique'),
        )
    )
    
    # 4) build the “Total” rows by summing across benefit_type
    if by:
        # sum across the last level (benefit_type)
        totals = p20.groupby(level=list(range(len(by)))).sum()
        # expand the index to include ‘Total’ as the last level
        new_tuples = [tuple(idx) + ('Total',) for idx in totals.index]
        totals.index = pd.MultiIndex.from_tuples(new_tuples, names=p20.index.names)
        # append and sort
        p20 = pd.concat([p20, totals]).sort_index()
    else:
        # if no `by`, just sum everything and add one Total row
        tot = p20.sum()
        p20.loc[('Total',), :] = tot
    
    # 5) count claims
    p20_claims = (
        self.mcr_df[cols_cnt]
        .groupby(grp_cols, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    # override hospital to only count IP claims
    ip_claims = (
        self.mcr_df[self.mcr_df['benefit'].str.contains('Day Centre|Surgeon', case=False)][cols_cnt]
        .groupby(grp_cols, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    mask_hosp = p20_claims.index.get_level_values('benefit_type') == 'Hospital'
    p20_claims.loc[mask_hosp, 'no_of_cases'] = ip_claims['no_of_cases']
    
    # 6) if you added Total rows above, reindex p20_claims to match
    p20_claims = p20_claims.reindex(p20.index, fill_value=0)
    
    # 7) join, compute ratios
    p20 = p20.join(p20_claims['no_of_cases'])
    p20['usage_ratio']     = p20.paid_amount   / p20.incurred_amount
    p20['incurred_per_case'] = p20.incurred_amount / p20.no_of_cases.replace(0, pd.NA)
    p20['paid_per_case']     = p20.paid_amount     / p20.no_of_cases.replace(0, pd.NA)
    
    # 8) reorder benefit_type and columns
    p20 = p20.reindex(benefit_type_order, level='benefit_type')
    p20 = p20[[
        'incurred_amount', 'paid_amount', 'usage_ratio',
        'no_of_cases', 'incurred_per_case', 'paid_per_case',
        'no_of_claimants', 'no_of_claim_id'
    ]]
    
    if annualize == True or ibnr == True:
      p20 = pd.merge(p20, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p20_factoring = p20.copy(deep=True)
      if annualize == True:
        p20_factoring['incurred_amount'] = p20_factoring['incurred_amount'] * (12 / p20_factoring['data_month'])
        p20_factoring['paid_amount']     = p20_factoring['paid_amount']     * (12 / p20_factoring['data_month'])
        p20_factoring['no_of_cases']     = p20_factoring['no_of_cases']     * (12 / p20_factoring['data_month'])
        p20_factoring['no_of_claimants'] = p20_factoring['no_of_claimants']
        p20_factoring['no_of_claim_id']  = p20_factoring['no_of_claim_id']  * (12 / p20_factoring['data_month'])
      if ibnr == True:
        p20_factoring['incurred_amount'] = p20_factoring['incurred_amount'] * (1 + p20_factoring['ibnr'])
        p20_factoring['paid_amount']     = p20_factoring['paid_amount']     * (1 + p20_factoring['ibnr'])
        p20_factoring['no_of_cases']     = p20_factoring['no_of_cases']     * (1 + p20_factoring['ibnr'])
        p20_factoring['no_of_claimants'] = p20_factoring['no_of_claimants']
        p20_factoring['no_of_claim_id']  = p20_factoring['no_of_claim_id']  * (1 + p20_factoring['ibnr'])

      p20_diff = p20_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] - p20[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
      temp_by = grp_cols.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p20_diff = p20_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p20.reset_index().set_index(temp_by, inplace=True)
      p20.loc[p20.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] = p20.loc[p20.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].add(p20_diff)
      temp_by = grp_cols.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p20 = p20.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p20 = p20.reindex(benefit_type_order, level='benefit_type')
      p20['usage_ratio'] = p20['paid_amount'] / p20['incurred_amount']
      p20['incurred_per_case'] = p20['incurred_amount'] / p20['no_of_cases'].replace(0, pd.NA)
      p20['paid_per_case']     = p20['paid_amount']     / p20['no_of_cases'].replace(0, pd.NA)

    grp_cols.remove('policy_id') if 'policy_id' in grp_cols else grp_cols
    p20 = p20.reset_index().set_index(grp_cols).dropna(subset=['incurred_amount', 'paid_amount'])
    p20 = p20[['incurred_amount', 'paid_amount', 'usage_ratio',
                'no_of_cases', 'incurred_per_case', 'paid_per_case',
                'no_of_claimants', 'no_of_claim_id']]

    if research_mode == True:
      p20 = p20.reset_index().groupby(by=grp_cols, dropna=False).sum()
      p20['usage_ratio'] = p20['paid_amount'] / p20['incurred_amount']
      p20['incurred_per_case'] = p20['incurred_amount'] / p20['no_of_cases'].replace(0, pd.NA)
      p20['paid_per_case']     = p20['paid_amount']     / p20['no_of_cases'].replace(0, pd.NA)
      p20 = p20[['incurred_amount', 'paid_amount', 'usage_ratio',
                'no_of_cases', 'incurred_per_case', 'paid_per_case',
                'no_of_claimants', 'no_of_claim_id']]
      p20 = p20.reindex(benefit_type_order, level='benefit_type')

    self.p20_benefit = p20
    self.p20         = p20
    return p20

  def mcr_p20_benefit_dep_type(self, by=None, benefit_type_order=['Hospital', 'Clinic', 'Dental', 'Optical', 'Maternity', 'Total'], annualize=False, ibnr=False, research_mode=False):

    base_by = list(by) if isinstance(by, list) else ([] if by is None else [by])
    if 'dep_type' in base_by:
      self.p20_benefit_dep_type = pd.DataFrame()
      return self.p20_benefit_dep_type

    dep_by = base_by + ['dep_type']
    cols_df   = dep_by + ['benefit_type', 'incurred_amount', 'paid_amount', 'claim_id', 'claimant']
    cols_cnt  = dep_by + ['benefit_type', 'incur_date']
    grp_cols  = dep_by.copy() + ['benefit_type']

    p20_dep = (
        self.mcr_df[cols_df]
        .groupby(grp_cols, dropna=False)
        .agg(
            incurred_amount=('incurred_amount', 'sum'),
            paid_amount    =('paid_amount',     'sum'),
            no_of_claimants=('claimant',        'nunique'),
            no_of_claim_id =('claim_id',        'nunique'),
        )
    )

    if dep_by:
      totals = p20_dep.groupby(level=list(range(len(dep_by)))).sum()
      new_tuples = [tuple(idx) + ('Total',) for idx in totals.index]
      totals.index = pd.MultiIndex.from_tuples(new_tuples, names=p20_dep.index.names)
      p20_dep = pd.concat([p20_dep, totals]).sort_index()
    else:
      tot = p20_dep.sum()
      p20_dep.loc[('Total',), :] = tot

    p20_claims = (
      self.mcr_df[cols_cnt]
        .groupby(grp_cols, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )

    ip_claims = (
      self.mcr_df[self.mcr_df['benefit'].str.contains('Day Centre|Surgeon', case=False)][cols_cnt]
        .groupby(grp_cols, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p20_claims = p20_claims.reindex(p20_dep.index, fill_value=0)
    mask_hosp = p20_claims.index.get_level_values('benefit_type') == 'Hospital'
    p20_claims.loc[mask_hosp, 'no_of_cases'] = ip_claims['no_of_cases']

    p20_dep = p20_dep.join(p20_claims['no_of_cases'])
    p20_dep['usage_ratio'] = p20_dep['paid_amount'] / p20_dep['incurred_amount']
    p20_dep['incurred_per_case'] = p20_dep['incurred_amount'] / p20_dep['no_of_cases'].replace(0, pd.NA)
    p20_dep['paid_per_case']     = p20_dep['paid_amount']     / p20_dep['no_of_cases'].replace(0, pd.NA)

    p20_dep = p20_dep.reindex(benefit_type_order, level='benefit_type')
    p20_dep = p20_dep[[
        'incurred_amount', 'paid_amount', 'usage_ratio',
        'no_of_cases', 'incurred_per_case', 'paid_per_case',
        'no_of_claimants', 'no_of_claim_id'
    ]]

    if annualize == True or ibnr == True:
      p20_dep = pd.merge(p20_dep, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p20_dep_factoring = p20_dep.copy(deep=True)
      if annualize == True:
        p20_dep_factoring['incurred_amount'] = p20_dep_factoring['incurred_amount'] * (12 / p20_dep_factoring['data_month'])
        p20_dep_factoring['paid_amount']     = p20_dep_factoring['paid_amount']     * (12 / p20_dep_factoring['data_month'])
        p20_dep_factoring['no_of_cases']     = p20_dep_factoring['no_of_cases']     * (12 / p20_dep_factoring['data_month'])
        p20_dep_factoring['no_of_claimants'] = p20_dep_factoring['no_of_claimants']
        p20_dep_factoring['no_of_claim_id']  = p20_dep_factoring['no_of_claim_id']  * (12 / p20_dep_factoring['data_month'])
      if ibnr == True:
        p20_dep_factoring['incurred_amount'] = p20_dep_factoring['incurred_amount'] * (1 + p20_dep_factoring['ibnr'])
        p20_dep_factoring['paid_amount']     = p20_dep_factoring['paid_amount']     * (1 + p20_dep_factoring['ibnr'])
        p20_dep_factoring['no_of_cases']     = p20_dep_factoring['no_of_cases']     * (1 + p20_dep_factoring['ibnr'])
        p20_dep_factoring['no_of_claimants'] = p20_dep_factoring['no_of_claimants']
        p20_dep_factoring['no_of_claim_id']  = p20_dep_factoring['no_of_claim_id']  * (1 + p20_dep_factoring['ibnr'])

      value_cols = ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']
      p20_dep_diff = p20_dep_factoring[value_cols] - p20_dep[value_cols]
      temp_by = grp_cols.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p20_dep_diff = p20_dep_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p20_dep.reset_index().set_index(temp_by, inplace=True)
      p20_dep.loc[p20_dep.factoring == True, value_cols] = p20_dep.loc[p20_dep.factoring == True, value_cols].add(p20_dep_diff)
      temp_by = grp_cols.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p20_dep = p20_dep.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p20_dep = p20_dep.reindex(benefit_type_order, level='benefit_type')
      p20_dep['usage_ratio'] = p20_dep['paid_amount'] / p20_dep['incurred_amount']
      p20_dep['incurred_per_case'] = p20_dep['incurred_amount'] / p20_dep['no_of_cases'].replace(0, pd.NA)
      p20_dep['paid_per_case']     = p20_dep['paid_amount']     / p20_dep['no_of_cases'].replace(0, pd.NA)

    grp_cols.remove('policy_id') if 'policy_id' in grp_cols else grp_cols
    p20_dep = p20_dep.reset_index().set_index(grp_cols).dropna(subset=['incurred_amount', 'paid_amount'])
    p20_dep = p20_dep[[
        'incurred_amount', 'paid_amount', 'usage_ratio',
        'no_of_cases', 'incurred_per_case', 'paid_per_case',
        'no_of_claimants', 'no_of_claim_id'
    ]]

    if research_mode == True:
      p20_dep = p20_dep.reset_index().groupby(by=grp_cols, dropna=False).sum()
      p20_dep['usage_ratio'] = p20_dep['paid_amount'] / p20_dep['incurred_amount']
      p20_dep['incurred_per_case'] = p20_dep['incurred_amount'] / p20_dep['no_of_cases'].replace(0, pd.NA)
      p20_dep['paid_per_case']     = p20_dep['paid_amount']     / p20_dep['no_of_cases'].replace(0, pd.NA)
      p20_dep = p20_dep[[
        'incurred_amount', 'paid_amount', 'usage_ratio',
        'no_of_cases', 'incurred_per_case', 'paid_per_case',
        'no_of_claimants', 'no_of_claim_id'
      ]]
      p20_dep = p20_dep.reindex(benefit_type_order, level='benefit_type')

    self.p20_benefit_dep_type = p20_dep
    return p20_dep

  def mcr_p20_panel(self, by=None, annualize=False, ibnr=False, research_mode=False):


    __p20_panel_df_col = by + ['panel', 'incurred_amount', 'paid_amount']
    __p20_panel_group_col = by.copy() + ['panel']
    
    p20_panel_df = self.mcr_df[__p20_panel_df_col].groupby(by=__p20_panel_group_col, dropna=False).sum()
    p20_panel_df['usage_ratio'] = p20_panel_df['paid_amount'] / p20_panel_df['incurred_amount']
    p20_panel_df = p20_panel_df.unstack().stack(dropna=False)

    if annualize == True or ibnr == True:
      p20_panel_df = pd.merge(p20_panel_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p20_panel_factoring = p20_panel_df.copy(deep=True)
      if annualize == True:
        p20_panel_factoring['incurred_amount'] = p20_panel_factoring['incurred_amount'] * (12 / p20_panel_factoring['data_month'])
        p20_panel_factoring['paid_amount']     = p20_panel_factoring['paid_amount']     * (12 / p20_panel_factoring['data_month'])
      if ibnr == True:
        p20_panel_factoring['incurred_amount'] = p20_panel_factoring['incurred_amount'] * (1 + p20_panel_factoring['ibnr'])
        p20_panel_factoring['paid_amount']     = p20_panel_factoring['paid_amount']     * (1 + p20_panel_factoring['ibnr'])
      p20_panel_diff = p20_panel_factoring[['incurred_amount', 'paid_amount',]] - p20_panel_df[['incurred_amount', 'paid_amount',]]
      temp_by = __p20_panel_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p20_panel_diff = p20_panel_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount',]].groupby(by=temp_by, dropna=False).sum()
      p20_panel_df.reset_index().set_index(temp_by, inplace=True)
      p20_panel_df.loc[p20_panel_df.factoring == True, ['incurred_amount', 'paid_amount',]] = p20_panel_df.loc[p20_panel_df.factoring == True, ['incurred_amount', 'paid_amount',]].add(p20_panel_diff)
      temp_by = __p20_panel_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p20_panel_df = p20_panel_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount',]].groupby(by=temp_by, dropna=False).sum()
      p20_panel_df['usage_ratio'] = p20_panel_df['paid_amount'] / p20_panel_df['incurred_amount']

    __p20_panel_group_col.remove('policy_id') if 'policy_id' in __p20_panel_group_col else __p20_panel_group_col
    p20_panel_df = p20_panel_df.reset_index().set_index(__p20_panel_group_col)
    p20_panel_df = p20_panel_df[['incurred_amount', 'paid_amount', 'usage_ratio',]]

    if research_mode == True:
      p20_panel_df = p20_panel_df.reset_index().groupby(by=__p20_panel_group_col, dropna=False).sum()
      p20_panel_df['usage_ratio'] = p20_panel_df['paid_amount'] / p20_panel_df['incurred_amount']
      p20_panel_df = p20_panel_df[['incurred_amount', 'paid_amount', 'usage_ratio',]]
    self.p20_panel = p20_panel_df
    return p20_panel_df
  
  def mcr_p20_panel_benefit(self, by=None, benefit_type_order=['Hospital', 'Clinic', 'Dental', 'Optical', 'Maternity', 'Total'], annualize=False, ibnr=False, research_mode=False):

    by = by or []
    # columns and grouping keys
    cols_df   = by + ['panel', 'benefit_type', 'incurred_amount', 'paid_amount', 'claim_id', 'claimant']
    cols_cnt  = by + ['panel', 'benefit_type', 'incur_date']
    grp_cols  = by.copy() + ['panel', 'benefit_type']

    # 1) sum incurred/paid and count unique claimants and claim_id
    p20 = (
        self.mcr_df[cols_df]
        .groupby(grp_cols, dropna=False)
        .agg(
            incurred_amount=('incurred_amount', 'sum'),
            paid_amount    =('paid_amount',     'sum'),
            no_of_claimants=('claimant',        'nunique'),
            no_of_claim_id =('claim_id',        'nunique'),
        )
    )

    # 2) build Total across benefit_type within each (by..., panel)
    if by or True:  # always true because we also have panel dim
        # sum across all levels except the last (benefit_type)
        totals = p20.groupby(level=list(range(len(by)+1))).sum()
        new_tuples = [tuple(idx) + ('Total',) for idx in totals.index]
        totals.index = pd.MultiIndex.from_tuples(new_tuples, names=p20.index.names)
        p20 = pd.concat([p20, totals]).sort_index()

    # 3) count cases (claim lines)
    p20_cases = (
        self.mcr_df[cols_cnt]
        .groupby(grp_cols, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )

    # override hospital cases to IP-only counts (Day Centre/Surgeon), as in mcr_p20_benefit
    ip_cases = (
        self.mcr_df[self.mcr_df['benefit'].str.contains('Day Centre|Surgeon', case=False)][cols_cnt]
        .groupby(grp_cols, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    # align to full index and replace hospital rows
    p20_cases = p20_cases.reindex(p20.index, fill_value=0)
    ip_cases  = ip_cases.reindex(p20.index, fill_value=0)
    mask_hosp = p20_cases.index.get_level_values('benefit_type') == 'Hospital'
    p20_cases.loc[mask_hosp, 'no_of_cases'] = ip_cases.loc[mask_hosp, 'no_of_cases']

    # 4) join, compute ratios and per-case metrics
    p20 = p20.join(p20_cases['no_of_cases'])
    p20['usage_ratio'] = p20['paid_amount'] / p20['incurred_amount']
    p20['incurred_per_case'] = p20['incurred_amount'] / p20['no_of_cases'].replace(0, pd.NA)
    p20['paid_per_case']     = p20['paid_amount']     / p20['no_of_cases'].replace(0, pd.NA)

    # 5) reorder the benefit_type level and keep empty groups
    p20 = p20.unstack().stack(dropna=False)
    p20 = p20.reindex(benefit_type_order, level='benefit_type')
    p20 = p20[[
        'incurred_amount', 'paid_amount', 'usage_ratio',
        'no_of_cases', 'incurred_per_case', 'paid_per_case',
        'no_of_claimants', 'no_of_claim_id'
    ]]

    # 6) factoring (annualize/ibnr) mirroring mcr_p20_benefit
    if annualize == True or ibnr == True:
      p20 = pd.merge(p20, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p20_factoring = p20.copy(deep=True)
      if annualize == True:
        p20_factoring['incurred_amount'] = p20_factoring['incurred_amount'] * (12 / p20_factoring['data_month'])
        p20_factoring['paid_amount']     = p20_factoring['paid_amount']     * (12 / p20_factoring['data_month'])
        p20_factoring['no_of_cases']     = p20_factoring['no_of_cases']     * (12 / p20_factoring['data_month'])
        p20_factoring['no_of_claimants'] = p20_factoring['no_of_claimants']
        p20_factoring['no_of_claim_id']  = p20_factoring['no_of_claim_id']  * (12 / p20_factoring['data_month'])
      if ibnr == True:
        p20_factoring['incurred_amount'] = p20_factoring['incurred_amount'] * (1 + p20_factoring['ibnr'])
        p20_factoring['paid_amount']     = p20_factoring['paid_amount']     * (1 + p20_factoring['ibnr'])
        p20_factoring['no_of_cases']     = p20_factoring['no_of_cases']     * (1 + p20_factoring['ibnr'])
        p20_factoring['no_of_claimants'] = p20_factoring['no_of_claimants']
        p20_factoring['no_of_claim_id']  = p20_factoring['no_of_claim_id']  * (1 + p20_factoring['ibnr'])

      p20_diff = p20_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] - p20[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
      temp_by = grp_cols.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p20_diff = p20_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p20.reset_index().set_index(temp_by, inplace=True)
      p20.loc[p20.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] = p20.loc[p20.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].add(p20_diff)
      temp_by = grp_cols.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p20 = p20.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p20 = p20.reindex(benefit_type_order, level='benefit_type')
      p20['usage_ratio'] = p20['paid_amount'] / p20['incurred_amount']
      p20['incurred_per_case'] = p20['incurred_amount'] / p20['no_of_cases'].replace(0, pd.NA)
      p20['paid_per_case']     = p20['paid_amount']     / p20['no_of_cases'].replace(0, pd.NA)

    # 7) clean index per previous convention and finalize columns
    grp_cols.remove('policy_id') if 'policy_id' in grp_cols else grp_cols
    p20 = p20.reset_index().set_index(grp_cols).dropna(subset=['incurred_amount', 'paid_amount'])
    p20 = p20[['incurred_amount', 'paid_amount', 'usage_ratio',
                'no_of_cases', 'incurred_per_case', 'paid_per_case',
                'no_of_claimants', 'no_of_claim_id']]

    if research_mode == True:
      p20 = p20.reset_index().groupby(by=grp_cols, dropna=False).sum()
      p20['usage_ratio'] = p20['paid_amount'] / p20['incurred_amount']
      p20['incurred_per_case'] = p20['incurred_amount'] / p20['no_of_cases'].replace(0, pd.NA)
      p20['paid_per_case']     = p20['paid_amount']     / p20['no_of_cases'].replace(0, pd.NA)
      p20 = p20[['incurred_amount', 'paid_amount', 'usage_ratio',
                'no_of_cases', 'incurred_per_case', 'paid_per_case',
                'no_of_claimants', 'no_of_claim_id']]
      p20 = p20.reindex(benefit_type_order, level='benefit_type')

    self.p20_panel_benefit = p20
    return p20
  
  def mcr_p20_day_procedure(self, by=None, annualize=False, ibnr=False, research_mode=False):

    base_by = list(by) if isinstance(by, list) else ([] if by is None else [by])
    cols_df   = base_by + ['day_procedure_flag', 'incurred_amount', 'paid_amount', 'claim_id']
    cols_cnt  = base_by + ['day_procedure_flag', 'incur_date']
    grp_cols  = base_by.copy() + ['day_procedure_flag']

    p20_dp_df = (
      self.mcr_df[cols_df]
      .groupby(grp_cols, dropna=False)
      .agg(
        incurred_amount=('incurred_amount', 'sum'),
        paid_amount    =('paid_amount',     'sum'),
        no_of_claim_id =('claim_id',        'nunique'),
      )
    )

    p20_dp_cases = (
      self.mcr_df[cols_cnt]
      .groupby(grp_cols, dropna=False)
      .count()
      .rename(columns={'incur_date': 'no_of_cases'})
    )
    ip_mask = self.mcr_df['benefit'].str.contains('Day Centre|Surgeon', case=False, na=False)
    p20_dp_case_overrides = (
      self.mcr_df[ip_mask][cols_cnt]
      .groupby(grp_cols, dropna=False)
      .count()
      .rename(columns={'incur_date': 'no_of_cases'})
    )
    p20_dp_cases = p20_dp_cases.reindex(p20_dp_df.index, fill_value=0)
    p20_dp_case_overrides = p20_dp_case_overrides.reindex(p20_dp_df.index, fill_value=0)
    if 'day_procedure_flag' in p20_dp_cases.index.names:
      flags = p20_dp_cases.index.get_level_values('day_procedure_flag')
      override_mask = flags.isin(["day_procedure","hospital_have_surgery", "hospital_no_surgery", "no_hospital_no_surgery"])
      p20_dp_cases.loc[override_mask, 'no_of_cases'] = p20_dp_case_overrides.loc[override_mask, 'no_of_cases']
    p20_dp_df = p20_dp_df.join(p20_dp_cases['no_of_cases'])

    p20_dp_df['usage_ratio'] = p20_dp_df['paid_amount'] / p20_dp_df['incurred_amount'].replace(0, pd.NA)
    p20_dp_df = p20_dp_df.unstack().stack(dropna=False)

    if annualize == True or ibnr == True:
      p20_dp_df = pd.merge(p20_dp_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p20_dp_factoring = p20_dp_df.copy(deep=True)
      if annualize == True:
        p20_dp_factoring['incurred_amount'] = p20_dp_factoring['incurred_amount'] * (12 / p20_dp_factoring['data_month'])
        p20_dp_factoring['paid_amount']     = p20_dp_factoring['paid_amount']     * (12 / p20_dp_factoring['data_month'])
        p20_dp_factoring['no_of_cases']     = p20_dp_factoring['no_of_cases']     * (12 / p20_dp_factoring['data_month'])
        p20_dp_factoring['no_of_claim_id']  = p20_dp_factoring['no_of_claim_id']  * (12 / p20_dp_factoring['data_month'])
      if ibnr == True:
        p20_dp_factoring['incurred_amount'] = p20_dp_factoring['incurred_amount'] * (1 + p20_dp_factoring['ibnr'])
        p20_dp_factoring['paid_amount']     = p20_dp_factoring['paid_amount']     * (1 + p20_dp_factoring['ibnr'])
        p20_dp_factoring['no_of_cases']     = p20_dp_factoring['no_of_cases']     * (1 + p20_dp_factoring['ibnr'])
        p20_dp_factoring['no_of_claim_id']  = p20_dp_factoring['no_of_claim_id']  * (1 + p20_dp_factoring['ibnr'])

      value_cols = ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']
      p20_panel_diff = p20_dp_factoring[value_cols] - p20_dp_df[value_cols]
      temp_by = grp_cols.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p20_panel_diff = p20_panel_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p20_dp_df.reset_index().set_index(temp_by, inplace=True)
      p20_dp_df.loc[p20_dp_df.factoring == True, value_cols] = p20_dp_df.loc[p20_dp_df.factoring == True, value_cols].add(p20_panel_diff)
      temp_by = grp_cols.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p20_dp_df = p20_dp_df.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p20_dp_df['usage_ratio'] = p20_dp_df['paid_amount'] / p20_dp_df['incurred_amount'].replace(0, pd.NA)

    final_index_cols = [col for col in grp_cols if col != 'policy_id']
    p20_dp_df = p20_dp_df.reset_index().set_index(final_index_cols)
    p20_dp_df = p20_dp_df.unstack().stack(dropna=False)
    p20_dp_df = p20_dp_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'no_of_claim_id']]

    if research_mode == True:
      p20_dp_df = p20_dp_df.reset_index().groupby(by=final_index_cols, dropna=False).sum()
      p20_dp_df['usage_ratio'] = p20_dp_df['paid_amount'] / p20_dp_df['incurred_amount'].replace(0, pd.NA)
      p20_dp_df = p20_dp_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'no_of_claim_id']]
    self.p20_dp = p20_dp_df
    return p20_dp_df

  def mcr_p20_day_procedure_class(self, by=None, annualize=False, ibnr=False, research_mode=False):

    base_by = list(by) if isinstance(by, list) else ([] if by is None else [by])
    cols_df   = base_by + ['class', 'day_procedure_flag', 'incurred_amount', 'paid_amount', 'claim_id']
    cols_cnt  = base_by + ['class', 'day_procedure_flag', 'incur_date']
    grp_cols  = base_by.copy() + ['class', 'day_procedure_flag']

    p20_dp_class_df = (
      self.mcr_df[cols_df]
      .groupby(grp_cols, dropna=False)
      .agg(
        incurred_amount=('incurred_amount', 'sum'),
        paid_amount    =('paid_amount',     'sum'),
        no_of_claim_id =('claim_id',        'nunique'),
      )
    )

    p20_dp_class_cases = (
      self.mcr_df[cols_cnt]
      .groupby(grp_cols, dropna=False)
      .count()
      .rename(columns={'incur_date': 'no_of_cases'})
    )
    ip_mask = self.mcr_df['benefit'].str.contains('Day Centre|Surgeon', case=False, na=False)
    p20_dp_class_overrides = (
      self.mcr_df[ip_mask][cols_cnt]
      .groupby(grp_cols, dropna=False)
      .count()
      .rename(columns={'incur_date': 'no_of_cases'})
    )
    p20_dp_class_cases = p20_dp_class_cases.reindex(p20_dp_class_df.index, fill_value=0)
    p20_dp_class_overrides = p20_dp_class_overrides.reindex(p20_dp_class_df.index, fill_value=0)
    if 'day_procedure_flag' in p20_dp_class_cases.index.names:
      flags = p20_dp_class_cases.index.get_level_values('day_procedure_flag')
      override_mask = flags.isin(["day_procedure","hospital_have_surgery", "hospital_no_surgery", "no_hospital_no_surgery"])
      p20_dp_class_cases.loc[override_mask, 'no_of_cases'] = p20_dp_class_overrides.loc[override_mask, 'no_of_cases']
    p20_dp_class_df = p20_dp_class_df.join(p20_dp_class_cases['no_of_cases'])

    p20_dp_class_df['usage_ratio'] = p20_dp_class_df['paid_amount'] / p20_dp_class_df['incurred_amount'].replace(0, pd.NA)
    p20_dp_class_df = p20_dp_class_df.unstack().stack(dropna=False)

    if annualize == True or ibnr == True:
      p20_dp_class_df = pd.merge(p20_dp_class_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p20_dp_class_factoring = p20_dp_class_df.copy(deep=True)
      if annualize == True:
        p20_dp_class_factoring['incurred_amount'] = p20_dp_class_factoring['incurred_amount'] * (12 / p20_dp_class_factoring['data_month'])
        p20_dp_class_factoring['paid_amount']     = p20_dp_class_factoring['paid_amount']     * (12 / p20_dp_class_factoring['data_month'])
        p20_dp_class_factoring['no_of_cases']     = p20_dp_class_factoring['no_of_cases']     * (12 / p20_dp_class_factoring['data_month'])
        p20_dp_class_factoring['no_of_claim_id']  = p20_dp_class_factoring['no_of_claim_id']  * (12 / p20_dp_class_factoring['data_month'])
      if ibnr == True:
        p20_dp_class_factoring['incurred_amount'] = p20_dp_class_factoring['incurred_amount'] * (1 + p20_dp_class_factoring['ibnr'])
        p20_dp_class_factoring['paid_amount']     = p20_dp_class_factoring['paid_amount']     * (1 + p20_dp_class_factoring['ibnr'])
        p20_dp_class_factoring['no_of_cases']     = p20_dp_class_factoring['no_of_cases']     * (1 + p20_dp_class_factoring['ibnr'])
        p20_dp_class_factoring['no_of_claim_id']  = p20_dp_class_factoring['no_of_claim_id']  * (1 + p20_dp_class_factoring['ibnr'])

      value_cols = ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']
      p20_dp_class_diff = p20_dp_class_factoring[value_cols] - p20_dp_class_df[value_cols]
      temp_by = grp_cols.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p20_dp_class_diff = p20_dp_class_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p20_dp_class_df.reset_index().set_index(temp_by, inplace=True)
      p20_dp_class_df.loc[p20_dp_class_df.factoring == True, value_cols] = p20_dp_class_df.loc[p20_dp_class_df.factoring == True, value_cols].add(p20_dp_class_diff)
      temp_by = grp_cols.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p20_dp_class_df = p20_dp_class_df.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p20_dp_class_df['usage_ratio'] = p20_dp_class_df['paid_amount'] / p20_dp_class_df['incurred_amount'].replace(0, pd.NA)

    final_index_cols = [col for col in grp_cols if col != 'policy_id']
    p20_dp_class_df = p20_dp_class_df.reset_index().set_index(final_index_cols)
    p20_dp_class_df = p20_dp_class_df.unstack().stack(dropna=False)
    p20_dp_class_df = p20_dp_class_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'no_of_claim_id']]

    if research_mode == True:
      p20_dp_class_df = p20_dp_class_df.reset_index().groupby(by=final_index_cols, dropna=False).sum()
      p20_dp_class_df['usage_ratio'] = p20_dp_class_df['paid_amount'] / p20_dp_class_df['incurred_amount'].replace(0, pd.NA)
      p20_dp_class_df = p20_dp_class_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'no_of_claim_id']]
    self.p20_dp_class = p20_dp_class_df
    return p20_dp_class_df

  def mcr_p21_class(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p21_df_col = by + ['class', 'incurred_amount', 'paid_amount', 'claimant']
    __p21_group_col = by + ['class']

    # self.mcr_df['year'] = self.mcr_df.policy_start_date.dt.year
    p21_class_df = self.mcr_df[__p21_df_col].groupby(by=__p21_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique'}).rename(columns={'claimant': 'no_of_claimants'})
    p21_class_df['usage_ratio'] = p21_class_df['paid_amount'] / p21_class_df['incurred_amount']
    p21_class_df = p21_class_df.unstack().stack(dropna=False)

    if annualize == True or ibnr == True:
      p21_class_df = pd.merge(p21_class_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p21_class_df_factoring = p21_class_df.copy(deep=True)
      if annualize == True:
        p21_class_df_factoring['incurred_amount'] = p21_class_df_factoring['incurred_amount'] * (12 / p21_class_df_factoring['data_month'])
        p21_class_df_factoring['paid_amount']     = p21_class_df_factoring['paid_amount']     * (12 / p21_class_df_factoring['data_month'])
        p21_class_df_factoring['no_of_claimants'] = p21_class_df_factoring['no_of_claimants']
      if ibnr == True:
        p21_class_df_factoring['incurred_amount'] = p21_class_df_factoring['incurred_amount'] * (1 + p21_class_df_factoring['ibnr'])
        p21_class_df_factoring['paid_amount']     = p21_class_df_factoring['paid_amount']     * (1 + p21_class_df_factoring['ibnr'])
        p21_class_df_factoring['no_of_claimants'] = p21_class_df_factoring['no_of_claimants']

      p21_diff = p21_class_df_factoring[['incurred_amount', 'paid_amount', 'no_of_claimants']] - p21_class_df[['incurred_amount', 'paid_amount', 'no_of_claimants']]
      temp_by = __p21_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p21_diff = p21_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p21_class_df.reset_index().set_index(temp_by, inplace=True)
      p21_class_df.loc[p21_class_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claimants']] = p21_class_df.loc[p21_class_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claimants']].add(p21_diff)
      temp_by = __p21_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p21_class_df = p21_class_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p21_class_df['usage_ratio'] = p21_class_df['paid_amount'] / p21_class_df['incurred_amount']
      p21_class_df = p21_class_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claimants']]

    __p21_group_col.remove('policy_id') if 'policy_id' in __p21_group_col else __p21_group_col
    p21_class_df = p21_class_df.reset_index().set_index(__p21_group_col)
    p21_class_df = p21_class_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claimants']]
    if research_mode == True:
      p21_class_df = p21_class_df.reset_index().groupby(by=__p21_group_col, dropna=False).sum()
      p21_class_df['usage_ratio'] = p21_class_df['paid_amount'] / p21_class_df['incurred_amount']
      p21_class_df = p21_class_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claimants']]
    self.p21_class = p21_class_df
    self.p21 = p21_class_df
    return p21_class_df

  def mcr_p22_class_benefit(self, by=None, benefit_type_order=None, annualize=False, ibnr=False, research_mode=False):

    by = by or []
    __p22_df_col = by + ['class', 'benefit_type', 'incurred_amount', 'paid_amount', 'claim_id', 'claimant']
    __p22_cnt_col = by + ['class', 'benefit_type', 'incur_date']
    __p22_group_col = by.copy() + ['class', 'benefit_type']

    # 1) sum incurred/paid and count unique claimants and claim ids
    p22_class_benefit_df = (
      self.mcr_df[__p22_df_col]
      .groupby(by=__p22_group_col, dropna=False)
      .agg(
        incurred_amount=('incurred_amount', 'sum'),
        paid_amount    =('paid_amount',     'sum'),
        no_of_claimants=('claimant',        'nunique'),
        no_of_claim_id =('claim_id',        'nunique'),
      )
    )

    # 2) add Total across benefit_type within each (by..., class)
    base_levels = list(range(len(by) + 1))  # by... + class
    if len(p22_class_benefit_df.index.names) > 0:
      totals = p22_class_benefit_df.groupby(level=base_levels).sum()
      new_tuples = [tuple(idx) + ('Total',) for idx in totals.index]
      totals.index = pd.MultiIndex.from_tuples(new_tuples, names=p22_class_benefit_df.index.names)
      p22_class_benefit_df = pd.concat([p22_class_benefit_df, totals]).sort_index()

    # 3) count cases and override Hospital with IP-specific counts (Day Centre|Surgeon)
    p22_cases = (
      self.mcr_df[__p22_cnt_col]
      .groupby(by=__p22_group_col, dropna=False)
      .count()
      .rename(columns={'incur_date': 'no_of_cases'})
    )

    ip_cases = (
      self.mcr_df[self.mcr_df['benefit'].str.contains('Day Centre|Surgeon', case=False)][__p22_cnt_col]
      .groupby(by=__p22_group_col, dropna=False)
      .count()
      .rename(columns={'incur_date': 'no_of_cases'})
    )

    # align cases to the main index (includes 'Total')
    p22_cases = p22_cases.reindex(p22_class_benefit_df.index, fill_value=0)
    mask_hosp = p22_cases.index.get_level_values('benefit_type') == 'Hospital'
    p22_cases.loc[mask_hosp, 'no_of_cases'] = ip_cases['no_of_cases']

    # 4) join and compute ratios
    p22_class_benefit_df = p22_class_benefit_df.join(p22_cases['no_of_cases'])
    p22_class_benefit_df['usage_ratio'] = p22_class_benefit_df['paid_amount'] / p22_class_benefit_df['incurred_amount']
    p22_class_benefit_df['incurred_per_case'] = p22_class_benefit_df['incurred_amount'] / p22_class_benefit_df['no_of_cases'].replace(0, pd.NA)
    p22_class_benefit_df['paid_per_case']     = p22_class_benefit_df['paid_amount']     / p22_class_benefit_df['no_of_cases'].replace(0, pd.NA)

    # 5) reorder benefit_type if order provided
    if benefit_type_order is not None:
      p22_class_benefit_df = p22_class_benefit_df.reindex(benefit_type_order, level='benefit_type')

    # 6) factoring / ibnr adjustments
    if annualize == True or ibnr == True:
      p22_class_benefit_df = pd.merge(p22_class_benefit_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p22_factoring = p22_class_benefit_df.copy(deep=True)
      if annualize == True:
        p22_factoring['incurred_amount'] = p22_factoring['incurred_amount'] * (12 / p22_factoring['data_month'])
        p22_factoring['paid_amount']     = p22_factoring['paid_amount']     * (12 / p22_factoring['data_month'])
        p22_factoring['no_of_cases']     = p22_factoring['no_of_cases']     * (12 / p22_factoring['data_month'])
        p22_factoring['no_of_claimants'] = p22_factoring['no_of_claimants']
        p22_factoring['no_of_claim_id']  = p22_factoring['no_of_claim_id']  * (12 / p22_factoring['data_month'])
      if ibnr == True:
        p22_factoring['incurred_amount'] = p22_factoring['incurred_amount'] * (1 + p22_factoring['ibnr'])
        p22_factoring['paid_amount']     = p22_factoring['paid_amount']     * (1 + p22_factoring['ibnr'])
        p22_factoring['no_of_cases']     = p22_factoring['no_of_cases']     * (1 + p22_factoring['ibnr'])
        p22_factoring['no_of_claimants'] = p22_factoring['no_of_claimants']
        p22_factoring['no_of_claim_id']  = p22_factoring['no_of_claim_id']  * (1 + p22_factoring['ibnr'])

      value_cols = ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']
      p22_diff = p22_factoring[value_cols] - p22_class_benefit_df[value_cols]
      temp_by = __p22_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p22_diff = p22_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p22_class_benefit_df.reset_index().set_index(temp_by, inplace=True)
      p22_class_benefit_df.loc[p22_class_benefit_df.factoring == True, value_cols] = p22_class_benefit_df.loc[p22_class_benefit_df.factoring == True, value_cols].add(p22_diff)
      temp_by = __p22_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p22_class_benefit_df = p22_class_benefit_df.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      if benefit_type_order is not None:
        p22_class_benefit_df = p22_class_benefit_df.reindex(benefit_type_order, level='benefit_type')
      p22_class_benefit_df['usage_ratio'] = p22_class_benefit_df['paid_amount'] / p22_class_benefit_df['incurred_amount']
      p22_class_benefit_df['incurred_per_case'] = p22_class_benefit_df['incurred_amount'] / p22_class_benefit_df['no_of_cases'].replace(0, pd.NA)
      p22_class_benefit_df['paid_per_case']     = p22_class_benefit_df['paid_amount']     / p22_class_benefit_df['no_of_cases'].replace(0, pd.NA)

    # 7) finalize index and columns
    __p22_group_col.remove('policy_id') if 'policy_id' in __p22_group_col else __p22_group_col
    p22_class_benefit_df = p22_class_benefit_df.reset_index().set_index(__p22_group_col).dropna(subset=['incurred_amount', 'paid_amount'])
    p22_class_benefit_df = p22_class_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio',
                                                'no_of_cases', 'incurred_per_case', 'paid_per_case',
                                                'no_of_claimants', 'no_of_claim_id']]

    if research_mode == True:
      p22_class_benefit_df = p22_class_benefit_df.reset_index().groupby(by=__p22_group_col, dropna=False).sum()
      p22_class_benefit_df['usage_ratio'] = p22_class_benefit_df['paid_amount'] / p22_class_benefit_df['incurred_amount']
      p22_class_benefit_df['incurred_per_case'] = p22_class_benefit_df['incurred_amount'] / p22_class_benefit_df['no_of_cases'].replace(0, pd.NA)
      p22_class_benefit_df['paid_per_case']     = p22_class_benefit_df['paid_amount']     / p22_class_benefit_df['no_of_cases'].replace(0, pd.NA)
      p22_class_benefit_df = p22_class_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio',
                                                  'no_of_cases', 'incurred_per_case', 'paid_per_case',
                                                  'no_of_claimants', 'no_of_claim_id']]
      if benefit_type_order is not None:
        p22_class_benefit_df = p22_class_benefit_df.reindex(benefit_type_order, level='benefit_type')

    self.p22_class_benefit = p22_class_benefit_df
    self.p22 = p22_class_benefit_df
    return p22_class_benefit_df

  def mcr_p22_class_benefit_dep_type(self, by=None, benefit_type_order=None, annualize=False, ibnr=False, research_mode=False):

    base_by = list(by) if isinstance(by, list) else ([] if by is None else [by])
    if 'dep_type' in base_by:
      self.p22_class_benefit_dep_type = pd.DataFrame()
      return self.p22_class_benefit_dep_type

    dep_by = base_by + ['dep_type']
    __p22_df_col = dep_by + ['class', 'benefit_type', 'incurred_amount', 'paid_amount', 'claim_id', 'claimant']
    __p22_cnt_col = dep_by + ['class', 'benefit_type', 'incur_date']
    __p22_group_col = dep_by.copy() + ['class', 'benefit_type']

    p22_dep_df = (
      self.mcr_df[__p22_df_col]
      .groupby(by=__p22_group_col, dropna=False)
      .agg(
        incurred_amount=('incurred_amount', 'sum'),
        paid_amount    =('paid_amount',     'sum'),
        no_of_claimants=('claimant',        'nunique'),
        no_of_claim_id =('claim_id',        'nunique'),
      )
    )

    base_levels = list(range(len(dep_by) + 1))
    if len(p22_dep_df.index.names) > 0:
      totals = p22_dep_df.groupby(level=base_levels).sum()
      new_tuples = [tuple(idx) + ('Total',) for idx in totals.index]
      totals.index = pd.MultiIndex.from_tuples(new_tuples, names=p22_dep_df.index.names)
      p22_dep_df = pd.concat([p22_dep_df, totals]).sort_index()

    p22_cases = (
      self.mcr_df[__p22_cnt_col]
      .groupby(by=__p22_group_col, dropna=False)
      .count()
      .rename(columns={'incur_date': 'no_of_cases'})
    )

    ip_cases = (
      self.mcr_df[self.mcr_df['benefit'].str.contains('Day Centre|Surgeon', case=False)][__p22_cnt_col]
      .groupby(by=__p22_group_col, dropna=False)
      .count()
      .rename(columns={'incur_date': 'no_of_cases'})
    )

    p22_cases = p22_cases.reindex(p22_dep_df.index, fill_value=0)
    mask_hosp = p22_cases.index.get_level_values('benefit_type') == 'Hospital'
    p22_cases.loc[mask_hosp, 'no_of_cases'] = ip_cases['no_of_cases']

    p22_dep_df = p22_dep_df.join(p22_cases['no_of_cases'])
    p22_dep_df['usage_ratio'] = p22_dep_df['paid_amount'] / p22_dep_df['incurred_amount']
    p22_dep_df['incurred_per_case'] = p22_dep_df['incurred_amount'] / p22_dep_df['no_of_cases'].replace(0, pd.NA)
    p22_dep_df['paid_per_case']     = p22_dep_df['paid_amount']     / p22_dep_df['no_of_cases'].replace(0, pd.NA)

    if benefit_type_order is not None:
      p22_dep_df = p22_dep_df.reindex(benefit_type_order, level='benefit_type')

    if annualize == True or ibnr == True:
      p22_dep_df = pd.merge(p22_dep_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p22_dep_factoring = p22_dep_df.copy(deep=True)
      if annualize == True:
        p22_dep_factoring['incurred_amount'] = p22_dep_factoring['incurred_amount'] * (12 / p22_dep_factoring['data_month'])
        p22_dep_factoring['paid_amount']     = p22_dep_factoring['paid_amount']     * (12 / p22_dep_factoring['data_month'])
        p22_dep_factoring['no_of_cases']     = p22_dep_factoring['no_of_cases']     * (12 / p22_dep_factoring['data_month'])
        p22_dep_factoring['no_of_claimants'] = p22_dep_factoring['no_of_claimants']
        p22_dep_factoring['no_of_claim_id']  = p22_dep_factoring['no_of_claim_id']  * (12 / p22_dep_factoring['data_month'])
      if ibnr == True:
        p22_dep_factoring['incurred_amount'] = p22_dep_factoring['incurred_amount'] * (1 + p22_dep_factoring['ibnr'])
        p22_dep_factoring['paid_amount']     = p22_dep_factoring['paid_amount']     * (1 + p22_dep_factoring['ibnr'])
        p22_dep_factoring['no_of_cases']     = p22_dep_factoring['no_of_cases']     * (1 + p22_dep_factoring['ibnr'])
        p22_dep_factoring['no_of_claimants'] = p22_dep_factoring['no_of_claimants']
        p22_dep_factoring['no_of_claim_id']  = p22_dep_factoring['no_of_claim_id']  * (1 + p22_dep_factoring['ibnr'])

      value_cols = ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']
      p22_dep_diff = p22_dep_factoring[value_cols] - p22_dep_df[value_cols]
      temp_by = __p22_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p22_dep_diff = p22_dep_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p22_dep_df.reset_index().set_index(temp_by, inplace=True)
      p22_dep_df.loc[p22_dep_df.factoring == True, value_cols] = p22_dep_df.loc[p22_dep_df.factoring == True, value_cols].add(p22_dep_diff)
      temp_by = __p22_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p22_dep_df = p22_dep_df.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      if benefit_type_order is not None:
        p22_dep_df = p22_dep_df.reindex(benefit_type_order, level='benefit_type')
      p22_dep_df['usage_ratio'] = p22_dep_df['paid_amount'] / p22_dep_df['incurred_amount']
      p22_dep_df['incurred_per_case'] = p22_dep_df['incurred_amount'] / p22_dep_df['no_of_cases'].replace(0, pd.NA)
      p22_dep_df['paid_per_case']     = p22_dep_df['paid_amount']     / p22_dep_df['no_of_cases'].replace(0, pd.NA)

    __p22_group_col.remove('policy_id') if 'policy_id' in __p22_group_col else __p22_group_col
    p22_dep_df = p22_dep_df.reset_index().set_index(__p22_group_col).dropna(subset=['incurred_amount', 'paid_amount'])
    p22_dep_df = p22_dep_df[['incurred_amount', 'paid_amount', 'usage_ratio',
                             'no_of_cases', 'incurred_per_case', 'paid_per_case',
                             'no_of_claimants', 'no_of_claim_id']]

    if research_mode == True:
      p22_dep_df = p22_dep_df.reset_index().groupby(by=__p22_group_col, dropna=False).sum()
      p22_dep_df['usage_ratio'] = p22_dep_df['paid_amount'] / p22_dep_df['incurred_amount']
      p22_dep_df['incurred_per_case'] = p22_dep_df['incurred_amount'] / p22_dep_df['no_of_cases'].replace(0, pd.NA)
      p22_dep_df['paid_per_case']     = p22_dep_df['paid_amount']     / p22_dep_df['no_of_cases'].replace(0, pd.NA)
      p22_dep_df = p22_dep_df[['incurred_amount', 'paid_amount', 'usage_ratio',
                               'no_of_cases', 'incurred_per_case', 'paid_per_case',
                               'no_of_claimants', 'no_of_claim_id']]
      if benefit_type_order is not None:
        p22_dep_df = p22_dep_df.reindex(benefit_type_order, level='benefit_type')

    self.p22_class_benefit_dep_type = p22_dep_df
    return p22_dep_df

  def mcr_p23_ip_benefit(self, by=None, annualize=False, ibnr=False, research_mode=False):
    self.ip_order = self.benefit_index.gum_benefit.loc[(self.benefit_index.gum_benefit_type == 'INPATIENT BENEFITS /HOSPITALIZATION') | (self.benefit_index.gum_benefit_type == 'MATERNITY') | (self.benefit_index.gum_benefit_type == 'SUPPLEMENTARY MAJOR MEDICAL')].drop_duplicates(keep='last').values.tolist()

    __p23_df_col = by + ['benefit', 'incurred_amount', 'paid_amount', 'claim_id', 'claimant']
    __p23_group_col = by.copy() + ['benefit']

    # self.mcr_df['year'] = self.mcr_df.policy_start_date.dt.year
    p23_ip_benefit_df = self.mcr_df[__p23_df_col].loc[self.mcr_df['benefit_type'] == 'Hospital'].groupby(by=__p23_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique', 'claim_id': 'nunique'}).rename(columns={'claimant': 'no_of_claimants', 'claim_id': 'no_of_claim_id'})

    # cases (count of rows/encounters per group)
    p23_ip_cases = (
      self.mcr_df[by + ['benefit', 'incur_date']]
        .loc[self.mcr_df['benefit_type'] == 'Hospital']
        .groupby(by=__p23_group_col, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p23_ip_cases = p23_ip_cases.reindex(p23_ip_benefit_df.index, fill_value=0)
    p23_ip_benefit_df = p23_ip_benefit_df.join(p23_ip_cases['no_of_cases'])

    # base metrics
    p23_ip_benefit_df['usage_ratio'] = p23_ip_benefit_df['paid_amount'] / p23_ip_benefit_df['incurred_amount']
    p23_ip_benefit_df['incurred_per_claim'] = p23_ip_benefit_df['incurred_amount'] / p23_ip_benefit_df['no_of_claim_id']
    p23_ip_benefit_df['paid_per_claim'] = p23_ip_benefit_df['paid_amount'] / p23_ip_benefit_df['no_of_claim_id']
    p23_ip_benefit_df['incurred_per_claimant'] = p23_ip_benefit_df['incurred_amount'] / p23_ip_benefit_df['no_of_claimants']
    p23_ip_benefit_df['paid_per_claimant'] = p23_ip_benefit_df['paid_amount'] / p23_ip_benefit_df['no_of_claimants']
    p23_ip_benefit_df['claim_frequency'] = p23_ip_benefit_df['no_of_claim_id'] / p23_ip_benefit_df['no_of_claimants']

    # fill missing and order benefits
    p23_ip_benefit_df = p23_ip_benefit_df.unstack().stack(dropna=False)
    p23_ip_benefit_df = p23_ip_benefit_df.reindex(self.ip_order, level='benefit')

    # per-case metrics
    p23_ip_benefit_df['incurred_per_case'] = p23_ip_benefit_df['incurred_amount'] / p23_ip_benefit_df['no_of_cases'].replace(0, pd.NA)
    p23_ip_benefit_df['paid_per_case']     = p23_ip_benefit_df['paid_amount']     / p23_ip_benefit_df['no_of_cases'].replace(0, pd.NA)

    # select output columns
    p23_ip_benefit_df = p23_ip_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim', 'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency']]

    # factoring
    if annualize == True or ibnr == True:
      p23_ip_benefit_df = pd.merge(p23_ip_benefit_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p23_ip_factoring = p23_ip_benefit_df.copy(deep=True)
      if annualize == True:
        p23_ip_factoring['incurred_amount'] = p23_ip_factoring['incurred_amount'] * (12 / p23_ip_factoring['data_month'])
        p23_ip_factoring['paid_amount']     = p23_ip_factoring['paid_amount']     * (12 / p23_ip_factoring['data_month'])
        p23_ip_factoring['no_of_cases']     = p23_ip_factoring['no_of_cases']     * (12 / p23_ip_factoring['data_month'])
        p23_ip_factoring['no_of_claim_id']  = p23_ip_factoring['no_of_claim_id']  * (12 / p23_ip_factoring['data_month'])
        p23_ip_factoring['no_of_claimants'] = p23_ip_factoring['no_of_claimants']
      if ibnr == True:
        p23_ip_factoring['incurred_amount'] = p23_ip_factoring['incurred_amount'] * (1 + p23_ip_factoring['ibnr'])
        p23_ip_factoring['paid_amount']     = p23_ip_factoring['paid_amount']     * (1 + p23_ip_factoring['ibnr'])
        p23_ip_factoring['no_of_cases']     = p23_ip_factoring['no_of_cases']     * (1 + p23_ip_factoring['ibnr'])
        p23_ip_factoring['no_of_claim_id']  = p23_ip_factoring['no_of_claim_id']  * (1 + p23_ip_factoring['ibnr'])
        p23_ip_factoring['no_of_claimants'] = p23_ip_factoring['no_of_claimants']

      p23_diff = p23_ip_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] - p23_ip_benefit_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
      temp_by = __p23_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p23_diff = p23_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p23_ip_benefit_df.reset_index().set_index(temp_by, inplace=True)
      p23_ip_benefit_df.loc[p23_ip_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] = p23_ip_benefit_df.loc[p23_ip_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].add(p23_diff)
      temp_by = __p23_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p23_ip_benefit_df = p23_ip_benefit_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p23_ip_benefit_df = p23_ip_benefit_df.reindex(self.ip_order, level='benefit')
      p23_ip_benefit_df['usage_ratio'] = p23_ip_benefit_df['paid_amount'] / p23_ip_benefit_df['incurred_amount'].replace(0, pd.NA)
      p23_ip_benefit_df['incurred_per_case'] = p23_ip_benefit_df['incurred_amount'] / p23_ip_benefit_df['no_of_cases'].replace(0, pd.NA)
      p23_ip_benefit_df['paid_per_case']     = p23_ip_benefit_df['paid_amount']     / p23_ip_benefit_df['no_of_cases'].replace(0, pd.NA)
      p23_ip_benefit_df['incurred_per_claim'] = p23_ip_benefit_df['incurred_amount'] / p23_ip_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23_ip_benefit_df['paid_per_claim'] = p23_ip_benefit_df['paid_amount'] / p23_ip_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23_ip_benefit_df['incurred_per_claimant'] = p23_ip_benefit_df['incurred_amount'] / p23_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23_ip_benefit_df['paid_per_claimant'] = p23_ip_benefit_df['paid_amount'] / p23_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23_ip_benefit_df['claim_frequency'] = p23_ip_benefit_df['no_of_claim_id'] / p23_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)

    __p23_group_col.remove('policy_id') if 'policy_id' in __p23_group_col else __p23_group_col
    p23_ip_benefit_df = p23_ip_benefit_df.reset_index().set_index(__p23_group_col)
    p23_ip_benefit_df = p23_ip_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim', 'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency']]

    if research_mode == True:
      p23_ip_benefit_df = p23_ip_benefit_df.reset_index().groupby(by=__p23_group_col, dropna=False).sum()
      p23_ip_benefit_df['usage_ratio'] = p23_ip_benefit_df['paid_amount'] / p23_ip_benefit_df['incurred_amount'].replace(0, pd.NA)
      p23_ip_benefit_df['incurred_per_case'] = p23_ip_benefit_df['incurred_amount'] / p23_ip_benefit_df['no_of_cases'].replace(0, pd.NA)
      p23_ip_benefit_df['paid_per_case']     = p23_ip_benefit_df['paid_amount']     / p23_ip_benefit_df['no_of_cases'].replace(0, pd.NA)
      p23_ip_benefit_df['incurred_per_claim'] = p23_ip_benefit_df['incurred_amount'] / p23_ip_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23_ip_benefit_df['paid_per_claim'] = p23_ip_benefit_df['paid_amount'] / p23_ip_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23_ip_benefit_df['incurred_per_claimant'] = p23_ip_benefit_df['incurred_amount'] / p23_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23_ip_benefit_df['paid_per_claimant'] = p23_ip_benefit_df['paid_amount'] / p23_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23_ip_benefit_df['claim_frequency'] = p23_ip_benefit_df['no_of_claim_id'] / p23_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23_ip_benefit_df = p23_ip_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim', 'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency']]
      p23_ip_benefit_df = p23_ip_benefit_df.reindex(self.ip_order, level='benefit')

    self.p23_ip_benefit = p23_ip_benefit_df
    self.p23 = p23_ip_benefit_df
    return p23_ip_benefit_df

  def mcr_p23a_class_ip_benefit(self, by=None, annualize=False, ibnr=False, research_mode=False):
    self.ip_order = self.benefit_index.gum_benefit.loc[(self.benefit_index.gum_benefit_type == 'INPATIENT BENEFITS /HOSPITALIZATION') | (self.benefit_index.gum_benefit_type == 'MATERNITY') | (self.benefit_index.gum_benefit_type == 'SUPPLEMENTARY MAJOR MEDICAL')].drop_duplicates(keep='last').values.tolist()

    __p23a_df_col = by + ['class', 'benefit', 'incurred_amount', 'paid_amount', 'claimant', 'claim_id']
    __p23a_cnt_col = by + ['class', 'benefit', 'incur_date']
    __p23a_group_col = by.copy() + ['class', 'benefit']

    p23_ip_class_benefit_df = (
      self.mcr_df[__p23a_df_col]
        .loc[self.mcr_df['benefit_type'] == 'Hospital']
        .groupby(by=__p23a_group_col, dropna=False)
        .agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique', 'claim_id': 'nunique'})
        .rename(columns={'claimant': 'no_of_claimants', 'claim_id': 'no_of_claim_id'})
    )

    p23_ip_class_cases = (
      self.mcr_df[__p23a_cnt_col]
        .loc[self.mcr_df['benefit_type'] == 'Hospital']
        .groupby(by=__p23a_group_col, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p23_ip_class_cases = p23_ip_class_cases.reindex(p23_ip_class_benefit_df.index, fill_value=0)
    p23_ip_class_benefit_df = p23_ip_class_benefit_df.join(p23_ip_class_cases['no_of_cases'])

    p23_ip_class_benefit_df['usage_ratio'] = p23_ip_class_benefit_df['paid_amount'] / p23_ip_class_benefit_df['incurred_amount']
    p23_ip_class_benefit_df['incurred_per_claim'] = p23_ip_class_benefit_df['incurred_amount'] / p23_ip_class_benefit_df['no_of_claim_id']
    p23_ip_class_benefit_df['paid_per_claim'] = p23_ip_class_benefit_df['paid_amount'] / p23_ip_class_benefit_df['no_of_claim_id']
    p23_ip_class_benefit_df['incurred_per_claimant'] = p23_ip_class_benefit_df['incurred_amount'] / p23_ip_class_benefit_df['no_of_claimants']
    p23_ip_class_benefit_df['paid_per_claimant'] = p23_ip_class_benefit_df['paid_amount'] / p23_ip_class_benefit_df['no_of_claimants']
    p23_ip_class_benefit_df['claim_frequency'] = p23_ip_class_benefit_df['no_of_claim_id'] / p23_ip_class_benefit_df['no_of_claimants']

    p23_ip_class_benefit_df = p23_ip_class_benefit_df.unstack().stack(dropna=False)
    p23_ip_class_benefit_df = p23_ip_class_benefit_df.reindex(self.ip_order, level='benefit')

    p23_ip_class_benefit_df['incurred_per_case'] = p23_ip_class_benefit_df['incurred_amount'] / p23_ip_class_benefit_df['no_of_cases'].replace(0, pd.NA)
    p23_ip_class_benefit_df['paid_per_case']     = p23_ip_class_benefit_df['paid_amount']     / p23_ip_class_benefit_df['no_of_cases'].replace(0, pd.NA)

    output_cols = ['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim', 'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency']
    p23_ip_class_benefit_df = p23_ip_class_benefit_df[output_cols]

    if annualize == True or ibnr == True:
      p23_ip_class_benefit_df = pd.merge(p23_ip_class_benefit_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p23_ip_class_benefit_factoring = p23_ip_class_benefit_df.copy(deep=True)
      if annualize == True:
        p23_ip_class_benefit_factoring['incurred_amount'] = p23_ip_class_benefit_factoring['incurred_amount'] * (12 / p23_ip_class_benefit_factoring['data_month'])
        p23_ip_class_benefit_factoring['paid_amount']     = p23_ip_class_benefit_factoring['paid_amount']     * (12 / p23_ip_class_benefit_factoring['data_month'])
        p23_ip_class_benefit_factoring['no_of_cases']     = p23_ip_class_benefit_factoring['no_of_cases']     * (12 / p23_ip_class_benefit_factoring['data_month'])
        p23_ip_class_benefit_factoring['no_of_claim_id']  = p23_ip_class_benefit_factoring['no_of_claim_id']  * (12 / p23_ip_class_benefit_factoring['data_month'])
        p23_ip_class_benefit_factoring['no_of_claimants'] = p23_ip_class_benefit_factoring['no_of_claimants']
      if ibnr == True:
        p23_ip_class_benefit_factoring['incurred_amount'] = p23_ip_class_benefit_factoring['incurred_amount'] * (1 + p23_ip_class_benefit_factoring['ibnr'])
        p23_ip_class_benefit_factoring['paid_amount']     = p23_ip_class_benefit_factoring['paid_amount']     * (1 + p23_ip_class_benefit_factoring['ibnr'])
        p23_ip_class_benefit_factoring['no_of_cases']     = p23_ip_class_benefit_factoring['no_of_cases']     * (1 + p23_ip_class_benefit_factoring['ibnr'])
        p23_ip_class_benefit_factoring['no_of_claim_id']  = p23_ip_class_benefit_factoring['no_of_claim_id']  * (1 + p23_ip_class_benefit_factoring['ibnr'])
        p23_ip_class_benefit_factoring['no_of_claimants'] = p23_ip_class_benefit_factoring['no_of_claimants']

      value_cols = ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']
      p23a_diff = p23_ip_class_benefit_factoring[value_cols] - p23_ip_class_benefit_df[value_cols]
      temp_by = __p23a_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p23a_diff = p23a_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p23_ip_class_benefit_df.reset_index().set_index(temp_by, inplace=True)
      p23_ip_class_benefit_df.loc[p23_ip_class_benefit_df.factoring == True, value_cols] = p23_ip_class_benefit_df.loc[p23_ip_class_benefit_df.factoring == True, value_cols].add(p23a_diff)
      temp_by = __p23a_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p23_ip_class_benefit_df = p23_ip_class_benefit_df.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p23_ip_class_benefit_df = p23_ip_class_benefit_df.reindex(self.ip_order, level='benefit')
      p23_ip_class_benefit_df['usage_ratio'] = p23_ip_class_benefit_df['paid_amount'] / p23_ip_class_benefit_df['incurred_amount'].replace(0, pd.NA)
      p23_ip_class_benefit_df['incurred_per_case'] = p23_ip_class_benefit_df['incurred_amount'] / p23_ip_class_benefit_df['no_of_cases'].replace(0, pd.NA)
      p23_ip_class_benefit_df['paid_per_case']     = p23_ip_class_benefit_df['paid_amount']     / p23_ip_class_benefit_df['no_of_cases'].replace(0, pd.NA)
      p23_ip_class_benefit_df['incurred_per_claim'] = p23_ip_class_benefit_df['incurred_amount'] / p23_ip_class_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23_ip_class_benefit_df['paid_per_claim'] = p23_ip_class_benefit_df['paid_amount'] / p23_ip_class_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23_ip_class_benefit_df['incurred_per_claimant'] = p23_ip_class_benefit_df['incurred_amount'] / p23_ip_class_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23_ip_class_benefit_df['paid_per_claimant'] = p23_ip_class_benefit_df['paid_amount'] / p23_ip_class_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23_ip_class_benefit_df['claim_frequency'] = p23_ip_class_benefit_df['no_of_claim_id'] / p23_ip_class_benefit_df['no_of_claimants'].replace(0, pd.NA)

    __p23a_group_col.remove('policy_id') if 'policy_id' in __p23a_group_col else __p23a_group_col
    p23_ip_class_benefit_df = p23_ip_class_benefit_df.reset_index().set_index(__p23a_group_col)
    p23_ip_class_benefit_df = p23_ip_class_benefit_df[output_cols]
    p23_ip_class_benefit_df = p23_ip_class_benefit_df.reindex(self.ip_order, level='benefit')

    if research_mode == True:
      p23_ip_class_benefit_df = p23_ip_class_benefit_df.reset_index().groupby(by=__p23a_group_col, dropna=False).sum()
      p23_ip_class_benefit_df['usage_ratio'] = p23_ip_class_benefit_df['paid_amount'] / p23_ip_class_benefit_df['incurred_amount'].replace(0, pd.NA)
      p23_ip_class_benefit_df['incurred_per_case'] = p23_ip_class_benefit_df['incurred_amount'] / p23_ip_class_benefit_df['no_of_cases'].replace(0, pd.NA)
      p23_ip_class_benefit_df['paid_per_case']     = p23_ip_class_benefit_df['paid_amount']     / p23_ip_class_benefit_df['no_of_cases'].replace(0, pd.NA)
      p23_ip_class_benefit_df['incurred_per_claim'] = p23_ip_class_benefit_df['incurred_amount'] / p23_ip_class_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23_ip_class_benefit_df['paid_per_claim'] = p23_ip_class_benefit_df['paid_amount'] / p23_ip_class_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23_ip_class_benefit_df['incurred_per_claimant'] = p23_ip_class_benefit_df['incurred_amount'] / p23_ip_class_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23_ip_class_benefit_df['paid_per_claimant'] = p23_ip_class_benefit_df['paid_amount'] / p23_ip_class_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23_ip_class_benefit_df['claim_frequency'] = p23_ip_class_benefit_df['no_of_claim_id'] / p23_ip_class_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23_ip_class_benefit_df = p23_ip_class_benefit_df[output_cols]
      p23_ip_class_benefit_df = p23_ip_class_benefit_df.reindex(self.ip_order, level='benefit')

    self.p23a_ip_class_benefit = p23_ip_class_benefit_df
    self.p23a = p23_ip_class_benefit_df
    return p23_ip_class_benefit_df
  
  def mcr_p23b_ip_common_diagnosis(self, by=None, annualize=False, ibnr=False, research_mode=False):
    self.ip_order = self.benefit_index.gum_benefit.loc[(self.benefit_index.gum_benefit_type == 'INPATIENT BENEFITS /HOSPITALIZATION') | (self.benefit_index.gum_benefit_type == 'MATERNITY') | (self.benefit_index.gum_benefit_type == 'SUPPLEMENTARY MAJOR MEDICAL')].drop_duplicates(keep='last').values.tolist()

    __p23b_df_col = by + ['common_diagnosis_flag', 'benefit', 'incurred_amount', 'paid_amount', 'claimant', 'claim_id']
    __p23b_group_col = by + ['common_diagnosis_flag', 'benefit']

    # self.mcr_df['year'] = self.mcr_df.policy_start_date.dt.year
    p23b_ip_benefit_df = self.mcr_df[__p23b_df_col].loc[self.mcr_df['benefit_type'] == 'Hospital'].groupby(by=__p23b_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique', 'claim_id': 'nunique'}).rename(columns={'claimant': 'no_of_claimants', 'claim_id': 'no_of_claim_id'})
    p23b_ip_benefit_df['usage_ratio'] = p23b_ip_benefit_df['paid_amount'] / p23b_ip_benefit_df['incurred_amount']
    p23b_ip_benefit_df['incurred_per_claim'] = p23b_ip_benefit_df['incurred_amount'] / p23b_ip_benefit_df['no_of_claim_id']
    p23b_ip_benefit_df['paid_per_claim'] = p23b_ip_benefit_df['paid_amount'] / p23b_ip_benefit_df['no_of_claim_id']
    p23b_ip_benefit_df['incurred_per_claimant'] = p23b_ip_benefit_df['incurred_amount'] / p23b_ip_benefit_df['no_of_claimants']
    p23b_ip_benefit_df['paid_per_claimant'] = p23b_ip_benefit_df['paid_amount'] / p23b_ip_benefit_df['no_of_claimants']
    p23b_ip_benefit_df['claim_frequency'] = p23b_ip_benefit_df['no_of_claim_id'] / p23b_ip_benefit_df['no_of_claimants']
    p23b_ip_benefit_df = p23b_ip_benefit_df.unstack().stack(dropna=False)
    p23b_ip_benefit_df = p23b_ip_benefit_df.reindex(self.ip_order, level='benefit')
    p23b_ip_benefit_df = p23b_ip_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim', 'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency']]
    if annualize == True or ibnr == True:
      p23b_ip_benefit_df = pd.merge(p23b_ip_benefit_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p23b_ip_benefit_factoring = p23b_ip_benefit_df.copy(deep=True)
      if annualize == True:
        p23b_ip_benefit_factoring['incurred_amount'] = p23b_ip_benefit_factoring['incurred_amount'] * (12 / p23b_ip_benefit_factoring['data_month'])
        p23b_ip_benefit_factoring['paid_amount']     = p23b_ip_benefit_factoring['paid_amount']     * (12 / p23b_ip_benefit_factoring['data_month'])
        p23b_ip_benefit_factoring['no_of_claim_id']  = p23b_ip_benefit_factoring['no_of_claim_id']  * (12 / p23b_ip_benefit_factoring['data_month'])
        p23b_ip_benefit_factoring['no_of_claimants'] = p23b_ip_benefit_factoring['no_of_claimants']
      if ibnr == True:
        p23b_ip_benefit_factoring['incurred_amount'] = p23b_ip_benefit_factoring['incurred_amount'] * (1 + p23b_ip_benefit_factoring['ibnr'])
        p23b_ip_benefit_factoring['paid_amount']     = p23b_ip_benefit_factoring['paid_amount']     * (1 + p23b_ip_benefit_factoring['ibnr'])
        p23b_ip_benefit_factoring['no_of_claim_id']  = p23b_ip_benefit_factoring['no_of_claim_id']  * (1 + p23b_ip_benefit_factoring['ibnr'])
        p23b_ip_benefit_factoring['no_of_claimants'] = p23b_ip_benefit_factoring['no_of_claimants']
      # p23b_ip_benefit_df = p23b_ip_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim', 'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency', 'data_month', 'ibnr']]
      p23b_diff = p23b_ip_benefit_factoring[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] - p23b_ip_benefit_df[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
      temp_by = __p23b_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p23b_diff = p23b_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p23b_ip_benefit_df.reset_index().set_index(temp_by, inplace=True)
      p23b_ip_benefit_df.loc[p23b_ip_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] = p23b_ip_benefit_df.loc[p23b_ip_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].add(p23b_diff)
      temp_by = __p23b_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p23b_ip_benefit_df = p23b_ip_benefit_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p23b_ip_benefit_df = p23b_ip_benefit_df.reset_index().set_index(temp_by)
      p23b_ip_benefit_df = p23b_ip_benefit_df.reindex(self.ip_order, level='benefit')
      p23b_ip_benefit_df['usage_ratio'] = p23b_ip_benefit_df['paid_amount'] / p23b_ip_benefit_df['incurred_amount'].replace(0, pd.NA)
      p23b_ip_benefit_df['incurred_per_claim'] = p23b_ip_benefit_df['incurred_amount'] / p23b_ip_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23b_ip_benefit_df['paid_per_claim'] = p23b_ip_benefit_df['paid_amount'] / p23b_ip_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23b_ip_benefit_df['incurred_per_claimant'] = p23b_ip_benefit_df['incurred_amount'] / p23b_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23b_ip_benefit_df['paid_per_claimant'] = p23b_ip_benefit_df['paid_amount'] / p23b_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23b_ip_benefit_df['claim_frequency'] = p23b_ip_benefit_df['no_of_claim_id'] / p23b_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)

    __p23b_group_col.remove('policy_id') if 'policy_id' in __p23b_group_col else __p23b_group_col
    p23b_ip_benefit_df = p23b_ip_benefit_df.reset_index().set_index(__p23b_group_col)
    p23b_ip_benefit_df = p23b_ip_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim', 'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency']]
    if research_mode == True:
      p23b_ip_benefit_df = p23b_ip_benefit_df.reset_index().groupby(by=__p23b_group_col, dropna=False).sum()
      p23b_ip_benefit_df['usage_ratio'] = p23b_ip_benefit_df['paid_amount'] / p23b_ip_benefit_df['incurred_amount'].replace(0, pd.NA)
      p23b_ip_benefit_df['incurred_per_claim'] = p23b_ip_benefit_df['incurred_amount'] / p23b_ip_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23b_ip_benefit_df['paid_per_claim'] = p23b_ip_benefit_df['paid_amount'] / p23b_ip_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p23b_ip_benefit_df['incurred_per_claimant'] = p23b_ip_benefit_df['incurred_amount'] / p23b_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23b_ip_benefit_df['paid_per_claimant'] = p23b_ip_benefit_df['paid_amount'] / p23b_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23b_ip_benefit_df['claim_frequency'] = p23b_ip_benefit_df['no_of_claim_id'] / p23b_ip_benefit_df['no_of_claimants'].replace(0, pd.NA)
      p23b_ip_benefit_df = p23b_ip_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim', 'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency']]
      p23b_ip_benefit_df = p23b_ip_benefit_df.reindex(self.ip_order, level='benefit')
    self.p23b_ip_common_diagnosis = p23b_ip_benefit_df
    return p23b_ip_benefit_df

  def mcr_p23b_ip_common_diagnosis_class(self, by=None, annualize=False, ibnr=False, research_mode=False):
    self.ip_order = self.benefit_index.gum_benefit.loc[(self.benefit_index.gum_benefit_type == 'INPATIENT BENEFITS /HOSPITALIZATION') | (self.benefit_index.gum_benefit_type == 'MATERNITY') | (self.benefit_index.gum_benefit_type == 'SUPPLEMENTARY MAJOR MEDICAL')].drop_duplicates(keep='last').values.tolist()

    __p23b_class_df_col = by + ['class', 'common_diagnosis_flag', 'benefit', 'incurred_amount', 'paid_amount', 'claimant', 'claim_id']
    __p23b_class_group_col = by + ['class', 'common_diagnosis_flag', 'benefit']

    p23b_ip_class_df = self.mcr_df[__p23b_class_df_col].loc[self.mcr_df['benefit_type'] == 'Hospital'].groupby(by=__p23b_class_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique', 'claim_id': 'nunique'}).rename(columns={'claimant': 'no_of_claimants', 'claim_id': 'no_of_claim_id'})
    p23b_ip_class_df['usage_ratio'] = p23b_ip_class_df['paid_amount'] / p23b_ip_class_df['incurred_amount']
    p23b_ip_class_df['incurred_per_claim'] = p23b_ip_class_df['incurred_amount'] / p23b_ip_class_df['no_of_claim_id']
    p23b_ip_class_df['paid_per_claim'] = p23b_ip_class_df['paid_amount'] / p23b_ip_class_df['no_of_claim_id']
    p23b_ip_class_df['incurred_per_claimant'] = p23b_ip_class_df['incurred_amount'] / p23b_ip_class_df['no_of_claimants']
    p23b_ip_class_df['paid_per_claimant'] = p23b_ip_class_df['paid_amount'] / p23b_ip_class_df['no_of_claimants']
    p23b_ip_class_df['claim_frequency'] = p23b_ip_class_df['no_of_claim_id'] / p23b_ip_class_df['no_of_claimants']
    p23b_ip_class_df = p23b_ip_class_df.unstack().stack(dropna=False)
    p23b_ip_class_df = p23b_ip_class_df.reindex(self.ip_order, level='benefit')
    p23b_ip_class_df = p23b_ip_class_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim', 'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency']]

    if annualize == True or ibnr == True:
      p23b_ip_class_df = pd.merge(p23b_ip_class_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p23b_ip_class_factoring = p23b_ip_class_df.copy(deep=True)
      if annualize == True:
        p23b_ip_class_factoring['incurred_amount'] = p23b_ip_class_factoring['incurred_amount'] * (12 / p23b_ip_class_factoring['data_month'])
        p23b_ip_class_factoring['paid_amount']     = p23b_ip_class_factoring['paid_amount']     * (12 / p23b_ip_class_factoring['data_month'])
        p23b_ip_class_factoring['no_of_claim_id']  = p23b_ip_class_factoring['no_of_claim_id']  * (12 / p23b_ip_class_factoring['data_month'])
        p23b_ip_class_factoring['no_of_claimants'] = p23b_ip_class_factoring['no_of_claimants']
      if ibnr == True:
        p23b_ip_class_factoring['incurred_amount'] = p23b_ip_class_factoring['incurred_amount'] * (1 + p23b_ip_class_factoring['ibnr'])
        p23b_ip_class_factoring['paid_amount']     = p23b_ip_class_factoring['paid_amount']     * (1 + p23b_ip_class_factoring['ibnr'])
        p23b_ip_class_factoring['no_of_claim_id']  = p23b_ip_class_factoring['no_of_claim_id']  * (1 + p23b_ip_class_factoring['ibnr'])
        p23b_ip_class_factoring['no_of_claimants'] = p23b_ip_class_factoring['no_of_claimants']
      p23b_class_diff = p23b_ip_class_factoring[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] - p23b_ip_class_df[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
      temp_by = __p23b_class_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p23b_class_diff = p23b_class_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p23b_ip_class_df.reset_index().set_index(temp_by, inplace=True)
      p23b_ip_class_df.loc[p23b_ip_class_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] = p23b_ip_class_df.loc[p23b_ip_class_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].add(p23b_class_diff)
      temp_by = __p23b_class_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p23b_ip_class_df = p23b_ip_class_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p23b_ip_class_df = p23b_ip_class_df.reindex(self.ip_order, level='benefit')
      p23b_ip_class_df['usage_ratio'] = p23b_ip_class_df['paid_amount'] / p23b_ip_class_df['incurred_amount'].replace(0, pd.NA)
      p23b_ip_class_df['incurred_per_claim'] = p23b_ip_class_df['incurred_amount'] / p23b_ip_class_df['no_of_claim_id'].replace(0, pd.NA)
      p23b_ip_class_df['paid_per_claim'] = p23b_ip_class_df['paid_amount'] / p23b_ip_class_df['no_of_claim_id'].replace(0, pd.NA)
      p23b_ip_class_df['incurred_per_claimant'] = p23b_ip_class_df['incurred_amount'] / p23b_ip_class_df['no_of_claimants'].replace(0, pd.NA)
      p23b_ip_class_df['paid_per_claimant'] = p23b_ip_class_df['paid_amount'] / p23b_ip_class_df['no_of_claimants'].replace(0, pd.NA)
      p23b_ip_class_df['claim_frequency'] = p23b_ip_class_df['no_of_claim_id'] / p23b_ip_class_df['no_of_claimants'].replace(0, pd.NA)

    __p23b_class_group_col.remove('policy_id') if 'policy_id' in __p23b_class_group_col else __p23b_class_group_col
    p23b_ip_class_df = p23b_ip_class_df.reset_index().set_index(__p23b_class_group_col)
    p23b_ip_class_df = p23b_ip_class_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim', 'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency']]
    if research_mode == True:
      p23b_ip_class_df = p23b_ip_class_df.reset_index().groupby(by=__p23b_class_group_col, dropna=False).sum()
      p23b_ip_class_df['usage_ratio'] = p23b_ip_class_df['paid_amount'] / p23b_ip_class_df['incurred_amount'].replace(0, pd.NA)
      p23b_ip_class_df['incurred_per_claim'] = p23b_ip_class_df['incurred_amount'] / p23b_ip_class_df['no_of_claim_id'].replace(0, pd.NA)
      p23b_ip_class_df['paid_per_claim'] = p23b_ip_class_df['paid_amount'] / p23b_ip_class_df['no_of_claim_id'].replace(0, pd.NA)
      p23b_ip_class_df['incurred_per_claimant'] = p23b_ip_class_df['incurred_amount'] / p23b_ip_class_df['no_of_claimants'].replace(0, pd.NA)
      p23b_ip_class_df['paid_per_claimant'] = p23b_ip_class_df['paid_amount'] / p23b_ip_class_df['no_of_claimants'].replace(0, pd.NA)
      p23b_ip_class_df['claim_frequency'] = p23b_ip_class_df['no_of_claim_id'] / p23b_ip_class_df['no_of_claimants'].replace(0, pd.NA)
      p23b_ip_class_df = p23b_ip_class_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim', 'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency']]
      p23b_ip_class_df = p23b_ip_class_df.reindex(self.ip_order, level='benefit')

    self.p23b_ip_common_diagnosis_class = p23b_ip_class_df
    return p23b_ip_class_df

  def mcr_p24_op_benefit(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p24_df_col = by + ['benefit', 'incurred_amount', 'paid_amount', 'claimant', 'claim_id']
    __p24_group_col = by + ['benefit']
    __p24_sort_col = by + ['paid_amount']
    __p24_sort_order = len(by) * [True] + [False]

    p24_op_benefit_df = self.mcr_df[__p24_df_col].loc[self.mcr_df['benefit_type'] == 'Clinic'].groupby(by=__p24_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    # cases
    p24_op_cases = (
      self.mcr_df[by + ['benefit', 'incur_date']]
        .loc[self.mcr_df['benefit_type'] == 'Clinic']
        .groupby(by=__p24_group_col, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p24_op_cases = p24_op_cases.reindex(p24_op_benefit_df.index, fill_value=0)
    p24_op_benefit_df = p24_op_benefit_df.join(p24_op_cases['no_of_cases'])
    # metrics
    p24_op_benefit_df['usage_ratio'] = p24_op_benefit_df['paid_amount'] / p24_op_benefit_df['incurred_amount']
    p24_op_benefit_df['incurred_per_claim'] = p24_op_benefit_df['incurred_amount'] / p24_op_benefit_df['no_of_claim_id']
    p24_op_benefit_df['paid_per_claim'] = p24_op_benefit_df['paid_amount'] / p24_op_benefit_df['no_of_claim_id']
    p24_op_benefit_df = p24_op_benefit_df.unstack().stack(dropna=False)
    p24_op_benefit_df['incurred_per_case'] = p24_op_benefit_df['incurred_amount'] / p24_op_benefit_df['no_of_cases'].replace(0, pd.NA)
    p24_op_benefit_df['paid_per_case']     = p24_op_benefit_df['paid_amount']     / p24_op_benefit_df['no_of_cases'].replace(0, pd.NA)
    p24_op_benefit_df = p24_op_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim']]

    if len(p24_op_benefit_df) > 0:
      p24_op_benefit_df.sort_values(by=__p24_sort_col, ascending=__p24_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p24_op_benefit_df = pd.merge(p24_op_benefit_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p24_op_factoring = p24_op_benefit_df.copy(deep=True)
        if annualize == True:
          p24_op_factoring['incurred_amount'] = p24_op_factoring['incurred_amount'] * (12 / p24_op_factoring['data_month'])
          p24_op_factoring['paid_amount']     = p24_op_factoring['paid_amount']     * (12 / p24_op_factoring['data_month'])
          p24_op_factoring['no_of_cases']     = p24_op_factoring['no_of_cases']     * (12 / p24_op_factoring['data_month'])
          p24_op_factoring['no_of_claim_id']  = p24_op_factoring['no_of_claim_id']  * (12 / p24_op_factoring['data_month'])
          p24_op_factoring['no_of_claimants'] = p24_op_factoring['no_of_claimants'] 
        if ibnr == True:
          p24_op_factoring['incurred_amount'] = p24_op_factoring['incurred_amount'] * (1 + p24_op_factoring['ibnr'])
          p24_op_factoring['paid_amount']     = p24_op_factoring['paid_amount']     * (1 + p24_op_factoring['ibnr'])
          p24_op_factoring['no_of_cases']     = p24_op_factoring['no_of_cases']     * (1 + p24_op_factoring['ibnr'])
          p24_op_factoring['no_of_claim_id']  = p24_op_factoring['no_of_claim_id']  * (1 + p24_op_factoring['ibnr'])
          p24_op_factoring['no_of_claimants'] = p24_op_factoring['no_of_claimants']
        p24_diff = p24_op_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] - p24_op_benefit_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
        temp_by = __p24_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p24_diff = p24_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
        p24_op_benefit_df.reset_index().set_index(temp_by, inplace=True)
        p24_op_benefit_df.loc[p24_op_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] = p24_op_benefit_df.loc[p24_op_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].add(p24_diff)
        temp_by = __p24_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p24_op_benefit_df = p24_op_benefit_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
        p24_op_benefit_df['usage_ratio'] = p24_op_benefit_df['paid_amount'] / p24_op_benefit_df['incurred_amount'].replace(0, pd.NA)
        p24_op_benefit_df['incurred_per_case'] = p24_op_benefit_df['incurred_amount'] / p24_op_benefit_df['no_of_cases'].replace(0, pd.NA)
        p24_op_benefit_df['paid_per_case']     = p24_op_benefit_df['paid_amount']     / p24_op_benefit_df['no_of_cases'].replace(0, pd.NA)
        p24_op_benefit_df['incurred_per_claim'] = p24_op_benefit_df['incurred_amount'] / p24_op_benefit_df['no_of_claim_id'].replace(0, pd.NA)
        p24_op_benefit_df['paid_per_claim'] = p24_op_benefit_df['paid_amount'] / p24_op_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      
    __p24_group_col.remove('policy_id') if 'policy_id' in __p24_group_col else __p24_group_col
    p24_op_benefit_df = p24_op_benefit_df.reset_index().set_index(__p24_group_col)
    p24_op_benefit_df = p24_op_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim']]
    if research_mode == True:
      p24_op_benefit_df = p24_op_benefit_df.reset_index().groupby(by=__p24_group_col, dropna=False).sum()
      p24_op_benefit_df['usage_ratio'] = p24_op_benefit_df['paid_amount'] / p24_op_benefit_df['incurred_amount'].replace(0, pd.NA)
      p24_op_benefit_df['incurred_per_case'] = p24_op_benefit_df['incurred_amount'] / p24_op_benefit_df['no_of_cases'].replace(0, pd.NA)
      p24_op_benefit_df['paid_per_case']     = p24_op_benefit_df['paid_amount']     / p24_op_benefit_df['no_of_cases'].replace(0, pd.NA)
      p24_op_benefit_df['incurred_per_claim'] = p24_op_benefit_df['incurred_amount'] / p24_op_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p24_op_benefit_df['paid_per_claim'] = p24_op_benefit_df['paid_amount'] / p24_op_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p24_op_benefit_df = p24_op_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim']]
    self.p24_op_benefit = p24_op_benefit_df
    self.p24 = p24_op_benefit_df
    return p24_op_benefit_df

  def mcr_p24a_op_class_benefit(self, by=None, annualize=False, ibnr=False, research_mode=False):
    __p24a_df_col = by + ['class', 'benefit', 'incurred_amount', 'paid_amount', 'claim_id']
    __p24a_group_col = by + ['class', 'benefit']
    __p24a_sort_col = by + ['class', 'paid_amount']
    __p24a_sort_order = len(by) * [True] + [True, False]

    p24_op_class_benefit_df = self.mcr_df[__p24a_df_col].loc[self.mcr_df['benefit_type'] == 'Clinic'].groupby(by=__p24a_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id'})
    # cases
    p24_op_class_cases = (
      self.mcr_df[by + ['class', 'benefit', 'incur_date']]
        .loc[self.mcr_df['benefit_type'] == 'Clinic']
        .groupby(by=__p24a_group_col, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p24_op_class_cases = p24_op_class_cases.reindex(p24_op_class_benefit_df.index, fill_value=0)
    p24_op_class_benefit_df = p24_op_class_benefit_df.join(p24_op_class_cases['no_of_cases'])
    p24_op_class_benefit_df['usage_ratio'] = p24_op_class_benefit_df['paid_amount'] / p24_op_class_benefit_df['incurred_amount']
    p24_op_class_benefit_df['incurred_per_claim'] = p24_op_class_benefit_df['incurred_amount'] / p24_op_class_benefit_df['no_of_claim_id']
    p24_op_class_benefit_df['paid_per_claim'] = p24_op_class_benefit_df['paid_amount'] / p24_op_class_benefit_df['no_of_claim_id']
    p24_op_class_benefit_df = p24_op_class_benefit_df.unstack().stack(dropna=False)
    p24_op_class_benefit_df['incurred_per_case'] = p24_op_class_benefit_df['incurred_amount'] / p24_op_class_benefit_df['no_of_cases'].replace(0, pd.NA)
    p24_op_class_benefit_df['paid_per_case']     = p24_op_class_benefit_df['paid_amount']     / p24_op_class_benefit_df['no_of_cases'].replace(0, pd.NA)
    p24_op_class_benefit_df = p24_op_class_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    if len(p24_op_class_benefit_df) > 0:
      p24_op_class_benefit_df.sort_values(by=__p24a_sort_col, ascending=__p24a_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p24_op_class_benefit_df = pd.merge(p24_op_class_benefit_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p24_op_class_factoring = p24_op_class_benefit_df.copy(deep=True)
        if annualize == True:
          p24_op_class_factoring['incurred_amount'] = p24_op_class_factoring['incurred_amount'] * (12 / p24_op_class_factoring['data_month'])
          p24_op_class_factoring['paid_amount']     = p24_op_class_factoring['paid_amount']     * (12 / p24_op_class_factoring['data_month'])
          p24_op_class_factoring['no_of_cases']     = p24_op_class_factoring['no_of_cases']     * (12 / p24_op_class_factoring['data_month'])
          p24_op_class_factoring['no_of_claim_id']  = p24_op_class_factoring['no_of_claim_id']  * (12 / p24_op_class_factoring['data_month'])
        if ibnr == True:
          p24_op_class_factoring['incurred_amount'] = p24_op_class_factoring['incurred_amount'] * (1 + p24_op_class_factoring['ibnr'])
          p24_op_class_factoring['paid_amount']     = p24_op_class_factoring['paid_amount']     * (1 + p24_op_class_factoring['ibnr'])
          p24_op_class_factoring['no_of_cases']     = p24_op_class_factoring['no_of_cases']     * (1 + p24_op_class_factoring['ibnr'])
          p24_op_class_factoring['no_of_claim_id']  = p24_op_class_factoring['no_of_claim_id']  * (1 + p24_op_class_factoring['ibnr'])
        p24a_diff = p24_op_class_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] - p24_op_class_benefit_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']]
        temp_by = __p24a_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p24a_diff = p24a_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
        p24_op_class_benefit_df.reset_index().set_index(temp_by, inplace=True)
        p24_op_class_benefit_df.loc[p24_op_class_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] = p24_op_class_benefit_df.loc[p24_op_class_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].add(p24a_diff)
        temp_by = __p24a_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p24_op_class_benefit_df = p24_op_class_benefit_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
        p24_op_class_benefit_df['usage_ratio'] = p24_op_class_benefit_df['paid_amount'] / p24_op_class_benefit_df['incurred_amount'].replace(0, pd.NA)
        p24_op_class_benefit_df['incurred_per_case'] = p24_op_class_benefit_df['incurred_amount'] / p24_op_class_benefit_df['no_of_cases'].replace(0, pd.NA)
        p24_op_class_benefit_df['paid_per_case']     = p24_op_class_benefit_df['paid_amount']     / p24_op_class_benefit_df['no_of_cases'].replace(0, pd.NA)
        p24_op_class_benefit_df['incurred_per_claim'] = p24_op_class_benefit_df['incurred_amount'] / p24_op_class_benefit_df['no_of_claim_id'].replace(0, pd.NA)
        p24_op_class_benefit_df['paid_per_claim'] = p24_op_class_benefit_df['paid_amount'] / p24_op_class_benefit_df['no_of_claim_id'].replace(0, pd.NA)


    __p24a_group_col.remove('policy_id') if 'policy_id' in __p24a_group_col else __p24a_group_col
    p24_op_class_benefit_df = p24_op_class_benefit_df.reset_index().set_index(__p24a_group_col)
    p24_op_class_benefit_df = p24_op_class_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    if research_mode == True:
      p24_op_class_benefit_df = p24_op_class_benefit_df.reset_index().groupby(by=__p24a_group_col, dropna=False).sum()
      p24_op_class_benefit_df['usage_ratio'] = p24_op_class_benefit_df['paid_amount'] / p24_op_class_benefit_df['incurred_amount'].replace(0, pd.NA)
      p24_op_class_benefit_df['incurred_per_case'] = p24_op_class_benefit_df['incurred_amount'] / p24_op_class_benefit_df['no_of_cases'].replace(0, pd.NA)
      p24_op_class_benefit_df['paid_per_case']     = p24_op_class_benefit_df['paid_amount']     / p24_op_class_benefit_df['no_of_cases'].replace(0, pd.NA)
      p24_op_class_benefit_df['incurred_per_claim'] = p24_op_class_benefit_df['incurred_amount'] / p24_op_class_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p24_op_class_benefit_df['paid_per_claim'] = p24_op_class_benefit_df['paid_amount'] / p24_op_class_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p24_op_class_benefit_df = p24_op_class_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    self.p24a_op_class_benefit = p24_op_class_benefit_df
    self.p24a = p24_op_class_benefit_df
    return p24_op_class_benefit_df

  def mcr_p24_dent_benefit(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p24d_df_col = by + ['benefit', 'incurred_amount', 'paid_amount', 'claim_id']
    __p24d_group_col = by + ['benefit']
    __p24d_sort_col = by + ['paid_amount']
    __p24d_sort_order = len(by) * [True] + [False]

    p24_dent_benefit_df = self.mcr_df[__p24d_df_col].loc[self.mcr_df['benefit_type'] == 'Dental'].groupby(by=__p24d_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id'})
    # cases
    p24_dent_cases = (
      self.mcr_df[by + ['benefit', 'incur_date']]
        .loc[self.mcr_df['benefit_type'] == 'Dental']
        .groupby(by=__p24d_group_col, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p24_dent_cases = p24_dent_cases.reindex(p24_dent_benefit_df.index, fill_value=0)
    p24_dent_benefit_df = p24_dent_benefit_df.join(p24_dent_cases['no_of_cases'])
    p24_dent_benefit_df['usage_ratio'] = p24_dent_benefit_df['paid_amount'] / p24_dent_benefit_df['incurred_amount']
    p24_dent_benefit_df['incurred_per_claim'] = p24_dent_benefit_df['incurred_amount'] / p24_dent_benefit_df['no_of_claim_id']
    p24_dent_benefit_df['paid_per_claim'] = p24_dent_benefit_df['paid_amount'] / p24_dent_benefit_df['no_of_claim_id']
    
    if len(p24_dent_benefit_df) > 0:
      p24_dent_benefit_df = p24_dent_benefit_df.unstack().stack(dropna=False)
    
    p24_dent_benefit_df['incurred_per_case'] = p24_dent_benefit_df['incurred_amount'] / p24_dent_benefit_df['no_of_cases'].replace(0, pd.NA)
    p24_dent_benefit_df['paid_per_case']     = p24_dent_benefit_df['paid_amount']     / p24_dent_benefit_df['no_of_cases'].replace(0, pd.NA)
    p24_dent_benefit_df = p24_dent_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    
    if len(p24_dent_benefit_df) > 0:
      p24_dent_benefit_df.sort_values(by=__p24d_sort_col, ascending=__p24d_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p24_dent_benefit_df = pd.merge(p24_dent_benefit_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p24_dent_factoring = p24_dent_benefit_df.copy(deep=True)
        if annualize == True:
          p24_dent_factoring['incurred_amount'] = p24_dent_factoring['incurred_amount'] * (12 / p24_dent_factoring['data_month'])
          p24_dent_factoring['paid_amount']     = p24_dent_factoring['paid_amount']     * (12 / p24_dent_factoring['data_month'])
          p24_dent_factoring['no_of_cases']     = p24_dent_factoring['no_of_cases']     * (12 / p24_dent_factoring['data_month'])
          p24_dent_factoring['no_of_claim_id']  = p24_dent_factoring['no_of_claim_id']  * (12 / p24_dent_factoring['data_month'])
        if ibnr == True:
          p24_dent_factoring['incurred_amount'] = p24_dent_factoring['incurred_amount'] * (1 + p24_dent_factoring['ibnr'])
          p24_dent_factoring['paid_amount']     = p24_dent_factoring['paid_amount']     * (1 + p24_dent_factoring['ibnr'])
          p24_dent_factoring['no_of_cases']     = p24_dent_factoring['no_of_cases']     * (1 + p24_dent_factoring['ibnr'])
          p24_dent_factoring['no_of_claim_id']  = p24_dent_factoring['no_of_claim_id']  * (1 + p24_dent_factoring['ibnr'])
        
        p24d_diff = p24_dent_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] - p24_dent_benefit_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']]
        temp_by = __p24d_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p24d_diff = p24d_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
        p24_dent_benefit_df.reset_index().set_index(temp_by, inplace=True)
        p24_dent_benefit_df.loc[p24_dent_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] = p24_dent_benefit_df.loc[p24_dent_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].add(p24d_diff)
        temp_by = __p24d_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p24_dent_benefit_df = p24_dent_benefit_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
        p24_dent_benefit_df['usage_ratio'] = p24_dent_benefit_df['paid_amount'] / p24_dent_benefit_df['incurred_amount'].replace(0, pd.NA)
        p24_dent_benefit_df['incurred_per_case'] = p24_dent_benefit_df['incurred_amount'] / p24_dent_benefit_df['no_of_cases'].replace(0, pd.NA)
        p24_dent_benefit_df['paid_per_case']     = p24_dent_benefit_df['paid_amount']     / p24_dent_benefit_df['no_of_cases'].replace(0, pd.NA)
        p24_dent_benefit_df['incurred_per_claim'] = p24_dent_benefit_df['incurred_amount'] / p24_dent_benefit_df['no_of_claim_id'].replace(0, pd.NA)
        p24_dent_benefit_df['paid_per_claim'] = p24_dent_benefit_df['paid_amount'] / p24_dent_benefit_df['no_of_claim_id'].replace(0, pd.NA)

    __p24d_group_col.remove('policy_id') if 'policy_id' in __p24d_group_col else __p24d_group_col
    p24_dent_benefit_df = p24_dent_benefit_df.reset_index().set_index(__p24d_group_col)
    p24_dent_benefit_df = p24_dent_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    if research_mode == True:
      p24_dent_benefit_df = p24_dent_benefit_df.reset_index().groupby(by=__p24d_group_col, dropna=False).sum()
      p24_dent_benefit_df['usage_ratio'] = p24_dent_benefit_df['paid_amount'] / p24_dent_benefit_df['incurred_amount'].replace(0, pd.NA)
      p24_dent_benefit_df['incurred_per_case'] = p24_dent_benefit_df['incurred_amount'] / p24_dent_benefit_df['no_of_cases'].replace(0, pd.NA)
      p24_dent_benefit_df['paid_per_case']     = p24_dent_benefit_df['paid_amount']     / p24_dent_benefit_df['no_of_cases'].replace(0, pd.NA)
      p24_dent_benefit_df['incurred_per_claim'] = p24_dent_benefit_df['incurred_amount'] / p24_dent_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p24_dent_benefit_df['paid_per_claim'] = p24_dent_benefit_df['paid_amount'] / p24_dent_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p24_dent_benefit_df = p24_dent_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    self.p24_dent_benefit = p24_dent_benefit_df
    return p24_dent_benefit_df
  
  def mcr_p24_wellness_benefit(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p24w_df_col = by + ['benefit', 'incurred_amount', 'paid_amount', 'claim_id']
    __p24w_group_col = by + ['benefit']
    __p24w_sort_col = by + ['paid_amount']
    __p24w_sort_order = len(by) * [True] + [False]

    p24_wellness_benefit_df = self.mcr_df[__p24w_df_col].loc[(self.mcr_df['benefit_type'] == 'Dental') | (self.mcr_df['benefit_type'] == 'Optical') | (self.mcr_df['benefit'].str.contains('vaccin|check|combined', case=False))].groupby(by=__p24w_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id'})
    # cases
    p24_wellness_cases = (
      self.mcr_df[by + ['benefit', 'incur_date']]
        .loc[(self.mcr_df['benefit_type'] == 'Dental') | (self.mcr_df['benefit_type'] == 'Optical') | (self.mcr_df['benefit'].str.contains('vaccin|check|combined', case=False))]
        .groupby(by=__p24w_group_col, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p24_wellness_cases = p24_wellness_cases.reindex(p24_wellness_benefit_df.index, fill_value=0)
    p24_wellness_benefit_df = p24_wellness_benefit_df.join(p24_wellness_cases['no_of_cases'])
    p24_wellness_benefit_df['usage_ratio'] = p24_wellness_benefit_df['paid_amount'] / p24_wellness_benefit_df['incurred_amount']
    p24_wellness_benefit_df['incurred_per_claim'] = p24_wellness_benefit_df['incurred_amount'] / p24_wellness_benefit_df['no_of_claim_id']
    p24_wellness_benefit_df['paid_per_claim'] = p24_wellness_benefit_df['paid_amount'] / p24_wellness_benefit_df['no_of_claim_id']
    if len(p24_wellness_benefit_df) > 0:
      p24_wellness_benefit_df = p24_wellness_benefit_df.unstack().stack(dropna=False)
    p24_wellness_benefit_df['incurred_per_case'] = p24_wellness_benefit_df['incurred_amount'] / p24_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
    p24_wellness_benefit_df['paid_per_case']     = p24_wellness_benefit_df['paid_amount']     / p24_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
    p24_wellness_benefit_df = p24_wellness_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    if len(p24_wellness_benefit_df) > 0:
      p24_wellness_benefit_df.sort_values(by=__p24w_sort_col, ascending=__p24w_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p24_wellness_benefit_df = pd.merge(p24_wellness_benefit_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p24_wellness_factoring = p24_wellness_benefit_df.copy(deep=True)
        if annualize == True:
          p24_wellness_factoring['incurred_amount'] = p24_wellness_factoring['incurred_amount'] * (12 / p24_wellness_factoring['data_month'])
          p24_wellness_factoring['paid_amount']     = p24_wellness_factoring['paid_amount']     * (12 / p24_wellness_factoring['data_month'])
          p24_wellness_factoring['no_of_cases']     = p24_wellness_factoring['no_of_cases']     * (12 / p24_wellness_factoring['data_month'])
          p24_wellness_factoring['no_of_claim_id']  = p24_wellness_factoring['no_of_claim_id']  * (12 / p24_wellness_factoring['data_month'])
        if ibnr == True:
          p24_wellness_factoring['incurred_amount'] = p24_wellness_factoring['incurred_amount'] * (1 + p24_wellness_factoring['ibnr'])
          p24_wellness_factoring['paid_amount']     = p24_wellness_factoring['paid_amount']     * (1 + p24_wellness_factoring['ibnr'])
          p24_wellness_factoring['no_of_cases']     = p24_wellness_factoring['no_of_cases']     * (1 + p24_wellness_factoring['ibnr'])
          p24_wellness_factoring['no_of_claim_id']  = p24_wellness_factoring['no_of_claim_id']  * (1 + p24_wellness_factoring['ibnr'])
        p24w_diff = p24_wellness_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] - p24_wellness_benefit_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']]
        temp_by = __p24w_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p24w_diff = p24w_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
        p24_wellness_benefit_df.reset_index().set_index(temp_by, inplace=True)
        p24_wellness_benefit_df.loc[p24_wellness_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] = p24_wellness_benefit_df.loc[p24_wellness_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].add(p24w_diff)
        temp_by = __p24w_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p24_wellness_benefit_df = p24_wellness_benefit_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
        p24_wellness_benefit_df['usage_ratio'] = p24_wellness_benefit_df['paid_amount'] / p24_wellness_benefit_df['incurred_amount'].replace(0, pd.NA)
        p24_wellness_benefit_df['incurred_per_case'] = p24_wellness_benefit_df['incurred_amount'] / p24_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
        p24_wellness_benefit_df['paid_per_case']     = p24_wellness_benefit_df['paid_amount']     / p24_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
        p24_wellness_benefit_df['incurred_per_claim'] = p24_wellness_benefit_df['incurred_amount'] / p24_wellness_benefit_df['no_of_claim_id'].replace(0, pd.NA)
        p24_wellness_benefit_df['paid_per_claim'] = p24_wellness_benefit_df['paid_amount'] / p24_wellness_benefit_df['no_of_claim_id'].replace(0, pd.NA)


    __p24w_group_col.remove('policy_id') if 'policy_id' in __p24w_group_col else __p24w_group_col
    p24_wellness_benefit_df = p24_wellness_benefit_df.reset_index().set_index(__p24w_group_col)
    p24_wellness_benefit_df = p24_wellness_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    if research_mode == True:
      p24_wellness_benefit_df = p24_wellness_benefit_df.reset_index().groupby(by=__p24w_group_col, dropna=False).sum()
      p24_wellness_benefit_df['usage_ratio'] = p24_wellness_benefit_df['paid_amount'] / p24_wellness_benefit_df['incurred_amount'].replace(0, pd.NA)
      p24_wellness_benefit_df['incurred_per_case'] = p24_wellness_benefit_df['incurred_amount'] / p24_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
      p24_wellness_benefit_df['paid_per_case']     = p24_wellness_benefit_df['paid_amount']     / p24_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
      p24_wellness_benefit_df['incurred_per_claim'] = p24_wellness_benefit_df['incurred_amount'] / p24_wellness_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p24_wellness_benefit_df['paid_per_claim'] = p24_wellness_benefit_df['paid_amount'] / p24_wellness_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p24_wellness_benefit_df = p24_wellness_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    self.p24_wellness_benefit = p24_wellness_benefit_df
    return p24_wellness_benefit_df
  
  def mcr_p24_class_wellness_benefit(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p24wc_df_col = by + ['class', 'benefit', 'incurred_amount', 'paid_amount', 'claim_id']
    __p24wc_group_col = by + ['class', 'benefit']
    __p24wc_sort_col = by + ['class', 'paid_amount']
    __p24wc_sort_order = len(by) * [True] + [True, False]

    p24_class_wellness_benefit_df = self.mcr_df[__p24wc_df_col].loc[(self.mcr_df['benefit_type'] == 'Dental') | (self.mcr_df['benefit_type'] == 'Optical') | (self.mcr_df['benefit'].str.contains('vaccin|check|combined', case=False))].groupby(by=__p24wc_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id'})
    # cases
    p24_class_wellness_cases = (
      self.mcr_df[by + ['class', 'benefit', 'incur_date']]
        .loc[(self.mcr_df['benefit_type'] == 'Dental') | (self.mcr_df['benefit_type'] == 'Optical') | (self.mcr_df['benefit'].str.contains('vaccin|check|combined', case=False))]
        .groupby(by=__p24wc_group_col, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p24_class_wellness_cases = p24_class_wellness_cases.reindex(p24_class_wellness_benefit_df.index, fill_value=0)
    p24_class_wellness_benefit_df = p24_class_wellness_benefit_df.join(p24_class_wellness_cases['no_of_cases'])
    p24_class_wellness_benefit_df['usage_ratio'] = p24_class_wellness_benefit_df['paid_amount'] / p24_class_wellness_benefit_df['incurred_amount']
    p24_class_wellness_benefit_df['incurred_per_claim'] = p24_class_wellness_benefit_df['incurred_amount'] / p24_class_wellness_benefit_df['no_of_claim_id']
    p24_class_wellness_benefit_df['paid_per_claim'] = p24_class_wellness_benefit_df['paid_amount'] / p24_class_wellness_benefit_df['no_of_claim_id']
    if len(p24_class_wellness_benefit_df) > 0:
      p24_class_wellness_benefit_df = p24_class_wellness_benefit_df.unstack().stack(dropna=False)
    p24_class_wellness_benefit_df['incurred_per_case'] = p24_class_wellness_benefit_df['incurred_amount'] / p24_class_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
    p24_class_wellness_benefit_df['paid_per_case']     = p24_class_wellness_benefit_df['paid_amount']     / p24_class_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
    p24_class_wellness_benefit_df = p24_class_wellness_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]

    if len(p24_class_wellness_benefit_df) > 0:
      p24_class_wellness_benefit_df.sort_values(by=__p24wc_sort_col, ascending=__p24wc_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p24_class_wellness_benefit_df = pd.merge(p24_class_wellness_benefit_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p24_class_wellness_factoring = p24_class_wellness_benefit_df.copy(deep=True)
        if annualize == True:
          p24_class_wellness_factoring['incurred_amount'] = p24_class_wellness_factoring['incurred_amount'] * (12 / p24_class_wellness_factoring['data_month'])
          p24_class_wellness_factoring['paid_amount']     = p24_class_wellness_factoring['paid_amount']     * (12 / p24_class_wellness_factoring['data_month'])
          p24_class_wellness_factoring['no_of_cases']     = p24_class_wellness_factoring['no_of_cases']     * (12 / p24_class_wellness_factoring['data_month'])
          p24_class_wellness_factoring['no_of_claim_id']  = p24_class_wellness_factoring['no_of_claim_id']  * (12 / p24_class_wellness_factoring['data_month'])
        if ibnr == True:
          p24_class_wellness_factoring['incurred_amount'] = p24_class_wellness_factoring['incurred_amount'] * (1 + p24_class_wellness_factoring['ibnr'])
          p24_class_wellness_factoring['paid_amount']     = p24_class_wellness_factoring['paid_amount']     * (1 + p24_class_wellness_factoring['ibnr'])
          p24_class_wellness_factoring['no_of_cases']     = p24_class_wellness_factoring['no_of_cases']     * (1 + p24_class_wellness_factoring['ibnr'])
          p24_class_wellness_factoring['no_of_claim_id']  = p24_class_wellness_factoring['no_of_claim_id']  * (1 + p24_class_wellness_factoring['ibnr'])
        p24cw_diff = p24_class_wellness_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] - p24_class_wellness_benefit_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']]
        temp_by = __p24wc_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p24cw_diff = p24cw_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
        p24_class_wellness_benefit_df.reset_index().set_index(temp_by, inplace=True)
        p24_class_wellness_benefit_df.loc[p24_class_wellness_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] = p24_class_wellness_benefit_df.loc[p24_class_wellness_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].add(p24cw_diff)
        temp_by = __p24wc_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p24_class_wellness_benefit_df = p24_class_wellness_benefit_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
        p24_class_wellness_benefit_df['usage_ratio'] = p24_class_wellness_benefit_df['paid_amount'] / p24_class_wellness_benefit_df['incurred_amount'].replace(0, pd.NA)
        p24_class_wellness_benefit_df['incurred_per_case'] = p24_class_wellness_benefit_df['incurred_amount'] / p24_class_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
        p24_class_wellness_benefit_df['paid_per_case']     = p24_class_wellness_benefit_df['paid_amount']     / p24_class_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
        p24_class_wellness_benefit_df['incurred_per_claim'] = p24_class_wellness_benefit_df['incurred_amount'] / p24_class_wellness_benefit_df['no_of_claim_id'].replace(0, pd.NA)
        p24_class_wellness_benefit_df['paid_per_claim'] = p24_class_wellness_benefit_df['paid_amount'] / p24_class_wellness_benefit_df['no_of_claim_id'].replace(0, pd.NA)


    __p24wc_group_col.remove('policy_id') if 'policy_id' in __p24wc_group_col else __p24wc_group_col
    p24_class_wellness_benefit_df = p24_class_wellness_benefit_df.reset_index().set_index(__p24wc_group_col)
    p24_class_wellness_benefit_df = p24_class_wellness_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    if research_mode == True:
      p24_class_wellness_benefit_df = p24_class_wellness_benefit_df.reset_index().groupby(by=__p24wc_group_col, dropna=False).sum()
      p24_class_wellness_benefit_df['usage_ratio'] = p24_class_wellness_benefit_df['paid_amount'] / p24_class_wellness_benefit_df['incurred_amount'].replace(0, pd.NA)
      p24_class_wellness_benefit_df['incurred_per_case'] = p24_class_wellness_benefit_df['incurred_amount'] / p24_class_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
      p24_class_wellness_benefit_df['paid_per_case']     = p24_class_wellness_benefit_df['paid_amount']     / p24_class_wellness_benefit_df['no_of_cases'].replace(0, pd.NA)
      p24_class_wellness_benefit_df['incurred_per_claim'] = p24_class_wellness_benefit_df['incurred_amount'] / p24_class_wellness_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p24_class_wellness_benefit_df['paid_per_claim'] = p24_class_wellness_benefit_df['paid_amount'] / p24_class_wellness_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p24_class_wellness_benefit_df = p24_class_wellness_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    self.p24_class_wellness_benefit = p24_class_wellness_benefit_df
    return p24_class_wellness_benefit_df

  def mcr_p25_class_panel_benefit(self, by=None, benefit_type_order=None, annualize=False, ibnr=False, research_mode=False):

    __p25_df_col = by + ['class', 'panel', 'benefit_type', 'incurred_amount', 'paid_amount', 'claim_id', 'claimant']
    __p25_group_col = by + ['class', 'panel', 'benefit_type']

    # self.mcr_df['year'] = self.mcr_df.policy_start_date.dt.year
    p25_class_panel_benefit_df = (
      self.mcr_df[__p25_df_col]
        .groupby(by=__p25_group_col, dropna=False)
        .agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'})
        .rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    )
    # cases
    p25_cases = (
      self.mcr_df[by + ['class', 'panel', 'benefit_type', 'incur_date']]
        .groupby(by=__p25_group_col, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p25_cases = p25_cases.reindex(p25_class_panel_benefit_df.index, fill_value=0)
    p25_class_panel_benefit_df = p25_class_panel_benefit_df.join(p25_cases['no_of_cases'])
    p25_class_panel_benefit_df['usage_ratio'] = p25_class_panel_benefit_df['paid_amount'] / p25_class_panel_benefit_df['incurred_amount']
    p25_class_panel_benefit_df = p25_class_panel_benefit_df.unstack().stack(dropna=False)
    p25_class_panel_benefit_df = p25_class_panel_benefit_df.reindex(benefit_type_order, level='benefit_type')
    p25_class_panel_benefit_df['incurred_per_claim'] = p25_class_panel_benefit_df['incurred_amount'] / p25_class_panel_benefit_df['no_of_claim_id']
    p25_class_panel_benefit_df['paid_per_claim'] = p25_class_panel_benefit_df['paid_amount'] / p25_class_panel_benefit_df['no_of_claim_id']
    p25_class_panel_benefit_df['incurred_per_case'] = p25_class_panel_benefit_df['incurred_amount'] / p25_class_panel_benefit_df['no_of_cases'].replace(0, pd.NA)
    p25_class_panel_benefit_df['paid_per_case']     = p25_class_panel_benefit_df['paid_amount']     / p25_class_panel_benefit_df['no_of_cases'].replace(0, pd.NA)
    p25_class_panel_benefit_df = p25_class_panel_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim', 'no_of_claimants']]
    if annualize == True or ibnr == True:
      p25_class_panel_benefit_df = pd.merge(p25_class_panel_benefit_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p25_class_panel_benefit_factoring = p25_class_panel_benefit_df.copy(deep=True)
      if annualize == True:
        p25_class_panel_benefit_factoring['incurred_amount'] = p25_class_panel_benefit_factoring['incurred_amount'] * (12 / p25_class_panel_benefit_factoring['data_month'])
        p25_class_panel_benefit_factoring['paid_amount']     = p25_class_panel_benefit_factoring['paid_amount']     * (12 / p25_class_panel_benefit_factoring['data_month'])
        p25_class_panel_benefit_factoring['no_of_cases']     = p25_class_panel_benefit_factoring['no_of_cases']     * (12 / p25_class_panel_benefit_factoring['data_month'])
        p25_class_panel_benefit_factoring['no_of_claim_id']  = p25_class_panel_benefit_factoring['no_of_claim_id']  * (12 / p25_class_panel_benefit_factoring['data_month'])
        p25_class_panel_benefit_factoring['no_of_claimants'] = p25_class_panel_benefit_factoring['no_of_claimants']
      if ibnr == True:
        p25_class_panel_benefit_factoring['incurred_amount'] = p25_class_panel_benefit_factoring['incurred_amount'] * (1 + p25_class_panel_benefit_factoring['ibnr'])
        p25_class_panel_benefit_factoring['paid_amount']     = p25_class_panel_benefit_factoring['paid_amount']     * (1 + p25_class_panel_benefit_factoring['ibnr'])
        p25_class_panel_benefit_factoring['no_of_cases']     = p25_class_panel_benefit_factoring['no_of_cases']     * (1 + p25_class_panel_benefit_factoring['ibnr'])
        p25_class_panel_benefit_factoring['no_of_claim_id']  = p25_class_panel_benefit_factoring['no_of_claim_id']  * (1 + p25_class_panel_benefit_factoring['ibnr'])
        p25_class_panel_benefit_factoring['no_of_claimants'] = p25_class_panel_benefit_factoring['no_of_claimants']
      p25_diff = p25_class_panel_benefit_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] - p25_class_panel_benefit_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
      temp_by = __p25_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p25_diff = p25_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p25_class_panel_benefit_df.reset_index().set_index(temp_by, inplace=True)
      p25_class_panel_benefit_df.loc[p25_class_panel_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] = (
        p25_class_panel_benefit_df.loc[p25_class_panel_benefit_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
        .add(p25_diff)
      )
      temp_by = __p25_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p25_class_panel_benefit_df = p25_class_panel_benefit_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p25_class_panel_benefit_df = p25_class_panel_benefit_df.reindex(benefit_type_order, level='benefit_type')
      p25_class_panel_benefit_df['usage_ratio'] = p25_class_panel_benefit_df['paid_amount'] / p25_class_panel_benefit_df['incurred_amount'].replace(0, pd.NA)
      p25_class_panel_benefit_df['incurred_per_case'] = p25_class_panel_benefit_df['incurred_amount'] / p25_class_panel_benefit_df['no_of_cases'].replace(0, pd.NA)
      p25_class_panel_benefit_df['paid_per_case']     = p25_class_panel_benefit_df['paid_amount']     / p25_class_panel_benefit_df['no_of_cases'].replace(0, pd.NA)
      p25_class_panel_benefit_df['incurred_per_claim'] = p25_class_panel_benefit_df['incurred_amount'] / p25_class_panel_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p25_class_panel_benefit_df['paid_per_claim'] = p25_class_panel_benefit_df['paid_amount'] / p25_class_panel_benefit_df['no_of_claim_id'].replace(0, pd.NA)


    __p25_group_col.remove('policy_id') if 'policy_id' in __p25_group_col else __p25_group_col
    p25_class_panel_benefit_df = p25_class_panel_benefit_df.reset_index().set_index(__p25_group_col)
    p25_class_panel_benefit_df = p25_class_panel_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim', 'no_of_claimants']]
    if research_mode == True:
      p25_class_panel_benefit_df = p25_class_panel_benefit_df.reset_index().groupby(by=__p25_group_col, dropna=False).sum()
      p25_class_panel_benefit_df = p25_class_panel_benefit_df.reindex(benefit_type_order, level='benefit_type')
      p25_class_panel_benefit_df['usage_ratio'] = p25_class_panel_benefit_df['paid_amount'] / p25_class_panel_benefit_df['incurred_amount'].replace(0, pd.NA)
      p25_class_panel_benefit_df['incurred_per_case'] = p25_class_panel_benefit_df['incurred_amount'] / p25_class_panel_benefit_df['no_of_cases'].replace(0, pd.NA)
      p25_class_panel_benefit_df['paid_per_case']     = p25_class_panel_benefit_df['paid_amount']     / p25_class_panel_benefit_df['no_of_cases'].replace(0, pd.NA)
      p25_class_panel_benefit_df['incurred_per_claim'] = p25_class_panel_benefit_df['incurred_amount'] / p25_class_panel_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p25_class_panel_benefit_df['paid_per_claim'] = p25_class_panel_benefit_df['paid_amount'] / p25_class_panel_benefit_df['no_of_claim_id'].replace(0, pd.NA)
      p25_class_panel_benefit_df = p25_class_panel_benefit_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim', 'no_of_claimants']]
      p25_class_panel_benefit_df = p25_class_panel_benefit_df.reindex(benefit_type_order, level='benefit_type')

    self.p25_class_panel_benefit = p25_class_panel_benefit_df
    self.p25 = p25_class_panel_benefit_df
    return p25_class_panel_benefit_df

  def mcr_p26_op_panel(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p26_df_col = by + ['panel', 'benefit', 'incurred_amount', 'paid_amount', 'claim_id', 'claimant']
    __p26_group_col = by + ['panel', 'benefit']
    __p26_sort_col = by + ['panel', 'paid_amount']
    __p26_sort_order = len(by) * [True] + [True, False]
    
    p26_op_panel_df = self.mcr_df[__p26_df_col].loc[self.mcr_df['benefit_type'] == 'Clinic'].groupby(by=__p26_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    # cases
    p26_cases = (
      self.mcr_df[by + ['panel', 'benefit', 'incur_date']]
        .loc[self.mcr_df['benefit_type'] == 'Clinic']
        .groupby(by=__p26_group_col, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p26_cases = p26_cases.reindex(p26_op_panel_df.index, fill_value=0)
    p26_op_panel_df = p26_op_panel_df.join(p26_cases['no_of_cases'])
    p26_op_panel_df['usage_ratio'] = p26_op_panel_df['paid_amount'] / p26_op_panel_df['incurred_amount']
    p26_op_panel_df['incurred_per_claim'] = p26_op_panel_df['incurred_amount'] / p26_op_panel_df['no_of_claim_id']
    p26_op_panel_df['paid_per_claim'] = p26_op_panel_df['paid_amount'] / p26_op_panel_df['no_of_claim_id']
    p26_op_panel_df = p26_op_panel_df.unstack().stack(dropna=False)
    p26_op_panel_df['incurred_per_case'] = p26_op_panel_df['incurred_amount'] / p26_op_panel_df['no_of_cases'].replace(0, pd.NA)
    p26_op_panel_df['paid_per_case']     = p26_op_panel_df['paid_amount']     / p26_op_panel_df['no_of_cases'].replace(0, pd.NA)
    p26_op_panel_df = p26_op_panel_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim', 'no_of_claimants']]
    if len(p26_op_panel_df.index) > 0:
      p26_op_panel_df.sort_values(by=__p26_sort_col, ascending=__p26_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p26_op_panel_df = pd.merge(p26_op_panel_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p26_op_factoring = p26_op_panel_df.copy(deep=True)
        if annualize == True:
          p26_op_factoring['incurred_amount'] = p26_op_factoring['incurred_amount'] * (12 / p26_op_factoring['data_month'])
          p26_op_factoring['paid_amount']     = p26_op_factoring['paid_amount']     * (12 / p26_op_factoring['data_month'])
          p26_op_factoring['no_of_cases']     = p26_op_factoring['no_of_cases']     * (12 / p26_op_factoring['data_month'])
          p26_op_factoring['no_of_claim_id']  = p26_op_factoring['no_of_claim_id']  * (12 / p26_op_factoring['data_month'])
          p26_op_factoring['no_of_claimants'] = p26_op_factoring['no_of_claimants']
        if ibnr == True:
          p26_op_factoring['incurred_amount'] = p26_op_factoring['incurred_amount'] * (1 + p26_op_factoring['ibnr'])
          p26_op_factoring['paid_amount']     = p26_op_factoring['paid_amount']     * (1 + p26_op_factoring['ibnr'])
          p26_op_factoring['no_of_cases']     = p26_op_factoring['no_of_cases']     * (1 + p26_op_factoring['ibnr'])
          p26_op_factoring['no_of_claim_id']  = p26_op_factoring['no_of_claim_id']  * (1 + p26_op_factoring['ibnr'])
          p26_op_factoring['no_of_claimants'] = p26_op_factoring['no_of_claimants']
        p26_diff = p26_op_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] - p26_op_panel_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
        temp_by = __p26_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p26_diff = p26_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
        p26_op_panel_df.reset_index().set_index(temp_by, inplace=True)
        p26_op_panel_df.loc[p26_op_panel_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] = p26_op_panel_df.loc[p26_op_panel_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].add(p26_diff)
        temp_by = __p26_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p26_op_panel_df = p26_op_panel_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
        p26_op_panel_df['usage_ratio'] = p26_op_panel_df['paid_amount'] / p26_op_panel_df['incurred_amount'].replace(0, pd.NA)
        p26_op_panel_df['incurred_per_case'] = p26_op_panel_df['incurred_amount'] / p26_op_panel_df['no_of_cases'].replace(0, pd.NA)
        p26_op_panel_df['paid_per_case']     = p26_op_panel_df['paid_amount']     / p26_op_panel_df['no_of_cases'].replace(0, pd.NA)
        p26_op_panel_df['incurred_per_claim'] = p26_op_panel_df['incurred_amount'] / p26_op_panel_df['no_of_claim_id'].replace(0, pd.NA)
        p26_op_panel_df['paid_per_claim'] = p26_op_panel_df['paid_amount'] / p26_op_panel_df['no_of_claim_id'].replace(0, pd.NA)

    __p26_group_col.remove('policy_id') if 'policy_id' in __p26_group_col else __p26_group_col
    p26_op_panel_df = p26_op_panel_df.reset_index().set_index(__p26_group_col)
    p26_op_panel_df = p26_op_panel_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim']]

    if research_mode == True:
      p26_op_panel_df = p26_op_panel_df.reset_index().groupby(by=__p26_group_col, dropna=False).sum()
      p26_op_panel_df['usage_ratio'] = p26_op_panel_df['paid_amount'] / p26_op_panel_df['incurred_amount'].replace(0, pd.NA)
      p26_op_panel_df['incurred_per_case'] = p26_op_panel_df['incurred_amount'] / p26_op_panel_df['no_of_cases'].replace(0, pd.NA)
      p26_op_panel_df['paid_per_case']     = p26_op_panel_df['paid_amount']     / p26_op_panel_df['no_of_cases'].replace(0, pd.NA)
      p26_op_panel_df['incurred_per_claim'] = p26_op_panel_df['incurred_amount'] / p26_op_panel_df['no_of_claim_id'].replace(0, pd.NA)
      p26_op_panel_df['paid_per_claim'] = p26_op_panel_df['paid_amount'] / p26_op_panel_df['no_of_claim_id'].replace(0, pd.NA)
      p26_op_panel_df = p26_op_panel_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'no_of_claimants', 'incurred_per_claim', 'paid_per_claim']]
    self.p26_op_panel = p26_op_panel_df
    self.p26 = p26_op_panel_df
    return p26_op_panel_df
  
  def mcr_p26_op_class_panel(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p26_df_col = by + ['class', 'panel', 'benefit', 'incurred_amount', 'paid_amount', 'claim_id']
    __p26_group_col = by + ['class', 'panel', 'benefit']
    __p26_sort_col = by + ['class', 'panel', 'paid_amount']
    __p26_sort_order = len(by) * [True] + [True, True, False]
    
    p26_op_class_panel_df = self.mcr_df[__p26_df_col].loc[self.mcr_df['benefit_type'] == 'Clinic'].groupby(by=__p26_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id'})
    # cases
    p26a_cases = (
      self.mcr_df[by + ['class', 'panel', 'benefit', 'incur_date']]
        .loc[self.mcr_df['benefit_type'] == 'Clinic']
        .groupby(by=__p26_group_col, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p26a_cases = p26a_cases.reindex(p26_op_class_panel_df.index, fill_value=0)
    p26_op_class_panel_df = p26_op_class_panel_df.join(p26a_cases['no_of_cases'])
    p26_op_class_panel_df['usage_ratio'] = p26_op_class_panel_df['paid_amount'] / p26_op_class_panel_df['incurred_amount']
    p26_op_class_panel_df['incurred_per_claim'] = p26_op_class_panel_df['incurred_amount'] / p26_op_class_panel_df['no_of_claim_id']
    p26_op_class_panel_df['paid_per_claim'] = p26_op_class_panel_df['paid_amount'] / p26_op_class_panel_df['no_of_claim_id']
    p26_op_class_panel_df = p26_op_class_panel_df.unstack().stack(dropna=False)
    p26_op_class_panel_df['incurred_per_case'] = p26_op_class_panel_df['incurred_amount'] / p26_op_class_panel_df['no_of_cases'].replace(0, pd.NA)
    p26_op_class_panel_df['paid_per_case']     = p26_op_class_panel_df['paid_amount']     / p26_op_class_panel_df['no_of_cases'].replace(0, pd.NA)
    p26_op_class_panel_df = p26_op_class_panel_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    
    if len(p26_op_class_panel_df.index) > 0:
      p26_op_class_panel_df.sort_values(by=__p26_sort_col, ascending=__p26_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p26_op_class_panel_df = pd.merge(p26_op_class_panel_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p26_op_class_factoring = p26_op_class_panel_df.copy(deep=True)
        if annualize == True:
          p26_op_class_factoring['incurred_amount'] = p26_op_class_factoring['incurred_amount'] * (12 / p26_op_class_factoring['data_month'])
          p26_op_class_factoring['paid_amount']     = p26_op_class_factoring['paid_amount']     * (12 / p26_op_class_factoring['data_month'])
          p26_op_class_factoring['no_of_cases']     = p26_op_class_factoring['no_of_cases']     * (12 / p26_op_class_factoring['data_month'])
          p26_op_class_factoring['no_of_claim_id']  = p26_op_class_factoring['no_of_claim_id']  * (12 / p26_op_class_factoring['data_month'])
        if ibnr == True:
          p26_op_class_factoring['incurred_amount'] = p26_op_class_factoring['incurred_amount'] * (1 + p26_op_class_factoring['ibnr'])
          p26_op_class_factoring['paid_amount']     = p26_op_class_factoring['paid_amount']     * (1 + p26_op_class_factoring['ibnr'])
          p26_op_class_factoring['no_of_cases']     = p26_op_class_factoring['no_of_cases']     * (1 + p26_op_class_factoring['ibnr'])
          p26_op_class_factoring['no_of_claim_id']  = p26_op_class_factoring['no_of_claim_id']  * (1 + p26_op_class_factoring['ibnr'])
        p26a_diff = p26_op_class_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] - p26_op_class_panel_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']]
        temp_by = __p26_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p26a_diff = p26a_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
        p26_op_class_panel_df.reset_index().set_index(temp_by, inplace=True)
        p26_op_class_panel_df.loc[p26_op_class_panel_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] = p26_op_class_panel_df.loc[p26_op_class_panel_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].add(p26a_diff)
        temp_by = __p26_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p26_op_class_panel_df = p26_op_class_panel_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
        p26_op_class_panel_df['usage_ratio'] = p26_op_class_panel_df['paid_amount'] / p26_op_class_panel_df['incurred_amount'].replace(0, pd.NA)
        p26_op_class_panel_df['incurred_per_case'] = p26_op_class_panel_df['incurred_amount'] / p26_op_class_panel_df['no_of_cases'].replace(0, pd.NA)
        p26_op_class_panel_df['paid_per_case']     = p26_op_class_panel_df['paid_amount']     / p26_op_class_panel_df['no_of_cases'].replace(0, pd.NA)
        p26_op_class_panel_df['incurred_per_claim'] = p26_op_class_panel_df['incurred_amount'] / p26_op_class_panel_df['no_of_claim_id'].replace(0, pd.NA)
        p26_op_class_panel_df['paid_per_claim'] = p26_op_class_panel_df['paid_amount'] / p26_op_class_panel_df['no_of_claim_id'].replace(0, pd.NA)

    __p26_group_col.remove('policy_id') if 'policy_id' in __p26_group_col else __p26_group_col
    p26_op_class_panel_df = p26_op_class_panel_df.reset_index().set_index(__p26_group_col)
    p26_op_class_panel_df = p26_op_class_panel_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    if research_mode == True:
      p26_op_class_panel_df = p26_op_class_panel_df.reset_index().groupby(by=__p26_group_col, dropna=False).sum()
      p26_op_class_panel_df['usage_ratio'] = p26_op_class_panel_df['paid_amount'] / p26_op_class_panel_df['incurred_amount'].replace(0, pd.NA)
      p26_op_class_panel_df['incurred_per_case'] = p26_op_class_panel_df['incurred_amount'] / p26_op_class_panel_df['no_of_cases'].replace(0, pd.NA)
      p26_op_class_panel_df['paid_per_case']     = p26_op_class_panel_df['paid_amount']     / p26_op_class_panel_df['no_of_cases'].replace(0, pd.NA)
      p26_op_class_panel_df['incurred_per_claim'] = p26_op_class_panel_df['incurred_amount'] / p26_op_class_panel_df['no_of_claim_id'].replace(0, pd.NA)
      p26_op_class_panel_df['paid_per_claim'] = p26_op_class_panel_df['paid_amount'] / p26_op_class_panel_df['no_of_claim_id'].replace(0, pd.NA)
      p26_op_class_panel_df = p26_op_class_panel_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    self.p26_op_class_panel = p26_op_class_panel_df

    return p26_op_class_panel_df
  
  def mcr_p26_ip_panel(self, by=None, annualize=False, ibnr=False, research_mode=False):
    self.ip_order = self.benefit_index.gum_benefit.loc[(self.benefit_index.gum_benefit_type == 'INPATIENT BENEFITS /HOSPITALIZATION') | (self.benefit_index.gum_benefit_type == 'MATERNITY') | (self.benefit_index.gum_benefit_type == 'SUPPLEMENTARY MAJOR MEDICAL')].drop_duplicates(keep='last').values.tolist()
    
    __p26_df_col = by + ['panel', 'benefit', 'incurred_amount', 'paid_amount', 'claim_id']
    __p26_group_col = by + ['panel', 'benefit']
  
    p26_ip_panel_df = self.mcr_df[__p26_df_col].loc[self.mcr_df['benefit_type'] == 'Hospital'].groupby(by=__p26_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id'})
    # cases
    p26b_cases = (
      self.mcr_df[by + ['panel', 'benefit', 'incur_date']]
        .loc[self.mcr_df['benefit_type'] == 'Hospital']
        .groupby(by=__p26_group_col, dropna=False)
        .count()
        .rename(columns={'incur_date': 'no_of_cases'})
    )
    p26b_cases = p26b_cases.reindex(p26_ip_panel_df.index, fill_value=0)
    p26_ip_panel_df = p26_ip_panel_df.join(p26b_cases['no_of_cases'])
    p26_ip_panel_df['usage_ratio'] = p26_ip_panel_df['paid_amount'] / p26_ip_panel_df['incurred_amount']
    p26_ip_panel_df['incurred_per_claim'] = p26_ip_panel_df['incurred_amount'] / p26_ip_panel_df['no_of_claim_id']
    p26_ip_panel_df['paid_per_claim'] = p26_ip_panel_df['paid_amount'] / p26_ip_panel_df['no_of_claim_id']
    p26_ip_panel_df = p26_ip_panel_df.unstack().stack(dropna=False)
    p26_ip_panel_df = p26_ip_panel_df.reindex(self.ip_order, level='benefit')
    p26_ip_panel_df['incurred_per_case'] = p26_ip_panel_df['incurred_amount'] / p26_ip_panel_df['no_of_cases'].replace(0, pd.NA)
    p26_ip_panel_df['paid_per_case']     = p26_ip_panel_df['paid_amount']     / p26_ip_panel_df['no_of_cases'].replace(0, pd.NA)
    if annualize == True or ibnr == True:
      p26_ip_panel_df = pd.merge(p26_ip_panel_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p26_ip_factoring = p26_ip_panel_df.copy(deep=True)
      if annualize == True:
        p26_ip_factoring['incurred_amount'] = p26_ip_factoring['incurred_amount'] * (12 / p26_ip_factoring['data_month'])
        p26_ip_factoring['paid_amount']     = p26_ip_factoring['paid_amount']     * (12 / p26_ip_factoring['data_month'])
        p26_ip_factoring['no_of_cases']     = p26_ip_factoring['no_of_cases']     * (12 / p26_ip_factoring['data_month'])
        p26_ip_factoring['no_of_claim_id']  = p26_ip_factoring['no_of_claim_id']  * (12 / p26_ip_factoring['data_month'])
      if ibnr == True:
        p26_ip_factoring['incurred_amount'] = p26_ip_factoring['incurred_amount'] * (1 + p26_ip_factoring['ibnr'])
        p26_ip_factoring['paid_amount']     = p26_ip_factoring['paid_amount']     * (1 + p26_ip_factoring['ibnr'])
        p26_ip_factoring['no_of_cases']     = p26_ip_factoring['no_of_cases']     * (1 + p26_ip_factoring['ibnr'])
        p26_ip_factoring['no_of_claim_id']  = p26_ip_factoring['no_of_claim_id']  * (1 + p26_ip_factoring['ibnr'])
      p26b_diff = p26_ip_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] - p26_ip_panel_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']]
      temp_by = __p26_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p26b_diff = p26b_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
      p26_ip_panel_df.reset_index().set_index(temp_by, inplace=True)
      p26_ip_panel_df.loc[p26_ip_panel_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']] = p26_ip_panel_df.loc[p26_ip_panel_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].add(p26b_diff)
      temp_by = __p26_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p26_ip_panel_df = p26_ip_panel_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id']].groupby(by=temp_by, dropna=False).sum()
      p26_ip_panel_df = p26_ip_panel_df.reindex(self.ip_order, level='benefit')
      p26_ip_panel_df['usage_ratio'] = p26_ip_panel_df['paid_amount'] / p26_ip_panel_df['incurred_amount'].replace(0, pd.NA)
      p26_ip_panel_df['incurred_per_case'] = p26_ip_panel_df['incurred_amount'] / p26_ip_panel_df['no_of_cases'].replace(0, pd.NA)
      p26_ip_panel_df['paid_per_case']     = p26_ip_panel_df['paid_amount']     / p26_ip_panel_df['no_of_cases'].replace(0, pd.NA)
      p26_ip_panel_df['incurred_per_claim'] = p26_ip_panel_df['incurred_amount'] / p26_ip_panel_df['no_of_claim_id'].replace(0, pd.NA)
      p26_ip_panel_df['paid_per_claim'] = p26_ip_panel_df['paid_amount'] / p26_ip_panel_df['no_of_claim_id'].replace(0, pd.NA)

    __p26_group_col.remove('policy_id') if 'policy_id' in __p26_group_col else __p26_group_col
    p26_ip_panel_df = p26_ip_panel_df.reset_index().set_index(__p26_group_col)
    p26_ip_panel_df = p26_ip_panel_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
    if research_mode == True:
      p26_ip_panel_df = p26_ip_panel_df.reset_index().groupby(by=__p26_group_col, dropna=False).sum()
      p26_ip_panel_df['usage_ratio'] = p26_ip_panel_df['paid_amount'] / p26_ip_panel_df['incurred_amount'].replace(0, pd.NA)
      p26_ip_panel_df['incurred_per_case'] = p26_ip_panel_df['incurred_amount'] / p26_ip_panel_df['no_of_cases'].replace(0, pd.NA)
      p26_ip_panel_df['paid_per_case']     = p26_ip_panel_df['paid_amount']     / p26_ip_panel_df['no_of_cases'].replace(0, pd.NA)
      p26_ip_panel_df['incurred_per_claim'] = p26_ip_panel_df['incurred_amount'] / p26_ip_panel_df['no_of_claim_id'].replace(0, pd.NA)
      p26_ip_panel_df['paid_per_claim'] = p26_ip_panel_df['paid_amount'] / p26_ip_panel_df['no_of_claim_id'].replace(0, pd.NA)
      p26_ip_panel_df = p26_ip_panel_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'incurred_per_case', 'paid_per_case', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']]
      p26_ip_panel_df = p26_ip_panel_df.reindex(self.ip_order, level='benefit')

    self.p26_ip_panel = p26_ip_panel_df
    return p26_ip_panel_df
  
  def mcr_p18a_top_diag_ip(self, by=None, annualize=False, ibnr=False, research_mode=False):

    if 'diagnosis' not in by:
      __p18_df_col = by + ['diagnosis', 'incurred_amount', 'paid_amount', 'claimant', 'claim_id']
      __p18_group_col = by + ['diagnosis']
    else:
      __p18_df_col = by + ['incurred_amount', 'paid_amount', 'claimant', 'claim_id']
      __p18_group_col = by

    __p18_sort_col = by + ['paid_amount']
    __p18_sort_order = len(by) * [True] + [False]

    p18a_df = self.mcr_df[__p18_df_col].loc[(self.mcr_df['benefit_type'] == 'Hospital')].groupby(by=__p18_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique', 'claim_id': 'nunique'}).rename(columns={'claimant': 'no_of_claimants', 'claim_id': 'no_of_claim_id'})
    case_mask = (
      self.mcr_df['benefit_type'].eq('Hospital')
      & self.mcr_df['benefit'].str.contains('Day Centre|Surgeon', case=False, na=False)
    )
    case_counts = (
      self.mcr_df.loc[case_mask, __p18_group_col + ['incur_date']]
      .groupby(__p18_group_col, dropna=False)['incur_date']
      .count()
      .rename('no_of_cases')
    )
    case_counts = case_counts.reindex(p18a_df.index, fill_value=0)
    p18a_df = p18a_df.join(case_counts)
    p18a_df = p18a_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
    if len(p18a_df) > 0:
      p18a_df.sort_values(by=__p18_sort_col, ascending=__p18_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p18a_df = pd.merge(p18a_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p18a_factoring = p18a_df.copy(deep=True)
        if annualize == True:
          p18a_factoring['incurred_amount'] = p18a_factoring['incurred_amount'] * (12 / p18a_factoring['data_month'])
          p18a_factoring['paid_amount']     = p18a_factoring['paid_amount']     * (12 / p18a_factoring['data_month'])
          p18a_factoring['no_of_cases']     = p18a_factoring['no_of_cases']     * (12 / p18a_factoring['data_month'])
          p18a_factoring['no_of_claim_id']  = p18a_factoring['no_of_claim_id']  * (12 / p18a_factoring['data_month'])
          p18a_factoring['no_of_claimants'] = p18a_factoring['no_of_claimants']
        if ibnr == True:
          p18a_factoring['incurred_amount'] = p18a_factoring['incurred_amount'] * (1 + p18a_factoring['ibnr'])
          p18a_factoring['paid_amount']     = p18a_factoring['paid_amount']     * (1 + p18a_factoring['ibnr'])
          p18a_factoring['no_of_cases']     = p18a_factoring['no_of_cases']     * (1 + p18a_factoring['ibnr'])
          p18a_factoring['no_of_claim_id']  = p18a_factoring['no_of_claim_id']  * (1 + p18a_factoring['ibnr'])
          p18a_factoring['no_of_claimants'] = p18a_factoring['no_of_claimants']
        p18a_diff = p18a_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] - p18a_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
        temp_by = __p18_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p18a_diff = p18a_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
        p18a_df.reset_index().set_index(temp_by, inplace=True)
        p18a_df.loc[p18a_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] = p18a_df.loc[p18a_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].add(p18a_diff)
        temp_by = __p18_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p18a_df = p18a_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p18_group_col.remove('policy_id') if 'policy_id' in __p18_group_col else __p18_group_col
    p18a_df = p18a_df.reset_index().set_index(__p18_group_col)
    p18a_df = p18a_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
    if research_mode == True:
      p18a_df = p18a_df.reset_index().groupby(by=__p18_group_col, dropna=False).sum()
      p18a_df['usage_ratio'] = p18a_df['paid_amount'] / p18a_df['incurred_amount'].replace(0, pd.NA)
      p18a_df = p18a_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
    self.p18a = p18a_df
    return p18a_df
  
  def mcr_p18a_top_diag_ip_class(self, by=None, annualize=False, ibnr=False, research_mode=False):

    if by is None:
      by = []

    __p18_group_col = by.copy()
    if 'class' not in __p18_group_col:
      __p18_group_col.append('class')
    if 'diagnosis' not in __p18_group_col:
      __p18_group_col.append('diagnosis')

    __p18_df_col = list(dict.fromkeys(__p18_group_col + ['incurred_amount', 'paid_amount', 'claimant', 'claim_id']))
    __p18_sort_col = __p18_group_col + ['paid_amount']
    __p18_sort_order = len(__p18_group_col) * [True] + [False]

    p18a_class_df = self.mcr_df[__p18_df_col].loc[(self.mcr_df['benefit_type'] == 'Hospital')].groupby(by=__p18_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique', 'claim_id': 'nunique'}).rename(columns={'claimant': 'no_of_claimants', 'claim_id': 'no_of_claim_id'})
    class_case_mask = (
      self.mcr_df['benefit_type'].eq('Hospital')
      & self.mcr_df['benefit'].str.contains('Day Centre|Surgeon', case=False, na=False)
    )
    class_case_counts = (
      self.mcr_df.loc[class_case_mask, __p18_group_col + ['incur_date']]
      .groupby(__p18_group_col, dropna=False)['incur_date']
      .count()
      .rename('no_of_cases')
    )
    class_case_counts = class_case_counts.reindex(p18a_class_df.index, fill_value=0)
    p18a_class_df = p18a_class_df.join(class_case_counts)
    p18a_class_df = p18a_class_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
    if len(p18a_class_df) > 0:
      p18a_class_df.sort_values(by=__p18_sort_col, ascending=__p18_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p18a_class_df = pd.merge(p18a_class_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p18a_class_factoring = p18a_class_df.copy(deep=True)
        if annualize == True:
          p18a_class_factoring['incurred_amount'] = p18a_class_factoring['incurred_amount'] * (12 / p18a_class_factoring['data_month'])
          p18a_class_factoring['paid_amount']     = p18a_class_factoring['paid_amount']     * (12 / p18a_class_factoring['data_month'])
          p18a_class_factoring['no_of_cases']     = p18a_class_factoring['no_of_cases']     * (12 / p18a_class_factoring['data_month'])
          p18a_class_factoring['no_of_claim_id']  = p18a_class_factoring['no_of_claim_id']  * (12 / p18a_class_factoring['data_month'])
          p18a_class_factoring['no_of_claimants'] = p18a_class_factoring['no_of_claimants']
        if ibnr == True:
          p18a_class_factoring['incurred_amount'] = p18a_class_factoring['incurred_amount'] * (1 + p18a_class_factoring['ibnr'])
          p18a_class_factoring['paid_amount']     = p18a_class_factoring['paid_amount']     * (1 + p18a_class_factoring['ibnr'])
          p18a_class_factoring['no_of_cases']     = p18a_class_factoring['no_of_cases']     * (1 + p18a_class_factoring['ibnr'])
          p18a_class_factoring['no_of_claim_id']  = p18a_class_factoring['no_of_claim_id']  * (1 + p18a_class_factoring['ibnr'])
          p18a_class_factoring['no_of_claimants'] = p18a_class_factoring['no_of_claimants']
        p18a_class_diff = p18a_class_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] - p18a_class_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
        temp_by = __p18_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p18a_class_diff = p18a_class_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
        p18a_class_df.reset_index().set_index(temp_by, inplace=True)
        p18a_class_df.loc[p18a_class_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] = p18a_class_df.loc[p18a_class_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].add(p18a_class_diff)
        temp_by = __p18_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p18a_class_df = p18a_class_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p18_group_col.remove('policy_id') if 'policy_id' in __p18_group_col else __p18_group_col
    p18a_class_df = p18a_class_df.reset_index().set_index(__p18_group_col)
    p18a_class_df = p18a_class_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
    if research_mode == True:
      p18a_class_df = p18a_class_df.reset_index().groupby(by=__p18_group_col, dropna=False).sum()
      p18a_class_df['usage_ratio'] = p18a_class_df['paid_amount'] / p18a_class_df['incurred_amount'].replace(0, pd.NA)
      p18a_class_df = p18a_class_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]

    self.p18a_class = p18a_class_df
    return p18a_class_df
  
  def mcr_p18b_top_diag_ip_day_procedure(self, by=None, annualize=False, ibnr=False, research_mode=False):

    if by is None:
      by = []

    __p18_group_col = by.copy()
    if 'diagnosis' not in __p18_group_col:
      __p18_group_col.append('diagnosis')
    if 'day_procedure_flag' not in __p18_group_col:
      __p18_group_col.append('day_procedure_flag')

    __p18_df_col = list(dict.fromkeys(__p18_group_col + ['incurred_amount', 'paid_amount', 'claimant', 'claim_id']))
    __p18_sort_col = __p18_group_col + ['paid_amount']
    __p18_sort_order = len(__p18_group_col) * [True] + [False]

    ip_mask = self.mcr_df['benefit_type'] == 'Hospital'
    p18b_dp_df = self.mcr_df.loc[ip_mask, __p18_df_col].groupby(by=__p18_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique', 'claim_id': 'nunique'}).rename(columns={'claimant': 'no_of_claimants', 'claim_id': 'no_of_claim_id'})
    case_mask = ip_mask & self.mcr_df['benefit'].str.contains('Day Centre|Surgeon', case=False, na=False)
    case_cols = list(dict.fromkeys(__p18_group_col + ['incur_date']))
    case_counts = self.mcr_df.loc[case_mask, case_cols].groupby(__p18_group_col, dropna=False)['incur_date'].count().rename('no_of_cases')
    case_counts = case_counts.reindex(p18b_dp_df.index, fill_value=0)
    p18b_dp_df = p18b_dp_df.join(case_counts)
    p18b_dp_df = p18b_dp_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
    if len(p18b_dp_df) > 0:
      p18b_dp_df.sort_values(by=__p18_sort_col, ascending=__p18_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p18b_dp_df = pd.merge(p18b_dp_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p18b_dp_factoring = p18b_dp_df.copy(deep=True)
        if annualize == True:
          p18b_dp_factoring['incurred_amount'] = p18b_dp_factoring['incurred_amount'] * (12 / p18b_dp_factoring['data_month'])
          p18b_dp_factoring['paid_amount']     = p18b_dp_factoring['paid_amount']     * (12 / p18b_dp_factoring['data_month'])
          p18b_dp_factoring['no_of_cases']     = p18b_dp_factoring['no_of_cases']     * (12 / p18b_dp_factoring['data_month'])
          p18b_dp_factoring['no_of_claim_id']  = p18b_dp_factoring['no_of_claim_id']  * (12 / p18b_dp_factoring['data_month'])
          p18b_dp_factoring['no_of_claimants'] = p18b_dp_factoring['no_of_claimants']
        if ibnr == True:
          p18b_dp_factoring['incurred_amount'] = p18b_dp_factoring['incurred_amount'] * (1 + p18b_dp_factoring['ibnr'])
          p18b_dp_factoring['paid_amount']     = p18b_dp_factoring['paid_amount']     * (1 + p18b_dp_factoring['ibnr'])
          p18b_dp_factoring['no_of_cases']     = p18b_dp_factoring['no_of_cases']     * (1 + p18b_dp_factoring['ibnr'])
          p18b_dp_factoring['no_of_claim_id']  = p18b_dp_factoring['no_of_claim_id']  * (1 + p18b_dp_factoring['ibnr'])
          p18b_dp_factoring['no_of_claimants'] = p18b_dp_factoring['no_of_claimants']
        p18b_dp_diff = p18b_dp_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] - p18b_dp_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
        temp_by = __p18_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p18b_dp_diff = p18b_dp_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
        p18b_dp_df.reset_index().set_index(temp_by, inplace=True)
        p18b_dp_df.loc[p18b_dp_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] = p18b_dp_df.loc[p18b_dp_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].add(p18b_dp_diff)
        temp_by = __p18_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p18b_dp_df = p18b_dp_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p18_group_col.remove('policy_id') if 'policy_id' in __p18_group_col else __p18_group_col
    p18b_dp_df = p18b_dp_df.reset_index().set_index(__p18_group_col)
    p18b_dp_df = p18b_dp_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
    if research_mode == True:
      p18b_dp_df = p18b_dp_df.reset_index().groupby(by=__p18_group_col, dropna=False).sum()
      p18b_dp_df['usage_ratio'] = p18b_dp_df['paid_amount'] / p18b_dp_df['incurred_amount'].replace(0, pd.NA)
      p18b_dp_df = p18b_dp_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
    self.p18b_ip_day_procedure = p18b_dp_df
    return p18b_dp_df
  
  def mcr_p18b_top_diag_ip_day_procedure_class(self, by=None, annualize=False, ibnr=False, research_mode=False):

    if by is None:
      by = []

    __p18_group_col = by.copy()
    if 'class' not in __p18_group_col:
      __p18_group_col.append('class')
    if 'diagnosis' not in __p18_group_col:
      __p18_group_col.append('diagnosis')
    if 'day_procedure_flag' not in __p18_group_col:
      __p18_group_col.append('day_procedure_flag')

    __p18_df_col = list(dict.fromkeys(__p18_group_col + ['incurred_amount', 'paid_amount', 'claimant', 'claim_id']))
    __p18_sort_col = __p18_group_col + ['paid_amount']
    __p18_sort_order = len(__p18_group_col) * [True] + [False]

    ip_mask = self.mcr_df['benefit_type'] == 'Hospital'
    p18b_dp_class_df = self.mcr_df.loc[ip_mask, __p18_df_col].groupby(by=__p18_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique', 'claim_id': 'nunique'}).rename(columns={'claimant': 'no_of_claimants', 'claim_id': 'no_of_claim_id'})
    class_case_mask = ip_mask & self.mcr_df['benefit'].str.contains('Day Centre|Surgeon', case=False, na=False)
    class_case_cols = list(dict.fromkeys(__p18_group_col + ['incur_date']))
    class_case_counts = self.mcr_df.loc[class_case_mask, class_case_cols].groupby(__p18_group_col, dropna=False)['incur_date'].count().rename('no_of_cases')
    class_case_counts = class_case_counts.reindex(p18b_dp_class_df.index, fill_value=0)
    p18b_dp_class_df = p18b_dp_class_df.join(class_case_counts)
    p18b_dp_class_df = p18b_dp_class_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
    if len(p18b_dp_class_df) > 0:
      p18b_dp_class_df.sort_values(by=__p18_sort_col, ascending=__p18_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p18b_dp_class_df = pd.merge(p18b_dp_class_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p18b_dp_class_factoring = p18b_dp_class_df.copy(deep=True)
        if annualize == True:
          p18b_dp_class_factoring['incurred_amount'] = p18b_dp_class_factoring['incurred_amount'] * (12 / p18b_dp_class_factoring['data_month'])
          p18b_dp_class_factoring['paid_amount']     = p18b_dp_class_factoring['paid_amount']     * (12 / p18b_dp_class_factoring['data_month'])
          p18b_dp_class_factoring['no_of_cases']     = p18b_dp_class_factoring['no_of_cases']     * (12 / p18b_dp_class_factoring['data_month'])
          p18b_dp_class_factoring['no_of_claim_id']  = p18b_dp_class_factoring['no_of_claim_id']  * (12 / p18b_dp_class_factoring['data_month'])
          p18b_dp_class_factoring['no_of_claimants'] = p18b_dp_class_factoring['no_of_claimants']
        if ibnr == True:
          p18b_dp_class_factoring['incurred_amount'] = p18b_dp_class_factoring['incurred_amount'] * (1 + p18b_dp_class_factoring['ibnr'])
          p18b_dp_class_factoring['paid_amount']     = p18b_dp_class_factoring['paid_amount']     * (1 + p18b_dp_class_factoring['ibnr'])
          p18b_dp_class_factoring['no_of_cases']     = p18b_dp_class_factoring['no_of_cases']     * (1 + p18b_dp_class_factoring['ibnr'])
          p18b_dp_class_factoring['no_of_claim_id']  = p18b_dp_class_factoring['no_of_claim_id']  * (1 + p18b_dp_class_factoring['ibnr'])
          p18b_dp_class_factoring['no_of_claimants'] = p18b_dp_class_factoring['no_of_claimants']
        p18b_dp_class_diff = p18b_dp_class_factoring[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] - p18b_dp_class_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
        temp_by = __p18_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p18b_dp_class_diff = p18b_dp_class_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
        p18b_dp_class_df.reset_index().set_index(temp_by, inplace=True)
        p18b_dp_class_df.loc[p18b_dp_class_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']] = p18b_dp_class_df.loc[p18b_dp_class_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].add(p18b_dp_class_diff)
        temp_by = __p18_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p18b_dp_class_df = p18b_dp_class_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p18_group_col.remove('policy_id') if 'policy_id' in __p18_group_col else __p18_group_col
    p18b_dp_class_df = p18b_dp_class_df.reset_index().set_index(__p18_group_col)
    p18b_dp_class_df = p18b_dp_class_df[['incurred_amount', 'paid_amount', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]
    if research_mode == True:
      p18b_dp_class_df = p18b_dp_class_df.reset_index().groupby(by=__p18_group_col, dropna=False).sum()
      p18b_dp_class_df['usage_ratio'] = p18b_dp_class_df['paid_amount'] / p18b_dp_class_df['incurred_amount'].replace(0, pd.NA)
      p18b_dp_class_df = p18b_dp_class_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_cases', 'no_of_claim_id', 'no_of_claimants']]

    self.p18b_ip_day_procedure_class = p18b_dp_class_df
    return p18b_dp_class_df
  
  def mcr_p18b_top_diag_op(self, by=None, annualize=False, ibnr=False, research_mode=False):

    if 'diagnosis' not in by:
      __p18_df_col = by + ['diagnosis', 'incurred_amount', 'paid_amount', 'claimant', 'claim_id']
      __p18_group_col = by + ['diagnosis']
    else:
      __p18_df_col = by + ['incurred_amount', 'paid_amount', 'claimant', 'claim_id']
      __p18_group_col = by
    __p18_sort_col = by + ['paid_amount']
    __p18_sort_order = len(by) * [True] + [False]

    p18b_df = self.mcr_df[__p18_df_col].loc[(self.mcr_df['benefit_type'] == 'Clinic')].groupby(by=__p18_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique', 'claim_id': 'nunique'}).rename(columns={'claimant': 'no_of_claimants', 'claim_id': 'no_of_claim_id'})
    p18b_df = p18b_df[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    if len(p18b_df) > 0:
      p18b_df.sort_values(by=__p18_sort_col, ascending=__p18_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p18b_df = pd.merge(p18b_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p18b_factoring = p18b_df.copy(deep=True)
        if annualize == True:
          p18b_factoring['incurred_amount'] = p18b_factoring['incurred_amount'] * (12 / p18b_factoring['data_month'])
          p18b_factoring['paid_amount']     = p18b_factoring['paid_amount']     * (12 / p18b_factoring['data_month'])
          p18b_factoring['no_of_claim_id']  = p18b_factoring['no_of_claim_id']  * (12 / p18b_factoring['data_month'])
          p18b_factoring['no_of_claimants'] = p18b_factoring['no_of_claimants']
        if ibnr == True:
          p18b_factoring['incurred_amount'] = p18b_factoring['incurred_amount'] * (1 + p18b_factoring['ibnr'])
          p18b_factoring['paid_amount']     = p18b_factoring['paid_amount']     * (1 + p18b_factoring['ibnr'])
          p18b_factoring['no_of_claim_id']  = p18b_factoring['no_of_claim_id']  * (1 + p18b_factoring['ibnr'])
          p18b_factoring['no_of_claimants'] = p18b_factoring['no_of_claimants']
        p18b_diff = p18b_factoring[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] - p18b_df[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
        temp_by = __p18_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p18b_diff = p18b_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
        p18b_df.reset_index().set_index(temp_by, inplace=True)
        p18b_df.loc[p18b_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] = p18b_df.loc[p18b_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].add(p18b_diff)
        temp_by = __p18_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p18b_df = p18b_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p18_group_col.remove('policy_id') if 'policy_id' in __p18_group_col else __p18_group_col
    p18b_df = p18b_df.reset_index().set_index(__p18_group_col)
    p18b_df = p18b_df[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    if research_mode == True:
      p18b_df = p18b_df.reset_index().groupby(by=__p18_group_col, dropna=False).sum()
      p18b_df['usage_ratio'] = p18b_df['paid_amount'] / p18b_df['incurred_amount'].replace(0, pd.NA)
      p18b_df = p18b_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'no_of_claimants']]
    self.p18b = p18b_df
    return p18b_df

  def mcr_p18b_top_diag_op_class(self, by=None, annualize=False, ibnr=False, research_mode=False):

    if by is None:
      by = []

    __p18_group_col = by.copy()
    if 'class' not in __p18_group_col:
      __p18_group_col.append('class')
    if 'diagnosis' not in __p18_group_col:
      __p18_group_col.append('diagnosis')

    __p18_df_col = list(dict.fromkeys(__p18_group_col + ['incurred_amount', 'paid_amount', 'claimant', 'claim_id']))
    __p18_sort_col = __p18_group_col + ['paid_amount']
    __p18_sort_order = len(__p18_group_col) * [True] + [False]

    p18b_class_df = self.mcr_df[__p18_df_col].loc[(self.mcr_df['benefit_type'] == 'Clinic')].groupby(by=__p18_group_col, dropna=False).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique', 'claim_id': 'nunique'}).rename(columns={'claimant': 'no_of_claimants', 'claim_id': 'no_of_claim_id'})
    p18b_class_df = p18b_class_df[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    if len(p18b_class_df) > 0:
      p18b_class_df.sort_values(by=__p18_sort_col, ascending=__p18_sort_order, inplace=True)
      if annualize == True or ibnr == True:
        p18b_class_df = pd.merge(p18b_class_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p18b_class_factoring = p18b_class_df.copy(deep=True)
        if annualize == True:
          p18b_class_factoring['incurred_amount'] = p18b_class_factoring['incurred_amount'] * (12 / p18b_class_factoring['data_month'])
          p18b_class_factoring['paid_amount']     = p18b_class_factoring['paid_amount']     * (12 / p18b_class_factoring['data_month'])
          p18b_class_factoring['no_of_claim_id']  = p18b_class_factoring['no_of_claim_id']  * (12 / p18b_class_factoring['data_month'])
          p18b_class_factoring['no_of_claimants'] = p18b_class_factoring['no_of_claimants']
        if ibnr == True:
          p18b_class_factoring['incurred_amount'] = p18b_class_factoring['incurred_amount'] * (1 + p18b_class_factoring['ibnr'])
          p18b_class_factoring['paid_amount']     = p18b_class_factoring['paid_amount']     * (1 + p18b_class_factoring['ibnr'])
          p18b_class_factoring['no_of_claim_id']  = p18b_class_factoring['no_of_claim_id']  * (1 + p18b_class_factoring['ibnr'])
          p18b_class_factoring['no_of_claimants'] = p18b_class_factoring['no_of_claimants']
        p18b_class_diff = p18b_class_factoring[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] - p18b_class_df[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
        temp_by = __p18_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p18b_class_diff = p18b_class_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
        p18b_class_df.reset_index().set_index(temp_by, inplace=True)
        p18b_class_df.loc[p18b_class_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] = p18b_class_df.loc[p18b_class_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].add(p18b_class_diff)
        temp_by = __p18_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p18b_class_df = p18b_class_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p18_group_col.remove('policy_id') if 'policy_id' in __p18_group_col else __p18_group_col
    p18b_class_df = p18b_class_df.reset_index().set_index(__p18_group_col)
    p18b_class_df = p18b_class_df[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    if research_mode == True:
      p18b_class_df = p18b_class_df.reset_index().groupby(by=__p18_group_col, dropna=False).sum()
      p18b_class_df['usage_ratio'] = p18b_class_df['paid_amount'] / p18b_class_df['incurred_amount'].replace(0, pd.NA)
      p18b_class_df = p18b_class_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'no_of_claimants']]

    self.p18b_class = p18b_class_df
    return p18b_class_df
  
  def mcr_p18ba_top_diag_network_op(self, by=None, annualize=False, ibnr=False, research_mode=False):
    
    if 'diagnosis' not in by:
      __p18_df_col = by + ['panel', 'diagnosis', 'incurred_amount', 'paid_amount', 'claimant', 'claim_id']
      __p18_group_col = by + ['panel','diagnosis']
    else:
      __p18_df_col = by + ['panel', 'incurred_amount', 'paid_amount', 'claimant', 'claim_id']
      __p18_group_col = by + ['panel']

    __p18_sort_col = by + ['paid_amount']
    __p18_sort_order = len(by) * [True] + [False]

    p18ba_df = self.mcr_df[__p18_df_col].loc[(self.mcr_df['benefit_type'] == 'Clinic')].groupby(by=__p18_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claimant': 'nunique', 'claim_id': 'nunique'}).rename(columns={'claimant': 'no_of_claimants', 'claim_id': 'no_of_claim_id'})
    p18ba_df = p18ba_df[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    if len(p18ba_df) > 0:
      p18ba_df.sort_values(by=__p18_sort_col, ascending=__p18_sort_order, inplace=True)

      if annualize == True or ibnr == True:
        p18ba_df = pd.merge(p18ba_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
        p18ba_factoring = p18ba_df.copy(deep=True)
        if annualize == True:
          p18ba_factoring['incurred_amount'] = p18ba_factoring['incurred_amount'] * (12 / p18ba_factoring['data_month'])
          p18ba_factoring['paid_amount']     = p18ba_factoring['paid_amount']     * (12 / p18ba_factoring['data_month'])
          p18ba_factoring['no_of_claim_id']  = p18ba_factoring['no_of_claim_id']  * (12 / p18ba_factoring['data_month'])
          p18ba_factoring['no_of_claimants'] = p18ba_factoring['no_of_claimants']
        if ibnr == True:
          p18ba_factoring['incurred_amount'] = p18ba_factoring['incurred_amount'] * (1 + p18ba_factoring['ibnr'])
          p18ba_factoring['paid_amount']     = p18ba_factoring['paid_amount']     * (1 + p18ba_factoring['ibnr'])
          p18ba_factoring['no_of_claim_id']  = p18ba_factoring['no_of_claim_id']  * (1 + p18ba_factoring['ibnr'])
          p18ba_factoring['no_of_claimants'] = p18ba_factoring['no_of_claimants']
        p18ba_diff = p18ba_factoring[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] - p18ba_df[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
        temp_by = __p18_group_col.copy()
        temp_by.remove('year') if 'year' in temp_by else temp_by
        temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
        p18ba_diff = p18ba_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
        p18ba_df.reset_index().set_index(temp_by, inplace=True)
        p18ba_df.loc[p18ba_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] = p18ba_df.loc[p18ba_df.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].add(p18ba_diff)
        temp_by = __p18_group_col.copy()
        temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
        p18ba_df = p18ba_df.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p18_group_col.remove('policy_id') if 'policy_id' in __p18_group_col else __p18_group_col
    p18ba_df = p18ba_df.reset_index().set_index(__p18_group_col)
    p18ba_df = p18ba_df[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    if research_mode == True:
      p18ba_df = p18ba_df.reset_index().groupby(by=__p18_group_col, dropna=False).sum()
      p18ba_df['usage_ratio'] = p18ba_df['paid_amount'] / p18ba_df['incurred_amount'].replace(0, pd.NA)
      p18ba_df['incurred_per_claim'] = p18ba_df['incurred_amount'] / p18ba_df['no_of_claim_id'].replace(0, pd.NA)
      p18ba_df['paid_per_claim'] = p18ba_df['paid_amount'] / p18ba_df['no_of_claim_id'].replace(0, pd.NA)
      p18ba_df = p18ba_df[['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'no_of_claimants']]
    
    self.p18ba = p18ba_df
    return p18ba_df

  def mcr_p27_ts(self, by=None, annualize=False, ibnr=False, research_mode=False):

    t_by = by.copy()
    t_by.remove('policy_id') if 'policy_id' in t_by else t_by

    __p27_df_col = t_by + ['incur_date', 'benefit_type', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p27_group_col = t_by + [pd.Grouper(key='incur_date', freq='MS'), 'benefit_type']
    __p27_sort_order = len(t_by) * [True] + [True]
    
    p27_df = self.mcr_df.dropna(subset=['incur_date'])
    p27_df['incur_date'] = pd.to_datetime(p27_df['incur_date'])
    p27_df = p27_df[__p27_df_col].groupby(by=__p27_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p27_df = p27_df.unstack()
    p27_df.sort_index(ascending=__p27_sort_order, inplace=True)
    self.p27 = p27_df
    return p27_df
  
  def mcr_p27_ts_op(self, by=None, annualize=False, ibnr=False, research_mode=False):

    t_by = by.copy()
    t_by.remove('policy_id') if 'policy_id' in t_by else t_by

    __p27_df_col = t_by + ['incur_date', 'benefit', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p27_group_col = t_by + [pd.Grouper(key='incur_date', freq='MS'), 'benefit']
    __p27_sort_order = len(t_by) * [True] + [True]

    p27_df_op = self.mcr_df.loc[(self.mcr_df['benefit_type'] == 'Dental') | (self.mcr_df['benefit_type'] == 'Optical') | (self.mcr_df['benefit'].str.contains('vaccin|check|combined', case=False)) | (self.mcr_df['benefit_type'] == 'Clinic')].copy(deep=True)
    p27_df_op = p27_df_op[__p27_df_col].groupby(by=__p27_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p27_df_op = p27_df_op.unstack()
    p27_df_op.sort_index(ascending=__p27_sort_order, inplace=True)
    self.p27_op = p27_df_op
    return p27_df_op
  
  def mcr_p27_ts_ip(self, by=None, annualize=False, ibnr=False, research_mode=False):

    t_by = by.copy()
    t_by.remove('policy_id') if 'policy_id' in t_by else t_by

    __p27_df_col = t_by + ['incur_date', 'benefit', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p27_group_col = t_by + [pd.Grouper(key='incur_date', freq='MS'), 'benefit']
    __p27_sort_order = len(t_by) * [True] + [True]

    p27_df_ip = self.mcr_df.loc[(self.mcr_df['benefit_type'] == 'Hospital')].copy(deep=True)
    p27_df_ip = p27_df_ip[__p27_df_col].groupby(by=__p27_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p27_df_ip = p27_df_ip.unstack()
    p27_df_ip.sort_index(ascending=__p27_sort_order, inplace=True)
    self.p27_ip = p27_df_ip
    return p27_df_ip

  def mcr_p28_hosp(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p28_df_col = by + ['hospital_name', 'benefit_type', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p28_group_col = by + ['hospital_name', 'benefit_type']
    __p28_sort_order = len(by) * [True] + [True]

    p28_df_hosp = self.mcr_df[__p28_df_col].copy(deep=True)
    p28_df_hosp['hospital_name'].fillna('No Hospital Name', inplace=True)
    p28_df_hosp = p28_df_hosp.groupby(by=__p28_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p28_df_hosp = p28_df_hosp.unstack()
    p28_df_hosp.columns = [c[0] + '_' + str(c[1]) for c in p28_df_hosp.columns]
    val_cols = p28_df_hosp.columns.to_list()
    
    if annualize == True or ibnr == True:
      p28_df_hosp = pd.merge(p28_df_hosp, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p28_df_hosp_factoring = p28_df_hosp.copy(deep=True)
      if annualize == True:
        for c in p28_df_hosp_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_hosp_factoring[c] = p28_df_hosp_factoring[c] * (12 / p28_df_hosp_factoring['data_month'])
          else:
            p28_df_hosp_factoring[c] = p28_df_hosp_factoring[c]
      if ibnr == True:
        for c in p28_df_hosp_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_hosp_factoring[c] = p28_df_hosp_factoring[c] * (1 + p28_df_hosp_factoring['ibnr'])
          else:
            p28_df_hosp_factoring[c] = p28_df_hosp_factoring[c]

      value_cols = [c for c in p28_df_hosp_factoring.columns if ('incurred_amount' in c) or ('paid_amount' in c) or ('no_of_claim_id' in c) or ('no_of_claimants' in c)]
      p28_df_hosp_diff = p28_df_hosp_factoring[value_cols] - p28_df_hosp[value_cols]
      temp_by = __p28_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      temp_by.remove('benefit_type') if 'benefit_type' in temp_by else temp_by
      p28_df_hosp_diff = p28_df_hosp_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p28_df_hosp.reset_index().set_index(temp_by, inplace=True)
      p28_df_hosp.loc[p28_df_hosp.factoring == True, value_cols] = p28_df_hosp.loc[p28_df_hosp.factoring == True, value_cols].add(p28_df_hosp_diff)
      temp_by = __p28_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      temp_by.remove('benefit_type') if 'benefit_type' in temp_by else temp_by
      p28_df_hosp = p28_df_hosp.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()

    __p28_group_col.remove('policy_id') if 'policy_id' in __p28_group_col else __p28_group_col
    __p28_group_col.remove('benefit_type') if 'benefit_type' in __p28_group_col else __p28_group_col
    p28_df_hosp = p28_df_hosp.reset_index().set_index(__p28_group_col)
    p28_df_hosp.sort_index(ascending=__p28_sort_order, inplace=True)
    p28_df_hosp = p28_df_hosp[val_cols]
    self.p28_hosp = p28_df_hosp
    return p28_df_hosp

  def mcr_p28_prov(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p28_df_prov_col = by + ['provider', 'benefit_type', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p28_group_col = by + ['provider', 'benefit_type']
    __p28_sort_order = len(by) * [True] + [True]

    p28_df_prov = self.mcr_df[__p28_df_prov_col].copy(deep=True)
    p28_df_prov['provider'].fillna('No Provider Name', inplace=True)
    p28_df_prov = p28_df_prov.groupby(by=__p28_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p28_df_prov = p28_df_prov.unstack()
    p28_df_prov.columns = [c[0] + '_' + str(c[1]) for c in p28_df_prov.columns]
    val_cols = p28_df_prov.columns.to_list()
    p28_df_prov.sort_index(ascending=__p28_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p28_df_prov = pd.merge(p28_df_prov, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p28_df_prov_factoring = p28_df_prov.copy(deep=True)
      if annualize == True:
        for c in p28_df_prov_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_prov_factoring[c] = p28_df_prov_factoring[c] * (12 / p28_df_prov_factoring['data_month'])
          else:
            p28_df_prov_factoring[c] = p28_df_prov_factoring[c]
      if ibnr == True:
        for c in p28_df_prov_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_prov_factoring[c] = p28_df_prov_factoring[c] * (1 + p28_df_prov_factoring['ibnr'])
          else:
            p28_df_prov_factoring[c] = p28_df_prov_factoring[c]
      value_cols = [c for c in p28_df_prov_factoring.columns if ('incurred_amount' in c) or ('paid_amount' in c) or ('no_of_claim_id' in c) or ('no_of_claimants' in c)]
      p28_df_prov_diff = p28_df_prov_factoring[value_cols] - p28_df_prov[value_cols]
      temp_by = __p28_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      temp_by.remove('benefit_type') if 'benefit_type' in temp_by else temp_by
      p28_df_prov_diff = p28_df_prov_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p28_df_prov.reset_index().set_index(temp_by, inplace=True)
      p28_df_prov.loc[p28_df_prov.factoring == True, value_cols] = p28_df_prov.loc[p28_df_prov.factoring == True, value_cols].add(p28_df_prov_diff)
      temp_by = __p28_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      temp_by.remove('benefit_type') if 'benefit_type' in temp_by else temp_by
      p28_df_prov = p28_df_prov.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()

    __p28_group_col.remove('policy_id') if 'policy_id' in __p28_group_col else __p28_group_col
    __p28_group_col.remove('benefit_type') if 'benefit_type' in __p28_group_col else __p28_group_col
    p28_df_prov = p28_df_prov.reset_index().set_index(__p28_group_col)
    p28_df_prov = p28_df_prov[val_cols]
    self.p28_prov = p28_df_prov
    return p28_df_prov

  def mcr_p28a_prov_benefit(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p28_df_prov_col = by + ['provider', 'benefit_type', 'benefit', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p28_group_col = by + ['provider', 'benefit_type', 'benefit']
    __p28_sort_order = len(by) * [True] + [True, True]

    p28_df_prov_benefit = self.mcr_df[__p28_df_prov_col].copy(deep=True)
    p28_df_prov_benefit['provider'].fillna('No Provider Name', inplace=True)
    p28_df_prov_benefit = p28_df_prov_benefit.groupby(by=__p28_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p28_df_prov_benefit.sort_index(ascending=__p28_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p28_df_prov_benefit = pd.merge(p28_df_prov_benefit, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p28_df_prov_benefit_factoring = p28_df_prov_benefit.copy(deep=True)
      if annualize == True:
        for c in p28_df_prov_benefit_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_prov_benefit_factoring[c] = p28_df_prov_benefit_factoring[c] * (12 / p28_df_prov_benefit_factoring['data_month'])
          else:
            p28_df_prov_benefit_factoring[c] = p28_df_prov_benefit_factoring[c]
      if ibnr == True:
        for c in p28_df_prov_benefit_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_prov_benefit_factoring[c] = p28_df_prov_benefit_factoring[c] * (1 + p28_df_prov_benefit_factoring['ibnr'])
          else:
            p28_df_prov_benefit_factoring[c] = p28_df_prov_benefit_factoring[c]
      value_cols = [c for c in p28_df_prov_benefit_factoring.columns if ('incurred_amount' in c) or ('paid_amount' in c) or ('no_of_claim_id' in c) or ('no_of_claimants' in c)]
      p28_df_prov_benefit_diff = p28_df_prov_benefit_factoring[value_cols] - p28_df_prov_benefit[value_cols]
      temp_by = __p28_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p28_df_prov_benefit_diff = p28_df_prov_benefit_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p28_df_prov_benefit.reset_index().set_index(temp_by, inplace=True)
      p28_df_prov_benefit.loc[p28_df_prov_benefit.factoring == True, value_cols] = p28_df_prov_benefit.loc[p28_df_prov_benefit.factoring == True, value_cols].add(p28_df_prov_benefit_diff)
      temp_by = __p28_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p28_df_prov_benefit = p28_df_prov_benefit.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()

    __p28_group_col.remove('policy_id') if 'policy_id' in __p28_group_col else __p28_group_col
    p28_df_prov_benefit = p28_df_prov_benefit.reset_index().set_index(__p28_group_col)
    p28_df_prov_benefit = p28_df_prov_benefit[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    self.p28_prov_benefit = p28_df_prov_benefit
    return p28_df_prov_benefit

  def mcr_p28_doct(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p28_df_doct_col = by + ['physician', 'benefit_type', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p28_group_col = by + ['physician', 'benefit_type']
    __p28_sort_order = len(by) * [True] + [True]

    p28_df_doct = self.mcr_df[__p28_df_doct_col].copy(deep=True)
    p28_df_doct['physician'].fillna('No Physician Name', inplace=True)
    p28_df_doct = p28_df_doct.groupby(by=__p28_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p28_df_doct = p28_df_doct.unstack()
    p28_df_doct.columns = [c[0] + '_' + str(c[1]) for c in p28_df_doct.columns]
    val_cols = p28_df_doct.columns.to_list()
    p28_df_doct.sort_index(ascending=__p28_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p28_df_doct = pd.merge(p28_df_doct, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p28_df_doct_factoring = p28_df_doct.copy(deep=True)
      if annualize == True:
        for c in p28_df_doct_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_doct_factoring[c] = p28_df_doct_factoring[c] * (12 / p28_df_doct_factoring['data_month'])
          else:
            p28_df_doct_factoring[c] = p28_df_doct_factoring[c]
      if ibnr == True:
        for c in p28_df_doct_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_doct_factoring[c] = p28_df_doct_factoring[c] * (1 + p28_df_doct_factoring['ibnr'])
          else:
            p28_df_doct_factoring[c] = p28_df_doct_factoring[c]
      value_cols = [c for c in p28_df_doct_factoring.columns if ('incurred_amount' in c) or ('paid_amount' in c) or ('no_of_claim_id' in c) or ('no_of_claimants' in c)]
      p28_df_doct_diff = p28_df_doct_factoring[value_cols] - p28_df_doct[value_cols]
      temp_by = __p28_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      temp_by.remove('benefit_type') if 'benefit_type' in temp_by else temp_by
      p28_df_doct_diff = p28_df_doct_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p28_df_doct.reset_index().set_index(temp_by, inplace=True)
      p28_df_doct.loc[p28_df_doct.factoring == True, value_cols] = p28_df_doct.loc[p28_df_doct.factoring == True, value_cols].add(p28_df_doct_diff)
      temp_by = __p28_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      temp_by.remove('benefit_type') if 'benefit_type' in temp_by else temp_by
      p28_df_doct = p28_df_doct.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()

    __p28_group_col.remove('policy_id') if 'policy_id' in __p28_group_col else __p28_group_col
    __p28_group_col.remove('benefit_type') if 'benefit_type' in __p28_group_col else __p28_group_col
    p28_df_doct = p28_df_doct.reset_index().set_index(__p28_group_col)
    p28_df_doct = p28_df_doct[val_cols]
    self.p28_doct = p28_df_doct
    return p28_df_doct
  
  def mcr_p28a_doct(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p28a_df_doct_col = by + ['physician', 'benefit_type', 'benefit', 'incurred_amount', 'paid_amount', 'claim_id', 'claimant']
    __p28a_group_col = by + ['physician', 'benefit_type', 'benefit']
    __p28a_sort_order = len(by) * [True] + [True]

    p28a_df_doct = self.mcr_df[__p28a_df_doct_col].copy(deep=True)
    p28a_df_doct['physician'].fillna('No Physician Name', inplace=True)
    p28a_df_doct = p28a_df_doct.groupby(by=__p28a_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p28a_df_doct['incurred_amount_per_claim'] = p28a_df_doct['incurred_amount'] / p28a_df_doct['no_of_claim_id']
    p28a_df_doct['paid_amount_per_claim'] = p28a_df_doct['paid_amount'] / p28a_df_doct['no_of_claim_id']
    p28a_df_doct.sort_index(ascending=__p28a_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p28a_df_doct = pd.merge(p28a_df_doct, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p28a_df_doct_factoring = p28a_df_doct.copy(deep=True)
      if annualize == True:
        p28a_df_doct_factoring['incurred_amount'] = p28a_df_doct_factoring['incurred_amount'] * (12 / p28a_df_doct_factoring['data_month'])
        p28a_df_doct_factoring['paid_amount']     = p28a_df_doct_factoring['paid_amount']     * (12 / p28a_df_doct_factoring['data_month'])
        p28a_df_doct_factoring['no_of_claim_id']  = p28a_df_doct_factoring['no_of_claim_id']  * (12 / p28a_df_doct_factoring['data_month'])
        p28a_df_doct_factoring['no_of_claimants'] = p28a_df_doct_factoring['no_of_claimants']
      if ibnr == True:
        p28a_df_doct_factoring['incurred_amount'] = p28a_df_doct_factoring['incurred_amount'] * (1 + p28a_df_doct_factoring['ibnr'])
        p28a_df_doct_factoring['paid_amount']     = p28a_df_doct_factoring['paid_amount']     * (1 + p28a_df_doct_factoring['ibnr'])
        p28a_df_doct_factoring['no_of_claim_id']  = p28a_df_doct_factoring['no_of_claim_id']  * (1 + p28a_df_doct_factoring['ibnr'])
        p28a_df_doct_factoring['no_of_claimants'] = p28a_df_doct_factoring['no_of_claimants']
      value_cols = [c for c in p28a_df_doct_factoring.columns if ('incurred_amount' in c) or ('paid_amount' in c) or ('no_of_claim_id' in c) or ('no_of_claimants' in c) or ('per_claim' in c)]
      p28a_df_doct_diff = p28a_df_doct_factoring[value_cols] - p28a_df_doct[value_cols]
      temp_by = __p28a_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p28a_df_doct_diff = p28a_df_doct_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p28a_df_doct.reset_index().set_index(temp_by, inplace=True)
      p28a_df_doct.loc[p28a_df_doct.factoring == True, value_cols] = p28a_df_doct.loc[p28a_df_doct.factoring == True, value_cols].add(p28a_df_doct_diff)
      temp_by = __p28a_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p28a_df_doct = p28a_df_doct.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()

    __p28a_group_col.remove('policy_id') if 'policy_id' in __p28a_group_col else __p28a_group_col
    p28a_df_doct = p28a_df_doct.reset_index().set_index(__p28a_group_col)
    p28a_df_doct = p28a_df_doct[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants', 'incurred_amount_per_claim', 'paid_amount_per_claim']]

    self.p28a_doct = p28a_df_doct
    return p28a_df_doct

  def mcr_p28_proce(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p28_df_proce_col = by + ['procedure', 'benefit_type', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p28_group_col = by + ['procedure', 'benefit_type']
    __p28_sort_order = len(by) * [True] + [True, True]

    p28_df_proce = self.mcr_df[__p28_df_proce_col].copy(deep=True)
    p28_df_proce['procedure'].fillna('No Procedure Name', inplace=True)
    p28_df_proce = p28_df_proce.groupby(by=__p28_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p28_df_proce = p28_df_proce.unstack()
    p28_df_proce.columns = [c[0] + '_' + str(c[1]) for c in p28_df_proce.columns]
    val_cols = p28_df_proce.columns.to_list()
    p28_df_proce.sort_index(ascending=__p28_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p28_df_proce = pd.merge(p28_df_proce, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p28_df_proce_factoring = p28_df_proce.copy(deep=True)
      if annualize == True:
        for c in p28_df_proce_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_proce_factoring[c] = p28_df_proce_factoring[c] * (12 / p28_df_proce_factoring['data_month'])
          else:
            p28_df_proce_factoring[c] = p28_df_proce_factoring[c]
      if ibnr == True:
        for c in p28_df_proce_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_proce_factoring[c] = p28_df_proce_factoring[c] * (1 + p28_df_proce_factoring['ibnr'])
          else: 
            p28_df_proce_factoring[c] = p28_df_proce_factoring[c]
      value_cols = [c for c in p28_df_proce_factoring.columns if ('incurred_amount' in c) or ('paid_amount' in c) or ('no_of_claim_id' in c) or ('no_of_claimants' in c)]
      p28_df_proce_diff = p28_df_proce_factoring[value_cols] - p28_df_proce[value_cols]
      temp_by = __p28_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      temp_by.remove('benefit_type') if 'benefit_type' in temp_by else temp_by
      p28_df_proce_diff = p28_df_proce_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p28_df_proce.reset_index().set_index(temp_by, inplace=True)
      p28_df_proce.loc[p28_df_proce.factoring == True, value_cols] = p28_df_proce.loc[p28_df_proce.factoring == True, value_cols].add(p28_df_proce_diff)
      temp_by = __p28_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      temp_by.remove('benefit_type') if 'benefit_type' in temp_by else temp_by
      p28_df_proce = p28_df_proce.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()

    __p28_group_col.remove('policy_id') if 'policy_id' in __p28_group_col else __p28_group_col
    __p28_group_col.remove('benefit_type') if 'benefit_type' in __p28_group_col else __p28_group_col
    p28_df_proce = p28_df_proce.reset_index().set_index(__p28_group_col)
    p28_df_proce = p28_df_proce[val_cols]
    self.p28_proce = p28_df_proce
    return p28_df_proce

  def mcr_p28_proce_diagnosis(self, by=None, annualize=False, ibnr=False, research_mode=False):

    if 'diagnosis' not in by:
      __p28_df_proce_diagnosis_col = by + ['diagnosis', 'procedure', 'benefit_type', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
      __p28_group_col = by + ['diagnosis', 'procedure', 'benefit_type']
      __p28_sort_order = len(by) * [True] + [True, True, True]
    else:
      __p28_df_proce_diagnosis_col = by + ['procedure', 'benefit_type', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
      __p28_group_col = by + ['procedure', 'benefit_type']
      __p28_sort_order = len(by) * [True] + [True, True]
    

    p28_df_proce_diagnosis = self.mcr_df[__p28_df_proce_diagnosis_col].copy(deep=True)
    p28_df_proce_diagnosis['procedure'].fillna('No Procedure Name', inplace=True)
    p28_df_proce_diagnosis = p28_df_proce_diagnosis.groupby(by=__p28_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p28_df_proce_diagnosis = p28_df_proce_diagnosis.unstack()
    p28_df_proce_diagnosis.columns = [c[0] + '_' + str(c[1]) for c in p28_df_proce_diagnosis.columns]
    val_cols = p28_df_proce_diagnosis.columns.to_list()
    p28_df_proce_diagnosis.sort_index(ascending=__p28_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p28_df_proce_diagnosis = pd.merge(p28_df_proce_diagnosis, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p28_df_proce_diagnosis_factoring = p28_df_proce_diagnosis.copy(deep=True)
      if annualize == True:
        for c in p28_df_proce_diagnosis_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_proce_diagnosis_factoring[c] = p28_df_proce_diagnosis_factoring[c] * (12 / p28_df_proce_diagnosis_factoring['data_month'])
          else:
            p28_df_proce_diagnosis_factoring[c] = p28_df_proce_diagnosis_factoring[c]
      if ibnr == True:
        for c in p28_df_proce_diagnosis_factoring.columns:
          if 'incurred_amount' in c or 'paid_amount' in c or 'no_of_claim_id' in c:
            p28_df_proce_diagnosis_factoring[c] = p28_df_proce_diagnosis_factoring[c] * (1 + p28_df_proce_diagnosis_factoring['ibnr'])
          else:
            p28_df_proce_diagnosis_factoring[c] = p28_df_proce_diagnosis_factoring[c]
      value_cols = [c for c in p28_df_proce_diagnosis_factoring.columns if ('incurred_amount' in c) or ('paid_amount' in c) or ('no_of_claim_id' in c) or ('no_of_claimants' in c)]
      p28_df_proce_diagnosis_diff = p28_df_proce_diagnosis_factoring[value_cols] - p28_df_proce_diagnosis[value_cols]
      temp_by = __p28_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      temp_by.remove('benefit_type') if 'benefit_type' in temp_by else temp_by
      p28_df_proce_diagnosis_diff = p28_df_proce_diagnosis_diff.reset_index(drop=False)[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()
      p28_df_proce_diagnosis.reset_index().set_index(temp_by, inplace=True)
      p28_df_proce_diagnosis.loc[p28_df_proce_diagnosis.factoring == True, value_cols] = p28_df_proce_diagnosis.loc[p28_df_proce_diagnosis.factoring == True, value_cols].add(p28_df_proce_diagnosis_diff)
      temp_by = __p28_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      temp_by.remove('benefit_type') if 'benefit_type' in temp_by else temp_by
      p28_df_proce_diagnosis = p28_df_proce_diagnosis.reset_index()[temp_by + value_cols].groupby(by=temp_by, dropna=False).sum()

    __p28_group_col.remove('policy_id') if 'policy_id' in __p28_group_col else __p28_group_col
    __p28_group_col.remove('benefit_type') if 'benefit_type' in __p28_group_col else __p28_group_col
    p28_df_proce_diagnosis = p28_df_proce_diagnosis.reset_index().set_index(__p28_group_col)
    p28_df_proce_diagnosis = p28_df_proce_diagnosis[val_cols]
    self.p28_proce_diagnosis = p28_df_proce_diagnosis
    return p28_df_proce_diagnosis

  def mcr_p28b_proce_network(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p28_df_proce_network_col = by + ['panel', 'procedure', 'benefit_type', 'benefit', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p28_group_col = by + ['panel', 'procedure', 'benefit_type', 'benefit']
    __p28_sort_order = len(by) * [True] + [True, True, True, True]

    p28_df_proce_network= self.mcr_df[__p28_df_proce_network_col].copy(deep=True)
    p28_df_proce_network['procedure'].fillna('No Procedure Name', inplace=True)
    p28_df_proce_network = p28_df_proce_network.groupby(by=__p28_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    # p28_df_proce_network = p28_df_proce_network.unstack()
    p28_df_proce_network.sort_index(ascending=__p28_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p28_df_proce_network = pd.merge(p28_df_proce_network, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p28_df_proce_network_factoring = p28_df_proce_network.copy(deep=True)
      if annualize == True:
        p28_df_proce_network_factoring['incurred_amount'] = p28_df_proce_network_factoring['incurred_amount'] * (12 / p28_df_proce_network_factoring['data_month'])
        p28_df_proce_network_factoring['paid_amount']     = p28_df_proce_network_factoring['paid_amount']     * (12 / p28_df_proce_network_factoring['data_month'])
        p28_df_proce_network_factoring['no_of_claim_id']  = p28_df_proce_network_factoring['no_of_claim_id']  * (12 / p28_df_proce_network_factoring['data_month'])
        p28_df_proce_network_factoring['no_of_claimants'] = p28_df_proce_network_factoring['no_of_claimants']
      if ibnr == True:
        p28_df_proce_network_factoring['incurred_amount'] = p28_df_proce_network_factoring['incurred_amount'] * (1 + p28_df_proce_network_factoring['ibnr'])
        p28_df_proce_network_factoring['paid_amount']     = p28_df_proce_network_factoring['paid_amount']     * (1 + p28_df_proce_network_factoring['ibnr'])
        p28_df_proce_network_factoring['no_of_claim_id']  = p28_df_proce_network_factoring['no_of_claim_id']  * (1 + p28_df_proce_network_factoring['ibnr'])
        p28_df_proce_network_factoring['no_of_claimants'] = p28_df_proce_network_factoring['no_of_claimants']
      val_cols = [c for c in p28_df_proce_network_factoring.columns if ('incurred_amount' in c) or ('paid_amount' in c) or ('no_of_claim_id' in c) or ('no_of_claimants' in c)]
      p28_df_proce_network_diff = p28_df_proce_network_factoring[val_cols] - p28_df_proce_network[val_cols]
      temp_by = __p28_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p28_df_proce_network_diff = p28_df_proce_network_diff.reset_index(drop=False)[temp_by + val_cols].groupby(by=temp_by, dropna=False).sum()
      p28_df_proce_network.reset_index().set_index(temp_by, inplace=True)
      p28_df_proce_network.loc[p28_df_proce_network.factoring == True, val_cols] = p28_df_proce_network.loc[p28_df_proce_network.factoring == True, val_cols].add(p28_df_proce_network_diff)
      temp_by = __p28_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p28_df_proce_network = p28_df_proce_network.reset_index()[temp_by + val_cols].groupby(by=temp_by, dropna=False).sum()

    __p28_group_col.remove('policy_id') if 'policy_id' in __p28_group_col else __p28_group_col
    p28_df_proce_network = p28_df_proce_network.reset_index().set_index(__p28_group_col)
    p28_df_proce_network.drop(columns=['policy_id'], inplace=True) if 'policy_id' in p28_df_proce_network.columns else None
    self.p28_proce_network = p28_df_proce_network
    return p28_df_proce_network

  def mcr_p29_sp_speciality(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p29_df_sp_specialty_col = by + ['speciality', 'benefit', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p29_group_col = by + ['speciality']
    __p29_sort_order = len(by) * [True] + [True]

    p29_df_sp_specialty = self.mcr_df[__p29_df_sp_specialty_col].copy(deep=True)
    p29_df_sp_specialty = p29_df_sp_specialty.loc[p29_df_sp_specialty['benefit'] == 'Specialist Consultation (SP)']
    p29_df_sp_specialty.drop(columns=['benefit'], inplace=True)
    p29_df_sp_specialty = p29_df_sp_specialty.groupby(by=__p29_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    # p29_df_sp_specialty = p29_df_sp_specialty.unstack()
    p29_df_sp_specialty.sort_index(ascending=__p29_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p29_df_sp_specialty = pd.merge(p29_df_sp_specialty, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p29_df_sp_specialty_factoring = p29_df_sp_specialty.copy(deep=True)
      if annualize == True:
        p29_df_sp_specialty_factoring['incurred_amount'] = p29_df_sp_specialty_factoring['incurred_amount'] * (12 / p29_df_sp_specialty_factoring['data_month'])
        p29_df_sp_specialty_factoring['paid_amount']     = p29_df_sp_specialty_factoring['paid_amount']     * (12 / p29_df_sp_specialty_factoring['data_month'])
        p29_df_sp_specialty_factoring['no_of_claim_id']  = p29_df_sp_specialty_factoring['no_of_claim_id']  * (12 / p29_df_sp_specialty_factoring['data_month'])
        p29_df_sp_specialty_factoring['no_of_claimants'] = p29_df_sp_specialty_factoring['no_of_claimants']
      if ibnr == True:
        p29_df_sp_specialty_factoring['incurred_amount'] = p29_df_sp_specialty_factoring['incurred_amount'] * (1 + p29_df_sp_specialty_factoring['ibnr'])
        p29_df_sp_specialty_factoring['paid_amount']     = p29_df_sp_specialty_factoring['paid_amount']     * (1 + p29_df_sp_specialty_factoring['ibnr'])
        p29_df_sp_specialty_factoring['no_of_claim_id']  = p29_df_sp_specialty_factoring['no_of_claim_id']  * (1 + p29_df_sp_specialty_factoring['ibnr'])
        p29_df_sp_specialty_factoring['no_of_claimants'] = p29_df_sp_specialty_factoring['no_of_claimants']
      p29_df_sp_specialty_diff = p29_df_sp_specialty_factoring[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] - p29_df_sp_specialty[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
      temp_by = __p29_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p29_df_sp_specialty_diff = p29_df_sp_specialty_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p29_df_sp_specialty.reset_index().set_index(temp_by, inplace=True)
      p29_df_sp_specialty.loc[p29_df_sp_specialty.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] = p29_df_sp_specialty.loc[p29_df_sp_specialty.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].add(p29_df_sp_specialty_diff)
      temp_by = __p29_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p29_df_sp_specialty = p29_df_sp_specialty.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p29_group_col.remove('policy_id') if 'policy_id' in __p29_group_col else __p29_group_col
    p29_df_sp_specialty = p29_df_sp_specialty.reset_index().set_index(__p29_group_col)
    p29_df_sp_specialty = p29_df_sp_specialty[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    self.p29_df_sp_specialty = p29_df_sp_specialty
    return p29_df_sp_specialty


  def mcr_p29_sp_speciality_diag(self, by=None, annualize=False, ibnr=False, research_mode=False):

    if 'diagnosis' not in by:
      __p29_df_sp_specialty_diag_col = by + ['speciality', 'diagnosis', 'benefit', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
      __p29_group_col = by + ['speciality', 'diagnosis']
      __p29_sort_order = len(by) * [True] + [True, True]
    else:
      __p29_df_sp_specialty_diag_col = by + ['speciality', 'benefit', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
      __p29_group_col = by + ['speciality']
      __p29_sort_order = len(by) * [True] + [True]

    p29_df_sp_specialty_diag = self.mcr_df[__p29_df_sp_specialty_diag_col].copy(deep=True)
    p29_df_sp_specialty_diag['diagnosis'].fillna('No Diagnosis Provided', inplace=True)
    p29_df_sp_specialty_diag = p29_df_sp_specialty_diag.loc[p29_df_sp_specialty_diag['benefit'] == 'Specialist Consultation (SP)']
    p29_df_sp_specialty_diag.drop(columns=['benefit'], inplace=True)
    p29_df_sp_specialty_diag = p29_df_sp_specialty_diag.groupby(by=__p29_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    # p29_df_sp_specialty = p29_df_sp_specialty.unstack()
    p29_df_sp_specialty_diag.sort_index(ascending=__p29_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p29_df_sp_specialty_diag = pd.merge(p29_df_sp_specialty_diag, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p29_df_sp_specialty_diag_factoring = p29_df_sp_specialty_diag.copy(deep=True)
      if annualize == True:
        p29_df_sp_specialty_diag_factoring['incurred_amount'] = p29_df_sp_specialty_diag_factoring['incurred_amount'] * (12 / p29_df_sp_specialty_diag_factoring['data_month'])
        p29_df_sp_specialty_diag_factoring['paid_amount']     = p29_df_sp_specialty_diag_factoring['paid_amount']     * (12 / p29_df_sp_specialty_diag_factoring['data_month'])
        p29_df_sp_specialty_diag_factoring['no_of_claim_id']  = p29_df_sp_specialty_diag_factoring['no_of_claim_id']  * (12 / p29_df_sp_specialty_diag_factoring['data_month'])
        p29_df_sp_specialty_diag_factoring['no_of_claimants'] = p29_df_sp_specialty_diag_factoring['no_of_claimants']
      if ibnr == True:
        p29_df_sp_specialty_diag_factoring['incurred_amount'] = p29_df_sp_specialty_diag_factoring['incurred_amount'] * (1 + p29_df_sp_specialty_diag_factoring['ibnr'])
        p29_df_sp_specialty_diag_factoring['paid_amount']     = p29_df_sp_specialty_diag_factoring['paid_amount']     * (1 + p29_df_sp_specialty_diag_factoring['ibnr'])
        p29_df_sp_specialty_diag_factoring['no_of_claim_id']  = p29_df_sp_specialty_diag_factoring['no_of_claim_id']  * (1 + p29_df_sp_specialty_diag_factoring['ibnr'])
        p29_df_sp_specialty_diag_factoring['no_of_claimants'] = p29_df_sp_specialty_diag_factoring['no_of_claimants']
      p29_df_sp_specialty_diag_diff = p29_df_sp_specialty_diag_factoring[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] - p29_df_sp_specialty_diag[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
      temp_by = __p29_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p29_df_sp_specialty_diag_diff = p29_df_sp_specialty_diag_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p29_df_sp_specialty_diag.reset_index().set_index(temp_by, inplace=True)
      p29_df_sp_specialty_diag.loc[p29_df_sp_specialty_diag.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] = p29_df_sp_specialty_diag.loc[p29_df_sp_specialty_diag.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].add(p29_df_sp_specialty_diag_diff)
      temp_by = __p29_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p29_df_sp_specialty_diag = p29_df_sp_specialty_diag.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p29_group_col.remove('policy_id') if 'policy_id' in __p29_group_col else __p29_group_col
    p29_df_sp_specialty_diag = p29_df_sp_specialty_diag.reset_index().set_index(__p29_group_col)
    p29_df_sp_specialty_diag = p29_df_sp_specialty_diag[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    self.p29_df_sp_specialty_diag = p29_df_sp_specialty_diag
    return p29_df_sp_specialty_diag

  def mcr_p29_grp_procedure(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p29_df_grp_procedure_col = by + ['minor_surgeries_grp', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p29_group_col = by + ['minor_surgeries_grp']
    __p29_sort_order = len(by) * [True] + [True]

    p29_df_grp_procedure = self.mcr_df[__p29_df_grp_procedure_col].copy(deep=True)
    p29_df_grp_procedure.dropna(subset=['minor_surgeries_grp'], inplace=True)

    p29_df_grp_procedure = p29_df_grp_procedure.groupby(by=__p29_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p29_df_grp_procedure.sort_index(ascending=__p29_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p29_df_grp_procedure = pd.merge(p29_df_grp_procedure, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p29_df_grp_procedure_factoring = p29_df_grp_procedure.copy(deep=True)
      if annualize == True:
        p29_df_grp_procedure_factoring['incurred_amount'] = p29_df_grp_procedure_factoring['incurred_amount'] * (12 / p29_df_grp_procedure_factoring['data_month'])
        p29_df_grp_procedure_factoring['paid_amount']     = p29_df_grp_procedure_factoring['paid_amount']     * (12 / p29_df_grp_procedure_factoring['data_month'])
        p29_df_grp_procedure_factoring['no_of_claim_id']  = p29_df_grp_procedure_factoring['no_of_claim_id']  * (12 / p29_df_grp_procedure_factoring['data_month'])
        p29_df_grp_procedure_factoring['no_of_claimants'] = p29_df_grp_procedure_factoring['no_of_claimants']
      if ibnr == True:
        p29_df_grp_procedure_factoring['incurred_amount'] = p29_df_grp_procedure_factoring['incurred_amount'] * (1 + p29_df_grp_procedure_factoring['ibnr'])
        p29_df_grp_procedure_factoring['paid_amount']     = p29_df_grp_procedure_factoring['paid_amount']     * (1 + p29_df_grp_procedure_factoring['ibnr'])
        p29_df_grp_procedure_factoring['no_of_claim_id']  = p29_df_grp_procedure_factoring['no_of_claim_id']  * (1 + p29_df_grp_procedure_factoring['ibnr'])
        p29_df_grp_procedure_factoring['no_of_claimants'] = p29_df_grp_procedure_factoring['no_of_claimants']
      p29_df_grp_procedure_diff = p29_df_grp_procedure_factoring[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] - p29_df_grp_procedure[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
      temp_by = __p29_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p29_df_grp_procedure_diff = p29_df_grp_procedure_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p29_df_grp_procedure.reset_index().set_index(temp_by, inplace=True)
      p29_df_grp_procedure.loc[p29_df_grp_procedure.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] = p29_df_grp_procedure.loc[p29_df_grp_procedure.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].add(p29_df_grp_procedure_diff)
      temp_by = __p29_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p29_df_grp_procedure = p29_df_grp_procedure.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p29_group_col.remove('policy_id') if 'policy_id' in __p29_group_col else __p29_group_col
    p29_df_grp_procedure = p29_df_grp_procedure.reset_index().set_index(__p29_group_col)
    p29_df_grp_procedure = p29_df_grp_procedure[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    self.p29_df_grp_procedure = p29_df_grp_procedure
    return p29_df_grp_procedure

  def mcr_p29_grp_procedure_org(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p29_df_grp_procedure_org_col = by + ['minor_surgeries_grp', 'organs', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p29_group_col = by + ['minor_surgeries_grp', 'organs']
    __p29_sort_order = len(by) * [True] + [True, True]

    p29_df_grp_procedure_org = self.mcr_df[__p29_df_grp_procedure_org_col].copy(deep=True)
    p29_df_grp_procedure_org.dropna(subset=['minor_surgeries_grp', 'organs'], inplace=True)

    p29_df_grp_procedure_org = p29_df_grp_procedure_org.groupby(by=__p29_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p29_df_grp_procedure_org.sort_index(ascending=__p29_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p29_df_grp_procedure_org = pd.merge(p29_df_grp_procedure_org, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p29_df_grp_procedure_org_factoring = p29_df_grp_procedure_org.copy(deep=True)
      if annualize == True:
        p29_df_grp_procedure_org_factoring['incurred_amount'] = p29_df_grp_procedure_org_factoring['incurred_amount'] * (12 / p29_df_grp_procedure_org_factoring['data_month'])
        p29_df_grp_procedure_org_factoring['paid_amount']     = p29_df_grp_procedure_org_factoring['paid_amount']     * (12 / p29_df_grp_procedure_org_factoring['data_month'])
        p29_df_grp_procedure_org_factoring['no_of_claim_id']  = p29_df_grp_procedure_org_factoring['no_of_claim_id']  * (12 / p29_df_grp_procedure_org_factoring['data_month'])
        p29_df_grp_procedure_org_factoring['no_of_claimants'] = p29_df_grp_procedure_org_factoring['no_of_claimants']
      if ibnr == True:
        p29_df_grp_procedure_org_factoring['incurred_amount'] = p29_df_grp_procedure_org_factoring['incurred_amount'] * (1 + p29_df_grp_procedure_org_factoring['ibnr'])
        p29_df_grp_procedure_org_factoring['paid_amount']     = p29_df_grp_procedure_org_factoring['paid_amount']     * (1 + p29_df_grp_procedure_org_factoring['ibnr'])
        p29_df_grp_procedure_org_factoring['no_of_claim_id']  = p29_df_grp_procedure_org_factoring['no_of_claim_id']  * (1 + p29_df_grp_procedure_org_factoring['ibnr'])
        p29_df_grp_procedure_org_factoring['no_of_claimants'] = p29_df_grp_procedure_org_factoring['no_of_claimants']
      p29_df_grp_procedure_org_diff = p29_df_grp_procedure_org_factoring[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] - p29_df_grp_procedure_org[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
      temp_by = __p29_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p29_df_grp_procedure_org_diff = p29_df_grp_procedure_org_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p29_df_grp_procedure_org.reset_index().set_index(temp_by, inplace=True)
      p29_df_grp_procedure_org.loc[p29_df_grp_procedure_org.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] = p29_df_grp_procedure_org.loc[p29_df_grp_procedure_org.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].add(p29_df_grp_procedure_org_diff)
      temp_by = __p29_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p29_df_grp_procedure_org = p29_df_grp_procedure_org.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p29_group_col.remove('policy_id') if 'policy_id' in __p29_group_col else __p29_group_col
    p29_df_grp_procedure_org = p29_df_grp_procedure_org.reset_index().set_index(__p29_group_col)
    p29_df_grp_procedure_org = p29_df_grp_procedure_org[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    self.p29_df_grp_procedure_org = p29_df_grp_procedure_org
    return p29_df_grp_procedure_org

  def mcr_p29b_grp_procedure_network(self, by=None, annualize=False, ibnr=False, research_mode=False):

    __p29_df_grp_procedure_network_col = by + ['minor_surgeries_grp', 'panel', 'incurred_amount', 'claim_id', 'paid_amount', 'claimant']
    __p29_group_col = by + ['minor_surgeries_grp', 'panel']
    __p29_sort_order = len(by) * [True] + [True, True]

    p29_df_grp_procedure_network = self.mcr_df[__p29_df_grp_procedure_network_col].copy(deep=True)
    p29_df_grp_procedure_network.dropna(subset=['minor_surgeries_grp'], inplace=True)

    p29_df_grp_procedure_network = p29_df_grp_procedure_network.groupby(by=__p29_group_col).agg({'incurred_amount': 'sum', 'paid_amount': 'sum', 'claim_id': 'nunique', 'claimant': 'nunique'}).rename(columns={'claim_id': 'no_of_claim_id', 'claimant': 'no_of_claimants'})
    p29_df_grp_procedure_network.sort_index(ascending=__p29_sort_order, inplace=True)
    if annualize == True or ibnr == True:
      p29_df_grp_procedure_network = pd.merge(p29_df_grp_procedure_network, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p29_df_grp_procedure_network_factoring = p29_df_grp_procedure_network.copy(deep=True)
      if annualize == True:
        p29_df_grp_procedure_network_factoring['incurred_amount'] = p29_df_grp_procedure_network_factoring['incurred_amount'] * (12 / p29_df_grp_procedure_network_factoring['data_month'])
        p29_df_grp_procedure_network_factoring['paid_amount']     = p29_df_grp_procedure_network_factoring['paid_amount']     * (12 / p29_df_grp_procedure_network_factoring['data_month'])
        p29_df_grp_procedure_network_factoring['no_of_claim_id']  = p29_df_grp_procedure_network_factoring['no_of_claim_id']  * (12 / p29_df_grp_procedure_network_factoring['data_month'])
        p29_df_grp_procedure_network_factoring['no_of_claimants'] = p29_df_grp_procedure_network_factoring['no_of_claimants']
      if ibnr == True:
        p29_df_grp_procedure_network_factoring['incurred_amount'] = p29_df_grp_procedure_network_factoring['incurred_amount'] * (1 + p29_df_grp_procedure_network_factoring['ibnr'])
        p29_df_grp_procedure_network_factoring['paid_amount']     = p29_df_grp_procedure_network_factoring['paid_amount']     * (1 + p29_df_grp_procedure_network_factoring['ibnr'])
        p29_df_grp_procedure_network_factoring['no_of_claim_id']  = p29_df_grp_procedure_network_factoring['no_of_claim_id']  * (1 + p29_df_grp_procedure_network_factoring['ibnr'])
        p29_df_grp_procedure_network_factoring['no_of_claimants'] = p29_df_grp_procedure_network_factoring['no_of_claimants']
      p29_df_grp_procedure_network_diff = p29_df_grp_procedure_network_factoring[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] - p29_df_grp_procedure_network[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
      temp_by = __p29_group_col.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p29_df_grp_procedure_network_diff = p29_df_grp_procedure_network_diff.reset_index(drop=False)[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p29_df_grp_procedure_network.reset_index().set_index(temp_by, inplace=True)
      p29_df_grp_procedure_network.loc[p29_df_grp_procedure_network.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']] = p29_df_grp_procedure_network.loc[p29_df_grp_procedure_network.factoring == True, ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].add(p29_df_grp_procedure_network_diff)
      temp_by = __p29_group_col.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p29_df_grp_procedure_network = p29_df_grp_procedure_network.reset_index()[temp_by + ['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    __p29_group_col.remove('policy_id') if 'policy_id' in __p29_group_col else __p29_group_col
    p29_df_grp_procedure_network = p29_df_grp_procedure_network.reset_index().set_index(__p29_group_col)
    p29_df_grp_procedure_network = p29_df_grp_procedure_network[['incurred_amount', 'paid_amount', 'no_of_claim_id', 'no_of_claimants']]
    self.p29_df_grp_procedure_network = p29_df_grp_procedure_network
    return p29_df_grp_procedure_network

  def mcr_p30_op_freq_claimant_analysis(self, by=None, annualize=False, ibnr=False, research_mode=False, min_cases=10):

    base_by = list(by) if isinstance(by, list) else ([] if by is None else [by])
    freq_benefits = ['General Consultation (GP)', 'Specialist Consultation (SP)', 'Chinese Med (CMT)', 'Chiro (CT)', 'Physio (PT)']

    cols_needed = base_by + ['class', 'dep_type', 'claimant', 'benefit']
    freq_df = self.mcr_df[cols_needed].copy(deep=True)
    freq_df = freq_df[freq_df['benefit'].isin(freq_benefits)]
    freq_df.dropna(subset=['claimant'], inplace=True)

    if freq_df.empty:
      empty_df = pd.DataFrame(columns=['no_of_claimants', 'no_of_cases'])
      empty_names = base_by + ['class', 'dep_type']
      if len(empty_names) > 0:
        empty_idx = pd.MultiIndex.from_arrays([[] for _ in empty_names], names=empty_names)
        empty_df = empty_df.reindex(empty_idx)
      self.p30_op_freq_claimant = empty_df
      return empty_df

    freq_df['dep_type'] = freq_df['dep_type'].fillna('Unknown')

    group_claimant_cols = base_by + ['class', 'dep_type', 'claimant']
    claimant_cases = (
      freq_df.groupby(by=group_claimant_cols, dropna=False)
      .size()
      .rename('no_of_cases')
      .to_frame()
    )

    claimant_cases = claimant_cases[claimant_cases['no_of_cases'] >= min_cases]

    if claimant_cases.empty:
      empty_df = pd.DataFrame(columns=['no_of_claimants', 'no_of_cases'])
      empty_names = base_by + ['class', 'dep_type']
      if len(empty_names) > 0:
        empty_idx = pd.MultiIndex.from_arrays([[] for _ in empty_names], names=empty_names)
        empty_df = empty_df.reindex(empty_idx)
      self.p30_op_freq_claimant = empty_df
      return empty_df

    claimant_cases = claimant_cases.reset_index()
    final_group_cols = base_by + ['class', 'dep_type']
    p30_df = claimant_cases.groupby(by=final_group_cols, dropna=False).agg({
      'claimant': 'nunique',
      'no_of_cases': 'sum'
    }).rename(columns={'claimant': 'no_of_claimants'})

    if annualize == True or ibnr == True:
      p30_df = pd.merge(p30_df, self.mcr_factor_df, how='left', left_index=True, right_index=True)
      p30_factoring = p30_df.copy(deep=True)
      if annualize == True:
        p30_factoring['no_of_cases'] = p30_factoring['no_of_cases'] * (12 / p30_factoring['data_month'])
        p30_factoring['no_of_claimants'] = p30_factoring['no_of_claimants']
      if ibnr == True:
        p30_factoring['no_of_cases'] = p30_factoring['no_of_cases'] * (1 + p30_factoring['ibnr'])
        p30_factoring['no_of_claimants'] = p30_factoring['no_of_claimants']

      p30_diff = p30_factoring[['no_of_cases', 'no_of_claimants']] - p30_df[['no_of_cases', 'no_of_claimants']]
      temp_by = final_group_cols.copy()
      temp_by.remove('year') if 'year' in temp_by else temp_by
      temp_by.remove('policy_number') if 'policy_number' in temp_by else temp_by
      p30_diff = p30_diff.reset_index(drop=False)[temp_by + ['no_of_cases', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()
      p30_df.reset_index().set_index(temp_by, inplace=True)
      p30_df.loc[p30_df.factoring == True, ['no_of_cases', 'no_of_claimants']] = p30_df.loc[p30_df.factoring == True, ['no_of_cases', 'no_of_claimants']].add(p30_diff)
      temp_by = final_group_cols.copy()
      temp_by.remove('policy_id') if 'policy_id' in temp_by else temp_by
      p30_df = p30_df.reset_index()[temp_by + ['no_of_cases', 'no_of_claimants']].groupby(by=temp_by, dropna=False).sum()

    final_index_cols = final_group_cols.copy()
    final_index_cols.remove('policy_id') if 'policy_id' in final_index_cols else final_index_cols
    p30_df = p30_df.reset_index().set_index(final_index_cols)
    p30_df = p30_df[['no_of_claimants', 'no_of_cases']]

    if research_mode == True:
      p30_df = p30_df.reset_index().groupby(by=final_index_cols, dropna=False).sum()
      p30_df = p30_df[['no_of_claimants', 'no_of_cases']]

    self.p30_op_freq_claimant = p30_df
    return p30_df

  def mcr_pages(self, by=None, export=False, benefit_type_order=['Hospital', 'Clinic', 'Dental', 'Optical', 'Maternity', 'Total'], year_incurred=False, annualize=False, ibnr=False, research_mode=False):
    
    if type(by) is not list and by != None: by = [by]
    

    self.mcr_df = self.df.copy(deep=True)
    freq_op_benefits = ['General Consultation (GP)', 'Specialist Consultation (SP)', 'Chinese Med (CMT)', 'Chiro (CT)', 'Physio (PT)']

    if year_incurred == False:
      self.mcr_df['year'] = self.mcr_df.policy_start_date.dt.year
      
      self.mcr_df['factoring'] = True
    else:
      self.mcr_df['year'] = self.mcr_df.incur_date.dt.year
      self.mcr_df['factoring'] = (self.mcr_df['year'] != self.mcr_df['policy_start_date'].dt.year) | (self.mcr_df['policy_start_date'].dt.year == self.mcr_df['policy_end_date'].dt.year) | (self.mcr_df['policy_start_date'].dt.year == self.mcr_df['policy_data_date'].dt.year)

    
    
    self.mcr_policy_info(by)
    self.mcr_p20_policy(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    self.mcr_p20_benefit(by, benefit_type_order, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if not (isinstance(by, list) and 'dep_type' in by):
      self.mcr_p20_benefit_dep_type(by, benefit_type_order, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    self.mcr_p20_panel(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    self.mcr_p20_panel_benefit(by, benefit_type_order, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    self.mcr_p20_day_procedure(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    self.mcr_p20_day_procedure_class(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    self.mcr_p21_class(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    self.mcr_p22_class_benefit(by, benefit_type_order, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if not (isinstance(by, list) and 'dep_type' in by):
      self.mcr_p22_class_benefit_dep_type(by, benefit_type_order, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit_type'].str.contains('hosp', case=False).any():
      self.mcr_p23_ip_benefit(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p23a_class_ip_benefit(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    self.mcr_p23b_ip_common_diagnosis(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    self.mcr_p23b_ip_common_diagnosis_class(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit_type'].str.contains('clin', case=False).any():
      self.mcr_p24_op_benefit(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p24a_op_class_benefit(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit_type'].str.contains('dent', case=False).any():
      self.mcr_p24_dent_benefit(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p24_wellness_benefit(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p24_class_wellness_benefit(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    self.mcr_p25_class_panel_benefit(by, benefit_type_order, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit_type'].str.contains('clin', case=False).any():
      self.mcr_p26_op_panel(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p26_op_class_panel(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit_type'].str.contains('hosp', case=False).any():
      self.mcr_p26_ip_panel(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    self.mcr_p27_ts(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit_type'].str.contains('hosp', case=False).any():
      self.mcr_p27_ts_ip(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit_type'].str.contains('clin', case=False).any():
      self.mcr_p27_ts_op(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit_type'].str.contains('hosp', case=False).any():
      self.mcr_p28_hosp(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p28_prov(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p28a_prov_benefit(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p28_doct(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p28a_doct(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p28_proce(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode) 
      self.mcr_p28_proce_diagnosis(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p28b_proce_network(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit'].str.contains('specialist', case=False).any():
      self.mcr_p29_sp_speciality(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p29_sp_speciality_diag(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit_type'].str.contains('hosp', case=False).any():
      self.mcr_p29_grp_procedure(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p29_grp_procedure_org(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p29b_grp_procedure_network(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    #self.mcr_p28_class_dep_ip_benefit(by)
    if self.mcr_df['benefit_type'].str.contains('hosp', case=False).any():
      self.mcr_p18a_top_diag_ip(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p18a_top_diag_ip_class(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p18b_top_diag_ip_day_procedure(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p18b_top_diag_ip_day_procedure_class(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit_type'].str.contains('clin', case=False).any():
      self.mcr_p18b_top_diag_op(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p18b_top_diag_op_class(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
      self.mcr_p18ba_top_diag_network_op(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)
    if self.mcr_df['benefit'].isin(freq_op_benefits).any():
      self.mcr_p30_op_freq_claimant_analysis(by, annualize=annualize, ibnr=ibnr, research_mode=research_mode)

    if export == True:
      from io import BytesIO
      output = BytesIO()
      # mcr_filename = 'mcr.xlsx'
      mcr_filename = output
      with pd.ExcelWriter(mcr_filename, engine='openpyxl') as writer:
        self.policy_info.to_excel(writer, sheet_name='Policy_Info', index=True, merge_cells=False)
        self.p20_policy.to_excel(writer, sheet_name='P.20_Policy', index=True, merge_cells=False)
        self.p20.to_excel(writer, sheet_name='P.20_BenefitType', index=True, merge_cells=False)
        if hasattr(self, 'p20_benefit_dep_type') and not self.p20_benefit_dep_type.empty:
          self.p20_benefit_dep_type.to_excel(writer, sheet_name='P.20_Benefit_DepType', index=True, merge_cells=False)
        self.p20_panel.to_excel(writer, sheet_name='P.20_Network', index=True, merge_cells=False)
        self.p20_panel_benefit.to_excel(writer, sheet_name='P.20_Network_BenefitType', index=True, merge_cells=False)
        self.p20_dp.to_excel(writer, sheet_name='P.20_Day_Prod', index=True, merge_cells=False)
        self.p20_dp_class.to_excel(writer, sheet_name='P.20_Day_Prod_Class', index=True, merge_cells=False)
        self.p21.to_excel(writer, sheet_name='P.21_Class', index=True, merge_cells=False)
        self.p22.to_excel(writer, sheet_name='P.22_Class_BenefitType', index=True, merge_cells=False)
        if hasattr(self, 'p22_class_benefit_dep_type') and not self.p22_class_benefit_dep_type.empty:
          self.p22_class_benefit_dep_type.to_excel(writer, sheet_name='P.22_Class_DepType', index=True, merge_cells=False)
        if self.mcr_df['benefit_type'].str.contains('hosp', case=False).any():
          self.p23.to_excel(writer, sheet_name='P.23_IP_Benefit', index=True, merge_cells=False)
          self.p23a.to_excel(writer, sheet_name='P.23a_Class_IP_Benefit', index=True, merge_cells=False)
          self.p23b_ip_common_diagnosis.to_excel(writer, sheet_name='P.23b_Common_Diagnosis_IP', index=True, merge_cells=False)
          self.p23b_ip_common_diagnosis_class.to_excel(writer, sheet_name='P.23b_Class_Common_Diag_IP', index=True, merge_cells=False)
        if self.mcr_df['benefit_type'].str.contains('clin', case=False).any():
          self.p24.to_excel(writer, sheet_name='P.24_OP_Benefit', index=True, merge_cells=False)
          self.p24a.to_excel(writer, sheet_name='P.24a_Class_OP_Benefit', index=True, merge_cells=False)
        if self.mcr_df['benefit_type'].str.contains('dent', case=False).any():
          self.p24_dent_benefit.to_excel(writer, sheet_name='P.24d_Dental', index=True, merge_cells=False)
          self.p24_wellness_benefit.to_excel(writer, sheet_name='P.24w_Wellness', index=True, merge_cells=False)
          self.p24_class_wellness_benefit.to_excel(writer, sheet_name='P.24wc_Class_Wellness', index=True, merge_cells=False)
        self.p25.to_excel(writer, sheet_name='P.25_Class_Panel_BenefitType', index=True, merge_cells=False)
        if self.mcr_df['benefit_type'].str.contains('clin', case=False).any():
          self.p26.to_excel(writer, sheet_name='P.26_OP_Panel_Benefit', index=True, merge_cells=False)
          self.p26_op_class_panel.to_excel(writer, sheet_name='P.26a_OP_Class_Panel_Benefit', index=True, merge_cells=False)
        if self.mcr_df['benefit_type'].str.contains('hosp', case=False).any():
          self.p26_ip_panel.to_excel(writer, sheet_name='P.26b_IP_Panel_Benefit', index=True, merge_cells=False)
        self.p27.to_excel(writer, sheet_name='P.27_TimeSeries', index=True, merge_cells=False)
        if self.mcr_df['benefit_type'].str.contains('hosp', case=False).any():
          self.p27_ip.to_excel(writer, sheet_name='P.27a_TimeSeries_IP', index=True, merge_cells=False)
        if self.mcr_df['benefit_type'].str.contains('clin', case=False).any():
          self.p27_op.to_excel(writer, sheet_name='P.27b_TimeSeries_OP', index=True, merge_cells=False)
        if self.mcr_df['benefit_type'].str.contains('hosp', case=False).any():
          self.p28_hosp.to_excel(writer, sheet_name='P.28_Hospital', index=True, merge_cells=False)
          self.p28_prov.to_excel(writer, sheet_name='P.28_Provider', index=True, merge_cells=False)
          self.p28_prov_benefit.to_excel(writer, sheet_name='P.28a_Provider_Benefit', index=True, merge_cells=False)
          self.p28_doct.to_excel(writer, sheet_name='P.28_Physician', index=True, merge_cells=False)
          self.p28a_doct.to_excel(writer, sheet_name='P.28a_Physician_Benefit', index=True, merge_cells=False)
          self.p28_proce.to_excel(writer, sheet_name='P.28_Procedures', index=True, merge_cells=False)
          self.p28_proce_diagnosis.to_excel(writer, sheet_name='P.28a_Procedures_Diag', index=True, merge_cells=False)
          self.p28_proce_network.to_excel(writer, sheet_name='P.28b_Procedures_Network', index=True, merge_cells=False)
        #self.p28.to_excel(writer, sheet_name='P.28_Class_Dep_IP_Benefit', index=True, merge_cells=False)
        if self.mcr_df['benefit'].str.contains('specialist', case=False).any():
          self.p29_df_sp_specialty.to_excel(writer, sheet_name='P.29_SP_Speciality', index=True, merge_cells=False)
          self.p29_df_sp_specialty_diag.to_excel(writer, sheet_name='P.29_SP_Spec_Diag', index=True, merge_cells=False)
        if self.mcr_df['benefit_type'].str.contains('hosp', case=False).any():
          self.p29_df_grp_procedure.to_excel(writer, sheet_name='P.29_Grp_Procedure', index=True, merge_cells=False)
          self.p29_df_grp_procedure_org.to_excel(writer, sheet_name='P.29_Grp_Proce_Org', index=True, merge_cells=False)
          self.p29_df_grp_procedure_network.to_excel(writer, sheet_name='P.29b_Grp_Proce_Network', index=True, merge_cells=False)
          self.p18a.to_excel(writer, sheet_name='P.18_TopHosDiag', index=True, merge_cells=False)
          self.p18a_class.to_excel(writer, sheet_name='P.18a_Class_TopHosDiag', index=True, merge_cells=False)
          self.p18b_ip_day_procedure.to_excel(writer, sheet_name='P.18b_IP_DayProc', index=True, merge_cells=False)
          self.p18b_ip_day_procedure_class.to_excel(writer, sheet_name='P.18b_Class_IP_DayProc', index=True, merge_cells=False)
        if self.mcr_df['benefit_type'].str.contains('clin', case=False).any():
          self.p18b.to_excel(writer, sheet_name='P.18b_TopClinDiag', index=True, merge_cells=False)
          self.p18b_class.to_excel(writer, sheet_name='P.18b_Class_TopClinDiag', index=True, merge_cells=False)
          self.p18ba.to_excel(writer, sheet_name='P.18b_TopNetClinDiag', index=True, merge_cells=False)
        if self.mcr_df['benefit'].isin(freq_op_benefits).any():
          self.p30_op_freq_claimant.to_excel(writer, sheet_name='P.30_OP_FreqClaimant', index=True, merge_cells=False)

        # Embed custom NamedStyle 'num' into the workbook
        wb = writer.book
        try:
          existing_names = [ns if isinstance(ns, str) else ns.name for ns in wb.named_styles]
        except Exception:
          existing_names = []

        if 'num' not in existing_names:
          num_style = NamedStyle(name='num')
          num_style.number_format = '#,##0'
          num_style.alignment = Alignment(horizontal='center', vertical='center')
          num_style.font = Font(name='Univers', size=14)
          num_style.border = Border()  # No border
          num_style.fill = PatternFill(fill_type=None)  # No shading
          num_style.protection = Protection(locked=False, hidden=False)  # No protection
          wb.add_named_style(num_style)


        # writer.add_named_style(per_format)
        writer.close()
        # processed_data = output.getvalue()


        return output.getvalue()


  def mcr_p20_ibnr(self, ibnr=1, no_month=9):

    self.p20_ibnr = self.p20['paid_amount']
    self.p20_ibnr['paid_ibnr'] = self.p20_ibnr['paid_amount'] * ibnr
    self.p20_ibnr['paid_amount_annu'] = self.p20_ibnr['paid_amount'] * 12 / no_month
    self.p20_ibnr['paid_ibnr_annu'] = self.p20_ibnr['paid_ibnr'] * 12 / no_month


  def time_comparison(self, start_m = None, policy_number=None, sub_policy=False):
    if start_m == None:
      start_m = str(self.df.policy_start_date.iloc[0].dt.month)
    order = [str(i) for i in np.arange(start_m, 13).tolist()] + [str(j) for j in np.arange(1, start_m).tolist()]
    # print(order)
    self.df['year'] = self.df.policy_start_date.dt.year
    self.df['month'] = self.df.incur_date.dt.month.astype(str)
    time_com = self.df[['policy_number', 'year', 'month', 'paid_amount']].groupby(by=['policy_number', 'year', 'month']).sum()

    if policy_number == None:
      policy_list = time_com.index.get_level_values(0).unique()
    else:
      policy_list = policy_number

    year_list = time_com.index.get_level_values(1).unique()

    self.by_time, self.by_time_ax = plt.subplots(figsize=(15, 5))

    for p in policy_list:
      for y in year_list:
        __tdf = time_com.loc[p, y].reindex(order)
        self.by_time_ax.plot(__tdf.index, __tdf.paid_amount, marker='o', label=f'{p}__{y}', linewidth=2, markersize=6)
        self.by_time_ax.set_title('Monthly Trend Claim Paid Comparison', fontsize=18)
        # self.by_time_ax.set_ylim(0, 1_500_000)
        self.by_time_ax.set_xticks(order)
        self.by_time_ax.set_xlabel('Month')
        self.by_time_ax.set_ylabel('Claim Paid Amount')
        self.by_time_ax.legend()



      # self.whatever_df.index.strftime('%b')
    # return self.by_time

  def frequent_claimant_analysis(self, export=True, sub_policy=None):
    freq_op_list = ['General Consultation (GP)', 'Specialist Consultation (SP)', 'Chinese Med (CMT)', 'Chiro (CT)', 'Physio (PT)', 'Diagnostic: X-Ray & Lab Test (DX)', 'Prescribed Medicine (PM)']
    total_visit_col = ['General Consultation (GP)', 'Specialist Consultation (SP)', 'Chinese Med (CMT)', 'Chiro (CT)', 'Physio (PT)']
    self.df.policy_start_date = pd.to_datetime(self.df.policy_start_date)
    self.df['year'] = self.df.policy_start_date.dt.year
    self.df.claimant = self.df.policy_id.str.cat(self.df.claimant, sep='_')
    __dep = self.df
    # __dep.claimant = __dep.policy_number.str.cat(__dep.year.astype(str), sep='_').str.cat(__dep.claimant, sep='_')
    __dep = __dep[['claimant', 'dep_type', 'class', 'suboffice', 'age', 'gender']]
    __dep.drop_duplicates(subset=['claimant', 'suboffice'], keep='first', inplace=True)
    __freq_df = self.df[['policy_number', 'year', 'claimant', 'benefit', 'incur_date', 'panel']].dropna()
    __freq_df_panel = __freq_df.loc[(__freq_df.benefit.isin(freq_op_list))&(__freq_df.panel == "Panel")].drop(columns='panel').groupby(['policy_number', 'year', 'claimant', 'benefit']).count()
    __freq_df_non_panel = __freq_df.loc[__freq_df.benefit.isin(freq_op_list)&(__freq_df.panel == "Non-Panel")].drop(columns='panel').groupby(['policy_number', 'year', 'claimant', 'benefit']).count()
    __freq_df = __freq_df.drop(columns='panel').loc[__freq_df.benefit.isin(freq_op_list)].groupby(['policy_number', 'year', 'claimant', 'benefit']).count()
    
    # print(__freq_df)
    __freq_df = __freq_df.unstack()
    __freq_df_panel = __freq_df_panel.unstack()
    __freq_df_non_panel = __freq_df_non_panel.unstack()
    __dep.index = __dep['claimant'].values.tolist()
    __freq_df.columns = __freq_df.columns.droplevel()
    __freq_df_panel.columns = __freq_df_panel.columns.droplevel()
    __freq_df_non_panel.columns = __freq_df_non_panel.columns.droplevel()

    for col in total_visit_col:
      if col not in __freq_df.columns:
        __freq_df[col] = np.nan
      if col not in __freq_df_panel.columns:
        __freq_df_panel[col] = np.nan
      if col not in __freq_df_non_panel.columns:
        __freq_df_non_panel[col] = np.nan
    __freq_df['total_claims'] = __freq_df[total_visit_col].sum(axis=1)
    __freq_df_panel['total_claims'] = __freq_df_panel[total_visit_col].sum(axis=1)
    __freq_df_non_panel['total_claims'] = __freq_df_non_panel[total_visit_col].sum(axis=1)
    # __freq_df = __freq_df.sort_values(by=['policy_number', 'year', 'total_claims'], ascending=False)
    # print(__dep)
    # print(__freq_df)
    #__freq_df = pd.merge(left=__freq_df, right=__dep, left_index=True, right_index=True, how='left')
    __freq_df = pd.merge(left=__freq_df, right=__dep, left_on='claimant', right_index=True, how='left').drop(columns=['claimant'])
    __freq_df = __freq_df.sort_values(by=['policy_number', 'year', 'total_claims'], ascending=[True, True, False])
    __freq_df_panel = pd.merge(left=__freq_df_panel, right=__dep, left_on='claimant', right_index=True, how='left').drop(columns=['claimant'])
    __freq_df_panel = __freq_df_panel.sort_values(by=['policy_number', 'year', 'total_claims'], ascending=[True, True, False])
    __freq_df_non_panel = pd.merge(left=__freq_df_non_panel, right=__dep, left_on='claimant', right_index=True, how='left').drop(columns=['claimant'])
    __freq_df_non_panel = __freq_df_non_panel.sort_values(by=['policy_number', 'year', 'total_claims'], ascending=[True, True, False])

    __freq_df['GP + SP'] = __freq_df[['General Consultation (GP)', 'Specialist Consultation (SP)']].sum(axis=1)
    __freq_df['Physio + Chiro'] = __freq_df[['Chiro (CT)', 'Physio (PT)']].sum(axis=1)
    __freq_df_panel['GP + SP'] = __freq_df_panel[['General Consultation (GP)', 'Specialist Consultation (SP)']].sum(axis=1)
    __freq_df_panel['Physio + Chiro'] = __freq_df_panel[['Chiro (CT)', 'Physio (PT)']].sum(axis=1)
    __freq_df_non_panel['GP + SP'] = __freq_df_non_panel[['General Consultation (GP)', 'Specialist Consultation (SP)']].sum(axis=1)
    __freq_df_non_panel['Physio + Chiro'] = __freq_df_non_panel[['Chiro (CT)', 'Physio (PT)']].sum(axis=1)

    __freq_df = __freq_df.reset_index()
    __freq_df_panel = __freq_df_panel.reset_index()
    __freq_df_non_panel = __freq_df_non_panel.reset_index()

    __freq_sum_benefit_df = self.df[['claimant', 'benefit', 'paid_amount']].loc[self.df.benefit.isin(total_visit_col)].groupby(['claimant', 'benefit']).mean().rename(columns={'paid_amount': 'total_paid'}).unstack().stack(dropna=False)
    __freq_sum_benefit_df = __freq_sum_benefit_df.reindex(total_visit_col, level='benefit').unstack()
    __freq_sum_benefit_df_panel = self.df[['claimant', 'benefit', 'paid_amount']].loc[(self.df.benefit.isin(total_visit_col))&(self.df.panel == "Panel")].groupby(['claimant', 'benefit']).mean().rename(columns={'paid_amount': 'total_paid'}).unstack().stack(dropna=False)
    __freq_sum_benefit_df_panel = __freq_sum_benefit_df_panel.reindex(total_visit_col, level='benefit').unstack()
    __freq_sum_benefit_df_non_panel = self.df[['claimant', 'benefit', 'paid_amount']].loc[self.df.benefit.isin(total_visit_col)&(self.df.panel == "Non-Panel")].groupby(['claimant', 'benefit']).mean().rename(columns={'paid_amount': 'total_paid'}).unstack().stack(dropna=False)
    __freq_sum_benefit_df_non_panel = __freq_sum_benefit_df_non_panel.reindex(total_visit_col, level='benefit').unstack()

    __freq_sum_benefit_df.columns = __freq_sum_benefit_df.columns.get_level_values(1).to_list()
    __freq_sum_benefit_df_panel.columns = __freq_sum_benefit_df_panel.columns.get_level_values(1).to_list()
    __freq_sum_benefit_df_non_panel.columns = __freq_sum_benefit_df_non_panel.columns.get_level_values(1).to_list()
    for col in total_visit_col:
      if col not in __freq_sum_benefit_df.columns:
        __freq_sum_benefit_df[col] = np.nan
      if col not in __freq_sum_benefit_df_panel.columns:
        __freq_sum_benefit_df_panel[col] = np.nan
      if col not in __freq_sum_benefit_df_non_panel.columns:
        __freq_sum_benefit_df_non_panel[col] = np.nan
    __freq_sum_benefit_df = __freq_sum_benefit_df[total_visit_col]
    __freq_sum_benefit_df_panel = __freq_sum_benefit_df_panel[total_visit_col]
    __freq_sum_benefit_df_non_panel = __freq_sum_benefit_df_non_panel[total_visit_col]

    __freq_sum_benefit_df.columns = [f'{b}_paid_per_claim' for b in total_visit_col]
    __freq_sum_benefit_df_panel.columns = [f'{b}_paid_per_claim' for b in total_visit_col]
    __freq_sum_benefit_df_non_panel.columns = [f'{b}_paid_per_claim' for b in total_visit_col]
    __freq_inc_benefit_df = self.df[['claimant', 'benefit', 'incurred_amount']].loc[self.df.benefit.isin(total_visit_col)].groupby(['claimant', 'benefit']).mean().rename(columns={'incurred_amount': 'total_incurred'}).unstack().stack(dropna=False)
    __freq_inc_benefit_df = __freq_inc_benefit_df.reindex(total_visit_col, level='benefit').unstack()
    __freq_inc_benefit_df_panel = self.df[['claimant', 'benefit', 'incurred_amount']].loc[(self.df.benefit.isin(total_visit_col))&(self.df.panel == "Panel")].groupby(['claimant', 'benefit']).mean().rename(columns={'incurred_amount': 'total_incurred'}).unstack().stack(dropna=False)
    __freq_inc_benefit_df_panel = __freq_inc_benefit_df_panel.reindex(total_visit_col, level='benefit').unstack()
    __freq_inc_benefit_df_non_panel = self.df[['claimant', 'benefit', 'incurred_amount']].loc[self.df.benefit.isin(total_visit_col)&(self.df.panel == "Non-Panel")].groupby(['claimant', 'benefit']).mean().rename(columns={'incurred_amount': 'total_incurred'}).unstack().stack(dropna=False)
    __freq_inc_benefit_df_non_panel = __freq_inc_benefit_df_non_panel.reindex(total_visit_col, level='benefit').unstack()

    __freq_inc_benefit_df.columns = __freq_inc_benefit_df.columns.get_level_values(1).to_list()
    __freq_inc_benefit_df_panel.columns = __freq_inc_benefit_df_panel.columns.get_level_values(1).to_list()
    __freq_inc_benefit_df_non_panel.columns = __freq_inc_benefit_df_non_panel.columns.get_level_values(1).to_list()
    for col in total_visit_col:
      if col not in __freq_inc_benefit_df.columns:
        __freq_inc_benefit_df[col] = np.nan
      if col not in __freq_inc_benefit_df_panel.columns:
        __freq_inc_benefit_df_panel[col] = np.nan
      if col not in __freq_inc_benefit_df_non_panel.columns:
        __freq_inc_benefit_df_non_panel[col] = np.nan
    __freq_inc_benefit_df = __freq_inc_benefit_df[total_visit_col]
    __freq_inc_benefit_df_panel = __freq_inc_benefit_df_panel[total_visit_col]
    __freq_inc_benefit_df_non_panel = __freq_inc_benefit_df_non_panel[total_visit_col]

    __freq_inc_benefit_df.columns = [f'{b}_incurred_per_claim' for b in total_visit_col]
    __freq_inc_benefit_df_panel.columns = [f'{b}_incurred_per_claim' for b in total_visit_col]
    __freq_inc_benefit_df_non_panel.columns = [f'{b}_incurred_per_claim' for b in total_visit_col]
    __freq_sum_df = self.df[['claimant', 'paid_amount']].loc[self.df.benefit.isin(total_visit_col)].groupby(['claimant']).sum().rename(columns={'paid_amount': 'total_paid'})
    __freq_sum_df_panel = self.df[['claimant', 'paid_amount']].loc[(self.df.benefit.isin(total_visit_col))&(self.df.panel == "Panel")].groupby(['claimant']).sum().rename(columns={'paid_amount': 'total_paid'})
    __freq_sum_df_non_panel = self.df[['claimant', 'paid_amount']].loc[(self.df.benefit.isin(total_visit_col))&(self.df.panel == "Non-Panel")].groupby(['claimant']).sum().rename(columns={'paid_amount': 'total_paid'})
    
    __freq_df = pd.merge(left=__freq_df, right=__freq_inc_benefit_df, left_on='claimant', right_on='claimant', how='left')
    __freq_df = pd.merge(left=__freq_df, right=__freq_sum_benefit_df, left_on='claimant', right_on='claimant', how='left')
    __freq_df = pd.merge(left=__freq_df, right=__freq_sum_df, left_on='claimant', right_on='claimant', how='left')

    __freq_df_panel = pd.merge(left=__freq_df_panel, right=__freq_inc_benefit_df_panel, left_on='claimant', right_on='claimant', how='left')
    __freq_df_panel = pd.merge(left=__freq_df_panel, right=__freq_sum_benefit_df_panel, left_on='claimant', right_on='claimant', how='left')
    __freq_df_panel = pd.merge(left=__freq_df_panel, right=__freq_sum_df_panel, left_on='claimant', right_on='claimant', how='left')

    __freq_df_non_panel = pd.merge(left=__freq_df_non_panel, right=__freq_inc_benefit_df_non_panel, left_on='claimant', right_on='claimant', how='left')
    __freq_df_non_panel = pd.merge(left=__freq_df_non_panel, right=__freq_sum_benefit_df_non_panel, left_on='claimant', right_on='claimant', how='left')
    __freq_df_non_panel = pd.merge(left=__freq_df_non_panel, right=__freq_sum_df_non_panel, left_on='claimant', right_on='claimant', how='left')

    __freq_dx_df = self.df[['claimant', 'incurred_amount', 'paid_amount']].loc[self.df.benefit.str.contains('DX', case=False)].groupby(['claimant']).agg({'incurred_amount': 'sum', 'paid_amount': 'sum'}).rename(columns={'incurred_amount': 'Diagnostic: X-Ray & Lab Test (DX)_incurred', 'paid_amount': 'Diagnostic: X-Ray & Lab Test (DX)_paid'})
    __freq_df = pd.merge(left=__freq_df, right=__freq_dx_df, left_on='claimant', right_on='claimant', how='left')
    __freq_pm_df = self.df[['claimant', 'incurred_amount', 'paid_amount']].loc[self.df.benefit.str.contains('PM', case=False)].groupby(['claimant']).agg({'incurred_amount': 'sum', 'paid_amount': 'sum'}).rename(columns={'incurred_amount': 'Prescribed Medicine (PM)_incurred', 'paid_amount': 'Prescribed Medicine (PM)_paid'})
    __freq_df = pd.merge(left=__freq_df, right=__freq_pm_df, left_on='claimant', right_on='claimant', how='left')
    __freq_df['paid_per_claim'] = __freq_df['total_paid'] / __freq_df['total_claims']

    __freq_dx_df_panel = self.df[['claimant', 'incurred_amount', 'paid_amount']].loc[(self.df.benefit.str.contains('DX', case=False))&(self.df.panel == "Panel")].groupby(['claimant']).agg({'incurred_amount': 'sum', 'paid_amount': 'sum'}).rename(columns={'incurred_amount': 'Diagnostic: X-Ray & Lab Test (DX)_incurred', 'paid_amount': 'Diagnostic: X-Ray & Lab Test (DX)_paid'})
    __freq_df_panel = pd.merge(left=__freq_df_panel, right=__freq_dx_df_panel, left_on='claimant', right_on='claimant', how='left')
    __freq_pm_df_panel = self.df[['claimant', 'incurred_amount', 'paid_amount']].loc[(self.df.benefit.str.contains('PM', case=False))&(self.df.panel == "Panel")].groupby(['claimant']).agg({'incurred_amount': 'sum', 'paid_amount': 'sum'}).rename(columns={'incurred_amount': 'Prescribed Medicine (PM)_incurred', 'paid_amount': 'Prescribed Medicine (PM)_paid'})
    __freq_df_panel = pd.merge(left=__freq_df_panel, right=__freq_pm_df_panel, left_on='claimant', right_on='claimant', how='left')
    __freq_df_panel['paid_per_claim'] = __freq_df_panel['total_paid'] / __freq_df_panel['total_claims']

    __freq_dx_df_non_panel = self.df[['claimant', 'incurred_amount', 'paid_amount']].loc[(self.df.benefit.str.contains('DX', case=False))&(self.df.panel == "Non-Panel")].groupby(['claimant']).agg({'incurred_amount': 'sum', 'paid_amount': 'sum'}).rename(columns={'incurred_amount': 'Diagnostic: X-Ray & Lab Test (DX)_incurred', 'paid_amount': 'Diagnostic: X-Ray & Lab Test (DX)_paid'})
    __freq_df_non_panel = pd.merge(left=__freq_df_non_panel, right=__freq_dx_df_non_panel, left_on='claimant', right_on='claimant', how='left')
    __freq_pm_df_non_panel = self.df[['claimant', 'incurred_amount', 'paid_amount']].loc[(self.df.benefit.str.contains('PM', case=False))&(self.df.panel == "Non-Panel")].groupby(['claimant']).agg({'incurred_amount': 'sum', 'paid_amount': 'sum'}).rename(columns={'incurred_amount': 'Prescribed Medicine (PM)_incurred', 'paid_amount': 'Prescribed Medicine (PM)_paid'})
    __freq_df_non_panel = pd.merge(left=__freq_df_non_panel, right=__freq_pm_df_non_panel, left_on='claimant', right_on='claimant', how='left')
    __freq_df_non_panel['paid_per_claim'] = __freq_df_non_panel['total_paid'] / __freq_df_non_panel['total_claims']

    if 'Diagnostic: X-Ray & Lab Test (DX)' in __freq_df.columns.to_list():
      __freq_df['Diagnostic: X-Ray & Lab Test (DX)_paid_per_claim'] = __freq_df['Diagnostic: X-Ray & Lab Test (DX)_paid'] / __freq_df['Diagnostic: X-Ray & Lab Test (DX)']
    else:
      __freq_df['Diagnostic: X-Ray & Lab Test (DX)_paid_per_claim'] = np.nan
      __freq_df['Diagnostic: X-Ray & Lab Test (DX)'] = np.nan
      __freq_df['Diagnostic: X-Ray & Lab Test (DX)_paid'] = np.nan
    
    if 'Prescribed Medicine (PM)' in __freq_df.columns.to_list():
      __freq_df['Prescribed Medicine (PM)_paid_per_claim'] = __freq_df['Prescribed Medicine (PM)_paid'] / __freq_df['Prescribed Medicine (PM)']
    else:
      __freq_df['Prescribed Medicine (PM)_paid_per_claim'] = np.nan
      __freq_df['Prescribed Medicine (PM)'] = np.nan
      __freq_df['Prescribed Medicine (PM)_paid'] = np.nan
    __freq_df = __freq_df.set_index(['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'gender'])

    if 'Diagnostic: X-Ray & Lab Test (DX)' in __freq_df_panel.columns.to_list():
      __freq_df_panel['Diagnostic: X-Ray & Lab Test (DX)_paid_per_claim'] = __freq_df_panel['Diagnostic: X-Ray & Lab Test (DX)_paid'] / __freq_df_panel['Diagnostic: X-Ray & Lab Test (DX)']
    else:
      __freq_df_panel['Diagnostic: X-Ray & Lab Test (DX)_paid_per_claim'] = np.nan
      __freq_df_panel['Diagnostic: X-Ray & Lab Test (DX)'] = np.nan
      __freq_df_panel['Diagnostic: X-Ray & Lab Test (DX)_paid'] = np.nan
    
    if 'Prescribed Medicine (PM)' in __freq_df_panel.columns.to_list():
      __freq_df_panel['Prescribed Medicine (PM)_paid_per_claim'] = __freq_df_panel['Prescribed Medicine (PM)_paid'] / __freq_df_panel['Prescribed Medicine (PM)']
    else:
      __freq_df_panel['Prescribed Medicine (PM)_paid_per_claim'] = np.nan
      __freq_df_panel['Prescribed Medicine (PM)'] = np.nan
      __freq_df_panel['Prescribed Medicine (PM)_paid'] = np.nan
    __freq_df_panel = __freq_df_panel.set_index(['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'gender'])

    if 'Diagnostic: X-Ray & Lab Test (DX)' in __freq_df_non_panel.columns.to_list():
      __freq_df_non_panel['Diagnostic: X-Ray & Lab Test (DX)_paid_per_claim'] = __freq_df_non_panel['Diagnostic: X-Ray & Lab Test (DX)_paid'] / __freq_df_non_panel['Diagnostic: X-Ray & Lab Test (DX)']
    else:
      __freq_df_non_panel['Diagnostic: X-Ray & Lab Test (DX)_paid_per_claim'] = np.nan
      __freq_df_non_panel['Diagnostic: X-Ray & Lab Test (DX)'] = np.nan
      __freq_df_non_panel['Diagnostic: X-Ray & Lab Test (DX)_paid'] = np.nan
    
    if 'Prescribed Medicine (PM)' in __freq_df_non_panel.columns.to_list():
      __freq_df_non_panel['Prescribed Medicine (PM)_paid_per_claim'] = __freq_df_non_panel['Prescribed Medicine (PM)_paid'] / __freq_df_non_panel['Prescribed Medicine (PM)']
    else:
      __freq_df_non_panel['Prescribed Medicine (PM)_paid_per_claim'] = np.nan
      __freq_df_non_panel['Prescribed Medicine (PM)'] = np.nan
      __freq_df_non_panel['Prescribed Medicine (PM)_paid'] = np.nan
    __freq_df_non_panel = __freq_df_non_panel.set_index(['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'gender'])

    # __freq_df.reset_index(inplace=True)
    __freq_df = __freq_df[['General Consultation (GP)', 'Specialist Consultation (SP)', 'Chinese Med (CMT)', 'Chiro (CT)', 'Physio (PT)',
                            'total_claims', 'total_paid', 'paid_per_claim', 'GP + SP', 'Physio + Chiro', 
                            'Diagnostic: X-Ray & Lab Test (DX)', 'Diagnostic: X-Ray & Lab Test (DX)_incurred', 'Diagnostic: X-Ray & Lab Test (DX)_paid', 'Diagnostic: X-Ray & Lab Test (DX)_paid_per_claim',
                            'Prescribed Medicine (PM)', 'Prescribed Medicine (PM)_incurred', 'Prescribed Medicine (PM)_paid', 'Prescribed Medicine (PM)_paid_per_claim',
                            'General Consultation (GP)_incurred_per_claim', 'Specialist Consultation (SP)_incurred_per_claim', 'Chinese Med (CMT)_incurred_per_claim', 
                            'Chiro (CT)_incurred_per_claim', 'Physio (PT)_incurred_per_claim',
                            'General Consultation (GP)_paid_per_claim', 'Specialist Consultation (SP)_paid_per_claim', 'Chinese Med (CMT)_paid_per_claim', 
                            'Chiro (CT)_paid_per_claim', 'Physio (PT)_paid_per_claim']]
    __freq_df_panel = __freq_df_panel[['General Consultation (GP)', 'Specialist Consultation (SP)', 'Chinese Med (CMT)', 'Chiro (CT)', 'Physio (PT)',
                            'total_claims', 'total_paid', 'paid_per_claim', 'GP + SP', 'Physio + Chiro', 
                            'Diagnostic: X-Ray & Lab Test (DX)', 'Diagnostic: X-Ray & Lab Test (DX)_incurred', 'Diagnostic: X-Ray & Lab Test (DX)_paid', 'Diagnostic: X-Ray & Lab Test (DX)_paid_per_claim',
                            'Prescribed Medicine (PM)', 'Prescribed Medicine (PM)_incurred', 'Prescribed Medicine (PM)_paid', 'Prescribed Medicine (PM)_paid_per_claim',
                            'General Consultation (GP)_incurred_per_claim', 'Specialist Consultation (SP)_incurred_per_claim', 'Chinese Med (CMT)_incurred_per_claim', 
                            'Chiro (CT)_incurred_per_claim', 'Physio (PT)_incurred_per_claim',
                            'General Consultation (GP)_paid_per_claim', 'Specialist Consultation (SP)_paid_per_claim', 'Chinese Med (CMT)_paid_per_claim', 
                            'Chiro (CT)_paid_per_claim', 'Physio (PT)_paid_per_claim']]
    __freq_df_non_panel = __freq_df_non_panel[['General Consultation (GP)', 'Specialist Consultation (SP)', 'Chinese Med (CMT)', 'Chiro (CT)', 'Physio (PT)',
                            'total_claims', 'total_paid', 'paid_per_claim', 'GP + SP', 'Physio + Chiro', 
                            'Diagnostic: X-Ray & Lab Test (DX)', 'Diagnostic: X-Ray & Lab Test (DX)_incurred', 'Diagnostic: X-Ray & Lab Test (DX)_paid', 'Diagnostic: X-Ray & Lab Test (DX)_paid_per_claim',
                            'Prescribed Medicine (PM)', 'Prescribed Medicine (PM)_incurred', 'Prescribed Medicine (PM)_paid', 'Prescribed Medicine (PM)_paid_per_claim',
                            'General Consultation (GP)_incurred_per_claim', 'Specialist Consultation (SP)_incurred_per_claim', 'Chinese Med (CMT)_incurred_per_claim', 
                            'Chiro (CT)_incurred_per_claim', 'Physio (PT)_incurred_per_claim',
                            'General Consultation (GP)_paid_per_claim', 'Specialist Consultation (SP)_paid_per_claim', 'Chinese Med (CMT)_paid_per_claim', 
                            'Chiro (CT)_paid_per_claim', 'Physio (PT)_paid_per_claim']]
    __freq_df_panel.columns = [f'panel_{col}' for col in __freq_df_panel.columns.to_list()]
    __freq_df_non_panel.columns = [f'non_panel_{col}' for col in __freq_df_non_panel.columns.to_list()]
    __freq_df = pd.merge(left = __freq_df, right=__freq_df_panel, left_index=True, right_index=True, how='left')
    __freq_df = pd.merge(left = __freq_df, right=__freq_df_non_panel, left_index=True, right_index=True, how='left')

    self.frequent_analysis = __freq_df

    def descriptive_stats(data):
    # Calculate statistics
      stats = data.describe()
      # Add median to the statistics
      stats.loc['median'] = data.median()
      stats.loc['90%'] = data.quantile(q=0.9)
      stats.loc['95%'] = data.quantile(q=0.95)
      stats.loc['97%'] = data.quantile(q=0.97)
      stats.loc['99%'] = data.quantile(q=0.99)
      # Reorder the statistics
      stats = stats.reindex(['count', 'mean', 'median', 'std', 'min', '25%', '50%', '75%', '90%', '95%', '97%', '99%', 'max'])
      return stats


    numeric_columns = __freq_df.select_dtypes(include=[np.number]).columns
    order_cols = ['year'] + numeric_columns.tolist()
    __freq_stat_df = pd.DataFrame()
    for y in __freq_df.index.get_level_values(level='year').unique():
      __temp_df = pd.DataFrame()
      for column in numeric_columns:
        # print(f"\nDescriptive Statistics for {column}:")
        __temp_df = pd.concat([__temp_df, descriptive_stats(__freq_df[column].loc[__freq_df.index.get_level_values(level='year') == y])], axis=1, ignore_index=False)
      __temp_df['year'] = y
      __temp_df = __temp_df[order_cols]
      __freq_stat_df = pd.concat([__freq_stat_df, __temp_df], axis=0)

    self.frequent_analysis_stat = __freq_stat_df

    hosp_mask = self.df.benefit_type.str.contains('hosp', case=False)

    self.ip_usage = self.df[['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'gender', 'benefit', 'diagnosis', 'paid_amount']] \
    .loc[hosp_mask] \
    .groupby(by=['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'gender', 'diagnosis', 'benefit']).sum().unstack()
    cols = [list(f)[-1] for f in self.ip_usage.columns]
    self.ip_usage.columns = cols

    self.ip_usage_incurred = self.df[['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'gender', 'benefit', 'diagnosis', 'incurred_amount']] \
    .loc[hosp_mask] \
    .groupby(by=['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'gender', 'diagnosis', 'benefit']).sum().unstack()
    cols = [list(f)[-1] for f in self.ip_usage_incurred.columns]
    self.ip_usage_incurred.columns = cols

    if self.df['discharge_date'].notna().any():
      days = self.df.loc[self.df.benefit.str.contains('room', case=False) & self.df['discharge_date'].notna()].copy()
      days['days_cover'] = (days['discharge_date'] - days['incur_date']).dt.days + 1

      hospital_days = days.loc[days['benefit_type'].str.contains('hosp', case=False)]

      days_cover = hospital_days[['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'gender', 'diagnosis', 'days_cover']] \
      .groupby(by=['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'gender', 'diagnosis']).sum()

      self.ip_usage_incurred = pd.concat([self.ip_usage_incurred, days_cover], axis=1, ignore_index=False)

      days_cover = hospital_days[['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'gender', 'benefit', 'diagnosis', 'days_cover']] \
      .groupby(by=['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'gender', 'diagnosis', 'benefit']).sum()

      # self.ip_usage_for_cal = self.df[['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'benefit', 'diagnosis', 'incurred_amount', 'paid_amount']] \
      # .loc[self.df.benefit_type.str.contains('hosp', case=False)] \
      # .groupby(by=['policy_number', 'year', 'suboffice', 'claimant', 'class', 'dep_type', 'age', 'diagnosis', 'benefit']).sum()

      # self.ip_usage_for_cal = pd.concat([self.ip_usage_for_cal, days_cover], axis=1, ignore_index=False)

    

    if export == True:
      from io import BytesIO
      output = BytesIO()
      # Write to an in-memory buffer using a deterministic engine for .xlsx
      with pd.ExcelWriter(output, engine='openpyxl') as writer_2:
        self.frequent_analysis.to_excel(writer_2, sheet_name='Claimant Visits', index=True, merge_cells=False)
        self.frequent_analysis_stat.to_excel(writer_2, sheet_name='Claimant Statistics', index=True, merge_cells=False)
        self.ip_usage.to_excel(writer_2, sheet_name='Claimant IP Usage', index=True, merge_cells=False)
        self.ip_usage_incurred.to_excel(writer_2, sheet_name='Claimant IP Usage Incurred', index=True, merge_cells=False)
        # self.ip_usage_for_cal.to_excel(writer_2, sheet_name='Claimant IP Usage for Calculation', index=True, merge_cells=False)
      return output.getvalue()

  def fair_ibnr_estimation(self, months=9):
    __fair_ibnr_df = self.df.copy(deep=True)
    __fair_ibnr_df['year']  = __fair_ibnr_df['policy_start_date'].dt.year.astype(str)
    __fair_ibnr_df['ibnr_date'] = __fair_ibnr_df['policy_start_date'] + pd.DateOffset(months=months)
    __fair_ibnr_df['ibnr_paid'] = 0
    __fair_ibnr_df['ibnr_paid'].loc[(__fair_ibnr_df['ibnr_date'] <= __fair_ibnr_df['submission_date']) & (__fair_ibnr_df['ibnr_date'] > __fair_ibnr_df['incur_date'])] = __fair_ibnr_df['paid_amount'].loc[(__fair_ibnr_df['ibnr_date'] <= __fair_ibnr_df['submission_date']) & (__fair_ibnr_df['ibnr_date'] > __fair_ibnr_df['incur_date'])]
    __fair_ibnr_df['reported_paid'] = 0
    __fair_ibnr_df['reported_paid'].loc[(__fair_ibnr_df['ibnr_date'] > __fair_ibnr_df['submission_date'])] = __fair_ibnr_df['paid_amount'].loc[(__fair_ibnr_df['ibnr_date'] > __fair_ibnr_df['submission_date'])]
    __fair_ibnr_df['hospital_ibnr_paid'], __fair_ibnr_df['hospital_reported_paid'], __fair_ibnr_df['clinic_ibnr_paid'], __fair_ibnr_df['clinic_reported_paid'] = 0, 0, 0, 0
    __fair_ibnr_df['hospital_ibnr_paid'].loc[__fair_ibnr_df.benefit_type == 'Hospital'] = __fair_ibnr_df['ibnr_paid'].loc[__fair_ibnr_df.benefit_type == 'Hospital']
    __fair_ibnr_df['hospital_reported_paid'].loc[__fair_ibnr_df.benefit_type == 'Hospital'] = __fair_ibnr_df['reported_paid'].loc[__fair_ibnr_df.benefit_type == 'Hospital']
    __fair_ibnr_df['clinic_ibnr_paid'].loc[__fair_ibnr_df.benefit_type == 'Clinic'] = __fair_ibnr_df['ibnr_paid'].loc[__fair_ibnr_df.benefit_type == 'Clinic']
    __fair_ibnr_df['clinic_reported_paid'].loc[__fair_ibnr_df.benefit_type == 'Clinic'] = __fair_ibnr_df['reported_paid'].loc[__fair_ibnr_df.benefit_type == 'Clinic']
    
    fair_ibnr = __fair_ibnr_df[['policy_number', 'year', 'reported_paid', 'ibnr_paid', 'hospital_reported_paid', 'hospital_ibnr_paid', 'clinic_reported_paid', 'clinic_ibnr_paid']].groupby(by=['policy_number', 'year']).sum()
    fair_ibnr['ibnr_rate'] = fair_ibnr['ibnr_paid'] / fair_ibnr['reported_paid']
    fair_ibnr['hospital_ibnr_rate'] = fair_ibnr['hospital_ibnr_paid'] / fair_ibnr['hospital_reported_paid']
    fair_ibnr['clinic_ibnr_rate'] = fair_ibnr['clinic_ibnr_paid'] / fair_ibnr['clinic_reported_paid']
    fair_ibnr = fair_ibnr[['reported_paid', 'ibnr_paid', 'ibnr_rate', 'hospital_reported_paid', 'hospital_ibnr_paid', 'hospital_ibnr_rate', 'clinic_reported_paid', 'clinic_ibnr_paid', 'clinic_ibnr_rate']]
    self.fair_ibnr = fair_ibnr
    return fair_ibnr

  def make_autopct(values):
    def my_autopct(pct):
        total = np.sum(values)
        val = int(round(pct*total/100.0))
        return '{p:.1f}%  ({v:,})'.format(p=pct,v=val)
    return my_autopct

  ################################################################################################
  # this function should modify into incident rate
  ################################################################################################

  def benefit_op_monthly(self, benefit=['GP', 'CMT', 'SP', 'PM', 'CT', 'PT', 'DX', 'Psy', 'Vac', 'Check']):
    self.df['year'] = self.df.policy_start_date.dt.year
    __benefit_df = self.df[['benefit', 'incur_date', 'policy_id']].loc[self.df['benefit'].str.contains('|'.join(benefit), case=True)].groupby(['benefit', pd.Grouper(key='incur_date', freq='m')]).count().rename(columns={'policy_id': 'no_of_claims'})

    __benefit_l = __benefit_df.index.get_level_values(level='benefit').unique().tolist()

    __tdf = pd.DataFrame()
    for b in __benefit_l:
      __tdf = pd.concat([__tdf, __benefit_df.loc[b].rename(columns={'no_of_claims': b})], axis=1, ignore_index=False)

    # self.benefit_trend = __benefit_fig

    
    df = __tdf
    fig = px.line(df, x=df.index, y=df.columns,
                #hover_data={df.index: "|%B %d, %Y"},
                markers=True,
                width=1200,
                height=600,
                title='Monthly Number of visit of Outpatient Benefit',
                )
    fig.update_layout(
      xaxis_title='Month',
      
      yaxis_title='No. of visits',
      yaxis=dict(
        tickmode='linear',
        tick0=0,
        dtick=25
      ),
      legend=dict(
        orientation='h',
        #yanchor='bottom',
        #y=1.02
      )
    )
    
    # fig.update_xaxes(
    # dtick="M1",
    # tickformat="%b\n%Y")

    return fig

  def benefit_op_yearly_bar(self, benefit=['GP', 'CMT', 'SP', 'PM', 'CT', 'PT', 'DX', 'Psy', 'Vac', 'Check']):
    self.df['year'] = self.df.policy_start_date.dt.year
    __benefit_df = self.df[['benefit', 'policy_id', 'year']].loc[self.df['benefit'].str.contains('|'.join(benefit), case=True)].groupby(['benefit', 'policy_id']).count().rename(columns={'year': 'no_of_claims'})
    __benefit_df  =__benefit_df.astype(int)
    __benefit_df.unstack().stack(dropna=False)
    __benefit_df.fillna(0, inplace=True)
    __benefit_l = __benefit_df.index.get_level_values(level='benefit').unique().tolist() #v
    __policy_id_l = __benefit_df.index.get_level_values(level='policy_id').unique().tolist() #x

    fig = go.Figure()
    for b in __benefit_l:
      fig.add_trace(go.Bar(
        x=__policy_id_l,
        y=__benefit_df['no_of_claims'].loc[b].values.tolist(),
        #y=[1,2,3],
        name=b,
        
      ))
      #print(__benefit_df.loc[b].values.tolist())


    fig.update_layout(
      barmode='group',
      title=dict(
        text='Number of visit of Outpatient Benefit',
        font=dict(size=28),
                 ),
      # title_x = 0.5,
      xaxis_title='Policy',
      yaxis_title='No. of Visit',
      yaxis=dict(
        tickmode='linear',
        tick0=0,
        dtick=100
      ),
      legend=dict(
        orientation='h',
        #yanchor='bottom',
        #y=1.02
      ),
      width=1200,
      height=600,
    )
    
    return fig


  def class_age_scatter(self):
    self.df['year'] = self.df.policy_start_date.dt.year
    __class_age_df = self.df[['class', 'age', 'incurred_amount']].loc[self.df['benefit_type'].str.contains('hosp|top|smm', case=False)]
    __class_age_df['age']  = __class_age_df['age'].astype(int)
    __class_l = __class_age_df['class'].unique()

    fig = go.Figure()
    for cls in __class_l:
      fig.add_trace(go.Scatter(
        x=__class_age_df['age'].loc[__class_age_df['class'] == cls],
        y=__class_age_df['incurred_amount'].loc[__class_age_df['class'] == cls],
        mode='markers',
        #y=[1,2,3],
        name=cls,
        
      ))
      #print(__benefit_df.loc[b].values.tolist())


    fig.update_layout(
      title=dict(
        text='Relationship of age and incurred amount of inpatient',
        font=dict(size=28),
                 ),
      # title_x = 0.5,
      xaxis_title='Age',
      yaxis_title='Incurred Amount',
      yaxis=dict(
        tickmode='linear',
        tick0=0,
        dtick=10_000
      ),
      legend=dict(
        orientation='h',
        #yanchor='bottom',
        #y=1.02
      ),
      width=1200,
      height=600,
    )
    
    return fig
  

  def paid_amount_by_dep_type_policy_year(self):
    
    self.df['year'] = self.df.policy_start_date.dt.year

    __temp_df = self.df[['policy_id','dep_type','paid_amount']].replace({'CH': 'DEP', 'SP': 'DEP'})
    __frequency_counts = self.df.groupby(['policy_id', 'dep_type']).count().rename(columns={'paid_amount': 'no_of_claims'})
    __paid_by_dep_plot_df = __temp_df.groupby(['policy_id', 'dep_type']).sum()
    __paid_by_dep_plot_df['no_of_claims'] = __frequency_counts['no_of_claims']
    __paid_by_dep_plot_df = __paid_by_dep_plot_df.reset_index()
    fig = px.bar(__paid_by_dep_plot_df, x='policy_id', y='paid_amount', color='dep_type', barmode='group', hover_data = {'no_of_claims': True})

    fig.update_layout(
      yaxis_title='Paid Amount',
      title=dict(
        text='Paid Amount by Dependent Type by Policy Year',
        font=dict(size=28),
        ),
      # title_x = 0.5,
      yaxis=dict(
        tickmode='linear',
        tick0=0,
        dtick=100_000
        ),
      xaxis_type='category',
      # xaxis=dict(
      #   tickmode='array',
      #   tickvals = __paid_by_dep_plot_df['class'].drop_duplicates(keep='first').to_list(),
      # ),
      legend=dict(
        orientation='h',
        
        #yanchor='bottom',
        #y=1.02
        ),
      legend_title_text='Dependent Type',
      width=900,
      height=600,
    )
    # fig.update_xaxes(type='category', categoryorder='category ascending')
    return fig
  
  def paid_amount_by_dep_type_class_policy_year(self):
    self.df['year'] = self.df.policy_start_date.dt.year

    __temp_df = self.df[['policy_id','class','dep_type','paid_amount']].replace({'CH': 'DEP', 'SP': 'DEP'})
    __frequency_counts = self.df.groupby(['policy_id', 'dep_type','class']).count().rename(columns={'paid_amount': 'no_of_claims'})
    __paid_by_dep_class_plot_df = __temp_df.groupby(['policy_id', 'dep_type' ,'class']).sum()
    __paid_by_dep_class_plot_df['no_of_claims'] = __frequency_counts['no_of_claims']
    __paid_by_dep_class_plot_df = __paid_by_dep_class_plot_df.reset_index()
    __paid_by_dep_class_plot_df['class'] = __paid_by_dep_class_plot_df['class'].astype(str)
    fig = px.bar(__paid_by_dep_class_plot_df, x='class', y='paid_amount', color='dep_type', barmode='group',facet_col='policy_id',hover_data = {'no_of_claims': True})

    fig.update_layout(
      yaxis_title='Paid Amount',
      title=dict(
        text='Paid Amount by Dependent Type by Class by Policy Year',
        font=dict(size=28),
                 ),
      # title_x = 0.5,
      yaxis=dict(
        tickmode='linear',
        tick0=0,
        dtick=50_000
      ),
      xaxis_type='category',
      # xaxis=dict(
      #   tickmode='array',
      #   tickvals = __paid_by_dep_plot_df['class'].drop_duplicates(keep='first').to_list(),
      # ),
      legend=dict(
        orientation='h',
        
        #yanchor='bottom',
        #y=1.02
      ),
      legend_title_text='Dependent Type',
      width=1500,
      height=600,
    )
    fig.update_xaxes(type='category', categoryorder='category ascending')
    return fig


  def gender_analysis_by_op(self):
    self.df['year'] = self.df.policy_start_date.dt.year
    sex_panel = self.df[['year', 'gender', 'panel', 'incur_date']].loc[(self.df.benefit == 'General Consultation (GP)') | (self.df.benefit == 'Chinese Med (CMT)') | (self.df.benefit == 'Specialist Consultation (SP)')]
    sex_panel = sex_panel.groupby(by=['year', 'gender', 'panel']).count().rename(columns={'incur_date': 'no_of_claims'})

    __year_l = self.df.year.unique().tolist()
    __gender_l = self.df.gender.unique().tolist()
    __benefit_l = ['General Consultation (GP)', 'Chinese Med (CMT)', 'Specialist Consultation (SP)']


    pie_fig, pie_ax = plt.subplots(nrows=len(__year_l), ncols=len(__gender_l), figsize=(6*len(__year_l), 6*len(__gender_l)))
    ax[0].pie(labels=_plt1.index, x=_plt1.no_of_claims, autopct=self.make_autopct(_plt1.no_of_claims), textprops={'fontsize': 12})
    ax[0].set_title(f'Male, Total no of visit {_plt1.no_of_claims.sum()}')
    ax[1].pie(labels=_plt2.index, x=_plt2.no_of_claims, autopct=self.make_autopct(_plt2.no_of_claims), textprops={'fontsize': 12})
    ax[1].set_title(f'Female, Total no of visit {_plt2.no_of_claims.sum()}')
    fig.suptitle('Panel Usage by Gender of General Consultation, Chinese Medicine, and Specialist')

  def export_database(self):
    return self.df.to_csv(index=False).encode('utf-8')


