import warnings
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

class MCRConvert():
    def __init__(self, input_file):
        try:
            self.input_file = input_file
            
            # template_file = "GMI_template.xlsx"
            self.template_file = "GMI_template.xlsx"
            
            self.template_wb = load_workbook(self.template_file, keep_links=False, data_only=False)
            self.previous_year_loss_ratio_df = None
            self.current_year_loss_ratio_df = None
            self.loss_ratio_history_by_policy = {}
            self.loss_ratio_records_by_policy = {}
            self.loss_ratio_history_by_policy = {}

            self.input_dtype = {
                "policy_number": str,
                "year": str,
                "benefit_type": str,
                "benefit": str,
                "panel": str,
                "class": str,
                "incurred_amount": float,
                "paid_amount": float,
                "usage_ratio": float,
                # "no_of_claims": int,
                # "no_of_claimants": int,
                "incurred_per_case": float,
                "paid_per_case": float,
                "incurred_per_claimant": float,
                "paid_per_claimant": float,
                "claim_frequency": float,
            }

            self.mcr_p20_policy = pd.read_excel(self.input_file, sheet_name='P.20_Policy', dtype=self.input_dtype)
            self.mcr_p20_benefit = pd.read_excel(self.input_file, sheet_name='P.20_BenefitType', dtype=self.input_dtype, na_values=0)
            self.mcr_p20_network = pd.read_excel(self.input_file, sheet_name='P.20_Network', dtype=self.input_dtype, na_values=0)
            self.mcr_p20_network_benefit = pd.read_excel(self.input_file, sheet_name='P.20_Network_BenefitType', dtype=self.input_dtype, na_values=0)
            self.mcr_p21_class = pd.read_excel(self.input_file, sheet_name='P.21_Class', dtype=self.input_dtype)
            self.mcr_p22_class_benefit = pd.read_excel(self.input_file, sheet_name='P.22_Class_BenefitType', dtype=self.input_dtype, na_values=0)
            self.mcr_p25_class_panel_benefit = pd.read_excel(self.input_file, sheet_name='P.25_Class_Panel_BenefitType', dtype=self.input_dtype, na_values=0)
            self.mcr_p26_op_panel_benefit = pd.read_excel(self.input_file, sheet_name='P.26_OP_Panel_Benefit', dtype=self.input_dtype, na_values=0)

            self.policy_info = self.mcr_p20_policy[["policy_number", "year"]]

        except FileNotFoundError:
            print(f"Error: The file '{self.input_file}' was not found. Please check the name and try again.")
        except Exception as e:
            print(f"An error occurred: {e}")

    def set_policy_input(self,
                         previous_policy_num, previous_year_start_date, previous_year_end_date, previous_year, 
                         current_policy_num, current_year_start_date, current_year_end_date, current_year
                         ):
        self.previous_year, self.current_year = previous_year, current_year
        self.previous_year_start_date, self.previous_year_end_date = previous_year_start_date, previous_year_end_date
        self.current_year_start_date, self.current_year_end_date = current_year_start_date, current_year_end_date
        self.previous_policy_num, self.current_policy_num = previous_policy_num, current_policy_num
        self.plan_info = self.mcr_p21_class.copy(deep=True)
        self.plan_info.dropna(inplace=True)
        self.plan_info = self.plan_info.loc[(self.plan_info["policy_number"] == self.current_policy_num) & 
                                            (self.plan_info["year"] == self.current_year), "class"].tolist()
        
    
    def claim_info(self):
        claim_info_worksheet = self.template_wb["ClaimInfo"]

        if self.previous_policy_num is not None:
            claim_info_worksheet.cell(row=4,column=3).value = self.previous_policy_num
            claim_info_worksheet.cell(row=6,column=3).value = self.previous_year_start_date
            claim_info_worksheet.cell(row=7,column=3).value = self.previous_year_end_date

        if self.current_policy_num is not None:
            claim_info_worksheet.cell(row=9,column=3).value = self.current_policy_num
            claim_info_worksheet.cell(row=11,column=3).value = self.current_year_start_date
            claim_info_worksheet.cell(row=12,column=3).value = self.current_year_end_date

        plan_start_row = 2
        for i in range(plan_start_row, len(self.plan_info) + 2):
            claim_info_worksheet.cell(row=i, column=8).value = self.plan_info[i-2]
        
                

    def P20_overall(self):
        input_p20 = self.mcr_p20_policy.loc[((self.mcr_p20_policy["policy_number"] == self.current_policy_num) & (self.mcr_p20_policy["year"] == self.current_year)) | 
                                            ((self.mcr_p20_policy["policy_number"] == self.previous_policy_num) & (self.mcr_p20_policy["year"] == self.previous_year))].dropna()
        template_p20 = self.template_wb["P20_Usage Overview"]

        previous_start_row = current_start_row = 10
        for index, row in input_p20.iterrows():
            if row['year'] == self.previous_year:
                previous_start_row += 1 
                for col, val in zip([2,3,4], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                    template_p20.cell(row=previous_start_row, column=1).value = self.previous_policy_num
                    template_p20.cell(row=previous_start_row, column=col).value = val
            
            elif row['year'] == self.current_year:
                current_start_row += 1
                for col, val in zip([6,7,8], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                    template_p20.cell(row=previous_start_row, column=5).value = self.current_policy_num
                    template_p20.cell(row=current_start_row, column=col).value = val


    def P20_benefittype(self):
        input_p20_btype = self.mcr_p20_benefit.loc[((self.mcr_p20_benefit["policy_number"] == self.current_policy_num) & (self.mcr_p20_benefit["year"] == self.current_year)) | 
                                                   ((self.mcr_p20_benefit["policy_number"] == self.previous_policy_num) & (self.mcr_p20_benefit["year"] == self.previous_year))].dropna()
        template_p20 = self.template_wb["P20_Usage Overview"]

        # by benefit table variables 
        previous_start_row = current_start_row = 15

        # number of claims table variables 
        previous_num_of_claim = current_num_of_claim = 33

        for index, row in input_p20_btype.iterrows():
            if row["benefit_type"] != 'Total':
                if row["year"] == self.previous_year:
                    previous_start_row +=1
                    previous_num_of_claim +=1
                    template_p20.cell(row=previous_num_of_claim, column=2).value = row['no_of_cases']
                    for col, val in zip([2,3,4], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):                      
                        template_p20.cell(row=previous_start_row, column=col).value = val

                elif row["year"] == self.current_year:
                    current_start_row += 1
                    current_num_of_claim += 1
                    template_p20.cell(row=current_num_of_claim, column=6).value = row['no_of_cases']
                    for col, val in zip([6,7,8], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                        template_p20.cell(row=current_start_row, column=col).value = val


    def P20_network(self):
        input_p20_btype = self.mcr_p20_network.loc[((self.mcr_p20_network["policy_number"] == self.current_policy_num) & (self.mcr_p20_network["year"] == self.current_year)) | 
                                                   ((self.mcr_p20_network["policy_number"] == self.previous_policy_num) & (self.mcr_p20_network["year"] == self.previous_year))]
        template_p20 = self.template_wb["P20_Usage Overview"]

        previous_start_row = current_start_row = 27
        for index, row in input_p20_btype.iterrows():
            # Handling the by network table in P20_overview worksheet
            row_offset = 1 if row['panel'] == "Panel" else 0
            if row['year'] == self.previous_year:
                for col, val in zip([2, 3, 4], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                    template_p20.cell(row=previous_start_row + row_offset, column=col).value = val

            elif row['year'] == self.current_year:
                for col, val in zip([6, 7, 8], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):                      
                    template_p20.cell(row=current_start_row + row_offset, column=col).value = val

        clinic_df = getattr(self, 'mcr_p20_network_benefit', None)
        if clinic_df is None or clinic_df.empty:
            return

        benefit_col = clinic_df.get('benefit_type')
        if benefit_col is None:
            return

        clinic_mask = benefit_col.astype(str).str.contains('clinic', case=False, na=False)
        clinic_df = clinic_df.loc[clinic_mask].copy()
        if clinic_df.empty:
            return

        clinic_df['panel'] = clinic_df['panel'].astype(str).str.strip().str.title()
        clinic_df['policy_number'] = clinic_df['policy_number'].astype(str)
        clinic_df['year'] = clinic_df['year'].astype(str)

        def _clinic_summary(policy_number, year_val):
            if policy_number is None or year_val is None:
                return None
            mask = (
                (clinic_df['policy_number'] == str(policy_number)) &
                (clinic_df['year'] == str(year_val))
            )
            subset = clinic_df.loc[mask]
            if subset.empty:
                return None
            return subset.groupby('panel').agg({
                'no_of_claim_id': 'sum',
                'paid_amount': 'sum'
            })

        prev_vals = _clinic_summary(self.previous_policy_num, self.previous_year)
        curr_vals = _clinic_summary(self.current_policy_num, self.current_year)

        def _value_or_none(group, panel_name, column):
            if group is None or panel_name not in group.index:
                return None
            value = group.loc[panel_name, column]
            try:
                return float(value)
            except (TypeError, ValueError):
                return None

        template_p20.cell(row=46, column=2).value = _value_or_none(prev_vals, 'Non-Panel', 'no_of_claim_id')
        template_p20.cell(row=47, column=2).value = _value_or_none(prev_vals, 'Panel', 'no_of_claim_id')
        template_p20.cell(row=46, column=6).value = _value_or_none(curr_vals, 'Non-Panel', 'no_of_claim_id')
        template_p20.cell(row=47, column=6).value = _value_or_none(curr_vals, 'Panel', 'no_of_claim_id')
        template_p20.cell(row=54, column=2).value = _value_or_none(prev_vals, 'Non-Panel', 'paid_amount')
        template_p20.cell(row=55, column=2).value = _value_or_none(prev_vals, 'Panel', 'paid_amount')
        template_p20.cell(row=54, column=6).value = _value_or_none(curr_vals, 'Non-Panel', 'paid_amount')
        template_p20.cell(row=55, column=6).value = _value_or_none(curr_vals, 'Panel', 'paid_amount')

    
    def P21_by_class(self):
        input_p21 = self.mcr_p21_class.loc[((self.mcr_p21_class["policy_number"] == self.current_policy_num) & (self.mcr_p21_class["year"] == self.current_year)) | 
                                           ((self.mcr_p21_class["policy_number"] == self.previous_policy_num) & (self.mcr_p21_class["year"] == self.previous_year))].dropna()
        template_p21 = self.template_wb["P21_Usage by Class"]

        current_start_row = previous_start_row = 7
        for index, row in input_p21.iterrows():
            if row['year'] == self.current_year:
                current_start_row  += 1 
                # template_p21.cell(row=current_start_row, column=1).value = row['class']
                for col, val in zip([6,7,8], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):                      
                    template_p21.cell(row=current_start_row, column=col).value = val

            elif row['year'] == self.previous_year:
                previous_start_row += 1
                for col, val in zip([3,4,5], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):                      
                    template_p21.cell(row=previous_start_row, column=col).value = val


    def P22_by_class(self):
        input_p22 = self.mcr_p22_class_benefit.loc[((self.mcr_p22_class_benefit["policy_number"] == self.current_policy_num) & (self.mcr_p22_class_benefit["year"] == self.current_year)) | 
                                                   ((self.mcr_p22_class_benefit["policy_number"] == self.previous_policy_num) & (self.mcr_p22_class_benefit["year"] == self.previous_year))]
        template_p22 = self.template_wb["P22_Usage by Class by Ben"]

        input_p22.fillna(0, inplace=True)

        plan_num = self.plan_info
        P22_format_row_num = [13,20,27,34,45,52,59,66,77,84,91,98]

        plan_num, P22_format_row_num = pd.Series(plan_num).astype(str), pd.Series(P22_format_row_num)
        current_plan_row_df = pd.concat([plan_num, P22_format_row_num],axis=1, ignore_index=True).dropna()
        current_plan_row_df.columns = ['plan', "start_row"]
        
        previous_row_dict = current_plan_row_df.set_index('plan')['start_row'].to_dict()
        current_row_dict = current_plan_row_df.set_index('plan')['start_row'].to_dict()

        for index, row in input_p22.iterrows():
            class_id = row['class']
            previous_start_row = previous_row_dict[class_id]
            current_start_row = current_row_dict[class_id]

            if row['year'] == self.previous_year:
                row_target = previous_start_row + 1
                cols = [3, 4, 5]
                previous_row_dict[class_id] = row_target
                for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                    template_p22.cell(row=row_target, column=col).value = val

            elif row['year'] == self.current_year:
                row_target = current_start_row + 1
                cols = [6, 7, 8]
                current_row_dict[class_id] = row_target
                for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                    template_p22.cell(row=row_target, column=col).value = val
                        
        
    def P25_by_plan(self):
        input_p25 = self.mcr_p25_class_panel_benefit.loc[((self.mcr_p25_class_panel_benefit["policy_number"] == self.current_policy_num) & (self.mcr_p25_class_panel_benefit["year"] == self.current_year)) | 
                                                         ((self.mcr_p25_class_panel_benefit["policy_number"] == self.previous_policy_num) & (self.mcr_p25_class_panel_benefit["year"] == self.previous_year))]
        template_p25 = self.template_wb["P25_UsageByplanBynetworkByben"]
        input_p25.fillna(0, inplace=True)
        plan_num = self.plan_info

        # creating dataframe for handling non-panel data
        P25_nonpanel_row_num = [13,27,41,60,74,88,107,121,135,154,168,182]
        plan_num, P25_nonpanel_row_num = pd.Series(plan_num).astype(str), pd.Series(P25_nonpanel_row_num)
        nonpanel_plan_row_df = pd.concat([plan_num, P25_nonpanel_row_num],axis=1, ignore_index=True).dropna()
        nonpanel_plan_row_df.columns = ['plan', "start_row"]

        previous_nonpanel_row_dict = nonpanel_plan_row_df.set_index('plan')['start_row'].to_dict()
        current_nonpanel_row_dict = nonpanel_plan_row_df.set_index('plan')['start_row'].to_dict()

        # creating dataframe for handling panel data
        P25_panel_row_num  = [20,34,48,67,81,95,114,128,142,161,175,189]
        plan_num, P25_panel_row_num = pd.Series(plan_num).astype(str), pd.Series(P25_panel_row_num)
        panel_plan_row_df = pd.concat([plan_num, P25_panel_row_num],axis=1, ignore_index=True).dropna()
        panel_plan_row_df.columns = ['plan', "start_row"]
        
        previous_panel_row_dict = panel_plan_row_df.set_index('plan')['start_row'].to_dict()
        current_panel_row_dict = panel_plan_row_df.set_index('plan')['start_row'].to_dict()
    
        for index, row in input_p25.iterrows():
            class_id = row['class']
            previous_start_row = previous_nonpanel_row_dict[class_id]
            current_start_row = current_nonpanel_row_dict[class_id]

            previous_panel_start_row = previous_panel_row_dict[class_id]
            current_panel_start_row = current_panel_row_dict[class_id]

            if row['panel'] == 'Non-Panel':
                if row['year'] == self.previous_year:
                    row_target = previous_start_row + 1
                    cols = [4, 5, 6]
                    previous_nonpanel_row_dict[class_id] = row_target
                    for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                        template_p25.cell(row=row_target, column=col).value = val
                elif row['year'] == self.current_year:
                    row_target = current_start_row + 1
                    cols = [7,8,9]
                    current_nonpanel_row_dict[class_id] = row_target
                    for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                        template_p25.cell(row=row_target, column=col).value = val
            
            elif row['panel'] == 'Panel':
                if row['year'] == self.previous_year:
                    row_target = previous_panel_start_row + 1
                    cols = [4, 5, 6]
                    previous_panel_row_dict[class_id] = row_target
                    for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                        template_p25.cell(row=row_target, column=col).value = val
                elif row['year'] == self.current_year:
                    row_target = current_panel_start_row + 1
                    cols = [7,8,9]
                    current_panel_row_dict[class_id] = row_target
                    for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                        template_p25.cell(row=row_target, column=col).value = val


    def P26_by_class(self):
        input_p26 = self.mcr_p26_op_panel_benefit.loc[((self.mcr_p26_op_panel_benefit["policy_number"] == self.current_policy_num) & (self.mcr_p26_op_panel_benefit["year"] == self.current_year)) | 
                                                      ((self.mcr_p26_op_panel_benefit["policy_number"] == self.previous_policy_num) & (self.mcr_p26_op_panel_benefit["year"] == self.previous_year))].dropna()
        template_p26 = self.template_wb["P26_Usage_Clinical by Network"]

        panel_previous_start_row = panel_current_start_row = 7
        nonpanel_previous_start_row = nonpanel_current_start_row = 28

        rows = {
            ('Panel', self.current_year): {'row': panel_previous_start_row, 'cols': [2, 9, 10, 11, 12, 13, 14], 'values': ['benefit', 'incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']},
            ('Panel', self.previous_year): {'row': panel_current_start_row, 'cols': [3, 4, 5, 6, 7, 8], 'values': ['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']},
            ('Non-Panel', self.current_year): {'row': nonpanel_previous_start_row, 'cols': [2, 9, 10, 11, 12, 13, 14], 'values': ['benefit', 'incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']},
            ('Non-Panel', self.previous_year): {'row': nonpanel_current_start_row, 'cols': [3, 4, 5, 6, 7, 8], 'values': ['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']}
        }

        current_rows = {
            ('Panel', self.current_year): panel_previous_start_row,
            ('Panel', self.previous_year): panel_current_start_row,
            ('Non-Panel', self.current_year): nonpanel_previous_start_row,
            ('Non-Panel', self.previous_year): nonpanel_current_start_row
        }

        for index, row in input_p26.iterrows():
            key = (row['panel'], row['year'])
            if key in rows:
                current_rows[key] += 1
                row_idx = current_rows[key]
                info = rows[key]
                values = [row[val] for val in info['values']]
                for col, val in zip(info['cols'], values):
                    template_p26.cell(row=row_idx, column=col).value = val

    def save(self):
        from io import BytesIO
        output_file = BytesIO()
        try:
            # self.template_wb.save(output_file)
            # print(f"File saved successfully as '{output_file}'")
            self.template_wb.save(output_file)
            return output_file.getvalue()
        except Exception as e:
            print(f"An error occurred while saving the file into the BytesIO: {e}")

    def load_loss_ratio_dataframe(self, loss_ratio_df, policy_number, year, period, loss_ratio_grouping_optical=True):
        subset = self._filter_loss_ratio_records(loss_ratio_df, policy_number, year)
        self._assign_loss_ratio_subset(subset, period, loss_ratio_grouping_optical)
        self._update_loss_ratio_history(loss_ratio_df, policy_number)
        self._append_loss_ratio_records(loss_ratio_df, policy_number)

    def _normalize_policy_number(self, value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        text = str(value).strip()
        if not text:
            return None
        normalized = text.lstrip('0')
        return normalized or '0'

    def _filter_loss_ratio_records(self, df, policy_number, year):
        if df is None or policy_number is None or year is None:
            return None

        try:
            year_int = int(year)
        except (TypeError, ValueError):
            return None

        subset = df.copy(deep=True)
        subset['policy_start_date'] = pd.to_datetime(subset.get('policy_start_date'), errors='coerce')
        subset['policy_number'] = subset.get('policy_number').astype(str)
        subset['year_key'] = subset['policy_start_date'].dt.year

        target_policy = self._normalize_policy_number(policy_number)
        subset['policy_number_norm'] = subset['policy_number'].apply(self._normalize_policy_number)

        mask = (
            (subset['policy_number'] == str(policy_number)) |
            (subset['policy_number_norm'] == target_policy)
        ) & (subset['year_key'] == year_int)

        subset = subset.loc[mask]
        if subset.empty:
            return None

        numeric_cols = ['actual_premium', 'actual_paid_w_ibnr', 'duration']
        for col in numeric_cols:
            subset[col] = pd.to_numeric(subset.get(col), errors='coerce')

        subset['benefit_type'] = subset.get('benefit_type').astype(str).str.strip()
        subset.replace([pd.NA, pd.NaT], None, inplace=True)
        subset = subset[['benefit_type', 'actual_premium', 'actual_paid_w_ibnr', 'duration']]
        subset.dropna(subset=['benefit_type', 'actual_premium', 'actual_paid_w_ibnr', 'duration'], inplace=True)
        subset = subset[subset['benefit_type'] != ""]
        return subset if not subset.empty else None

    def _assign_loss_ratio_subset(self, subset_df, period, loss_ratio_grouping_optical):
        target_attr = 'previous_year_loss_ratio_df' if period == 'previous' else 'current_year_loss_ratio_df'
        if subset_df is None or subset_df.empty:
            setattr(self, target_attr, None)
            return

        df = subset_df.copy(deep=True)
        df = df.groupby('benefit_type', as_index=False).agg({
            'actual_premium': 'sum',
            'actual_paid_w_ibnr': 'sum',
            'duration': 'first'
        })
        df['duration'] = df['duration'].replace(0, pd.NA)
        df.dropna(subset=['duration'], inplace=True)

        if loss_ratio_grouping_optical and not df.empty:
            optical_mask = df['benefit_type'].str.lower() == 'optical'
            if optical_mask.any():
                clinical_mask = df['benefit_type'].str.lower() == 'clinical'
                optical_premium = df.loc[optical_mask, 'actual_premium'].sum()
                optical_paid = df.loc[optical_mask, 'actual_paid_w_ibnr'].sum()
                optical_duration = df.loc[optical_mask, 'duration'].iloc[0]

                if clinical_mask.any():
                    idx = df.index[clinical_mask][0]
                    df.at[idx, 'actual_premium'] += optical_premium
                    df.at[idx, 'actual_paid_w_ibnr'] += optical_paid
                    if pd.isna(df.at[idx, 'duration']):
                        df.at[idx, 'duration'] = optical_duration
                    df = df.loc[~optical_mask]
                else:
                    df.loc[optical_mask, 'benefit_type'] = 'Clinical'
                    df = df.groupby('benefit_type', as_index=False).agg({
                        'actual_premium': 'sum',
                        'actual_paid_w_ibnr': 'sum',
                        'duration': 'first'
                    })

        setattr(self, target_attr, df if not df.empty else None)

    def _update_loss_ratio_history(self, loss_ratio_df, policy_number):
        if loss_ratio_df is None or policy_number is None:
            return

        history_df = self._prepare_loss_ratio_history(loss_ratio_df, policy_number)
        if history_df is None or history_df.empty:
            return

        key_raw = str(policy_number)
        key_norm = self._normalize_policy_number(policy_number) or key_raw
        for key in {key_raw, key_norm}:
            if key:
                self.loss_ratio_history_by_policy[key] = history_df

    def _prepare_loss_ratio_history(self, df, policy_number):
        if df is None or policy_number is None:
            return None

        work = df.copy(deep=True)
        work['policy_number'] = work.get('policy_number').astype(str)
        target_norm = self._normalize_policy_number(policy_number)
        work['policy_number_norm'] = work['policy_number'].apply(self._normalize_policy_number)
        mask = (
            (work['policy_number'] == str(policy_number)) |
            (work['policy_number_norm'] == target_norm)
        )
        work = work.loc[mask]
        if work.empty:
            return None

        def _to_year(series):
            if series is None:
                return pd.Series([pd.NA] * len(work))
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                parsed = pd.to_datetime(series, errors='coerce', dayfirst=True)
                missing = parsed.isna()
                if missing.any():
                    parsed_alt = pd.to_datetime(series, errors='coerce', dayfirst=False)
                    parsed = parsed.fillna(parsed_alt)
            return parsed.dt.year

        year_series = _to_year(work.get('policy_start_date'))
        year_series = year_series.fillna(_to_year(work.get('policy_end_date')))
        year_series = year_series.fillna(_to_year(work.get('data_as_of')))
        if 'policy_id' in work.columns:
            policy_id_year = pd.to_numeric(
                work['policy_id'].astype(str).str.extract(r"_(\d{4})")[0],
                errors='coerce'
            )
            year_series = year_series.fillna(policy_id_year)

        work['year_key'] = year_series
        work.dropna(subset=['year_key'], inplace=True)
        if work.empty:
            return None

        work['year_key'] = work['year_key'].astype(int)
        work['benefit_type'] = work.get('benefit_type', '').astype(str).str.strip()
        work['benefit_type_norm'] = work['benefit_type'].str.lower()

        numeric_cols = ['actual_premium', 'actual_paid_w_ibnr', 'duration']
        for col in numeric_cols:
            work[col] = pd.to_numeric(work.get(col), errors='coerce')

        headcount_col = None
        for candidate in [
            'headcount', 'head_count', 'member_count', 'no_of_members',
            'no_of_employees', 'employee_count', 'no_of_lives', 'lives'
        ]:
            if candidate in work.columns:
                headcount_col = candidate
                work[headcount_col] = pd.to_numeric(work[headcount_col], errors='coerce')
                break

        total_mask = work['benefit_type_norm'].isin({'total', 'grand total'})
        work = work.loc[total_mask].copy() if total_mask.any() else work.copy()
        if work.empty:
            return None

        valid_duration = work['duration'].where(work['duration'] > 0)
        annual_factor = 12.0 / valid_duration
        work['premium_annualized'] = work['actual_premium'] * annual_factor
        work['claim_annualized'] = work['actual_paid_w_ibnr'] * annual_factor

        fallback_mask = annual_factor.isna()
        work.loc[fallback_mask, 'premium_annualized'] = work.loc[fallback_mask, 'actual_premium']
        work.loc[fallback_mask, 'claim_annualized'] = work.loc[fallback_mask, 'actual_paid_w_ibnr']

        agg_dict = {
            'premium_annualized': 'sum',
            'claim_annualized': 'sum'
        }
        if headcount_col:
            agg_dict[headcount_col] = 'sum'

        yearly = work.groupby('year_key', dropna=False).agg(agg_dict).reset_index()
        yearly.rename(columns={'year_key': 'year'}, inplace=True)

        if headcount_col:
            yearly.rename(columns={headcount_col: 'headcount'}, inplace=True)
        else:
            yearly['headcount'] = None

        def _safe_div(numerator, denominator):
            try:
                if numerator is None or denominator is None:
                    return None
                if pd.isna(numerator) or pd.isna(denominator):
                    return None
                denominator = float(denominator)
                if denominator == 0:
                    return None
                return float(numerator) / denominator
            except (TypeError, ValueError, ZeroDivisionError):
                return None

        yearly['loss_ratio'] = [
            _safe_div(paid, prem)
            for paid, prem in zip(yearly['claim_annualized'], yearly['premium_annualized'])
        ]
        yearly['premium_per_head'] = [
            _safe_div(prem, head)
            for prem, head in zip(yearly['premium_annualized'], yearly['headcount'])
        ]
        yearly['claim_per_head'] = [
            _safe_div(claim, head)
            for claim, head in zip(yearly['claim_annualized'], yearly['headcount'])
        ]

        yearly = yearly.dropna(subset=['premium_annualized', 'claim_annualized'], how='all')
        if yearly.empty:
            return None

        yearly.sort_values('year', ascending=False, inplace=True)
        yearly.reset_index(drop=True, inplace=True)
        return yearly

    def _get_loss_ratio_history_dataframe(self):
        for policy in [self.current_policy_num, self.previous_policy_num]:
            if policy is None:
                continue
            key_raw = str(policy)
            key_norm = self._normalize_policy_number(policy)
            for key in filter(None, {key_raw, key_norm}):
                if key in self.loss_ratio_history_by_policy:
                    df = self.loss_ratio_history_by_policy[key]
                    if df is not None and not df.empty:
                        return df
        for df in self.loss_ratio_history_by_policy.values():
            if df is not None and not df.empty:
                return df
        return None

    def _populate_loss_ratio_history_table(self, worksheet):
        history_df = self._get_loss_ratio_history_dataframe()
        rows = []
        if history_df is not None and not history_df.empty:
            ordered = history_df.copy()
            if 'year' in ordered.columns:
                ordered = ordered.sort_values(
                    by='year',
                    ascending=True,
                    key=lambda col: pd.to_numeric(col, errors='coerce'),
                    na_position='last'
                )
            ordered = ordered.tail(3)
            rows = ordered.to_dict('records')
        start_row = 5
        col_map = [
            ('year', 10),
            ('premium_annualized', 11),
            ('claim_annualized', 12),
            ('loss_ratio', 13),
            ('headcount', 14),
            ('premium_per_head', 15),
            ('claim_per_head', 16)
        ]

        for offset in range(3):
            row_idx = start_row + offset
            row_data = rows[offset] if offset < len(rows) else {}
            for key, col_idx in col_map:
                value = row_data.get(key)
                if value is not None and pd.isna(value):
                    value = None
                worksheet.cell(row=row_idx, column=col_idx).value = value

    def _append_loss_ratio_records(self, loss_ratio_df, policy_number):
        if loss_ratio_df is None or policy_number is None:
            return

        work = loss_ratio_df.copy(deep=True)
        work['policy_number'] = work.get('policy_number').astype(str)
        target_norm = self._normalize_policy_number(policy_number)
        work['policy_number_norm'] = work['policy_number'].apply(self._normalize_policy_number)
        mask = (
            (work['policy_number'] == str(policy_number)) |
            (work['policy_number_norm'] == target_norm)
        )
        subset = work.loc[mask].drop(columns=['policy_number_norm'], errors='ignore')
        if subset.empty:
            return

        key_raw = str(policy_number)
        key_norm = target_norm or key_raw
        for key in filter(None, {key_raw, key_norm}):
            self.loss_ratio_records_by_policy[key] = subset.copy(deep=True)

    def _get_loss_ratio_records_dataframe(self):
        frames = []
        seen = set()
        for policy in [self.current_policy_num, self.previous_policy_num]:
            if policy is None:
                continue
            key_raw = str(policy)
            key_norm = self._normalize_policy_number(policy)
            for key in filter(None, {key_raw, key_norm}):
                if key in seen:
                    continue
                df = self.loss_ratio_records_by_policy.get(key)
                if df is not None and not df.empty:
                    frames.append(df.copy(deep=True))
                    seen.add(key)
        if not frames:
            return None
        combined = pd.concat(frames, ignore_index=True)
        combined.drop_duplicates(inplace=True)
        return combined if not combined.empty else None

    def _normalize_benefit_label(self, value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        text = str(value).strip()
        if not text:
            return None
        cleaned = ''.join(ch for ch in text.lower() if ch.isalnum())
        alias_map = {
            'hospital': 'Hospital',
            'clinical': 'Clinical',
            'dental': 'Dental',
            'optical': 'Optical',
            'maternity': 'Maternity',
            'topupsmm': 'Top-Up/ SMM',
            'topupsmmrl': 'Top-Up/ SMM',
            'topupsmml': 'Top-Up/ SMM',
            'topupsmmw': 'Top-Up/ SMM',
            'topupsmm': 'Top-Up/ SMM',
            'topup': 'Top-Up/ SMM',
            'total': 'Total',
            'grandtotal': 'Total',
            'grandtotals': 'Total'
        }
        return alias_map.get(cleaned, text)

    def _parse_dates_with_fallback(self, series):
        if series is None:
            return pd.Series([pd.NaT] * 0)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            parsed = pd.to_datetime(series, errors='coerce', dayfirst=True)
            missing = parsed.isna()
            if missing.any():
                parsed_alt = pd.to_datetime(series, errors='coerce', dayfirst=False)
                parsed = parsed.fillna(parsed_alt)
        return parsed

    def _populate_loss_ratio_detail_table(self, worksheet):
        detail_df = self._get_loss_ratio_records_dataframe()
        start_row = 14
        clear_rows = 200
        year_column_index = 10 + 12  # Column reserved for Year formula
        for row in range(start_row, start_row + clear_rows):
            for col in range(10, 26):
                if col == year_column_index:
                    continue
                cell = worksheet.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        if detail_df is None or detail_df.empty:
            return

        df = detail_df.copy(deep=True)
        required_cols = [
            'policy_id', 'policy_number', 'client_name', 'policy_start_date', 'policy_end_date',
            'duration', 'ibnr', 'data_as_of', 'benefit_type', 'actual_premium', 'actual_paid_w_ibnr', 'loss_ratio'
        ]
        for col in required_cols:
            if col not in df.columns:
                df[col] = None

        df['benefit_display'] = df['benefit_type'].apply(self._normalize_benefit_label)
        df.dropna(subset=['benefit_display'], inplace=True)
        if df.empty:
            return

        benefit_priority = ['Hospital', 'Clinical', 'Dental', 'Optical', 'Maternity', 'Top-Up/ SMM', 'Top-Up/SMM', 'Total']
        order_map = {name: idx for idx, name in enumerate(benefit_priority)}
        df['benefit_order'] = df['benefit_display'].map(order_map).fillna(len(order_map))

        df['duration_numeric'] = pd.to_numeric(df['duration'], errors='coerce')
        df['actual_premium_numeric'] = pd.to_numeric(df['actual_premium'], errors='coerce')
        df['actual_paid_numeric'] = pd.to_numeric(df['actual_paid_w_ibnr'], errors='coerce')

        start_series = df.get('policy_start_date')
        if start_series is None:
            start_series = pd.Series([pd.NaT] * len(df))
        start_dates = self._parse_dates_with_fallback(start_series)
        df['policy_start_sort'] = start_dates

        data_as_of_series = df.get('data_as_of')
        if data_as_of_series is None:
            data_as_of_series = pd.Series([pd.NaT] * len(df))
        df['data_as_of_sort'] = self._parse_dates_with_fallback(data_as_of_series)

        duration_factor = 12.0 / df['duration_numeric']
        duration_factor.replace([pd.NA, float('inf'), -float('inf')], pd.NA, inplace=True)
        df['ann_premium'] = df['actual_premium_numeric'] * duration_factor
        df['ann_paid'] = df['actual_paid_numeric'] * duration_factor

        fallback_mask = df['ann_premium'].isna()
        df.loc[fallback_mask, 'ann_premium'] = df.loc[fallback_mask, 'actual_premium_numeric']
        fallback_mask = df['ann_paid'].isna()
        df.loc[fallback_mask, 'ann_paid'] = df.loc[fallback_mask, 'actual_paid_numeric']

        loss_ratio_series = df['loss_ratio'].astype(str).str.strip()
        cleaned = loss_ratio_series.str.replace('%', '', regex=False).str.replace(',', '', regex=False)
        numeric_loss = pd.to_numeric(cleaned, errors='coerce')
        percent_mask = loss_ratio_series.str.contains('%', regex=False)
        mask_valid = percent_mask & numeric_loss.notna()
        numeric_loss.loc[mask_valid] = numeric_loss.loc[mask_valid] / 100.0
        fallback_lr = df.apply(
            lambda row: (row['actual_paid_numeric'] / row['actual_premium_numeric'])
            if row['actual_premium_numeric'] and not pd.isna(row['actual_premium_numeric']) else pd.NA,
            axis=1
        )
        df['loss_ratio_numeric'] = numeric_loss.fillna(fallback_lr)

        ann_lr = df.apply(
            lambda row: (row['ann_paid'] / row['ann_premium'])
            if row['ann_premium'] and not pd.isna(row['ann_premium']) else pd.NA,
            axis=1
        )
        df['ann_loss_ratio'] = ann_lr.fillna(df['loss_ratio_numeric'])

        df.sort_values(
            by=['benefit_order', 'policy_start_sort', 'data_as_of_sort'],
            ascending=[True, True, True],
            inplace=True
        )

        column_order = [
            ('policy_id', 'policy_id'),
            ('policy_number', 'policy_number'),
            ('client_name', 'client_name'),
            ('policy_start_date', 'policy_start_date'),
            ('policy_end_date', 'policy_end_date'),
            ('duration', 'duration'),
            ('ibnr', 'ibnr'),
            ('data_as_of', 'data_as_of'),
            ('benefit_type', 'benefit_display'),
            ('actual_premium', 'actual_premium_numeric'),
            ('actual_paid_w_ibnr', 'actual_paid_numeric'),
            ('loss_ratio', 'loss_ratio_numeric'),
            ('Year', None),
            ('Annl. Premium', 'ann_premium'),
            ('Annl. Paid /w IBNR', 'ann_paid'),
            ('Loss Ratio', 'ann_loss_ratio')
        ]

        row_idx = start_row
        end_col = 25
        def _next_available_row(current_row):
            row = current_row
            limit = start_row + clear_rows
            while row < limit:
                has_merged = False
                for col in range(10, end_col + 1):
                    if isinstance(worksheet.cell(row=row, column=col), MergedCell):
                        has_merged = True
                        break
                if not has_merged:
                    return row
                row += 1
            return row

        for _, row in df.iterrows():
            row_idx = _next_available_row(row_idx)
            for offset, (label, key) in enumerate(column_order):
                if label == 'Year':
                    continue
                value = row.get(key)
                if isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()
                if pd.isna(value):
                    value = None
                worksheet.cell(row=row_idx, column=10 + offset).value = value
            row_idx += 1


    def loss_ratio_text_convert(self, previous_yr_loss_ratio_text=None, current_yr_loss_ratio_text = None, loss_ratio_grouping_optical=True):
        print("previous_yr_loss_ratio_text:", previous_yr_loss_ratio_text)
        print("current_yr_loss_ratio_text:", current_yr_loss_ratio_text)
        if previous_yr_loss_ratio_text:
            previous_yr_loss_ratio_text = previous_yr_loss_ratio_text.splitlines()
            data_l = [previous_yr_loss_ratio_text[i].split(",") for i in range(len(previous_yr_loss_ratio_text))]
            df_prev = pd.DataFrame(data_l[1:], columns=data_l[0])
            self._assign_loss_ratio_subset(df_prev, 'previous', loss_ratio_grouping_optical)
        if current_yr_loss_ratio_text:
            current_yr_loss_ratio_text = current_yr_loss_ratio_text.splitlines()
            data_l = [current_yr_loss_ratio_text[i].split(",") for i in range(len(current_yr_loss_ratio_text))]
            df_curr = pd.DataFrame(data_l[1:], columns=data_l[0])
            self._assign_loss_ratio_subset(df_curr, 'current', loss_ratio_grouping_optical)


    def p16_LR_by_benefits(self):
        template_p16 = self.template_wb["P16_LR by Benefits"]
        cols = [3,4,5]
        previous_start_row, current_start_row = 5,12
        
        
        if self.previous_year_loss_ratio_df is not None:
            for i in range(previous_start_row, len(self.previous_year_loss_ratio_df) + previous_start_row):
                if self.previous_year_loss_ratio_df.iloc[i-previous_start_row]['benefit_type'] != "Total":
                    template_p16.cell(row=i, column=3).value = float(self.previous_year_loss_ratio_df.iloc[i-previous_start_row]['actual_premium']) * 12 / int(self.previous_year_loss_ratio_df.iloc[i-previous_start_row]["duration"])
                    template_p16.cell(row=i, column=4).value = float(self.previous_year_loss_ratio_df.iloc[i-previous_start_row]['actual_paid_w_ibnr']) * 12 / int(self.previous_year_loss_ratio_df.iloc[i-previous_start_row]["duration"])

        if self.current_year_loss_ratio_df is not None:
            for i in range(current_start_row, len(self.current_year_loss_ratio_df) + current_start_row):
                if self.current_year_loss_ratio_df.iloc[i-current_start_row]['benefit_type'] != "Total":
                    template_p16.cell(row=i, column=3).value = float(self.current_year_loss_ratio_df.iloc[i-current_start_row]['actual_premium']) * 12 / int(self.current_year_loss_ratio_df.iloc[i-current_start_row]["duration"])
                    template_p16.cell(row=i, column=4).value = float(self.current_year_loss_ratio_df.iloc[i-current_start_row]['actual_paid_w_ibnr']) * 12 / int(self.current_year_loss_ratio_df.iloc[i-current_start_row]["duration"])

        self._populate_loss_ratio_history_table(template_p16)
        self._populate_loss_ratio_detail_table(template_p16)

    def convert_all(self):
        self.claim_info()
        self.p16_LR_by_benefits()
        self.P20_overall()
        self.P20_benefittype()
        self.P20_network()
        self.P21_by_class()
        self.P22_by_class()
        self.P25_by_plan()
        self.P26_by_class()
    
