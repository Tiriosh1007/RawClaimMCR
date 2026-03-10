import pandas as pd
import numpy as np
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from dateutil.relativedelta import relativedelta

class IBNRTool():
    def __init__(self, input_path, number_of_months=8, month_to_extraction=0, col_setup=None):
        if col_setup is None:
            col_setup = ['policy_number', 'insurer', 'policy_start_date', 'policy_end_date', 'incur_date', 'pay_date', 'benefit_type', 'incurred_amount', 'paid_amount']

        self.input_path = input_path
        self.number_of_months = number_of_months
        self.month_to_extraction = month_to_extraction
        self.col_setup = col_setup
        self.benefit_map = {'Hospital': 'IP', 'Clinic': 'OP', 'Dental': 'Dental', 'Maternity': 'Maternity'}

        self.claim_data = None
        self.global_triangle_start = None
        

    def read_claim_data(self):
        self.claim_data = pd.read_csv(self.input_path, 
                                      usecols=self.col_setup, 
                                      dtype={'policy_number': str, 'insurer': str, 'policy_start_date': str, 
                                             'policy_end_date': str, 'incur_date': str, 'pay_date': str, 
                                             'benefit_type': str, 'incurred_amount': float, 'paid_amount': float}, 
                                      low_memory=False
                                      )

    def apply_benefit_map(self):
        self.claim_data['benefit_type'] = self.claim_data['benefit_type'].map(self.benefit_map).fillna(self.claim_data['benefit_type'])

    def _to_month_string(self, value):
        return pd.to_datetime(value).strftime('%Y%m')

    def _to_month_timestamp(self, value):
        return pd.to_datetime(str(value), format='%Y%m')

    def _generate_month_index(self, start_month, end_month):
        return pd.date_range(
            start=self._to_month_timestamp(start_month),
            end=self._to_month_timestamp(end_month),
            freq='MS'
        ).to_period('M').astype(str).str.replace('-', '')

    def _is_observed_cell(self, policy_month, lag_month, observed_through_month):
        policy_month_date = self._to_month_timestamp(policy_month)
        observed_through_date = self._to_month_timestamp(observed_through_month)
        return policy_month_date + pd.DateOffset(months=int(lag_month)) <= observed_through_date + pd.DateOffset(months=self.month_to_extraction)

    def _build_triangle(self, filtered_data, end_month):
        all_months = self._generate_month_index(self.global_triangle_start, end_month)
        max_lag_month = int(filtered_data['lag_month'].max()) if filtered_data['lag_month'].notna().any() else 0
        lag_months = list(range(0, max(13, max_lag_month + 1)))

        loss_triangle = filtered_data.pivot_table(
            index='policy_month',
            columns='lag_month',
            values='paid_amount',
            aggfunc='sum',
            fill_value=0
        )
        loss_triangle = loss_triangle.reindex(all_months, fill_value=0)
        loss_triangle = loss_triangle.reindex(columns=lag_months, fill_value=0)
        loss_triangle = loss_triangle.cumsum(axis=1).astype(float)
        return loss_triangle

    def _mask_unobserved_triangle(self, base_triangle, observed_through_month):
        masked_triangle = base_triangle.copy()
        projection_mask = pd.DataFrame(False, index=base_triangle.index, columns=base_triangle.columns)

        for policy_month in base_triangle.index:
            for lag_month in base_triangle.columns:
                if not self._is_observed_cell(policy_month, lag_month, observed_through_month):
                    masked_triangle.loc[policy_month, lag_month] = np.nan
                    projection_mask.loc[policy_month, lag_month] = True

        return masked_triangle, projection_mask

    def _calculate_development_metrics(self, actual_triangle):
        development_factors = []
        total_new_paid = []

        for idx in range(len(actual_triangle.columns) - 1):
            current_col = actual_triangle.iloc[:, idx].dropna()
            next_col = actual_triangle.iloc[:, idx + 1].dropna()
            valid_indices = current_col.index.intersection(next_col.index)

            current_data = current_col.loc[valid_indices].tail(12)
            next_data = next_col.loc[valid_indices].tail(12)
            next_data_non_zero = next_data[next_data > 0]

            if current_data.sum() == 0:
                factor = 1
                new_paid = 0
            else:
                factor = next_data.sum() / current_data.sum()
                increment_count = len(next_data_non_zero)
                if increment_count == 0:
                    new_paid = 0
                else:
                    new_paid = (next_data.sum() - current_data.sum()) / increment_count

            development_factors.append(factor)
            total_new_paid.append(new_paid)

        return development_factors, total_new_paid

    def _project_triangle(self, actual_triangle, observed_through_month, development_factors):
        projected_triangle = actual_triangle.copy()

        for policy_month in projected_triangle.index:
            for lag_month in projected_triangle.columns:
                if self._is_observed_cell(policy_month, lag_month, observed_through_month):
                    continue

                if lag_month == 0:
                    projected_triangle.loc[policy_month, lag_month] = np.nan
                    continue

                previous_value = projected_triangle.loc[policy_month, lag_month - 1]
                if pd.isna(previous_value):
                    projected_triangle.loc[policy_month, lag_month] = np.nan
                else:
                    projected_triangle.loc[policy_month, lag_month] = previous_value * development_factors[lag_month - 1]

        return projected_triangle

    def _carry_forward_actual_triangle(self, actual_triangle, observed_through_month):
        carried_triangle = actual_triangle.copy()

        for policy_month in carried_triangle.index:
            for lag_month in carried_triangle.columns:
                if self._is_observed_cell(policy_month, lag_month, observed_through_month):
                    continue

                if lag_month == 0:
                    carried_triangle.loc[policy_month, lag_month] = np.nan
                else:
                    carried_triangle.loc[policy_month, lag_month] = carried_triangle.loc[policy_month, lag_month - 1]

        return carried_triangle

    def _calculate_current_ibnr_factor(self, projected_triangle, carried_triangle):
        actual_paid_col = carried_triangle.iloc[-self.number_of_months:, -1]
        estimated_paid_col = projected_triangle.iloc[-self.number_of_months:, -1]

        total_actual_paid = actual_paid_col.sum(skipna=True)
        total_estimated_paid = estimated_paid_col.sum(skipna=True)

        if total_actual_paid == 0:
            return np.nan

        return total_estimated_paid / total_actual_paid - 1

    def _calculate_historical_ibnr_factors(self, actual_triangle, projected_triangle, observed_through_month):
        historical_factors = []
        latest_month = self._to_month_timestamp(observed_through_month)
        earliest_month = self._to_month_timestamp(actual_triangle.index[0])
        latest_renewal_month = latest_month - pd.DateOffset(months=self.number_of_months - 1)

        renewal_month = latest_renewal_month - pd.DateOffset(years=1)
        while renewal_month >= earliest_month:
            policy_year_months = [renewal_month + pd.DateOffset(months=offset) for offset in range(12)]

            partial_paid_values = []
            annual_claim_values = []
            used_cells = []
            valid_partial = True
            valid_annual = True

            for offset in range(self.number_of_months):
                row_month = renewal_month + pd.DateOffset(months=offset)
                row_key = row_month.strftime('%Y%m')
                lag_month = self.number_of_months - 1 - offset

                if row_key not in actual_triangle.index or lag_month not in actual_triangle.columns:
                    valid_partial = False
                    break

                cell_value = actual_triangle.loc[row_key, lag_month]
                if pd.isna(cell_value):
                    valid_partial = False
                    break

                partial_paid_values.append(cell_value)
                used_cells.append((row_key, lag_month))

            for row_month in policy_year_months:
                row_key = row_month.strftime('%Y%m')

                if row_key not in projected_triangle.index or 12 not in projected_triangle.columns:
                    valid_annual = False
                    break

                cell_value = projected_triangle.loc[row_key, 12]
                if pd.isna(cell_value):
                    valid_annual = False
                    break

                annual_claim_values.append(cell_value)

            if valid_partial and valid_annual:
                historical_paid = float(np.sum(partial_paid_values))
                annual_claim = float(np.sum(annual_claim_values))
                annualized_paid = historical_paid * 12 / self.number_of_months

                if annualized_paid != 0:
                    ibnr_factor = annual_claim / annualized_paid - 1
                else:
                    ibnr_factor = np.nan

                historical_factors.append({
                    'renewal_month': renewal_month.strftime('%Y%m'),
                    'historical_paid': historical_paid,
                    'annual_claim': annual_claim,
                    'annualized_paid': annualized_paid,
                    'ibnr_factor': ibnr_factor,
                    'used_cells': used_cells,
                })

            renewal_month = renewal_month - pd.DateOffset(years=1)

        return historical_factors

    def lag_month(self, row):
        if pd.isna(row['incur_date']) or pd.isna(row['pay_date']):
            return None
        incur = pd.to_datetime(row['incur_date'])
        pay = pd.to_datetime(row['pay_date'])
        diff = relativedelta(pay, incur)
        if diff.years == 0 and diff.months == 0 and diff.days < 0: # definition on lag_month = 0
            return 0
        return diff.years * 12 + diff.months + (0 if diff.days >= 0 else -1)
    
    def prepare_claim_data(self):
        self.claim_data['policy_start_date'] = pd.to_datetime(self.claim_data['policy_start_date'], errors='coerce')
        self.claim_data['policy_end_date'] = pd.to_datetime(self.claim_data['policy_end_date'], errors='coerce')
        self.claim_data['incur_date'] = pd.to_datetime(self.claim_data['incur_date'], errors='coerce')
        self.claim_data['pay_date'] = pd.to_datetime(self.claim_data['pay_date'], errors='coerce')
        self.claim_data['paid_amount'] = self.claim_data['paid_amount'].fillna(0)

        self.claim_data['lag_month'] = self.claim_data.apply(self.lag_month, axis=1)
        self.claim_data = self.claim_data[(self.claim_data['lag_month'].isna()) | (self.claim_data['lag_month'] >= 0)]
        self.claim_data = self.claim_data.dropna(subset=['incur_date'])
        self.claim_data['policy_month'] = self.claim_data['incur_date'].dt.strftime('%Y%m')
        self.claim_data['effective_month'] = self.claim_data['policy_start_date'].dt.month
        self.apply_benefit_map()

        if self.claim_data.empty:
            return

        self.global_triangle_start = self.claim_data['policy_month'].min()

    def _analyze_subset(self, filtered_data, end_month):
        if filtered_data.empty:
            return None

        base_triangle = self._build_triangle(filtered_data, end_month)
        actual_triangle, projection_mask = self._mask_unobserved_triangle(base_triangle, end_month)
        development_factors, total_new_paid = self._calculate_development_metrics(actual_triangle)
        projected_triangle = self._project_triangle(actual_triangle, end_month, development_factors)
        carried_triangle = self._carry_forward_actual_triangle(actual_triangle, end_month)
        current_ibnr_factor = self._calculate_current_ibnr_factor(projected_triangle, carried_triangle)
        historical_factors = self._calculate_historical_ibnr_factors(actual_triangle, projected_triangle, end_month)

        return {
            'loss_triangle': projected_triangle,
            'development_factors': development_factors,
            'total_new_paid': total_new_paid,
            'ibnr_factor': current_ibnr_factor,
            'historical_factors': historical_factors,
            'projection_mask': projection_mask,
        }

    def run(self):
        self.read_claim_data()
        self.prepare_claim_data()
        from io import BytesIO
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_written = False

            if self.claim_data is None or self.claim_data.empty:
                pd.DataFrame({'Info': ['No data available for selected policies']}).to_excel(writer, sheet_name='NoData')
                return output.getvalue()

            policy_no = self.claim_data['policy_number'].dropna().unique()

            # Analyze Each Policy Data
            for pno in policy_no:
                policy_data = self.claim_data[self.claim_data['policy_number'] == pno]
                if policy_data.empty:
                    continue
                policy_end_month = policy_data['policy_month'].max()

                for bnf in policy_data['benefit_type'].unique():
                    filtered_data = policy_data[policy_data['benefit_type'] == bnf]
                    if filtered_data.empty:
                        continue
                    analysis = self._analyze_subset(filtered_data, policy_end_month)
                    self._write_sheets_and_format(writer, filtered_data, analysis, pno, bnf)
                    sheet_written = True

                overall_analysis = self._analyze_subset(policy_data, policy_end_month)
                self._write_sheets_and_format(writer, policy_data, overall_analysis, pno, 'Overall')
                sheet_written = True

            # Analyze Overall Data
            for bnf in self.claim_data['benefit_type'].unique():
                bnf_claim_data = self.claim_data[self.claim_data['benefit_type'] == bnf]
                if bnf_claim_data.empty:
                    continue
                overall_end_month = self.claim_data['policy_month'].max()
                analysis = self._analyze_subset(bnf_claim_data, overall_end_month)
                self._write_sheets_and_format(writer, bnf_claim_data, analysis, 'Overall', bnf)
                sheet_written = True

            overall_overall_analysis = self._analyze_subset(self.claim_data, self.claim_data['policy_month'].max())
            self._write_sheets_and_format(writer, self.claim_data, overall_overall_analysis, 'Overall', 'Overall')
            sheet_written = True

            if not sheet_written:
                pd.DataFrame({'Info': ['No data available for selected policies']}).to_excel(writer, sheet_name='NoData')

        return output.getvalue()


    def _write_sheets_and_format(self, writer, filtered_data, analysis, pno, bnf):
        sheet_name = f'Triangle_{pno}_{bnf}'
        loss_triangle = analysis['loss_triangle']
        development_factors = analysis['development_factors']
        total_new_paid = analysis['total_new_paid']
        IBNR_factor = analysis['ibnr_factor']
        historical_factors = analysis['historical_factors']
        projection_mask = analysis['projection_mask']
        policy_boundary_months = set(filtered_data['policy_start_date'].dropna().dt.strftime('%Y%m').unique())

        loss_triangle.to_excel(writer, sheet_name=sheet_name)
        filtered_data.to_excel(writer, sheet_name=f'RawData_{pno}_{bnf}', index=False)

        wb = writer.book
        ws = wb[sheet_name]
        last_row = ws.max_row

        dev_factors_title = ws.cell(row=last_row + 3, column=1, value='Development Factors')
        dev_factors_title.font = Font(bold=True)
        dev_factors_title.alignment = Alignment(horizontal='center')

        for idx, factor in enumerate(development_factors):
            lag_month_title = ws.cell(row=last_row + 2, column=idx + 2, value=f'{idx}/{idx + 1}')
            lag_month_title.font = Font(bold=True)
            lag_month_title.alignment = Alignment(horizontal='center')

            avg_paid_title = ws.cell(row=last_row + 4, column=1, value='Avg New Paid')
            avg_paid_title.font = Font(bold=True)

            dev_factors_display = ws.cell(row=last_row + 3, column=idx + 2, value=factor)
            dev_factors_display.number_format = '#,##0.000'

            avg_paid_display = ws.cell(row=last_row + 4, column=idx + 2, value=total_new_paid[idx])
            avg_paid_display.number_format = '#,##0'

        current_ibnr_row = last_row + 6

        IBNR_title = ws.cell(row=current_ibnr_row, column=1, value='IBNR Factor')
        IBNR_title.font = Font(bold=True)

        IBNR_display = ws.cell(row=current_ibnr_row, column=2, value=IBNR_factor)
        IBNR_display.number_format = '0.00%'

        for idx, historical_factor in enumerate(historical_factors, start=1):
            historical_row = current_ibnr_row + idx
            renewal_month_label = pd.to_datetime(historical_factor['renewal_month'], format='%Y%m').strftime('%Y-%b')
            historical_title = ws.cell(
                row=historical_row,
                column=1,
                value=f"{renewal_month_label}_IBNR (incl. seasonality)"
            )
            historical_title.font = Font(bold=True)

            historical_display = ws.cell(
                row=historical_row,
                column=2,
                value=historical_factor['ibnr_factor']
            )
            historical_display.number_format = '0.00%'

        projection_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
        historical_fill = PatternFill(fill_type="solid", fgColor="FFC6EFCE")
        boundary_fill = PatternFill(fill_type="solid", fgColor="FFD9EAF7")
        boundary_side = Side(style='thick', color='FF5B9BD5')
        boundary_border = Border(top=boundary_side)

        used_historical_cells = set()
        for historical_factor in historical_factors:
            used_historical_cells.update(historical_factor.get('used_cells', []))

        for policy_month in loss_triangle.index:
            excel_row = loss_triangle.index.get_loc(policy_month) + 2

            if policy_month in policy_boundary_months:
                ws.cell(row=excel_row, column=1).fill = boundary_fill
                ws.cell(row=excel_row, column=1).font = Font(bold=True)
                for column_index in range(1, len(loss_triangle.columns) + 2):
                    ws.cell(row=excel_row, column=column_index).border = boundary_border

            for lag_month in loss_triangle.columns:
                cell = ws.cell(
                    row=excel_row,
                    column=loss_triangle.columns.get_loc(lag_month) + 2
                )

                if (policy_month, lag_month) in used_historical_cells:
                    cell.fill = historical_fill

                if projection_mask.loc[policy_month, lag_month]:
                    cell.fill = projection_fill




    