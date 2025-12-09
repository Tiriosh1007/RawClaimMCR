import warnings
from datetime import datetime, date
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

DATE_PARSE_FORMATS = (
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%m-%d-%Y",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%Y.%m.%d",
    "%d.%m.%Y",
    "%m.%d.%Y",
    "%d %b %Y",
    "%d %B %Y",
    "%b %d %Y",
    "%B %d %Y",
    "%d %b, %Y",
    "%d %B, %Y",
    "%b %d, %Y",
    "%B %d, %Y"
)

class PresentationConvert():
    def __init__(self, input_file):
        try:
            self.input_file = input_file
            
            # template_file = "GMI__Presentation_template.xlsx"
            self.template_file = "GMI_Presentation_template.xlsx"
            
            self.template_wb = load_workbook(self.template_file, keep_links=False, data_only=False)
            self.plan_info = []
            self.previous_year_loss_ratio_df = None
            self.current_year_loss_ratio_df = None
            self.loss_ratio_history_by_policy = {}
            self.loss_ratio_records_by_policy = {}

            self.input_dtype = {
                "policy_number": str,
                "year": str,
                "benefit_type": str,
                "benefit": str,
                "panel": str,
                "class": str,
                "incurred_amount": float,
                "paid_amount": float,
                "usage_ratio": float,
                "no_of_claim_id": float,
                "no_of_claimants": float,
                "no_of_cases": float,
                "member_count": float,
                "incurred_per_claim": float, 
                "paid_per_claim": float,
                "incurred_per_case": float,
                "paid_per_case": float,
                "incurred_per_claimant": float,
                "paid_per_claimant": float,
                "claim_frequency": float,
                "data_month": float,
                "ibnr": float,
            }

            sheet_specs = [
                ("mcr_p20_policy", "P.20_Policy", {}),
                ("mcr_p20_benefit", "P.20_BenefitType", {"na_values": 0}),
                ("mcr_p20_benefit_dep_type", "P.20_Benefit_DepType", {"na_values": 0}),
                ("mcr_p20_network", "P.20_Network", {"na_values": 0}),
                ("mcr_p20_network_benefit", "P.20_Network_BenefitType", {"na_values": 0}),
                ("mcr_p21_class", "P.21_Class", {}),
                ("mcr_p22_class_benefit", "P.22_Class_BenefitType", {"na_values": 0}),
                ("mcr_policy_info", "Policy_Info", {"na_values": 0}),
                ("mcr_p25_class_panel_benefit", "P.25_Class_Panel_BenefitType", {"na_values": 0}),
                ("mcr_p23_usage_by_hosp_benefit", "P.23_IP_Benefit", {"na_values": 0}),
                ("mcr_p24_usage_by_clinic_benefit", "P.24_OP_Benefit", {"na_values": 0}),
                ("mcr_p26_op_panel_benefit", "P.26_OP_Panel_Benefit", {"na_values": 0}),
                ("mcr_p18_top_hosp_diag", "P.18_TopHosDiag", {"na_values": 0}),
                ("mcr_p19_top_clinic_diag", "P.18b_TopClinDiag", {"na_values": 0}),
                ("mcr_p30_op_freqclaimant", "P.30_OP_FreqClaimant", {"na_values": 0}),
            ]

            self.sheet_sources = {attr: sheet for attr, sheet, _ in sheet_specs}

            with pd.ExcelFile(self.input_file) as excel_file:
                self.available_sheets = set(excel_file.sheet_names)

                for attr, sheet_name, extra_kwargs in sheet_specs:
                    if sheet_name in self.available_sheets:
                        read_kwargs = {"sheet_name": sheet_name, "dtype": self.input_dtype}
                        read_kwargs.update(extra_kwargs)
                        setattr(self, attr, excel_file.parse(**read_kwargs))
                    else:
                        setattr(self, attr, None)
                        print(f"Warning: worksheet '{sheet_name}' not found in '{self.input_file}'. Dependent presentation pages will be skipped.")

            if self.mcr_p20_policy is not None:
                self.policy_info = self.mcr_p20_policy[["policy_number", "year"]]
            else:
                self.policy_info = pd.DataFrame(columns=["policy_number", "year"])

        except FileNotFoundError:
            print(f"Error: The file '{self.input_file}' was not found. Please check the name and try again.")
        except Exception as e:
            print(f"An error occurred: {e}")

    def _get_dataframe(self, attr_name, context_label):
        df = getattr(self, attr_name, None)
        if df is None:
            sheet = self.sheet_sources.get(attr_name, attr_name) if hasattr(self, 'sheet_sources') else attr_name
            print(f"Skipping {context_label}: worksheet '{sheet}' is unavailable in '{self.input_file}'.")
        return df

    def _populate_benefit_order(self, worksheet):
        df = self._get_dataframe("mcr_p20_benefit", "ClaimInfo benefit order")
        if df is None or 'benefit_type' not in df.columns:
            return

        ordered = []
        for value in df['benefit_type']:
            if pd.isna(value):
                continue
            benefit = str(value).strip()
            if not benefit or benefit.lower() == 'total':
                continue
            if benefit not in ordered:
                ordered.append(benefit)

        if not ordered:
            return

        col_idx = 6
        start_row = 2
        clear_limit = max(len(ordered) + 5, 20)
        for row in range(start_row, start_row + clear_limit):
            worksheet.cell(row=row, column=col_idx).value = None

        for offset, benefit in enumerate(ordered):
            worksheet.cell(row=start_row + offset, column=col_idx).value = benefit

    def set_policy_input(self,
                         previous_policy_num, previous_year_start_date, previous_year_end_date, previous_year, 
                         current_policy_num, current_year_start_date, current_year_end_date, current_year
                         ):
        
        self.previous_year, self.current_year = previous_year, current_year
        self.previous_year_start_date, self.previous_year_end_date = previous_year_start_date, previous_year_end_date
        self.current_year_start_date, self.current_year_end_date = current_year_start_date, current_year_end_date
        self.previous_policy_num, self.current_policy_num = previous_policy_num, current_policy_num
        class_df = self._get_dataframe("mcr_p21_class", "set_policy_input (plan setup)")
        if class_df is None:
            self.plan_info = []
            print("Plan list unavailable: worksheet 'P.21_Class' missing; claim info plan section will be empty.")
        else:
            class_df = class_df.copy(deep=True)
            class_df.dropna(subset=["class"], inplace=True)

            current_classes = class_df.loc[(class_df["policy_number"] == self.current_policy_num) & 
                                           (class_df["year"] == self.current_year), "class"].tolist()

            previous_classes = class_df.loc[(class_df["policy_number"] == self.previous_policy_num) & 
                                             (class_df["year"] == self.previous_year), "class"].tolist()

            combined_classes = current_classes + previous_classes
            deduped = []
            for class_name in combined_classes:
                if pd.isna(class_name):
                    continue
                class_str = str(class_name)
                if class_str not in deduped:
                    deduped.append(class_str)

            self.plan_info = deduped
        
    
    def claim_info(self):
        claim_info_worksheet = self.template_wb["ClaimInfo"]

        if self.previous_policy_num is not None:
            claim_info_worksheet.cell(row=4,column=3).value = self.previous_policy_num
            claim_info_worksheet.cell(row=6,column=3).value = self.previous_year_start_date
            claim_info_worksheet.cell(row=7,column=3).value = self.previous_year_end_date

        if self.current_policy_num is not None:
            claim_info_worksheet.cell(row=9,column=3).value = self.current_policy_num
            claim_info_worksheet.cell(row=11,column=3).value = self.current_year_start_date
            claim_info_worksheet.cell(row=12,column=3).value = self.current_year_end_date

        plan_start_row = 2
        for i in range(plan_start_row, len(self.plan_info) + 2):
            claim_info_worksheet.cell(row=i, column=8).value = self.plan_info[i-2]

        self._populate_benefit_order(claim_info_worksheet)
        
                

    def P20_overall(self):
        df = self._get_dataframe("mcr_p20_policy", "P20_overall")
        if df is None:
            return

        filtered_df = df.loc[
            ((df["policy_number"] == self.current_policy_num) & (df["year"] == self.current_year)) |
            ((df["policy_number"] == self.previous_policy_num) & (df["year"] == self.previous_year))
        ]

        required_cols = [
            col for col in ["incurred_amount", "paid_amount", "usage_ratio"]
            if col in filtered_df.columns
        ]

        input_p20 = filtered_df.dropna(subset=required_cols) if required_cols else filtered_df
        template_p20 = self.template_wb["P20_Usage Overview"]

        previous_start_row = current_start_row = 10
        for index, row in input_p20.iterrows():
            if row['year'] == self.previous_year:
                previous_start_row += 1 
                for col, val in zip([2,3,4], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                    template_p20.cell(row=previous_start_row, column=1).value = self.previous_policy_num
                    template_p20.cell(row=previous_start_row, column=col).value = val
            
            elif row['year'] == self.current_year:
                current_start_row += 1
                for col, val in zip([6,7,8], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                    template_p20.cell(row=previous_start_row, column=5).value = self.current_policy_num
                    template_p20.cell(row=current_start_row, column=col).value = val


    def P20_benefittype(self):
        df = self._get_dataframe("mcr_p20_benefit", "P20_benefittype")
        if df is None:
            return

        filtered_df = df.loc[
            ((df["policy_number"] == self.current_policy_num) & (df["year"] == self.current_year)) |
            ((df["policy_number"] == self.previous_policy_num) & (df["year"] == self.previous_year))
        ]

        required_cols = [
            col for col in ["benefit_type", "incurred_amount", "paid_amount", "usage_ratio", "no_of_claim_id"]
            if col in filtered_df.columns
        ]

        input_p20_btype = filtered_df.dropna(subset=required_cols) if required_cols else filtered_df
        template_p20 = self.template_wb["P20_Usage Overview"]

        # by benefit table variables 
        previous_start_row = current_start_row = 15

        # number of claims table variables 
        previous_num_of_claim = current_num_of_claim = 33

        for index, row in input_p20_btype.iterrows():
            if row["benefit_type"] != 'Total':
                if row["year"] == self.previous_year:
                    previous_start_row +=1
                    previous_num_of_claim +=1
                    template_p20.cell(row=previous_num_of_claim, column=2).value = row['no_of_claim_id']
                    for col, val in zip([2,3,4], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):                      
                        template_p20.cell(row=previous_start_row, column=col).value = val

                elif row["year"] == self.current_year:
                    current_start_row += 1
                    current_num_of_claim += 1
                    template_p20.cell(row=current_num_of_claim, column=6).value = row['no_of_claim_id']
                    for col, val in zip([6,7,8], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                        template_p20.cell(row=current_start_row, column=col).value = val


    def P20_network(self):
        df = self._get_dataframe("mcr_p20_network", "P20_network")
        if df is None:
            return

        input_p20_btype = df.loc[((df["policy_number"] == self.current_policy_num) & (df["year"] == self.current_year)) | 
                                  ((df["policy_number"] == self.previous_policy_num) & (df["year"] == self.previous_year))]
        template_p20 = self.template_wb["P20_Usage Overview"]

        previous_start_row = current_start_row = 27
        for index, row in input_p20_btype.iterrows():
            # Handling the by network table in P20_overview worksheet
            row_offset = 1 if row['panel'] == "Panel" else 0
            if row['year'] == self.previous_year:
                for col, val in zip([2, 3, 4], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                    template_p20.cell(row=previous_start_row + row_offset, column=col).value = val

            elif row['year'] == self.current_year:
                for col, val in zip([6, 7, 8], [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):                      
                    template_p20.cell(row=current_start_row + row_offset, column=col).value = val

        # Clinic panel/non-panel summaries from P.20_Network_BenefitType
        df_network_benefit = self._get_dataframe("mcr_p20_network_benefit", "P20_network_benefit")
        if df_network_benefit is None or df_network_benefit.empty:
            return

        benefit_col = df_network_benefit.get("benefit_type")
        if benefit_col is None:
            return
        clinic_mask = benefit_col.astype(str).str.contains("clinic", case=False, na=False)
        clinic_df = df_network_benefit.loc[clinic_mask].copy()
        if clinic_df.empty:
            return

        clinic_df["panel"] = clinic_df["panel"].astype(str).str.strip().str.title()
        clinic_df["policy_number"] = clinic_df["policy_number"].astype(str)
        clinic_df["year"] = clinic_df["year"].astype(str)

        def _clinic_summary(policy_number, year_value):
            if policy_number is None or year_value is None:
                return None
            mask = (
                (clinic_df["policy_number"] == str(policy_number)) &
                (clinic_df["year"] == str(year_value))
            )
            subset = clinic_df.loc[mask]
            if subset.empty:
                return None
            return subset.groupby("panel").agg({
                "no_of_claim_id": "sum",
                "paid_amount": "sum"
            })

        prev_vals = _clinic_summary(self.previous_policy_num, self.previous_year)
        curr_vals = _clinic_summary(self.current_policy_num, self.current_year)

        def _value_or_none(group, panel_name, column):
            if group is None or panel_name not in group.index:
                return None
            value = group.loc[panel_name, column]
            try:
                return float(value)
            except (TypeError, ValueError):
                return None

        template_p20.cell(row=46, column=2).value = _value_or_none(prev_vals, "Non-Panel", "no_of_claim_id")
        template_p20.cell(row=47, column=2).value = _value_or_none(prev_vals, "Panel", "no_of_claim_id")
        template_p20.cell(row=46, column=6).value = _value_or_none(curr_vals, "Non-Panel", "no_of_claim_id")
        template_p20.cell(row=47, column=6).value = _value_or_none(curr_vals, "Panel", "no_of_claim_id")
        template_p20.cell(row=54, column=2).value = _value_or_none(prev_vals, "Non-Panel", "paid_amount")
        template_p20.cell(row=55, column=2).value = _value_or_none(prev_vals, "Panel", "paid_amount")
        template_p20.cell(row=54, column=6).value = _value_or_none(curr_vals, "Non-Panel", "paid_amount")
        template_p20.cell(row=55, column=6).value = _value_or_none(curr_vals, "Panel", "paid_amount")

    
    def P21_by_class(self):
        df = self._get_dataframe("mcr_p21_class", "P21_by_class")
        if df is None:
            return

        input_p21 = df.loc[((df["policy_number"] == self.current_policy_num) & (df["year"] == self.current_year)) | 
                            ((df["policy_number"] == self.previous_policy_num) & (df["year"] == self.previous_year))]

        if input_p21.empty:
            return

        input_p21 = input_p21.dropna(subset=["class"]).fillna("N/A")

        template_p21 = self.template_wb["P21_Usage by Class"]

        prev_df = input_p21.loc[input_p21['year'] == self.previous_year].copy(deep=True)
        curr_df = input_p21.loc[input_p21['year'] == self.current_year].copy(deep=True)

        prev_df['class_str'] = prev_df['class'].astype(str)
        curr_df['class_str'] = curr_df['class'].astype(str)

        class_order = pd.Index(prev_df['class_str']).append(pd.Index(curr_df['class_str']))
        deduped_classes = []
        for class_name in class_order:
            if class_name not in deduped_classes:
                deduped_classes.append(class_name)

        if not deduped_classes:
            return

        row_pointer = 7
        for class_name in deduped_classes:
            row_pointer += 1
            template_p21.cell(row=row_pointer, column=1).value = class_name

            prev_row = prev_df.loc[prev_df['class_str'] == class_name]
            if not prev_row.empty:
                values = prev_row.iloc[0][['incurred_amount', 'paid_amount', 'usage_ratio']]
                for col, val in zip([3, 4, 5], values):
                    template_p21.cell(row=row_pointer, column=col).value = val

            curr_row = curr_df.loc[curr_df['class_str'] == class_name]
            if not curr_row.empty:
                values = curr_row.iloc[0][['incurred_amount', 'paid_amount', 'usage_ratio']]
                for col, val in zip([6, 7, 8], values):
                    template_p21.cell(row=row_pointer, column=col).value = val


    def P22_by_class(self):
        df = self._get_dataframe("mcr_p22_class_benefit", "P22_by_class")
        if df is None:
            return

        input_p22 = df.loc[((df["policy_number"] == self.current_policy_num) & (df["year"] == self.current_year)) | 
                            ((df["policy_number"] == self.previous_policy_num) & (df["year"] == self.previous_year))].fillna("N/A")
        template_p22 = self.template_wb["P22_Usage by Class by Ben"]

        input_p22.fillna(0, inplace=True)

        plan_num = self.plan_info
        P22_format_row_num = [13,20,27,34,45,52,59,66,77,84,91,98]

        plan_num, P22_format_row_num = pd.Series(plan_num).astype(str), pd.Series(P22_format_row_num)
        current_plan_row_df = pd.concat([plan_num, P22_format_row_num],axis=1, ignore_index=True).fillna("N/A")
        current_plan_row_df.columns = ['plan', "start_row"]
        
        previous_row_dict = current_plan_row_df.set_index('plan')['start_row'].to_dict()
        current_row_dict = current_plan_row_df.set_index('plan')['start_row'].to_dict()

        for index, row in input_p22.iterrows():
            if row.get('benefit_type') == 'Total':
                continue
            class_id = row['class']
            previous_start_row = previous_row_dict[class_id]
            current_start_row = current_row_dict[class_id]

            if row['year'] == self.previous_year:
                row_target = previous_start_row + 1
                cols = [3, 4, 5]
                previous_row_dict[class_id] = row_target
                for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                    template_p22.cell(row=row_target, column=col).value = val

            elif row['year'] == self.current_year:
                row_target = current_start_row + 1
                cols = [6, 7, 8]
                current_row_dict[class_id] = row_target
                for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                    template_p22.cell(row=row_target, column=col).value = val
                        

        
    def P23_Usage_HospByBnf(self):
        df = self._get_dataframe("mcr_p23_usage_by_hosp_benefit", "P23_Usage_HospByBnf")
        if df is None:
            return

        value_cols = [
            "benefit", "incurred_amount", "paid_amount", "usage_ratio", "no_of_claim_id", "no_of_claimants",
            "incurred_per_claim", "paid_per_claim", "incurred_per_claimant", "paid_per_claimant", "claim_frequency"
        ]

        prev_mask = (
            (df["policy_number"] == self.previous_policy_num) &
            (df["year"] == self.previous_year)
        )
        curr_mask = (
            (df["policy_number"] == self.current_policy_num) &
            (df["year"] == self.current_year)
        )

        def _normalize_benefit(value):
            if pd.isna(value):
                return None
            text = str(value).strip()
            return text if text else None

        def _prepare_subset(subset):
            if subset is None or subset.empty:
                return pd.DataFrame(columns=value_cols)
            prepared = subset.copy(deep=True)
            prepared = prepared.dropna(subset=["benefit"])
            if prepared.empty:
                return prepared
            prepared["benefit"] = prepared["benefit"].astype(str).str.strip()
            prepared = prepared[prepared["benefit"] != ""]
            numeric_cols = [col for col in prepared.columns if col != "benefit"]
            if numeric_cols:
                prepared.loc[:, numeric_cols] = prepared.loc[:, numeric_cols].fillna(0)
            return prepared

        def _is_smm_benefit(name: str) -> bool:
            lowered = name.lower()
            if lowered == "smm":
                return True
            has_smm = "smm" in lowered
            has_top = any(token in lowered for token in ["top-up", "top up", "topup", "top up/smm", "top"])
            return has_smm and has_top

        prev_df = _prepare_subset(df.loc[prev_mask, value_cols])
        curr_df = _prepare_subset(df.loc[curr_mask, value_cols])

        benefit_order_source = df.loc[prev_mask | curr_mask, ["benefit"]]
        ordered_benefits = []
        for raw_value in benefit_order_source["benefit"]:
            norm = _normalize_benefit(raw_value)
            if norm and norm not in ordered_benefits:
                ordered_benefits.append(norm)

        for source_df in (prev_df, curr_df):
            if source_df.empty:
                continue
            for norm in source_df["benefit"]:
                if norm not in ordered_benefits:
                    ordered_benefits.append(norm)

        smm_benefits = [name for name in ordered_benefits if name and _is_smm_benefit(name)]
        ordered_regular = [name for name in ordered_benefits if name not in smm_benefits]

        prev_lookup = prev_df.drop_duplicates(subset=["benefit"]).set_index("benefit") if not prev_df.empty else pd.DataFrame()
        curr_lookup = curr_df.drop_duplicates(subset=["benefit"]).set_index("benefit") if not curr_df.empty else pd.DataFrame()

        template_p23 = self.template_wb["P23_Usage_HospByBnf"]

        start_row = 7
        reserved_rows = {32, 33, 34}
        prev_cols = [2,3,4,5,6,7,8,9,10,11]
        curr_cols = [12,13,14,15,16,17,18,19,20,21]
        metric_fields = [
            'incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id',
            'no_of_claimants', 'incurred_per_claim', 'paid_per_claim',
            'incurred_per_claimant', 'paid_per_claimant', 'claim_frequency'
        ]

        def _write_values(lookup_df, benefit_name, target_row, col_targets):
            if lookup_df is None or lookup_df.empty or benefit_name not in lookup_df.index:
                return
            data_row = lookup_df.loc[benefit_name]
            values = [data_row[field] for field in metric_fields]
            for col, val in zip(col_targets, values):
                template_p23.cell(row=target_row, column=col).value = val

        def _next_available_row(row_idx):
            while row_idx in reserved_rows:
                row_idx += 1
            return row_idx

        next_row = start_row
        for benefit in ordered_regular:
            next_row = _next_available_row(next_row)
            template_p23.cell(row=next_row, column=1).value = benefit
            _write_values(prev_lookup, benefit, next_row, prev_cols)
            _write_values(curr_lookup, benefit, next_row, curr_cols)
            next_row += 1

        if smm_benefits:
            smm_name = smm_benefits[0]
            smm_row = 33
            _write_values(prev_lookup, smm_name, smm_row, prev_cols)
            _write_values(curr_lookup, smm_name, smm_row, curr_cols)

    def P24_Usage_ClinicByBnf(self):
        df = self._get_dataframe("mcr_p24_usage_by_clinic_benefit", "P24_Usage_ClinicByBnf")
        if df is None:
            return

        prev_df = df.loc[
            (df["policy_number"] == self.previous_policy_num) & 
            (df["year"] == self.previous_year)
        ][["benefit", "incurred_amount", "paid_amount", "usage_ratio", "no_of_claim_id", "incurred_per_claim",
        "paid_per_claim"]].dropna()

        curr_df = df.loc[
            (df["policy_number"] == self.current_policy_num) & 
            (df["year"] == self.current_year)
        ][["benefit", "incurred_amount", "paid_amount", "usage_ratio", "no_of_claim_id", "incurred_per_claim",
        "paid_per_claim"]].dropna()

        merged_df = prev_df.merge(curr_df, on="benefit", how="outer", suffixes=('_prev', '_curr')).fillna(0)
        merged_df['paid_amount_curr'] = pd.to_numeric(merged_df.get('paid_amount_curr'), errors='coerce').fillna(0)
        merged_df.sort_values('paid_amount_curr', ascending=False, inplace=True)
        merged_df.reset_index(drop=True, inplace=True)

        template_p24 = self.template_wb["P24_Usage_ClinicByBnf"]

        start_row = 8
        for i, row in merged_df.iterrows():
            excel_row = start_row + i
            # Write benefit name in col 2
            template_p24.cell(row=excel_row, column=2).value = row["benefit"]

            # Write previous year data columns 3-8
            prev_cols = [3,4,5,6,7,8]
            prev_values = [row[col] for col in [
                "incurred_amount_prev", "paid_amount_prev", "usage_ratio_prev", "no_of_claim_id_prev", 
                "incurred_per_claim_prev", "paid_per_claim_prev"]]

            for col, val in zip(prev_cols, prev_values):
                template_p24.cell(row=excel_row, column=col).value = val

            # Write current year data columns 9-14
            curr_cols = [9,10,11,12,13,14]
            curr_values = [row[col] for col in [
                "incurred_amount_curr", "paid_amount_curr", "usage_ratio_curr", "no_of_claim_id_curr",
                "incurred_per_claim_curr", "paid_per_claim_curr"]]

            for col, val in zip(curr_cols, curr_values):
                template_p24.cell(row=excel_row, column=col).value = val


    def P25_by_plan(self):
        df = self._get_dataframe("mcr_p25_class_panel_benefit", "P25_by_plan")
        if df is None:
            return

        input_p25 = df.loc[((df["policy_number"] == self.current_policy_num) & (df["year"] == self.current_year)) | 
                            ((df["policy_number"] == self.previous_policy_num) & (df["year"] == self.previous_year))]
        template_p25 = self.template_wb["P25_UsageByplanBynetworkByben"]
        input_p25.fillna(0, inplace=True)
        plan_num = self.plan_info

        # creating dataframe for handling non-panel data
        P25_nonpanel_row_num = [13,27,41,60,74,88,107,121,135,154,168,182]
        plan_num, P25_nonpanel_row_num = pd.Series(plan_num).astype(str), pd.Series(P25_nonpanel_row_num)
        nonpanel_plan_row_df = pd.concat([plan_num, P25_nonpanel_row_num],axis=1, ignore_index=True).fillna("N/A")
        nonpanel_plan_row_df.columns = ['plan', "start_row"]

        previous_nonpanel_row_dict = nonpanel_plan_row_df.set_index('plan')['start_row'].to_dict()
        current_nonpanel_row_dict = nonpanel_plan_row_df.set_index('plan')['start_row'].to_dict()

        # creating dataframe for handling panel data
        P25_panel_row_num  = [20,34,48,67,81,95,114,128,142,161,175,189]
        plan_num, P25_panel_row_num = pd.Series(plan_num).astype(str), pd.Series(P25_panel_row_num)
        panel_plan_row_df = pd.concat([plan_num, P25_panel_row_num],axis=1, ignore_index=True).fillna("N/A")
        panel_plan_row_df.columns = ['plan', "start_row"]
        
        previous_panel_row_dict = panel_plan_row_df.set_index('plan')['start_row'].to_dict()
        current_panel_row_dict = panel_plan_row_df.set_index('plan')['start_row'].to_dict()
    
        for index, row in input_p25.iterrows():
            class_id = row['class']
            previous_start_row = previous_nonpanel_row_dict[class_id]
            current_start_row = current_nonpanel_row_dict[class_id]

            previous_panel_start_row = previous_panel_row_dict[class_id]
            current_panel_start_row = current_panel_row_dict[class_id]

            if row['panel'] == 'Non-Panel':
                if row['year'] == self.previous_year:
                    row_target = previous_start_row + 1
                    cols = [4, 5, 6]
                    previous_nonpanel_row_dict[class_id] = row_target
                    for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                        template_p25.cell(row=row_target, column=col).value = val
                elif row['year'] == self.current_year:
                    row_target = current_start_row + 1
                    cols = [7,8,9]
                    current_nonpanel_row_dict[class_id] = row_target
                    for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                        template_p25.cell(row=row_target, column=col).value = val
            
            elif row['panel'] == 'Panel':
                if row['year'] == self.previous_year:
                    row_target = previous_panel_start_row + 1
                    cols = [4, 5, 6]
                    previous_panel_row_dict[class_id] = row_target
                    for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                        template_p25.cell(row=row_target, column=col).value = val
                elif row['year'] == self.current_year:
                    row_target = current_panel_start_row + 1
                    cols = [7,8,9]
                    current_panel_row_dict[class_id] = row_target
                    for col, val in zip(cols, [row['incurred_amount'], row['paid_amount'], row['usage_ratio']]):
                        template_p25.cell(row=row_target, column=col).value = val


    def P26_by_class(self):
        df = self._get_dataframe("mcr_p26_op_panel_benefit", "P26_by_class")
        if df is None:
            return

        input_p26 = df.loc[((df["policy_number"] == self.current_policy_num) & (df["year"] == self.current_year)) | 
                            ((df["policy_number"] == self.previous_policy_num) & (df["year"] == self.previous_year))].fillna("N/A")
        template_p26 = self.template_wb["P26_Usage_Clinical by Network"]

        panel_previous_start_row = panel_current_start_row = 7
        nonpanel_previous_start_row = nonpanel_current_start_row = 28

        rows = {
            ('Panel', self.current_year): {
                'row': panel_previous_start_row,
                'cols': [2, 9, 10, 11, 12, 13, 14],
                'values': ['benefit', 'incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']
            },
            ('Panel', self.previous_year): {
                'row': panel_current_start_row,
                'cols': [3, 4, 5, 6, 7, 8],
                'values': ['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']
            },
            ('Non-Panel', self.current_year): {
                'row': nonpanel_previous_start_row,
                'cols': [2, 9, 10, 11, 12, 13, 14],
                'values': ['benefit', 'incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']
            },
            ('Non-Panel', self.previous_year): {
                'row': nonpanel_current_start_row,
                'cols': [3, 4, 5, 6, 7, 8],
                'values': ['incurred_amount', 'paid_amount', 'usage_ratio', 'no_of_claim_id', 'incurred_per_claim', 'paid_per_claim']
            }
        }

        current_rows = {
            ('Panel', self.current_year): panel_previous_start_row,
            ('Panel', self.previous_year): panel_current_start_row,
            ('Non-Panel', self.current_year): nonpanel_previous_start_row,
            ('Non-Panel', self.previous_year): nonpanel_current_start_row
        }

        for index, row in input_p26.iterrows():
            key = (row['panel'], row['year'])
            if key in rows:
                current_rows[key] += 1
                row_idx = current_rows[key]
                info = rows[key]
                values = [row[val] for val in info['values']]
                for col, val in zip(info['cols'], values):
                    template_p26.cell(row=row_idx, column=col).value = val


    def P17_ClaimPattern(self):
        class_benefit_df = self._get_dataframe("mcr_p22_class_benefit", "P17_ClaimPattern")
        if class_benefit_df is None or class_benefit_df.empty:
            return

        template_p17 = self.template_wb["P17_ClaimPattern"]

        clinic_panel_df = self._get_dataframe("mcr_p25_class_panel_benefit", "P17_clinic_panel_source")
        policy_info_df = self._get_dataframe("mcr_policy_info", "P17_policy_info")
        policy_factor_cache = {}

        def _get_policy_factor(policy, year):
            if policy_info_df is None or policy_info_df.empty or policy is None or year is None:
                return {}
            key = (str(policy), str(year))
            if key in policy_factor_cache:
                return policy_factor_cache[key]
            work = policy_info_df.copy()
            if 'policy_number' not in work.columns or 'year' not in work.columns:
                policy_factor_cache[key] = {}
                return {}
            mask = (
                work['policy_number'].astype(str) == str(policy)
            ) & (
                work['year'].astype(str) == str(year)
            )
            subset = work.loc[mask].copy()
            if subset.empty:
                policy_factor_cache[key] = {}
                return {}
            subset['data_month'] = pd.to_numeric(subset.get('data_month'), errors='coerce')
            subset['ibnr'] = pd.to_numeric(subset.get('ibnr'), errors='coerce')
            row = subset.iloc[0]
            factor = {
                'data_month': row.get('data_month'),
                'ibnr': row.get('ibnr')
            }
            policy_factor_cache[key] = factor
            return factor

        def _apply_incidence_adjustment(value, policy, year):
            if value is None or pd.isna(value):
                return value
            factors = _get_policy_factor(policy, year)
            data_month = factors.get('data_month') if factors else None
            ibnr_val = factors.get('ibnr') if factors else None
            try:
                value_float = float(value)
            except (TypeError, ValueError):
                return value
            try:
                data_month = float(data_month)
            except (TypeError, ValueError):
                data_month = None
            try:
                ibnr_val = float(ibnr_val) if ibnr_val is not None else 0.0
            except (TypeError, ValueError):
                ibnr_val = 0.0
            if not data_month or data_month <= 0:
                return value_float
            return value_float * 12 / data_month * (1 + ibnr_val)

        def _write_factor_cells():
            template_p17.cell(row=29, column=2).value = None
            template_p17.cell(row=30, column=2).value = None
            candidates = [
                (self.current_policy_num, self.current_year),
                (self.previous_policy_num, self.previous_year)
            ]
            for policy, year in candidates:
                factors = _get_policy_factor(policy, year)
                if factors:
                    template_p17.cell(row=29, column=2).value = factors.get('data_month')
                    template_p17.cell(row=30, column=2).value = factors.get('ibnr')
                    break

        metric_candidates = {
            "member": ["member_count", "no_of_members", "no_of_insured"],
            "incidence": ["incidence_rate_case", "incidence_rate"],
            "incurred": ["incurred_per_case", "incurred_per_claim", "incurred_amount_per_case"],
            "paid": ["paid_per_case", "paid_per_claim", "paid_amount_per_case"]
        }

        numeric_columns = set()
        for cols in metric_candidates.values():
            numeric_columns.update(cols)
        numeric_columns.update(["paid_amount", "claim_paid", "paid", "no_of_claim_id", "no_of_cases"])
        numeric_columns = list(numeric_columns)

        def _coerce_numeric(frame):
            if frame is None or frame.empty:
                return frame
            work = frame.copy()
            for col in numeric_columns:
                if col in work.columns:
                    work[col] = pd.to_numeric(work[col], errors='coerce')
            return work

        def _benefit_subset(source_df, benefit_key, include_context=False):
            if source_df is None or source_df.empty or 'benefit_type' not in source_df.columns:
                return (pd.DataFrame(), (None, None)) if include_context else pd.DataFrame()
            work = source_df.copy()
            work['benefit_type'] = work['benefit_type'].astype(str).str.strip().str.lower()
            for col in ('policy_number', 'year'):
                if col in work.columns:
                    work[col] = work[col].astype(str)

            def _match(policy, year):
                if policy is None or year is None:
                    return work.iloc[0:0].copy()
                mask = (
                    (work['benefit_type'] == benefit_key) &
                    (work['policy_number'] == str(policy)) &
                    (work['year'] == str(year))
                )
                return work.loc[mask].copy()

            selected_policy = None
            selected_year = None
            subset = _match(self.current_policy_num, self.current_year)
            if not subset.empty:
                selected_policy = self.current_policy_num
                selected_year = self.current_year
            else:
                subset = _match(self.previous_policy_num, self.previous_year)
                if not subset.empty:
                    selected_policy = self.previous_policy_num
                    selected_year = self.previous_year
            if include_context:
                return subset, (selected_policy, selected_year)
            return subset

        def _clinic_panel_subset(include_context=False):
            if clinic_panel_df is None or clinic_panel_df.empty:
                return (pd.DataFrame(), (None, None)) if include_context else pd.DataFrame()
            work = clinic_panel_df.copy()
            for col in ('policy_number', 'year'):
                if col in work.columns:
                    work[col] = work[col].astype(str)
            if 'panel' in work.columns:
                work['panel_key'] = work['panel'].astype(str).str.strip().str.lower()
            else:
                work['panel_key'] = ''
            if 'benefit_type' in work.columns:
                work['benefit_type_key'] = work['benefit_type'].astype(str).str.strip().str.lower()
            else:
                work['benefit_type_key'] = ''
            if 'class' in work.columns:
                work['class'] = work['class'].astype(str).str.strip()

            def _match(policy, year):
                if policy is None or year is None:
                    return work.iloc[0:0].copy()
                mask = (
                    (work['policy_number'] == str(policy)) &
                    (work['year'] == str(year)) &
                    (work['panel_key'] == 'non-panel') &
                    (work['benefit_type_key'] == 'clinic')
                )
                return work.loc[mask].copy()

            selected_policy = None
            selected_year = None
            subset = _match(self.current_policy_num, self.current_year)
            if not subset.empty:
                selected_policy = self.current_policy_num
                selected_year = self.current_year
            else:
                subset = _match(self.previous_policy_num, self.previous_year)
                if not subset.empty:
                    selected_policy = self.previous_policy_num
                    selected_year = self.previous_year
            if include_context:
                return subset, (selected_policy, selected_year)
            return subset

        def _build_clinic_avg_lookup(subset):
            lookup = {}
            if subset is None or subset.empty or 'class' not in subset.columns:
                return lookup
            work = subset.copy()
            work['class_key'] = work['class'].astype(str).str.strip().str.lower()
            for _, row in work.iterrows():
                key = row['class_key']
                incurred_val = row.get('incurred_per_case')
                paid_val = row.get('paid_per_case')
                if pd.isna(incurred_val) and pd.isna(paid_val):
                    continue
                lookup[key] = (incurred_val, paid_val)
            return lookup

        def _compute_clinic_overall(subset):
            if subset is None or subset.empty:
                return None
            work = subset.copy()
            work['incurred_per_case'] = pd.to_numeric(work.get('incurred_per_case'), errors='coerce')
            work['paid_per_case'] = pd.to_numeric(work.get('paid_per_case'), errors='coerce')
            count_series = None
            if 'no_of_claim_id' in work.columns:
                count_series = pd.to_numeric(work.get('no_of_claim_id'), errors='coerce')
            elif 'no_of_cases' in work.columns:
                count_series = pd.to_numeric(work.get('no_of_cases'), errors='coerce')
            if count_series is None:
                incurred_avg = work['incurred_per_case'].mean(skipna=True)
                paid_avg = work['paid_per_case'].mean(skipna=True)
                if pd.isna(incurred_avg) and pd.isna(paid_avg):
                    return None
                return (incurred_avg, paid_avg)
            total_available = count_series.sum(min_count=1)
            if total_available is None or total_available == 0 or pd.isna(total_available):
                incurred_avg = work['incurred_per_case'].mean(skipna=True)
                paid_avg = work['paid_per_case'].mean(skipna=True)
                if pd.isna(incurred_avg) and pd.isna(paid_avg):
                    return None
                return (incurred_avg, paid_avg)
            weights = count_series.fillna(0)
            total_weight = weights.sum()
            if not total_weight:
                return None
            incurred_avg = float((work['incurred_per_case'] * weights).sum(min_count=1) / total_weight) if 'incurred_per_case' in work.columns else None
            paid_avg = float((work['paid_per_case'] * weights).sum(min_count=1) / total_weight) if 'paid_per_case' in work.columns else None
            if pd.isna(incurred_avg) and pd.isna(paid_avg):
                return None
            return (incurred_avg, paid_avg)

        def _pick_value(row, candidates):
            for col in candidates:
                if col in row.index:
                    value = row[col]
                    if pd.isna(value):
                        continue
                    return value
            return None

        def _clear_section(start_col):
            for row in range(6, 26):
                for col in range(start_col, start_col + 5):
                    template_p17.cell(row=row, column=col).value = None

        clinic_panel_info = _clinic_panel_subset(include_context=True)
        clinic_panel_subset = clinic_panel_info[0] if clinic_panel_info else pd.DataFrame()
        if clinic_panel_subset is not None and not clinic_panel_subset.empty:
            clinic_panel_subset = _coerce_numeric(clinic_panel_subset)
        clinic_avg_lookup = _build_clinic_avg_lookup(clinic_panel_subset)
        clinic_overall_avg = _compute_clinic_overall(clinic_panel_subset)

        def _write_section(subset_info, start_col, avg_override_lookup=None):
            subset, context = subset_info
            policy_used, year_used = context
            if subset is None or subset.empty or 'class' not in subset.columns:
                _clear_section(start_col)
                return
            work = subset.dropna(subset=['class']).copy()
            work['class'] = work['class'].astype(str).str.strip()
            work = work[work['class'] != ""]
            work = work.drop_duplicates(subset=['class'], keep='first')
            work = _coerce_numeric(work)
            _clear_section(start_col)

            row_pointer = 6
            max_rows = 20
            for _, row in work.iterrows():
                if row_pointer >= 6 + max_rows:
                    break
                template_p17.cell(row=row_pointer, column=start_col).value = row['class']
                template_p17.cell(row=row_pointer, column=start_col + 1).value = _pick_value(row, metric_candidates['member'])
                incidence_value = _pick_value(row, metric_candidates['incidence'])
                template_p17.cell(row=row_pointer, column=start_col + 2).value = _apply_incidence_adjustment(incidence_value, policy_used, year_used)
                class_key = str(row['class']).strip().lower()
                incurred_override = paid_override = None
                if avg_override_lookup:
                    override_vals = avg_override_lookup.get(class_key)
                    if override_vals is not None:
                        incurred_override, paid_override = override_vals
                incurred_value = incurred_override if (incurred_override is not None and not pd.isna(incurred_override)) else _pick_value(row, metric_candidates['incurred'])
                paid_value = paid_override if (paid_override is not None and not pd.isna(paid_override)) else _pick_value(row, metric_candidates['paid'])
                template_p17.cell(row=row_pointer, column=start_col + 3).value = incurred_value
                template_p17.cell(row=row_pointer, column=start_col + 4).value = paid_value
                row_pointer += 1

        hospital_subset = _benefit_subset(class_benefit_df, 'hospital', include_context=True)
        clinic_subset = _benefit_subset(class_benefit_df, 'clinic', include_context=True)
        _write_section(hospital_subset, 1)
        _write_section(clinic_subset, 6, avg_override_lookup=clinic_avg_lookup)

        benefit_totals_df = self._get_dataframe("mcr_p20_benefit", "P17_ClaimPattern_totals")
        if benefit_totals_df is not None and not benefit_totals_df.empty:
            hospital_total_info = _benefit_subset(benefit_totals_df, 'hospital', include_context=True)
            clinic_total_info = _benefit_subset(benefit_totals_df, 'clinic', include_context=True)

            hospital_total = _coerce_numeric(hospital_total_info[0]) if hospital_total_info else None
            clinic_total = _coerce_numeric(clinic_total_info[0]) if clinic_total_info else None
            hospital_context = hospital_total_info[1] if hospital_total_info else (None, None)
            clinic_context = clinic_total_info[1] if clinic_total_info else (None, None)

            def _set_if_empty(row_idx, col_idx, value):
                cell = template_p17.cell(row=row_idx, column=col_idx)
                if cell.value in (None, ""):
                    cell.value = value

            def _write_totals(row_data, start_col, label_col, label_text, context, avg_override=None):
                if row_data is None or row_data.empty:
                    return
                row = row_data.iloc[0]
                _set_if_empty(26, label_col, label_text)
                template_p17.cell(row=26, column=start_col).value = _pick_value(row, metric_candidates['member'])
                incidence_val = _pick_value(row, metric_candidates['incidence'])
                policy_used, year_used = context
                adjusted_incidence = _apply_incidence_adjustment(incidence_val, policy_used, year_used)
                template_p17.cell(row=26, column=start_col + 1).value = adjusted_incidence
                incurred_value = _pick_value(row, metric_candidates['incurred'])
                paid_value = _pick_value(row, metric_candidates['paid'])
                if avg_override is not None:
                    override_incurred, override_paid = avg_override
                    if override_incurred is not None and not pd.isna(override_incurred):
                        incurred_value = override_incurred
                    if override_paid is not None and not pd.isna(override_paid):
                        paid_value = override_paid
                template_p17.cell(row=26, column=start_col + 2).value = incurred_value
                template_p17.cell(row=26, column=start_col + 3).value = paid_value

            _write_totals(hospital_total, 2, 1, "Hospital Overall", hospital_context)
            clinic_avg_override = clinic_overall_avg if clinic_overall_avg is not None else None
            _write_totals(clinic_total, 7, 6, "Clinic Overall", clinic_context, avg_override=clinic_avg_override)

        dep_type_df = self._get_dataframe("mcr_p20_benefit_dep_type", "P17_ClaimPattern_dep")
        if dep_type_df is None or dep_type_df.empty:
            return

        required_cols = {'policy_number', 'year', 'dep_type'}
        if not required_cols.issubset(dep_type_df.columns):
            return

        dep_columns_prev = {'employee': 2, 'spouse': 3, 'child': 4}
        dep_columns_curr = {'employee': 5, 'spouse': 6, 'child': 7}
        dep_paid_candidates = ["paid_amount", "claim_paid", "paid"]

        dep_aliases = {
            'employee': 'employee',
            'emp': 'employee',
            'ee': 'employee',
            'self': 'employee',
            'staff': 'employee',
            'member': 'employee',
            'spouse': 'spouse',
            'sp': 'spouse',
            'wife': 'spouse',
            'husband': 'spouse',
            'partner': 'spouse',
            'child': 'child',
            'children': 'child',
            'ch': 'child',
            'kid': 'child',
            'son': 'child',
            'daughter': 'child'
        }

        dep_work = _coerce_numeric(dep_type_df)
        for col in ('policy_number', 'year', 'benefit_type', 'dep_type'):
            if col not in dep_work.columns:
                if col in ('benefit_type', 'dep_type'):
                    return
                continue
            dep_work[col] = dep_work[col].astype(str).str.strip()

        dep_work['benefit_type_key'] = dep_work.get('benefit_type', "").astype(str).str.strip().str.lower()
        dep_work['dep_type_key'] = dep_work['dep_type'].astype(str).str.strip().str.lower()
        dep_work['dep_type_key'] = dep_work['dep_type_key'].map(lambda x: dep_aliases.get(x, x))

        def _subset_dep(policy, year, benefit_key=None):
            if policy is None or year is None:
                return pd.DataFrame()
            mask = (
                (dep_work['policy_number'] == str(policy)) &
                (dep_work['year'] == str(year))
            )
            if benefit_key:
                mask &= (dep_work['benefit_type_key'] == benefit_key)
            return dep_work.loc[mask].copy()

        def _aggregate_dep_values(subset, candidates):
            if subset is None or subset.empty:
                return {}
            value_col = next((col for col in candidates if col in subset.columns), None)
            if value_col is None:
                return {}
            working = subset.copy()
            working['_value'] = pd.to_numeric(working[value_col], errors='coerce')
            grouped = working.groupby('dep_type_key')['_value'].sum(min_count=1)
            result = {}
            for dep_key, total in grouped.items():
                if dep_key not in dep_columns_prev or pd.isna(total):
                    continue
                result[dep_key] = float(total)
            return result

        def _dep_values(policy, year, benefit_key, value_candidates, allow_fallback=False):
            subset = _subset_dep(policy, year, benefit_key)
            if subset.empty and allow_fallback:
                subset = _subset_dep(policy, year, None)
            return _aggregate_dep_values(subset, value_candidates)

        dep_value_cols = list(dep_columns_prev.values()) + list(dep_columns_curr.values())

        def _populate_dep_row(row_idx, prev_vals, curr_vals):
            for col in dep_value_cols:
                template_p17.cell(row=row_idx, column=col).value = None
            for dep_key, col in dep_columns_prev.items():
                template_p17.cell(row=row_idx, column=col).value = prev_vals.get(dep_key)
            for dep_key, col in dep_columns_curr.items():
                template_p17.cell(row=row_idx, column=col).value = curr_vals.get(dep_key)

        member_prev = _dep_values(self.previous_policy_num, self.previous_year, 'total', metric_candidates['member'], allow_fallback=True)
        member_curr = _dep_values(self.current_policy_num, self.current_year, 'total', metric_candidates['member'], allow_fallback=True)
        _populate_dep_row(43, member_prev, member_curr)

        benefit_rows = [
            ('total', 45),
            ('hospital', 46),
            ('clinic', 47),
            ('dental', 48),
            ('maternity', 49),
            ('optical', 50)
        ]

        for benefit_key, row_idx in benefit_rows:
            allow_fallback = (benefit_key == 'total')
            prev_vals = _dep_values(self.previous_policy_num, self.previous_year, benefit_key, dep_paid_candidates, allow_fallback=allow_fallback)
            curr_vals = _dep_values(self.current_policy_num, self.current_year, benefit_key, dep_paid_candidates, allow_fallback=allow_fallback)
            _populate_dep_row(row_idx, prev_vals, curr_vals)

        freq_df = self._get_dataframe("mcr_p30_op_freqclaimant", "P17_freq_claimant")
        if freq_df is not None and not freq_df.empty:
            rename_map = {}
            for col in freq_df.columns:
                key = str(col).strip().lower().replace(" ", "_")
                if key == "policy_number":
                    rename_map[col] = "policy_number"
                elif key == "year":
                    rename_map[col] = "year"
                elif key in {"class", "plan"}:
                    rename_map[col] = "class"
                elif key in {"dep_type", "dep", "dep_type_group"}:
                    rename_map[col] = "dep_type"
                elif key in {"no_of_cases", "cases"}:
                    rename_map[col] = "no_of_cases"
                elif key in {"no_of_claimants", "no_of_claimant", "claimants"}:
                    rename_map[col] = "no_of_claimants"
                elif key in {"member_count", "members", "member"}:
                    rename_map[col] = "member_count"
            work = freq_df.rename(columns=rename_map).copy()

            required_cols = {"policy_number", "year", "class", "dep_type", "no_of_cases", "no_of_claimants", "member_count"}
            if required_cols.issubset(work.columns):
                for col in ("policy_number", "year", "class", "dep_type"):
                    work[col] = work[col].astype(str).str.strip()
                numeric_cols = ["no_of_cases", "no_of_claimants", "member_count"]
                for col in numeric_cols:
                    work[col] = pd.to_numeric(work[col], errors="coerce")

                current_mask = (work["policy_number"] == str(self.current_policy_num)) & (work["year"] == str(self.current_year))
                subset = work.loc[current_mask].copy()
                if subset.empty:
                    prev_mask = (work["policy_number"] == str(self.previous_policy_num)) & (work["year"] == str(self.previous_year))
                    subset = work.loc[prev_mask].copy()

                headers = [
                    "Plan",
                    "Dep Type",
                    "No. of Claims",
                    "No. of Claimants",
                    "times/EE",
                    "Member",
                    "% of Member",
                ]
                for idx, header in enumerate(headers, start=1):
                    template_p17.cell(row=71, column=idx).value = header

                for row_idx in range(72, 103):
                    for col_idx in range(1, 8):
                        template_p17.cell(row=row_idx, column=col_idx).value = None

                if not subset.empty:
                    subset["no_of_claimants"] = subset["no_of_claimants"].replace({0: np.nan})
                    subset["member_count"] = subset["member_count"].replace({0: np.nan})
                    subset["times_per_ee"] = subset["no_of_cases"] / subset["no_of_claimants"]
                    subset["pct_members"] = subset["no_of_claimants"] / subset["member_count"]
                    subset.replace([np.inf, -np.inf], np.nan, inplace=True)
                    subset = subset.dropna(subset=["class", "dep_type", "no_of_claimants"]).copy()

                    dep_order = {"EE": 0, "SP": 1, "CH": 2}
                    subset["dep_type_key"] = subset["dep_type"].astype(str).str.upper()
                    subset["dep_sort"] = subset["dep_type_key"].map(dep_order)
                    subset = subset.sort_values(by=["class", "dep_sort", "dep_type_key"])

                    write_cols = [
                        "class",
                        "dep_type",
                        "no_of_cases",
                        "no_of_claimants",
                        "times_per_ee",
                        "member_count",
                        "pct_members",
                    ]
                    row_pointer = 72
                    for _, row in subset.iterrows():
                        if row_pointer > 102:
                            break
                        for offset, col_name in enumerate(write_cols, start=0):
                            template_p17.cell(row=row_pointer, column=1 + offset).value = row.get(col_name)
                        row_pointer += 1

        _write_factor_cells()


    def P18_TopHospDiag(self):
        df = self._get_dataframe("mcr_p18_top_hosp_diag", "P18_TopHospDiag")
        if df is None:
            return
        # Filter data for previous and current years separately
        required_cols = [
            'diagnosis', 'incurred_amount', 'paid_amount',
            'no_of_claim_id', 'no_of_claimants'
        ]

        prev_df = df.loc[
            (df["policy_number"] == self.previous_policy_num) & 
            (df["year"] == self.previous_year)
        ].copy()

        curr_df = df.loc[
            (df["policy_number"] == self.current_policy_num) & 
            (df["year"] == self.current_year)
        ].copy()

        prev_df = prev_df.dropna(subset=required_cols)
        curr_df = curr_df.dropna(subset=required_cols)

        # Get top 10 by paid amount
        top_prev = prev_df.nlargest(10, 'paid_amount').reset_index(drop=True)
        top_curr = curr_df.nlargest(10, 'paid_amount').reset_index(drop=True)

        template_p18 = self.template_wb["P.18_TopHospDiag"]

        start_row = 3

        # Write previous year data
        for i, row in top_prev.iterrows():
            excel_row = start_row + i
            rank = i + 1
            template_p18.cell(row=excel_row, column=1).value = rank                           # Rank
            template_p18.cell(row=excel_row, column=2).value = row['diagnosis']               # Diagnosis
            template_p18.cell(row=excel_row, column=3).value = row['incurred_amount']         # incurred_amount
            template_p18.cell(row=excel_row, column=4).value = row['paid_amount']             # paid_amount
            template_p18.cell(row=excel_row, column=5).value = row['no_of_claim_id']          # no_of_claim_id
            template_p18.cell(row=excel_row, column=6).value = row['no_of_claimants']         # no_of_claimants

            # Calculated columns
            incurred_per_claim = (row['incurred_amount'] / row['no_of_claim_id']) if row['no_of_claim_id'] else 0
            paid_per_claim = (row['paid_amount'] / row['no_of_claim_id']) if row['no_of_claim_id'] else 0
            incurred_per_claimant = (row['incurred_amount'] / row['no_of_claimants']) if row['no_of_claimants'] else 0
            paid_per_claimant = (row['paid_amount'] / row['no_of_claimants']) if row['no_of_claimants'] else 0

            template_p18.cell(row=excel_row, column=7).value = incurred_per_claim
            template_p18.cell(row=excel_row, column=8).value = paid_per_claim
            template_p18.cell(row=excel_row, column=9).value = incurred_per_claimant
            template_p18.cell(row=excel_row, column=10).value = paid_per_claimant

        # Write current year data
        for i, row in top_curr.iterrows():
            excel_row = start_row + i
            rank = i + 1
            template_p18.cell(row=excel_row, column=11).value = row['diagnosis']              # Diagnosis
            template_p18.cell(row=excel_row, column=12).value = row['incurred_amount']        # incurred_amount
            template_p18.cell(row=excel_row, column=13).value = row['paid_amount']            # paid_amount
            template_p18.cell(row=excel_row, column=14).value = row['no_of_claim_id']         # no_of_claim_id
            template_p18.cell(row=excel_row, column=15).value = row['no_of_claimants']        # no_of_claimant

            # Calculated columns
            incurred_per_claim = (row['incurred_amount'] / row['no_of_claim_id']) if row['no_of_claim_id'] else 0
            paid_per_claim = (row['paid_amount'] / row['no_of_claim_id']) if row['no_of_claim_id'] else 0
            incurred_per_claimant = (row['incurred_amount'] / row['no_of_claimants']) if row['no_of_claimants'] else 0
            paid_per_claimant = (row['paid_amount'] / row['no_of_claimants']) if row['no_of_claimants'] else 0

            template_p18.cell(row=excel_row, column=16).value = incurred_per_claim
            template_p18.cell(row=excel_row, column=17).value = paid_per_claim
            template_p18.cell(row=excel_row, column=18).value = incurred_per_claimant
            template_p18.cell(row=excel_row, column=19).value = paid_per_claimant

    def p16_LR_by_benefits(self):
        template_p16 = self.template_wb["P16_LR by Benefits"]
        cols = [3,4,5]
        previous_start_row, current_start_row = 5,12
        
        
        if self.previous_year_loss_ratio_df is not None:
            for i in range(previous_start_row, len(self.previous_year_loss_ratio_df) + previous_start_row):
                if self.previous_year_loss_ratio_df.iloc[i-previous_start_row]['benefit_type'] != "Total":
                    template_p16.cell(row=i, column=3).value = float(self.previous_year_loss_ratio_df.iloc[i-previous_start_row]['actual_premium']) * 12 / int(self.previous_year_loss_ratio_df.iloc[i-previous_start_row]["duration"])
                    template_p16.cell(row=i, column=4).value = float(self.previous_year_loss_ratio_df.iloc[i-previous_start_row]['actual_paid_w_ibnr']) * 12 / int(self.previous_year_loss_ratio_df.iloc[i-previous_start_row]["duration"])

        if self.current_year_loss_ratio_df is not None:
            for i in range(current_start_row, len(self.current_year_loss_ratio_df) + current_start_row):
                if self.current_year_loss_ratio_df.iloc[i-current_start_row]['benefit_type'] != "Total":
                    template_p16.cell(row=i, column=3).value = float(self.current_year_loss_ratio_df.iloc[i-current_start_row]['actual_premium']) * 12 / int(self.current_year_loss_ratio_df.iloc[i-current_start_row]["duration"])
                    template_p16.cell(row=i, column=4).value = float(self.current_year_loss_ratio_df.iloc[i-current_start_row]['actual_paid_w_ibnr']) * 12 / int(self.current_year_loss_ratio_df.iloc[i-current_start_row]["duration"])

        self._populate_loss_ratio_history_table(template_p16)
        self._populate_loss_ratio_detail_table(template_p16)


    def P19_TopClinicDiag(self):
        df = self._get_dataframe("mcr_p19_top_clinic_diag", "P19_TopClinicDiag")
        if df is None:
            return
        # Filter data for previous and current years separately
        required_cols = [
            'diagnosis', 'incurred_amount', 'paid_amount',
            'no_of_claim_id', 'no_of_claimants'
        ]

        prev_df = df.loc[
            (df["policy_number"] == self.previous_policy_num) & 
            (df["year"] == self.previous_year)
        ].copy()
        curr_df = df.loc[
            (df["policy_number"] == self.current_policy_num) & 
            (df["year"] == self.current_year)
        ].copy()

        prev_df = prev_df.dropna(subset=required_cols)
        curr_df = curr_df.dropna(subset=required_cols)

        # Get top 10 by paid amount
        top_prev = prev_df.nlargest(10, 'paid_amount').reset_index(drop=True)
        top_curr = curr_df.nlargest(10, 'paid_amount').reset_index(drop=True)

        template_p19 = self.template_wb["P.19_TopClinDiag"]

        start_row = 3

        # Write previous year data
        for i, row in top_prev.iterrows():
            excel_row = start_row + i
            rank = i + 1
            template_p19.cell(row=excel_row, column=1).value = rank                           # Rank
            template_p19.cell(row=excel_row, column=2).value = row['diagnosis']               # Diagnosis
            template_p19.cell(row=excel_row, column=3).value = row['incurred_amount']         # incurred_amount
            template_p19.cell(row=excel_row, column=4).value = row['paid_amount']             # paid_amount
            template_p19.cell(row=excel_row, column=5).value = row['no_of_claim_id']          # no_of_claim_id
            template_p19.cell(row=excel_row, column=6).value = row['no_of_claimants']         # no_of_claimant

            # Calculated columns
            incurred_per_claim = (row['incurred_amount'] / row['no_of_claim_id']) if row['no_of_claim_id'] else 0
            paid_per_claim = (row['paid_amount'] / row['no_of_claim_id']) if row['no_of_claim_id'] else 0
            incurred_per_claimant = (row['incurred_amount'] / row['no_of_claimants']) if row['no_of_claimants'] else 0
            paid_per_claimant = (row['paid_amount'] / row['no_of_claimants']) if row['no_of_claimants'] else 0

            template_p19.cell(row=excel_row, column=7).value = incurred_per_claim
            template_p19.cell(row=excel_row, column=8).value = paid_per_claim
            template_p19.cell(row=excel_row, column=9).value = incurred_per_claimant
            template_p19.cell(row=excel_row, column=10).value = paid_per_claimant

        # Write current year data
        for i, row in top_curr.iterrows():
            excel_row = start_row + i
            rank = i + 1
            template_p19.cell(row=excel_row, column=11).value = row['diagnosis']              # Diagnosis
            template_p19.cell(row=excel_row, column=12).value = row['incurred_amount']        # incurred_amount
            template_p19.cell(row=excel_row, column=13).value = row['paid_amount']            # paid_amount
            template_p19.cell(row=excel_row, column=14).value = row['no_of_claim_id']         # no_of_claim_id
            template_p19.cell(row=excel_row, column=15).value = row['no_of_claimants']        # no_of_claimant

            # Calculated columns
            incurred_per_claim = (row['incurred_amount'] / row['no_of_claim_id']) if row['no_of_claim_id'] else 0
            paid_per_claim = (row['paid_amount'] / row['no_of_claim_id']) if row['no_of_claim_id'] else 0
            incurred_per_claimant = (row['incurred_amount'] / row['no_of_claimants']) if row['no_of_claimants'] else 0
            paid_per_claimant = (row['paid_amount'] / row['no_of_claimants']) if row['no_of_claimants'] else 0

            template_p19.cell(row=excel_row, column=16).value = incurred_per_claim
            template_p19.cell(row=excel_row, column=17).value = paid_per_claim
            template_p19.cell(row=excel_row, column=18).value = incurred_per_claimant
            template_p19.cell(row=excel_row, column=19).value = paid_per_claimant



    def save(self):
        from io import BytesIO
        output_file = BytesIO()
        try:
            self.template_wb.save(output_file)
            output_file.seek(0)
            data = output_file.getvalue()
            if not data:
                raise ValueError("Generated workbook stream is empty.")
            return data
        except Exception as e:
            print(f"An error occurred while saving the file into the BytesIO: {e}")
            return b""
        finally:
            output_file.close()

    def load_loss_ratio_dataframe(self, loss_ratio_df, policy_number, year, period, loss_ratio_grouping_optical=True):
        normalized_df = self._normalize_loss_ratio_dates(loss_ratio_df)
        subset = self._filter_loss_ratio_records(normalized_df, policy_number, year)
        self._assign_loss_ratio_subset(subset, period, loss_ratio_grouping_optical)
        self._append_loss_ratio_records(normalized_df, policy_number)
        self._update_loss_ratio_history(normalized_df, policy_number)

    def _format_date_yyyy_mm_dd(self, value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
        if isinstance(value, (pd.Timestamp, datetime, date)):
            parsed = pd.to_datetime(value, errors='coerce')
        else:
            text = str(value).strip()
            if not text:
                return None
            parsed = None
            for fmt in DATE_PARSE_FORMATS:
                try:
                    parsed = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    continue
            if parsed is None:
                parsed = pd.to_datetime(text, errors='coerce', dayfirst=False)
                if pd.isna(parsed):
                    parsed = pd.to_datetime(text, errors='coerce', dayfirst=True)
        parsed_ts = pd.to_datetime(parsed, errors='coerce')
        if parsed_ts is None or pd.isna(parsed_ts):
            return None
        return parsed_ts.strftime('%Y-%m-%d')

    def _normalize_loss_ratio_dates(self, df):
        if df is None or not isinstance(df, pd.DataFrame) or df.empty:
            return df
        date_columns = [
            col for col in (
                'policy_start_date', 'policy_end_date', 'data_as_of',
                'policy_effective_date', 'policy_expiry_date'
            )
            if col in df.columns
        ]
        if not date_columns:
            return df
        work = df.copy(deep=True)
        for col in date_columns:
            work[col] = work[col].apply(self._format_date_yyyy_mm_dd)
        return work

    def _normalize_policy_number(self, value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        text = str(value).strip()
        if not text:
            return None
        normalized = text.lstrip('0')
        return normalized or '0'

    def _filter_loss_ratio_records(self, df, policy_number, year):
        if df is None or policy_number is None or year is None:
            return None

        try:
            year_int = int(year)
        except (TypeError, ValueError):
            return None

        subset = df.copy(deep=True)

        def _parse_dt_series(series):
            if series is None:
                return pd.Series(pd.NaT, index=subset.index)
            parsed = pd.to_datetime(series, errors='coerce', dayfirst=False)
            needs_retry = parsed.isna() & series.notna()
            if needs_retry.any():
                retry = pd.to_datetime(series[needs_retry], errors='coerce', dayfirst=True)
                parsed.loc[needs_retry] = retry
            return parsed

        subset['policy_start_date'] = _parse_dt_series(subset.get('policy_start_date'))
        subset['policy_end_date'] = _parse_dt_series(subset.get('policy_end_date')) if 'policy_end_date' in subset.columns else pd.Series(pd.NaT, index=subset.index)
        subset['data_as_of'] = _parse_dt_series(subset.get('data_as_of')) if 'data_as_of' in subset.columns else pd.Series(pd.NaT, index=subset.index)
        subset['policy_number'] = subset.get('policy_number').astype(str)

        def _years_from_row(row):
            years = set()
            for col in ['policy_start_date', 'policy_end_date', 'data_as_of']:
                val = row.get(col)
                if isinstance(val, (pd.Timestamp, datetime)) and not pd.isna(val):
                    years.add(val.year)
            raw_year = row.get('year') if 'year' in row.index else None
            try:
                if raw_year is not None and str(raw_year).strip():
                    years.add(int(str(raw_year)))
            except Exception:
                pass
            return years

        subset['year_keys'] = subset.apply(_years_from_row, axis=1)

        target_policy = self._normalize_policy_number(policy_number)
        subset['policy_number_norm'] = subset['policy_number'].apply(self._normalize_policy_number)

        mask = (
            (subset['policy_number'] == str(policy_number)) |
            (subset['policy_number_norm'] == target_policy)
        ) & (subset['year_keys'].apply(lambda ys: year_int in ys if isinstance(ys, set) else False))

        subset = subset.loc[mask]

        if subset.empty:
            # Fallback: match only on policy and a plain 'year' column if present
            if 'year' in df.columns:
                year_col = df['year'].astype(str).str.extract(r"(\d{2,4})", expand=False)
                def _normalize_year_str(y):
                    if y is None or pd.isna(y):
                        return None
                    try:
                        val = int(y)
                        if val < 100:
                            val += 2000
                        return val
                    except Exception:
                        return None
                year_col = year_col.map(_normalize_year_str)
                df_work = df.copy(deep=True)
                df_work['year_col_norm'] = year_col
                df_work['policy_number'] = df_work.get('policy_number').astype(str)
                df_work['policy_number_norm'] = df_work['policy_number'].apply(self._normalize_policy_number)
                mask_fb = (
                    (df_work['policy_number'] == str(policy_number)) |
                    (df_work['policy_number_norm'] == target_policy)
                ) & (df_work['year_col_norm'] == year_int)
                subset = df_work.loc[mask_fb, ['benefit_type', 'actual_premium', 'actual_paid_w_ibnr', 'duration']].copy()

        if subset.empty:
            return None

        numeric_cols = ['actual_premium', 'actual_paid_w_ibnr', 'duration']
        for col in numeric_cols:
            subset[col] = pd.to_numeric(subset.get(col), errors='coerce')

        subset['benefit_type'] = subset.get('benefit_type').astype(str).str.strip()
        subset = subset[['benefit_type', 'actual_premium', 'actual_paid_w_ibnr', 'duration']]
        subset.dropna(subset=['benefit_type', 'actual_premium', 'actual_paid_w_ibnr', 'duration'], inplace=True)
        subset = subset[subset['benefit_type'] != ""]
        return subset if not subset.empty else None

    def _assign_loss_ratio_subset(self, subset_df, period, loss_ratio_grouping_optical):
        target_attr = 'previous_year_loss_ratio_df' if period == 'previous' else 'current_year_loss_ratio_df'
        if subset_df is None or subset_df.empty:
            setattr(self, target_attr, None)
            return

        df = subset_df.copy(deep=True)
        df = df.groupby('benefit_type', as_index=False).agg({
            'actual_premium': 'sum',
            'actual_paid_w_ibnr': 'sum',
            'duration': 'first'
        })
        df['duration'] = df['duration'].replace(0, pd.NA)
        df.dropna(subset=['duration'], inplace=True)

        if loss_ratio_grouping_optical and not df.empty:
            optical_mask = df['benefit_type'].str.lower() == 'optical'
            if optical_mask.any():
                clinical_mask = df['benefit_type'].str.lower() == 'clinical'
                optical_premium = df.loc[optical_mask, 'actual_premium'].sum()
                optical_paid = df.loc[optical_mask, 'actual_paid_w_ibnr'].sum()
                optical_duration = df.loc[optical_mask, 'duration'].iloc[0]

                if clinical_mask.any():
                    idx = df.index[clinical_mask][0]
                    df.at[idx, 'actual_premium'] += optical_premium
                    df.at[idx, 'actual_paid_w_ibnr'] += optical_paid
                    if pd.isna(df.at[idx, 'duration']):
                        df.at[idx, 'duration'] = optical_duration
                    df = df.loc[~optical_mask]
                else:
                    df.loc[optical_mask, 'benefit_type'] = 'Clinical'
                    df = df.groupby('benefit_type', as_index=False).agg({
                        'actual_premium': 'sum',
                        'actual_paid_w_ibnr': 'sum',
                        'duration': 'first'
                    })

        setattr(self, target_attr, df if not df.empty else None)

    def _update_loss_ratio_history(self, loss_ratio_df, policy_number):
        if loss_ratio_df is None or policy_number is None:
            return

        history_df = self._prepare_loss_ratio_history(loss_ratio_df, policy_number)
        if history_df is None or history_df.empty:
            return

        key_raw = str(policy_number)
        key_norm = self._normalize_policy_number(policy_number) or key_raw
        for key in {key_raw, key_norm}:
            if key:
                self.loss_ratio_history_by_policy[key] = history_df

    def _prepare_loss_ratio_history(self, df, policy_number):
        if df is None or policy_number is None:
            return None

        work = df.copy(deep=True)
        work['policy_number'] = work.get('policy_number').astype(str)
        target_norm = self._normalize_policy_number(policy_number)
        work['policy_number_norm'] = work['policy_number'].apply(self._normalize_policy_number)
        mask = (
            (work['policy_number'] == str(policy_number)) |
            (work['policy_number_norm'] == target_norm)
        )
        work = work.loc[mask]
        if work.empty:
            return None

        def _to_year(series):
            if series is None:
                return pd.Series([pd.NA] * len(work))
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                parsed = pd.to_datetime(series, errors='coerce', dayfirst=True)
                missing = parsed.isna()
                if missing.any():
                    parsed_alt = pd.to_datetime(series, errors='coerce', dayfirst=False)
                    parsed = parsed.fillna(parsed_alt)
            return parsed.dt.year

        year_series = _to_year(work.get('policy_start_date'))
        year_series = year_series.fillna(_to_year(work.get('policy_end_date')))
        year_series = year_series.fillna(_to_year(work.get('data_as_of')))
        if 'policy_id' in work.columns:
            policy_id_year = pd.to_numeric(
                work['policy_id'].astype(str).str.extract(r"_(\d{4})")[0],
                errors='coerce'
            )
            year_series = year_series.fillna(policy_id_year)

        work['year_key'] = year_series
        work.dropna(subset=['year_key'], inplace=True)
        if work.empty:
            return None

        work['year_key'] = work['year_key'].astype(int)
        work['benefit_type'] = work.get('benefit_type', '').astype(str).str.strip()
        work['benefit_type_norm'] = work['benefit_type'].str.lower()

        numeric_cols = ['actual_premium', 'actual_paid_w_ibnr', 'duration']
        for col in numeric_cols:
            work[col] = pd.to_numeric(work.get(col), errors='coerce')

        headcount_col = None
        for candidate in [
            'headcount', 'head_count', 'member_count', 'no_of_members',
            'no_of_employees', 'employee_count', 'no_of_lives', 'lives'
        ]:
            if candidate in work.columns:
                headcount_col = candidate
                work[headcount_col] = pd.to_numeric(work[headcount_col], errors='coerce')
                break

        total_mask = work['benefit_type_norm'].isin({'total', 'grand total'})
        work = work.loc[total_mask].copy() if total_mask.any() else work.copy()
        if work.empty:
            return None

        valid_duration = work['duration'].where(work['duration'] > 0)
        annual_factor = 12.0 / valid_duration
        work['premium_annualized'] = work['actual_premium'] * annual_factor
        work['claim_annualized'] = work['actual_paid_w_ibnr'] * annual_factor

        fallback_mask = annual_factor.isna()
        work.loc[fallback_mask, 'premium_annualized'] = work.loc[fallback_mask, 'actual_premium']
        work.loc[fallback_mask, 'claim_annualized'] = work.loc[fallback_mask, 'actual_paid_w_ibnr']

        agg_dict = {
            'premium_annualized': 'sum',
            'claim_annualized': 'sum'
        }
        if headcount_col:
            agg_dict[headcount_col] = 'sum'

        yearly = work.groupby('year_key', dropna=False).agg(agg_dict).reset_index()
        yearly.rename(columns={'year_key': 'year'}, inplace=True)

        if headcount_col:
            yearly.rename(columns={headcount_col: 'headcount'}, inplace=True)
        else:
            yearly['headcount'] = None

        def _safe_div(numerator, denominator):
            try:
                if numerator is None or denominator is None:
                    return None
                if pd.isna(numerator) or pd.isna(denominator):
                    return None
                denominator = float(denominator)
                if denominator == 0:
                    return None
                return float(numerator) / denominator
            except (TypeError, ValueError, ZeroDivisionError):
                return None

        yearly['loss_ratio'] = [
            _safe_div(paid, prem)
            for paid, prem in zip(yearly['claim_annualized'], yearly['premium_annualized'])
        ]
        yearly['premium_per_head'] = [
            _safe_div(prem, head)
            for prem, head in zip(yearly['premium_annualized'], yearly['headcount'])
        ]
        yearly['claim_per_head'] = [
            _safe_div(claim, head)
            for claim, head in zip(yearly['claim_annualized'], yearly['headcount'])
        ]

        yearly = yearly.dropna(subset=['premium_annualized', 'claim_annualized'], how='all')
        if yearly.empty:
            return None

        yearly.sort_values('year', ascending=False, inplace=True)
        yearly.reset_index(drop=True, inplace=True)
        return yearly

    def _get_loss_ratio_history_dataframe(self):
        for policy in [self.current_policy_num, self.previous_policy_num]:
            if policy is None:
                continue
            key_raw = str(policy)
            key_norm = self._normalize_policy_number(policy)
            for key in filter(None, {key_raw, key_norm}):
                if key in self.loss_ratio_history_by_policy:
                    df = self.loss_ratio_history_by_policy[key]
                    if df is not None and not df.empty:
                        return df
        for df in self.loss_ratio_history_by_policy.values():
            if df is not None and not df.empty:
                return df
        return None

    def _get_member_count_lookup(self):
        df = self._get_dataframe("mcr_p20_policy", "policy headcount lookup")
        if df is None or 'member_count' not in df.columns or 'policy_number' not in df.columns:
            return {}

        work = df.copy(deep=True)
        work['member_count'] = pd.to_numeric(work['member_count'], errors='coerce')
        year_col = next((col for col in ['year', 'policy_year', 'policy_year_start', 'policy_year_end'] if col in work.columns), None)
        if year_col is None:
            return {}

        work['year_numeric'] = pd.to_numeric(work[year_col], errors='coerce')
        work['policy_number'] = work['policy_number'].astype(str)
        work['policy_number_norm'] = work['policy_number'].apply(self._normalize_policy_number)

        mask = pd.Series(False, index=work.index)
        for policy in [self.current_policy_num, self.previous_policy_num]:
            if policy is None:
                continue
            policy_str = str(policy).strip()
            if policy_str:
                mask |= work['policy_number'] == policy_str
            normalized = self._normalize_policy_number(policy)
            if normalized:
                mask |= work['policy_number_norm'] == normalized

        if not mask.any():
            return {}

        work = work.loc[mask].copy()
        work.dropna(subset=['year_numeric', 'member_count'], inplace=True)
        if work.empty:
            return {}

        work['year_numeric'] = work['year_numeric'].astype(int)
        grouped = work.groupby('year_numeric')['member_count'].sum()
        return grouped.to_dict()

    def _populate_loss_ratio_history_table(self, worksheet):
        history_df = self._get_loss_ratio_history_dataframe()
        rows = []
        if history_df is not None and not history_df.empty:
            ordered = history_df.copy()
            if 'year' in ordered.columns:
                ordered = ordered.sort_values(
                    by='year',
                    ascending=True,
                    key=lambda col: pd.to_numeric(col, errors='coerce'),
                    na_position='last'
                )
            ordered = ordered.tail(3)
            rows = ordered.to_dict('records')

        member_counts = self._get_member_count_lookup()
        if member_counts and rows:
            def _safe_divide(numerator, denominator):
                try:
                    if numerator is None or denominator is None:
                        return None
                    if pd.isna(numerator) or pd.isna(denominator):
                        return None
                    denominator = float(denominator)
                    if denominator == 0:
                        return None
                    return float(numerator) / denominator
                except (TypeError, ValueError, ZeroDivisionError):
                    return None

            updated_rows = []
            for row in rows:
                year_numeric = pd.to_numeric(row.get('year'), errors='coerce')
                if pd.isna(year_numeric):
                    updated_rows.append(row)
                    continue
                member_value = member_counts.get(int(year_numeric))
                if member_value is None or pd.isna(member_value):
                    updated_rows.append(row)
                    continue
                new_row = dict(row)
                new_row['headcount'] = member_value
                new_row['premium_per_head'] = _safe_divide(new_row.get('premium_annualized'), member_value)
                new_row['claim_per_head'] = _safe_divide(new_row.get('claim_annualized'), member_value)
                updated_rows.append(new_row)
            rows = updated_rows

        start_row = 5
        col_map = [
            ('year', 10),
            ('premium_annualized', 11),
            ('claim_annualized', 12),
            ('loss_ratio', 13),
            ('headcount', 14),
            ('premium_per_head', 15),
            ('claim_per_head', 16)
        ]

        for offset in range(3):
            row_idx = start_row + offset
            row_data = rows[offset] if offset < len(rows) else {}
            for key, col_idx in col_map:
                value = row_data.get(key)
                if value is not None and pd.isna(value):
                    value = None
                worksheet.cell(row=row_idx, column=col_idx).value = value

    def _append_loss_ratio_records(self, loss_ratio_df, policy_number):
        if loss_ratio_df is None or policy_number is None:
            return

        work = loss_ratio_df.copy(deep=True)
        work['policy_number'] = work.get('policy_number').astype(str)
        target_norm = self._normalize_policy_number(policy_number)
        work['policy_number_norm'] = work['policy_number'].apply(self._normalize_policy_number)
        mask = (
            (work['policy_number'] == str(policy_number)) |
            (work['policy_number_norm'] == target_norm)
        )
        subset = work.loc[mask].drop(columns=['policy_number_norm'], errors='ignore')
        if subset.empty:
            return

        key_raw = str(policy_number)
        key_norm = target_norm or key_raw
        for key in filter(None, {key_raw, key_norm}):
            self.loss_ratio_records_by_policy[key] = subset.copy(deep=True)

    def _get_loss_ratio_records_dataframe(self):
        frames = []
        seen = set()
        for policy in [self.current_policy_num, self.previous_policy_num]:
            if policy is None:
                continue
            key_raw = str(policy)
            key_norm = self._normalize_policy_number(policy)
            for key in filter(None, {key_raw, key_norm}):
                if key in seen:
                    continue
                df = self.loss_ratio_records_by_policy.get(key)
                if df is not None and not df.empty:
                    frames.append(df.copy(deep=True))
                    seen.add(key)
        if not frames:
            return None
        combined = pd.concat(frames, ignore_index=True)
        combined.drop_duplicates(inplace=True)
        return combined if not combined.empty else None

    def _normalize_benefit_label(self, value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        text = str(value).strip()
        if not text:
            return None
        cleaned = ''.join(ch for ch in text.lower() if ch.isalnum())
        alias_map = {
            'hospital': 'Hospital',
            'clinical': 'Clinical',
            'dental': 'Dental',
            'optical': 'Optical',
            'maternity': 'Maternity',
            'topupsmm': 'Top-Up/ SMM',
            'topupsmmrl': 'Top-Up/ SMM',
            'topupsmml': 'Top-Up/ SMM',
            'topupsmmw': 'Top-Up/ SMM',
            'topupsmm': 'Top-Up/ SMM',
            'topup': 'Top-Up/ SMM',
            'total': 'Total',
            'grandtotal': 'Total',
            'grandtotals': 'Total'
        }
        return alias_map.get(cleaned, text)

    def _parse_dates_with_fallback(self, series):
        if series is None:
            return pd.Series([pd.NaT] * 0)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            parsed = pd.to_datetime(series, errors='coerce', dayfirst=True)
            missing = parsed.isna()
            if missing.any():
                parsed_alt = pd.to_datetime(series, errors='coerce', dayfirst=False)
                parsed = parsed.fillna(parsed_alt)
        return parsed

    def _populate_loss_ratio_detail_table(self, worksheet):
        detail_df = self._get_loss_ratio_records_dataframe()
        start_row = 14
        clear_rows = 200
        year_column_index = 10 + 12  # Column reserved for Year formula
        for row in range(start_row, start_row + clear_rows):
            for col in range(10, 26):
                if col == year_column_index:
                    continue
                cell = worksheet.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        if detail_df is None or detail_df.empty:
            return

        df = detail_df.copy(deep=True)
        required_cols = [
            'policy_id', 'policy_number', 'client_name', 'policy_start_date', 'policy_end_date',
            'duration', 'ibnr', 'data_as_of', 'benefit_type', 'actual_premium', 'actual_paid_w_ibnr', 'loss_ratio'
        ]
        for col in required_cols:
            if col not in df.columns:
                df[col] = None

        df['benefit_display'] = df['benefit_type'].apply(self._normalize_benefit_label)
        df.dropna(subset=['benefit_display'], inplace=True)
        if df.empty:
            return

        benefit_priority = ['Hospital', 'Clinical', 'Dental', 'Optical', 'Maternity', 'Top-Up/ SMM', 'Top-Up/SMM', 'Total']
        order_map = {name: idx for idx, name in enumerate(benefit_priority)}
        df['benefit_order'] = df['benefit_display'].map(order_map).fillna(len(order_map))

        df['duration_numeric'] = pd.to_numeric(df['duration'], errors='coerce')
        df['actual_premium_numeric'] = pd.to_numeric(df['actual_premium'], errors='coerce')
        df['actual_paid_numeric'] = pd.to_numeric(df['actual_paid_w_ibnr'], errors='coerce')

        start_series = df.get('policy_start_date')
        if start_series is None:
            start_series = pd.Series([pd.NaT] * len(df))
        start_dates = self._parse_dates_with_fallback(start_series)
        df['policy_start_sort'] = start_dates

        data_as_of_series = df.get('data_as_of')
        if data_as_of_series is None:
            data_as_of_series = pd.Series([pd.NaT] * len(df))
        df['data_as_of_sort'] = self._parse_dates_with_fallback(data_as_of_series)

        duration_factor = 12.0 / df['duration_numeric']
        duration_factor.replace([pd.NA, float('inf'), -float('inf')], pd.NA, inplace=True)
        df['ann_premium'] = df['actual_premium_numeric'] * duration_factor
        df['ann_paid'] = df['actual_paid_numeric'] * duration_factor

        fallback_mask = df['ann_premium'].isna()
        df.loc[fallback_mask, 'ann_premium'] = df.loc[fallback_mask, 'actual_premium_numeric']
        fallback_mask = df['ann_paid'].isna()
        df.loc[fallback_mask, 'ann_paid'] = df.loc[fallback_mask, 'actual_paid_numeric']

        loss_ratio_series = df['loss_ratio'].astype(str).str.strip()
        cleaned = loss_ratio_series.str.replace('%', '', regex=False).str.replace(',', '', regex=False)
        numeric_loss = pd.to_numeric(cleaned, errors='coerce')
        percent_mask = loss_ratio_series.str.contains('%', regex=False)
        numeric_loss.loc[percent_mask & numeric_loss.notna()] = numeric_loss.loc[percent_mask & numeric_loss.notna()] / 100.0
        fallback_lr = df.apply(
            lambda row: (row['actual_paid_numeric'] / row['actual_premium_numeric'])
            if row['actual_premium_numeric'] and not pd.isna(row['actual_premium_numeric']) else pd.NA,
            axis=1
        )
        df['loss_ratio_numeric'] = numeric_loss.fillna(fallback_lr)

        ann_lr = df.apply(
            lambda row: (row['ann_paid'] / row['ann_premium'])
            if row['ann_premium'] and not pd.isna(row['ann_premium']) else pd.NA,
            axis=1
        )
        df['ann_loss_ratio'] = ann_lr.fillna(df['loss_ratio_numeric'])

        df.sort_values(
            by=['benefit_order', 'policy_start_sort', 'data_as_of_sort'],
            ascending=[True, True, True],
            inplace=True
        )

        column_order = [
            ('policy_id', 'policy_id'),
            ('policy_number', 'policy_number'),
            ('client_name', 'client_name'),
            ('policy_start_date', 'policy_start_date'),
            ('policy_end_date', 'policy_end_date'),
            ('duration', 'duration'),
            ('ibnr', 'ibnr'),
            ('data_as_of', 'data_as_of'),
            ('benefit_type', 'benefit_display'),
            ('actual_premium', 'actual_premium_numeric'),
            ('actual_paid_w_ibnr', 'actual_paid_numeric'),
            ('loss_ratio', 'loss_ratio_numeric'),
            ('Year', None),
            ('Annl. Premium', 'ann_premium'),
            ('Annl. Paid /w IBNR', 'ann_paid'),
            ('Loss Ratio', 'ann_loss_ratio')
        ]

        row_idx = start_row
        end_col = 25
        def _next_available_row(current_row):
            row = current_row
            limit = start_row + clear_rows
            while row < limit:
                has_merged = False
                for col in range(10, end_col + 1):
                    if isinstance(worksheet.cell(row=row, column=col), MergedCell):
                        has_merged = True
                        break
                if not has_merged:
                    return row
                row += 1
            return row

        for _, row in df.iterrows():
            row_idx = _next_available_row(row_idx)
            for offset, (label, key) in enumerate(column_order):
                if label == 'Year':
                    continue
                value = row.get(key)
                if isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()
                if pd.isna(value):
                    value = None
                worksheet.cell(row=row_idx, column=10 + offset).value = value
            row_idx += 1


    def loss_ratio_text_convert(self, previous_yr_loss_ratio_text=None, current_yr_loss_ratio_text = None, loss_ratio_grouping_optical=True):
        print("previous_yr_loss_ratio_text:", previous_yr_loss_ratio_text)
        print("current_yr_loss_ratio_text:", current_yr_loss_ratio_text)
        if previous_yr_loss_ratio_text:
            previous_yr_loss_ratio_text = previous_yr_loss_ratio_text.splitlines()
            data_l = [previous_yr_loss_ratio_text[i].split(",") for i in range(len(previous_yr_loss_ratio_text))]
            df_prev = pd.DataFrame(data_l[1:], columns=data_l[0])
            self._assign_loss_ratio_subset(df_prev, 'previous', loss_ratio_grouping_optical)
        if current_yr_loss_ratio_text:
            current_yr_loss_ratio_text = current_yr_loss_ratio_text.splitlines()
            data_l = [current_yr_loss_ratio_text[i].split(",") for i in range(len(current_yr_loss_ratio_text))]
            df_curr = pd.DataFrame(data_l[1:], columns=data_l[0])
            self._assign_loss_ratio_subset(df_curr, 'current', loss_ratio_grouping_optical)

    

    def convert_all(self):
        self.claim_info()
        self.p16_LR_by_benefits()
        gated_steps = [
            (self.P17_ClaimPattern, "mcr_p22_class_benefit"),
            (self.P20_overall, "mcr_p20_policy"),
            (self.P20_benefittype, "mcr_p20_benefit"),
            (self.P20_network, "mcr_p20_network"),
            (self.P21_by_class, "mcr_p21_class"),
            (self.P22_by_class, "mcr_p22_class_benefit"),
            (self.P23_Usage_HospByBnf, "mcr_p23_usage_by_hosp_benefit"),
            (self.P24_Usage_ClinicByBnf, "mcr_p24_usage_by_clinic_benefit"),
            (self.P25_by_plan, "mcr_p25_class_panel_benefit"),
            (self.P26_by_class, "mcr_p26_op_panel_benefit"),
            (self.P18_TopHospDiag, "mcr_p18_top_hosp_diag"),
            (self.P19_TopClinicDiag, "mcr_p19_top_clinic_diag"),
        ]

        for func, attr in gated_steps:
            if getattr(self, attr, None) is not None:
                func()
            else:
                sheet = self.sheet_sources.get(attr, attr) if hasattr(self, 'sheet_sources') else attr
                print(f"convert_all: Skipping {func.__name__} because worksheet '{sheet}' is unavailable.")